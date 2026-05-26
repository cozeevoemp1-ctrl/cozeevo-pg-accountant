"""
Opex benchmarking report — Cozeevo vs Babai PG (competitor reference).
Output: data/reports/Opex_Comparison_2026_05.xlsx

Imports OPEX data from pnl_builder — never duplicates figures.
Uses April 2026 (most recent complete month).

Run:
  venv/Scripts/python scripts/export_opex_comparison.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.reports.pnl_builder import OPEX, MONTHS
from src.utils.inr_format import INR_NUMBER_FORMAT

OUT = Path(__file__).parent.parent / "data" / "reports" / "Opex_Comparison_2026_05.xlsx"

# ── Config ────────────────────────────────────────────────────────────────────
APR_IDX         = MONTHS.index("Apr'26")
OUR_TOTAL_BEDS  = 297
OUR_OCC_BEDS    = 270   # ~91% Apr'26
AVG_RENT_BED    = 14500 # Apr'26 true revenue / occupied beds
RENT_PER_ROOM   = 13500
ROOMS           = 164   # rooms in lease (164 × 13,500 = 22,14,000)

BABAI_TOTAL_BEDS = 340  # 200 rooms × ~1.7 beds
BABAI_OCC_BEDS   = 306  # ~90% occ

# Babai reference figures (from shared photo, per month)
BABAI = {
    "Food & Groceries":     375000,  # milk+curd+paneer+chicken+eggs+groceries+vegetables
    "Staff & Labour":       225000,
    "Electricity":          200000,
    "Water":                225000,
    "Gas / Fuel":            35000,
    "Maintenance":           50000,
    "Housekeeping":          50000,
    "Waste Disposal":         5000,
    "Internet":              33000,
    "Police":                 3000,
    "Miscellaneous":         50000,
}
BABAI_TOTAL = sum(BABAI.values())

# ── Our Apr'26 figures mapped to comparable categories ───────────────────────
def _apr(key):
    return OPEX[key][APR_IDX]

OUR_MAPPED = {
    "Food & Groceries":  _apr("Food & Groceries"),
    "Staff & Labour":    _apr("Staff & Labour") + _apr("Partner Reimbursable (Personal Acct SBI 0167)"),
    "Electricity":       _apr("Electricity"),
    "Water":             _apr("Water (bank tankers + Manoj cash; Mar bill paid Apr)"),
    "Gas / Fuel":        _apr("Fuel & Diesel"),
    "Maintenance":       _apr("Maintenance & Repairs"),
    "Housekeeping":      _apr("Cleaning Supplies"),
    "Waste Disposal":    _apr("Waste Disposal (Pavan Rs.3.5K/mo)"),
    "Internet":          _apr("IT & Software"),   # prepaid exhausted in Apr; IT proxy
    "Police":            _apr("Govt & Regulatory (incl Police Rs.3K accrual Jan+)"),
    "Miscellaneous":     _apr("Shopping & Supplies") + _apr("HULK — Operational Expenses")
                         + _apr("Marketing") + _apr("Bank Charges") + _apr("Other Expenses"),
}

# One-time CAPEX items folded into Furniture & Supplies this month
OUR_ONETIME = {
    "Chairs":            47000,
    "Atta machine":      21060,
    "Kitchen vessels":   37500,
    "Mirrors / decor":   10258,  # Shopping & Supplies Apr = volipi.l mirrors
}
OUR_ONETIME_TOTAL = sum(OUR_ONETIME.values())

# F&F ongoing (from Furniture & Supplies minus one-time items)
_furniture_apr = _apr("Furniture & Supplies")
OUR_FF_ONGOING = _furniture_apr - OUR_ONETIME_TOTAL  # 43,119
OUR_MAPPED["F&F ongoing + Amazon"] = OUR_FF_ONGOING

OUR_RECURRING_TOTAL = sum(OUR_MAPPED.values())
OUR_RENT = ROOMS * RENT_PER_ROOM  # 22,14,000

# ── Fixed vs variable split ───────────────────────────────────────────────────
# (fraction that is FIXED regardless of occupancy)
FIXED_FRAC = {
    "Food & Groceries":      0.00,
    "Staff & Labour":        1.00,
    "Electricity":           0.80,
    "Water":                 0.75,
    "Gas / Fuel":            0.70,
    "Maintenance":           0.80,
    "Housekeeping":          0.60,
    "Waste Disposal":        1.00,
    "Internet":              1.00,
    "Police":                1.00,
    "Miscellaneous":         1.00,
    "F&F ongoing + Amazon":  0.80,
}

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT  = Font(bold=True, color="FFFFFF")
BOLD      = Font(bold=True)
TOTAL_FILL= PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
RED_FONT  = Font(bold=True, color="9C0006")
GRN_FONT  = Font(bold=True, color="375623")
GRN_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YLW_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CTR       = Alignment(horizontal="center")

def _hdr(ws, row_vals):
    ws.append(row_vals)
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR

def _fmt(ws, col_start=2, col_end=None):
    """Apply INR format to numeric cells in last row."""
    row = ws[ws.max_row]
    end = col_end or len(row)
    for c in row[col_start - 1:end]:
        if isinstance(c.value, (int, float)):
            c.number_format = INR_NUMBER_FORMAT

def _total_row(ws, label, vals, fill=TOTAL_FILL, font=None):
    ws.append([label] + vals)
    for c in ws[ws.max_row]:
        c.fill = fill
        if font:
            c.font = font
        elif isinstance(c.value, (int, float)):
            c.font = BOLD
    _fmt(ws)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — Side-by-side comparison
# ─────────────────────────────────────────────────────────────────────────────
def _tab_comparison(wb):
    ws = wb.active
    ws.title = "Comparison"

    _hdr(ws, [
        "Category",
        f"Babai (total)\n{BABAI_TOTAL_BEDS} beds",
        f"Babai /bed\n({BABAI_TOTAL_BEDS} total)",
        f"Babai /occ bed\n({BABAI_OCC_BEDS} occ)",
        f"Us Apr'26 (total)\n{OUR_TOTAL_BEDS} beds",
        f"Us /bed\n({OUR_TOTAL_BEDS} total)",
        f"Us /occ bed\n({OUR_OCC_BEDS} occ)",
        "Diff /bed\n(Us − Babai)",
    ])

    categories = list(OUR_MAPPED.keys())

    for cat in categories:
        b_total = BABAI.get(cat, 0)
        o_total = OUR_MAPPED[cat]
        b_per   = round(b_total / BABAI_TOTAL_BEDS) if b_total else None
        b_occ   = round(b_total / BABAI_OCC_BEDS)   if b_total else None
        o_per   = round(o_total / OUR_TOTAL_BEDS)
        o_occ   = round(o_total / OUR_OCC_BEDS)
        diff    = o_per - b_per if b_per else None

        ws.append([cat, b_total or None, b_per, b_occ, o_total, o_per, o_occ, diff])
        _fmt(ws)

        # Colour diff column
        diff_cell = ws[ws.max_row][7]
        if diff is not None:
            if diff > 0:
                diff_cell.fill = YLW_FILL; diff_cell.font = RED_FONT
            else:
                diff_cell.fill = GRN_FILL; diff_cell.font = GRN_FONT

    # Totals
    ws.append([])
    _total_row(ws, "Recurring ops (excl. rent)", [
        BABAI_TOTAL,
        round(BABAI_TOTAL / BABAI_TOTAL_BEDS),
        round(BABAI_TOTAL / BABAI_OCC_BEDS),
        OUR_RECURRING_TOTAL,
        round(OUR_RECURRING_TOTAL / OUR_TOTAL_BEDS),
        round(OUR_RECURRING_TOTAL / OUR_OCC_BEDS),
        round(OUR_RECURRING_TOTAL / OUR_TOTAL_BEDS) - round(BABAI_TOTAL / BABAI_TOTAL_BEDS),
    ])

    # One-time CAPEX section
    ws.append([])
    ws.append(["ONE-TIME CAPEX (Apr'26 only — will normalise to ₹0)"])
    ws[ws.max_row][0].font = Font(bold=True, italic=True, color="595959")
    for item, amt in OUR_ONETIME.items():
        ws.append([f"  {item}", None, None, None, amt, round(amt / OUR_TOTAL_BEDS), round(amt / OUR_OCC_BEDS)])
        _fmt(ws)
    _total_row(ws, "One-time total", [
        None, None, None,
        OUR_ONETIME_TOTAL,
        round(OUR_ONETIME_TOTAL / OUR_TOTAL_BEDS),
        round(OUR_ONETIME_TOTAL / OUR_OCC_BEDS),
    ], fill=YLW_FILL)

    # Notes
    ws.append([])
    notes = [
        "Notes:",
        f"Babai reference: shared photo. 200 rooms, ~{BABAI_TOTAL_BEDS} beds assumed, ~{BABAI_OCC_BEDS} occupied (90%).",
        f"Our data: Apr'26 actuals. {OUR_TOTAL_BEDS} revenue beds, {OUR_OCC_BEDS} occupied (~91%).",
        "Babai owns their building — no rent cost. Our rent (₹22.14L/mo) excluded from this table.",
        "Internet Apr = ₹0 (prepaid exhausted). Annual cost ~₹2.3L / 12 = ₹19K/mo.",
        "F&F ongoing = Furniture & Supplies Apr minus identified one-time items (chairs/atta machine/kitchen vessels).",
    ]
    for n in notes:
        ws.append([n])
        ws[ws.max_row][0].font = Font(italic=True, color="595959")

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — Fixed vs Variable
# ─────────────────────────────────────────────────────────────────────────────
def _tab_fixed_variable(wb):
    ws = wb.create_sheet("Fixed vs Variable")

    _hdr(ws, ["Category", "Apr'26 total", "Fixed %", "Fixed ₹", "Variable ₹", "Variable /occ bed"])

    total_fixed = 0
    total_var   = 0

    for cat, amt in OUR_MAPPED.items():
        frac  = FIXED_FRAC.get(cat, 0.5)
        fixed = round(amt * frac)
        var   = amt - fixed
        v_bed = round(var / OUR_OCC_BEDS)
        total_fixed += fixed
        total_var   += var
        ws.append([cat, amt, f"{int(frac*100)}%", fixed, var, v_bed])
        _fmt(ws, col_start=2)

    ws.append([])
    _total_row(ws, "Ops sub-total (excl. rent)", [
        OUR_RECURRING_TOTAL, "",
        total_fixed, total_var,
        round(total_var / OUR_OCC_BEDS),
    ])

    # Add rent as fully fixed
    ws.append(["Property Rent", OUR_RENT, "100%", OUR_RENT, 0, 0])
    _fmt(ws, col_start=2)

    grand_fixed = total_fixed + OUR_RENT
    grand_var   = total_var
    ws.append([])
    _total_row(ws, "TOTAL (incl. rent)", [
        OUR_RECURRING_TOTAL + OUR_RENT, "",
        grand_fixed, grand_var,
        round(grand_var / OUR_OCC_BEDS),
    ], font=BOLD)

    # Contribution per bed
    contribution = AVG_RENT_BED - round(grand_var / OUR_OCC_BEDS)
    breakeven    = round(grand_fixed / contribution)
    ws.append([])
    ws.append(["Avg revenue / occ bed",    None, None, None, None, AVG_RENT_BED])
    ws.append(["Variable cost / occ bed",  None, None, None, None, round(grand_var / OUR_OCC_BEDS)])
    ws.append(["Contribution / occ bed",   None, None, None, None, contribution])
    for r in ws[ws.max_row - 2: ws.max_row + 1]:
        for c in r:
            if isinstance(c.value, int):
                c.number_format = INR_NUMBER_FORMAT
    ws.append(["Break-even beds", f"= ₹{grand_fixed:,} fixed ÷ ₹{contribution:,} contribution = {breakeven} beds ({round(breakeven/OUR_TOTAL_BEDS*100)}% occ)"])
    ws[ws.max_row][0].font = BOLD

    ws.column_dimensions["A"].width = 32
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "B2"


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — Profit scenarios
# ─────────────────────────────────────────────────────────────────────────────
def _tab_scenarios(wb):
    ws = wb.create_sheet("Profit Scenarios")

    # Derive fixed / variable from tab 2 calcs
    total_fixed_ops = sum(round(v * FIXED_FRAC.get(k, 0.5)) for k, v in OUR_MAPPED.items())
    total_var_ops   = OUR_RECURRING_TOTAL - total_fixed_ops
    var_per_bed     = total_var_ops / OUR_OCC_BEDS  # at 270 beds

    grand_fixed = total_fixed_ops + OUR_RENT

    _hdr(ws, ["Occupancy %", "Occupied beds", "Revenue", "Variable ops", "Fixed costs", "Total costs", "Profit / mo", "Profit / yr", "Margin %"])

    scenarios = [0.65, 0.70, 0.75, 0.80, 0.85, 0.91, 0.95, 1.00]
    for occ in scenarios:
        beds    = round(OUR_TOTAL_BEDS * occ)
        rev     = beds * AVG_RENT_BED
        var     = round(beds * var_per_bed)
        total_c = grand_fixed + var
        profit  = rev - total_c
        margin  = profit / rev if rev else 0

        row = [
            f"{int(occ*100)}%",
            beds,
            rev,
            var,
            grand_fixed,
            total_c,
            profit,
            profit * 12,
            f"{margin*100:.1f}%",
        ]
        ws.append(row)
        _fmt(ws, col_start=3, col_end=8)

        # Highlight current and break-even
        row_ref = ws[ws.max_row]
        if occ == 0.91:
            for c in row_ref:
                c.fill = GRN_FILL
        if profit < 0:
            for c in row_ref:
                c.fill = YLW_FILL
                c.font = RED_FONT

    # Key callouts
    ws.append([])
    ws.append([f"Fixed cost base: ₹{grand_fixed:,.0f}/mo  |  Variable: ₹{var_per_bed:,.0f}/occ bed  |  Contribution/bed: ₹{AVG_RENT_BED - var_per_bed:,.0f}"])
    ws[ws.max_row][0].font = BOLD
    ws.append([f"Break-even: ~{round(grand_fixed / (AVG_RENT_BED - var_per_bed))} beds ({round(grand_fixed / (AVG_RENT_BED - var_per_bed) / OUR_TOTAL_BEDS * 100)}% occ)  |  Rent alone = {round(OUR_RENT/OUR_TOTAL_BEDS)} /total bed, {round(OUR_RENT/(OUR_TOTAL_BEDS*0.91))} /occ bed"])
    ws.append(["Cost base barely moves between 80% and 100% occ — every extra bed above break-even contributes ~₹13K/mo directly to profit."])
    ws[ws.max_row][0].font = Font(italic=True, color="595959")

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# Build & save
# ─────────────────────────────────────────────────────────────────────────────
def build_comparison_workbook():
    wb = openpyxl.Workbook()
    _tab_comparison(wb)
    _tab_fixed_variable(wb)
    _tab_scenarios(wb)
    return wb


if __name__ == "__main__":
    OUT.parent.mkdir(parents=True, exist_ok=True)
    build_comparison_workbook().save(OUT)
    print(f"Saved: {OUT}")
