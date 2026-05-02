"""
Income reconciliation — Nov'25 through Apr'26
Three sources stitched together:
  1. DB  — verified rent cash + UPI by month (the authoritative income record)
  2. Bank statements (Yes Bank …961) — all credits classified
  3. DB deposits / bookings — working capital, NOT income (must be refunded)

Output: data/reports/income_reconciliation.xlsx
  Tab 1  Income P&L          — rent income vs bank UPI collections per month
  Tab 2  Bank Credits Detail — every credit row classified
  Tab 3  Working Capital     — deposits + bookings held (liability, not profit)
"""
import asyncio, os, csv, datetime
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from collections import defaultdict
from dotenv import load_dotenv
load_dotenv()

import asyncpg, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.inr_format import INR_NUMBER_FORMAT

# ── helpers ───────────────────────────────────────────────────────────────────
def parse_amt(v):
    if v is None or str(v).strip() == '': return 0.0
    try: return float(str(v).replace(',', '').replace('INR', '').strip())
    except: return 0.0

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try: return datetime.datetime.strptime(s[:10], fmt).date()
        except: pass
    return None

def month_key(d):
    return f'{d.year}-{d.month:02d}' if d else None

# ── Credit classification ─────────────────────────────────────────────────────
# Returns (category, note)
def classify_credit(desc: str, ref: str, amt: float) -> tuple[str, str]:
    d = (desc or '').lower()
    r = (ref or '').lower()

    # Yes Bank UPI collection settlement = all tenant UPI payments batched daily
    if 'upi collection settlement' in d or 'cbs payable acc' in d:
        return 'Tenant UPI Collection', 'Yes Bank daily UPI batch settlement'

    # Capital infusion / personal transfers from admin phones
    if any(p in d for p in ['7358341775', '7845952289', 'kiran kumar', 'kiran pemm']):
        return 'Personal / Capital Transfer', 'Admin phone'

    # RTGS / NEFT — usually loans or capital from investors
    if 'rtgs' in d or 'rtgs' in r or 'neft' in d or 'neft' in r:
        return 'Capital Infusion (RTGS/NEFT)', desc[:60]

    # Deposit refund received (rare — you usually pay out, not receive)
    if 'refund' in d and amt > 0:
        return 'Refund Received', desc[:60]

    # UPI from unknown — could be a tenant paying via personal QR not covered by collection
    if 'upi' in d or 'upi' in r:
        return 'UPI - Individual (unidentified)', desc[:60]

    return 'Other Credit', desc[:40]

# ── Statement reader — handles xlsx and csv, skips 5-row metadata header ──────
def read_statement(path: str) -> list[dict]:
    rows = []
    if path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        # Find header row (contains "Transaction Date")
        hrow = next(
            i for i, r in enumerate(all_rows)
            if r and any(str(c or '').strip().lower() == 'transaction date' for c in r)
        )
        headers = [str(c or '').strip() for c in all_rows[hrow]]
        for r in all_rows[hrow + 1:]:
            if not any(r): continue
            rows.append(dict(zip(headers, r)))
    else:
        with open(path, encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        hrow = next(
            i for i, r in enumerate(all_rows)
            if r and r[0].strip().lower() == 'transaction date'
        )
        headers = [c.strip() for c in all_rows[hrow]]
        for r in all_rows[hrow + 1:]:
            if not any(c.strip() for c in r): continue
            rows.append(dict(zip(headers, r)))
    return rows

# ── Read all statements ────────────────────────────────────────────────────────
STATEMENTS = [
    '2025 statement.xlsx',   # Oct-Dec 2025
    '2026 statment.csv',      # Jan-Mar 2026
    'april month.csv',        # Apr 2026
]
MONTHS_ORDER = ['2025-10','2025-11','2025-12','2026-01','2026-02','2026-03','2026-04']
MONTH_LABELS = {'2025-10':"Oct'25",'2025-11':"Nov'25",'2025-12':"Dec'25",
                '2026-01':"Jan'26",'2026-02':"Feb'26",'2026-03':"Mar'26",'2026-04':"Apr'26"}

credit_rows = []   # (date, description, ref, amount, category, note, month)
bank_by_month = defaultdict(lambda: defaultdict(float))   # month → category → total

for path in STATEMENTS:
    for row in read_statement(path):
        dt = parse_date(row.get('Transaction Date'))
        dep = parse_amt(row.get('Deposits', ''))
        if not dt or dep == 0:
            continue
        desc = str(row.get('Description', '') or '')
        ref  = str(row.get('Reference Number', '') or row.get('Reference', '') or '')
        cat, note = classify_credit(desc, ref, dep)
        mk = month_key(dt)
        credit_rows.append((dt, desc[:80], ref[:30], dep, cat, note, mk))
        bank_by_month[mk][cat] += dep

# ── DB data ────────────────────────────────────────────────────────────────────
async def fetch_db():
    conn = await asyncpg.connect(os.getenv('DATABASE_URL').replace('postgresql+asyncpg','postgresql'))

    # Rent by month + mode
    rent_rows = await conn.fetch(
        "SELECT TO_CHAR(period_month,'YYYY-MM') as m, payment_mode, SUM(amount) as total "
        "FROM payments WHERE for_type='rent' AND is_void=false "
        "GROUP BY m, payment_mode ORDER BY m, payment_mode")

    # All deposit and booking totals (working capital)
    wc_rows = await conn.fetch(
        "SELECT for_type, payment_mode, SUM(amount) as total, COUNT(*) as n "
        "FROM payments WHERE for_type IN ('deposit','booking') AND is_void=false "
        "GROUP BY for_type, payment_mode")

    # Deposits by when-collected (period_month may be NULL for deposits)
    dep_detail = await conn.fetch(
        "SELECT t.name, te.checkin_date, p.amount, p.payment_mode, p.for_type "
        "FROM payments p "
        "JOIN tenancies te ON p.tenancy_id=te.id "
        "JOIN tenants t ON te.tenant_id=t.id "
        "WHERE p.for_type IN ('deposit','booking') AND p.is_void=false "
        "ORDER BY te.checkin_date NULLS LAST, t.name")

    await conn.close()
    return rent_rows, wc_rows, dep_detail

rent_rows, wc_rows, dep_detail = asyncio.run(fetch_db())

db_rent = defaultdict(lambda: {'cash': 0, 'upi': 0})
for r in rent_rows:
    db_rent[r['m']][r['payment_mode']] = int(r['total'])

wc_summary = {}
for r in wc_rows:
    k = (r['for_type'], r['payment_mode'])
    wc_summary[k] = {'total': int(r['total']), 'n': r['n']}

# Deposit refunds paid (from expense classification table already in memory)
DEPOSIT_REFUNDS_PAID = 146024   # from expense_classified_full.xlsx "Tenant Deposit Refund"

# ── Excel workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Styles ────────────────────────────────────────────────────────────────────
def hdr(ws, row, col, val, bg='1F4E79', fg='FFFFFF', bold=True, sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=bold, size=sz)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return c

def num(ws, row, col, val, fmt=None, bg=None):
    if fmt is None: fmt = INR_NUMBER_FORMAT
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = fmt
    c.alignment = Alignment(horizontal='right')
    if bg: c.fill = PatternFill('solid', fgColor=bg)
    return c

def lbl(ws, row, col, val, bold=False, bg=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=10)
    if bg: c.fill = PatternFill('solid', fgColor=bg)
    return c

LIGHT_BLUE  = 'DEEAF1'
LIGHT_GREEN = 'E2EFDA'
LIGHT_RED   = 'FCE4D6'
LIGHT_GRAY  = 'F2F2F2'
YELLOW      = 'FFF2CC'

thin = Side(style='thin', color='CCCCCC')
border = Border(bottom=thin)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — Income P&L
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet('Income P&L')
ws1.freeze_panes = 'B3'
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells('A1:J1')
t = ws1['A1']
t.value = 'COZEEVO — INCOME RECONCILIATION   (Nov\'25 – Apr\'26)'
t.font = Font(bold=True, size=13, color='1F4E79')
t.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 24

months = [m for m in MONTHS_ORDER if m >= '2025-11']

# Column headers row 2
COLS = ['Category'] + [MONTH_LABELS[m] for m in months] + ['TOTAL']
for ci, h in enumerate(COLS, 1):
    hdr(ws1, 2, ci, h)
ws1.row_dimensions[2].height = 20

SECT = [
    # (label, db_key, bg, bold)
    ('── RENT INCOME (from DB — authoritative) ──', None, '1F4E79', 'FFFFFF'),
    ('Rent — Cash',    'cash',  LIGHT_GREEN, None),
    ('Rent — UPI',     'upi',   LIGHT_GREEN, None),
    ('Total Rent Income', 'both', 'C6EFCE', None),
    ('', None, None, None),
    ('── BANK STATEMENT CREDITS (Yes Bank …961) ──', None, '1F4E79', 'FFFFFF'),
    ('Tenant UPI Collection', 'bank_tenant', LIGHT_BLUE, None),
    ('Personal / Capital Transfer', 'bank_personal', LIGHT_GRAY, None),
    ('Capital Infusion (RTGS/NEFT)', 'bank_rtgs', LIGHT_GRAY, None),
    ('UPI - Individual (unidentified)', 'bank_indiv', YELLOW, None),
    ('Other Credit', 'bank_other', YELLOW, None),
    ('Total Bank Credits', 'bank_total', LIGHT_BLUE, None),
    ('', None, None, None),
    ('── RECONCILIATION ──', None, '1F4E79', 'FFFFFF'),
    ('Tenant UPI Collection (bank)', 'bank_tenant', LIGHT_BLUE, None),
    ('Less: DB Rent UPI', 'upi_neg', None, None),
    ('Gap (deposits/bookings via UPI + unrecorded)', 'gap', YELLOW, None),
]

row = 3
for label, key, bg, fg in SECT:
    if label == '':
        row += 1
        continue
    is_section = key is None
    ws1.row_dimensions[row].height = 18

    c = ws1.cell(row=row, column=1, value=label)
    c.font = Font(bold=(is_section or 'Total' in label), size=10,
                  color=(fg or '000000') if is_section else '000000')
    if bg: c.fill = PatternFill('solid', fgColor=fg if is_section else bg)
    c.alignment = Alignment(horizontal='left', vertical='center')

    if is_section:
        for ci in range(2, len(COLS) + 1):
            ws1.cell(row=row, column=ci).fill = PatternFill('solid', fgColor=fg if is_section else bg)
        row += 1
        continue

    col_total = 0
    for ci, m in enumerate(months, 2):
        v = 0
        if key == 'cash':   v = db_rent[m]['cash']
        elif key == 'upi':  v = db_rent[m]['upi']
        elif key == 'both': v = db_rent[m]['cash'] + db_rent[m]['upi']
        elif key == 'bank_tenant':   v = int(bank_by_month[m].get('Tenant UPI Collection', 0))
        elif key == 'bank_personal': v = int(bank_by_month[m].get('Personal / Capital Transfer', 0))
        elif key == 'bank_rtgs':     v = int(bank_by_month[m].get('Capital Infusion (RTGS/NEFT)', 0))
        elif key == 'bank_indiv':    v = int(bank_by_month[m].get('UPI - Individual (unidentified)', 0))
        elif key == 'bank_other':    v = int(bank_by_month[m].get('Other Credit', 0))
        elif key == 'bank_total':    v = int(sum(bank_by_month[m].values()))
        elif key == 'upi_neg':       v = -db_rent[m]['upi']
        elif key == 'gap':
            v = int(bank_by_month[m].get('Tenant UPI Collection', 0)) - db_rent[m]['upi']
        col_total += v
        cell = num(ws1, row, ci, v if v != 0 else None,
                   bg=bg if v else None)
        if key == 'gap' and v != 0:
            cell.fill = PatternFill('solid', fgColor='FCE4D6' if v > 5000 else LIGHT_GREEN)

    # Total column
    num(ws1, row, len(COLS), col_total if col_total != 0 else None,
        bg='E2EFDA' if 'Total Rent' in label else (LIGHT_BLUE if 'Total Bank' in label else None))
    row += 1

# Note row
row += 1
n = ws1.cell(row=row, column=1,
    value='NOTE: Gap = money collected via tenant UPI but not classified as rent in DB '
          '(security deposits paid via UPI, booking advances, payments before bot was used).')
n.font = Font(italic=True, size=9, color='595959')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))

# Column widths
ws1.column_dimensions['A'].width = 44
for ci in range(2, len(COLS) + 1):
    ws1.column_dimensions[get_column_letter(ci)].width = 13

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — Working Capital (Deposits)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Working Capital')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:E1')
t2 = ws2['A1']
t2.value = 'WORKING CAPITAL — DEPOSITS & BOOKING ADVANCES (NOT profit — must be refunded)'
t2.font = Font(bold=True, size=13, color='C00000')
t2.alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[1].height = 24

# Summary table
headers2 = ['Item', 'Mode', 'Count', 'Amount (₹)', 'Notes']
for ci, h in enumerate(headers2, 1):
    hdr(ws2, 3, ci, h, bg='C00000')

wc_data = [
    ('Security Deposits collected',  'cash',
     wc_summary.get(('deposit','cash'), {}).get('n', 0),
     wc_summary.get(('deposit','cash'), {}).get('total', 0),
     'Held until checkout. Refundable (minus damage). Pure LIABILITY.'),
    ('Security Deposits collected',  'upi',
     wc_summary.get(('deposit','upi'), {}).get('n', 0),
     wc_summary.get(('deposit','upi'), {}).get('total', 0),
     ''),
    ('Booking Advances collected',   'cash',
     wc_summary.get(('booking','cash'), {}).get('n', 0),
     wc_summary.get(('booking','cash'), {}).get('total', 0),
     'Applied to first month rent when tenant moves in. Deferred income.'),
    ('Booking Advances collected',   'upi',
     wc_summary.get(('booking','upi'), {}).get('n', 0),
     wc_summary.get(('booking','upi'), {}).get('total', 0),
     ''),
]

total_collected = sum(r[3] for r in wc_data)

row2 = 4
for item, mode, n, amt, note in wc_data:
    if n == 0 and amt == 0: continue
    lbl(ws2, row2, 1, item)
    lbl(ws2, row2, 2, mode)
    num(ws2, row2, 3, n, INR_NUMBER_FORMAT)
    num(ws2, row2, 4, amt, INR_NUMBER_FORMAT, bg=LIGHT_RED)
    lbl(ws2, row2, 5, note)
    row2 += 1

# Refunds paid out row
row2 += 1
lbl(ws2, row2, 1, 'Less: Deposits already refunded to tenants', bold=True)
lbl(ws2, row2, 2, 'cash+upi')
num(ws2, row2, 4, -DEPOSIT_REFUNDS_PAID, INR_NUMBER_FORMAT, bg=LIGHT_GREEN)
lbl(ws2, row2, 5, 'From expense classification (bank outflows tagged Tenant Deposit Refund)')

row2 += 1
lbl(ws2, row2, 1, 'NET WORKING CAPITAL OWED TO TENANTS', bold=True)
num(ws2, row2, 4, total_collected - DEPOSIT_REFUNDS_PAID, INR_NUMBER_FORMAT, bg='FCE4D6')
lbl(ws2, row2, 5, 'This is a liability on your balance sheet — NOT profit')
ws2.row_dimensions[row2].height = 20

row2 += 3
ws2.cell(row=row2, column=1,
    value='IMPORTANT: This money was collected in cash (mostly). It does not appear in the bank statement.').font = Font(italic=True, color='595959')
row2 += 1
ws2.cell(row=row2, column=1,
    value='When a tenant checks out and you refund their deposit, that cash outflow reduces this liability.').font = Font(italic=True, color='595959')

row2 += 3
hdr(ws2, row2, 1, 'Tenant', bg='C00000')
hdr(ws2, row2, 2, 'Check-in', bg='C00000')
hdr(ws2, row2, 3, 'Type', bg='C00000')
hdr(ws2, row2, 4, 'Mode', bg='C00000')
hdr(ws2, row2, 5, 'Amount', bg='C00000')
row2 += 1

for r in dep_detail:
    lbl(ws2, row2, 1, r['name'])
    lbl(ws2, row2, 2, str(r['checkin_date'] or ''))
    lbl(ws2, row2, 3, r['for_type'])
    lbl(ws2, row2, 4, r['payment_mode'])
    num(ws2, row2, 5, int(r['amount']))
    row2 += 1

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 55

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — Bank Credits Detail
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Bank Credits Detail')
ws3.sheet_view.showGridLines = False

headers3 = ['Date', 'Amount (₹)', 'Category', 'Description', 'Reference', 'Note']
for ci, h in enumerate(headers3, 1):
    hdr(ws3, 1, ci, h)

CAT_COLORS = {
    'Tenant UPI Collection':              LIGHT_BLUE,
    'Personal / Capital Transfer':        LIGHT_GRAY,
    'Capital Infusion (RTGS/NEFT)':       LIGHT_GRAY,
    'UPI - Individual (unidentified)':    YELLOW,
    'Refund Received':                    LIGHT_GREEN,
    'Other Credit':                       YELLOW,
}

for ri, (dt, desc, ref, amt, cat, note, mk) in enumerate(
        sorted(credit_rows, key=lambda x: x[0]), 2):
    bg = CAT_COLORS.get(cat, None)
    lbl(ws3, ri, 1, str(dt))
    num(ws3, ri, 2, amt, INR_NUMBER_FORMAT, bg=bg)
    lbl(ws3, ri, 3, cat)
    lbl(ws3, ri, 4, desc)
    lbl(ws3, ri, 5, ref)
    lbl(ws3, ri, 6, note)

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 13
ws3.column_dimensions['C'].width = 33
ws3.column_dimensions['D'].width = 65
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 40

# ── Save ─────────────────────────────────────────────────────────────────────
OUT = 'data/reports/income_reconciliation.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')

# ── Console summary ───────────────────────────────────────────────────────────
print()
print(f'{"":44s}', end='')
for m in months: print(f'  {MONTH_LABELS[m]:>10s}', end='')
print(f'  {"TOTAL":>10s}')
print('-' * (44 + (len(months)+1) * 12))

rows_to_print = [
    ('Rent Cash (DB)',          lambda m: db_rent[m]['cash']),
    ('Rent UPI (DB)',           lambda m: db_rent[m]['upi']),
    ('Total Rent Income',       lambda m: db_rent[m]['cash'] + db_rent[m]['upi']),
    ('---', None),
    ('Bank: Tenant UPI Collect',lambda m: int(bank_by_month[m].get('Tenant UPI Collection', 0))),
    ('Bank: Capital/Personal',  lambda m: int(bank_by_month[m].get('Personal / Capital Transfer', 0) +
                                              bank_by_month[m].get('Capital Infusion (RTGS/NEFT)', 0))),
    ('Bank: Total Credits',     lambda m: int(sum(bank_by_month[m].values()))),
    ('---', None),
    ('Gap (Tenant UPI - Rent)', lambda m: int(bank_by_month[m].get('Tenant UPI Collection', 0)) - db_rent[m]['upi']),
]

for label, fn in rows_to_print:
    if label == '---':
        print()
        continue
    vals = [fn(m) for m in months]
    print(f'{label:44s}', end='')
    for v in vals: print(f'  {v:>10,}', end='')
    print(f'  {sum(vals):>10,}')

print()
dep_total   = wc_summary.get(('deposit','cash'),{}).get('total',0) + wc_summary.get(('deposit','upi'),{}).get('total',0)
book_total  = wc_summary.get(('booking','cash'),{}).get('total',0) + wc_summary.get(('booking','upi'),{}).get('total',0)
print(f'Security deposits held (liability): {dep_total:>12,}')
print(f'Booking advances held (deferred):   {book_total:>12,}')
print(f'Less: deposits refunded:            {-DEPOSIT_REFUNDS_PAID:>12,}')
print(f'NET working capital owed to tenants:{dep_total + book_total - DEPOSIT_REFUNDS_PAID:>12,}')
