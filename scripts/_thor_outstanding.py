"""
THOR Outstanding Dues — May 2026

Sources:
  MAY 2026 Google Sheet tab (THOR tenants only): rent_due, prev_due, cash, sheet_upi
  thor bank statement april til now.xlsx (May 1+ rows): UPI paid to THOR account
  Hulk may 11th bank statement.xlsx: cross-building payments (THOR tenants who paid HULK)

Logic:
  bank_total = sum of all THOR+HULK May bank matches by phone/name
  effective_upi = max(sheet_upi, bank_total)
  balance = rent_due + prev_due - effective_upi - cash

Note: prev_due in May Sheet = April carry-over entered by receptionist.
      Using effective_upi vs sheet_upi flags cases where bank shows more than Sheet recorded.

Output: data/reports/THOR_Outstanding_May2026.xlsx
"""
import re, datetime, os
import openpyxl, gspread
from google.oauth2.service_account import Credentials
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

MAY_START = datetime.date(2026, 5, 1)

# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_phone(vpa):
    m = re.match(r'^(\d{10})(?:-\d+)?@', str(vpa or ''))
    return m.group(1) if m else None

def normalize_phone(ph):
    if not ph: return None
    ph = re.sub(r'\D', '', str(ph))
    if ph.startswith('91') and len(ph) == 12: ph = ph[2:]
    return ph if len(ph) == 10 else None

def name_tokens(s):
    return [p for p in re.sub(r'[^A-Z0-9 ]', ' ', s.upper()).split() if len(p) > 1]

def safe_float(v):
    try: return float(str(v).replace(',', '').replace('Rs.', '').strip() or 0)
    except: return 0

def find_total_in_pool(tenant_name, tenant_phone, pool):
    """Sum all matching entries in pool by phone, then by name, then fuzzy."""
    tphone = normalize_phone(tenant_phone)
    name_up = tenant_name.strip().upper()
    tparts = name_tokens(name_up)

    # Phase 1: phone match — sum all
    if tphone:
        total = 0
        for e in pool:
            if e['used']: continue
            if normalize_phone(e['phone']) == tphone:
                total += e['amt']
                e['used'] = True
        if total > 0:
            return total, 'phone'

    # Phase 2: exact name
    for e in pool:
        if not e['used'] and e['name'] == name_up:
            e['used'] = True
            return e['amt'], 'name'

    # Phase 3: fuzzy name (>=2 token hits, score >=0.6)
    if len(tparts) >= 2:
        best, bi = 0, None
        for i, e in enumerate(pool):
            if e['used']: continue
            bparts = name_tokens(e['name'])
            hits = sum(1 for tp in tparts if any(tp in bp or bp in tp for bp in bparts))
            score = hits / len(tparts)
            if hits >= 2 and score >= 0.6 and score > best:
                best, bi = score, i
        if bi is not None:
            pool[bi]['used'] = True
            return pool[bi]['amt'], 'fuzzy'

    return 0, None

# ── Load bank statements ──────────────────────────────────────────────────────

thor_pool = []
seen_rrn = set()
wb = openpyxl.load_workbook('thor bank statement april til now.xlsx')
for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
    if row[11] != 'SUCCESS' or not row[3]: continue
    rrn = row[0]
    if rrn in seen_rrn: continue
    seen_rrn.add(rrn)
    txn_date = row[1].date() if isinstance(row[1], datetime.datetime) else row[1]
    if txn_date and txn_date >= MAY_START:
        thor_pool.append({
            'amt': float(row[3]), 'phone': extract_phone(row[7]),
            'name': str(row[8]).strip().upper(), 'used': False, 'bank': 'THOR',
        })

hulk_pool = []
seen_rrn_h = set()
wb2 = openpyxl.load_workbook('Hulk may 11th bank statement.xlsx')
for row in wb2.worksheets[0].iter_rows(min_row=2, values_only=True):
    if row[11] != 'SUCCESS' or not row[3]: continue
    rrn = row[0]
    if rrn in seen_rrn_h: continue
    seen_rrn_h.add(rrn)
    hulk_pool.append({
        'amt': float(row[3]), 'phone': extract_phone(row[7]),
        'name': str(row[8]).strip().upper(), 'used': False, 'bank': 'HULK',
    })

print(f'THOR May bank entries: {len(thor_pool)}  (total Rs.{sum(e["amt"] for e in thor_pool):,.0f})')
print(f'HULK May bank entries: {len(hulk_pool)}  (total Rs.{sum(e["amt"] for e in hulk_pool):,.0f})')

# ── Read MAY 2026 Google Sheet — THOR tenants ─────────────────────────────────

creds = Credentials.from_service_account_file(
    'credentials/gsheets_service_account.json',
    scopes=['https://www.googleapis.com/auth/spreadsheets']
)
gc = gspread.authorize(creds)
sh = gc.open_by_key('1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw')
ws = sh.worksheet('MAY 2026')
rows = ws.get_all_values()
H = {h: i for i, h in enumerate(rows[7])}

tenants = []
for row in rows[8:]:
    if len(row) <= H.get('UPI', 999): continue
    room = row[H['Room']].strip()
    name = row[H['Name']].strip()
    if not room or not name: continue
    building = row[H['Building']].strip().upper()
    if 'THOR' not in building: continue
    rent_due = safe_float(row[H['Rent Due']])
    if rent_due <= 0: continue
    prev_due = safe_float(row[H['Prev Due']]) if 'Prev Due' in H and len(row) > H['Prev Due'] else 0
    phone = row[H['Phone']].strip() if 'Phone' in H else ''
    checkin_raw = row[H['Check-in']].strip() if 'Check-in' in H and len(row) > H['Check-in'] else ''
    checkin = None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            checkin = datetime.datetime.strptime(checkin_raw, fmt).date(); break
        except: pass
    rent = safe_float(row[H['Rent']]) if 'Rent' in H else 0
    tenants.append({
        'room': room, 'name': name, 'phone': phone,
        'rent': rent, 'rent_due': rent_due, 'prev_due': prev_due,
        'cash': safe_float(row[H['Cash']]),
        'sheet_upi': safe_float(row[H['UPI']]),
        'checkin': checkin,
    })

print(f'THOR tenants in May Sheet with dues: {len(tenants)}')

# ── Match each tenant against bank pools ─────────────────────────────────────

results = []
for t in tenants:
    thor_amt, thor_method = find_total_in_pool(t['name'], t['phone'], thor_pool)
    hulk_amt, hulk_method = find_total_in_pool(t['name'], t['phone'], hulk_pool)
    bank_total = thor_amt + hulk_amt
    effective_upi = max(t['sheet_upi'], bank_total)
    true_total = effective_upi + t['cash']
    balance = t['rent_due'] + t['prev_due'] - true_total

    # Build verification note
    parts = []
    if thor_amt > 0:
        parts.append(f'THOR bank Rs.{thor_amt:,.0f} ({thor_method})')
    if hulk_amt > 0:
        parts.append(f'HULK bank Rs.{hulk_amt:,.0f} ({hulk_method})')
    if bank_total < t['sheet_upi']:
        parts.append('Sheet UPI higher than bank — used sheet')
    if not parts:
        parts.append('No bank match found')
    note = '; '.join(parts)

    results.append({**t,
        'thor_bank': thor_amt, 'hulk_bank': hulk_amt,
        'bank_total': bank_total, 'effective_upi': effective_upi,
        'true_total': true_total, 'balance': balance,
        'note': note,
    })

# ── Categorize ────────────────────────────────────────────────────────────────

no_payment  = sorted([r for r in results if r['balance'] > 0 and r['true_total'] == 0],  key=lambda x: -x['balance'])
partial     = sorted([r for r in results if r['balance'] > 0 and r['true_total'] > 0],   key=lambda x: -x['balance'])
cleared     = sorted([r for r in results if r['balance'] <= 0],                           key=lambda x: x['name'])

print(f'\nNO PAYMENT: {len(no_payment)}')
print(f'PARTIAL:    {len(partial)}')
print(f'CLEARED (per bank / sheet): {len(cleared)}')

total_outstanding = sum(r['balance'] for r in no_payment + partial)
print(f'Total outstanding: Rs.{total_outstanding:,.0f}')

print('\nUnmatched THOR May bank entries (amt >= 1000):')
for e in sorted([e for e in thor_pool if not e['used'] and e['amt'] >= 1000], key=lambda x: -x['amt']):
    print(f'  Rs.{e["amt"]:>8,.0f}  {e["name"]}  {e["phone"] or ""}')

print('\nUnmatched HULK May bank entries >= 1000 (could be THOR cross-payments or already entered in sheet):')
for e in sorted([e for e in hulk_pool if not e['used'] and e['amt'] >= 1000], key=lambda x: -x['amt']):
    print(f'  Rs.{e["amt"]:>8,.0f}  {e["name"]}  {e["phone"] or ""}')

# ── Excel styles ──────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill('solid', fgColor='1F3864')
HDR_FONT    = Font(bold=True, color='FFFFFF', size=10)
NOPAY_FILL  = PatternFill('solid', fgColor='FCE4D6')   # red-ish
PARTIAL_FILL= PatternFill('solid', fgColor='FFF2CC')   # yellow
CLEAR_FILL  = PatternFill('solid', fgColor='E2EFDA')   # green
CROSS_FILL  = PatternFill('solid', fgColor='DDEBF7')   # blue: HULK cross-payment
TOTAL_FILL  = PatternFill('solid', fgColor='BDD7EE')
TOTAL_FONT  = Font(bold=True, size=10)
NORM_FONT   = Font(size=10)
thin        = Side(style='thin', color='BBBBBB')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT       = Alignment(horizontal='right',  vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')

COLS = ['#', 'Room', 'Name', 'Check-in', 'Rent', 'Rent Due', 'Prev Due',
        'Cash', 'Sheet UPI', 'THOR Bank', 'HULK Bank', 'Eff. UPI',
        'Total Paid', 'Balance', 'Status', 'Bank Notes']
WIDTHS = [4, 6, 28, 12, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 18, 45]
NUM_COLS = {5, 6, 7, 8, 9, 10, 11, 12, 13, 14}  # 1-indexed

def write_section_header(ws, row_n, label, color='595959'):
    ws.merge_cells(f'A{row_n}:{get_column_letter(len(COLS))}{row_n}')
    c = ws.cell(row=row_n, column=1, value=label)
    c.font = Font(bold=True, italic=True, size=10, color=color)
    c.alignment = LEFT

def write_data_row(ws, row_n, no, r, fill):
    vals = [
        no, r['room'], r['name'], r['checkin'],
        r['rent'], r['rent_due'], r['prev_due'],
        r['cash'], r['sheet_upi'], r['thor_bank'], r['hulk_bank'],
        r['effective_upi'], r['true_total'], r['balance'],
        ('No Payment' if r['true_total'] == 0 else
         'Cleared' if r['balance'] <= 0 else
         f"Owes Rs.{int(r['balance']):,}"),
        r['note'],
    ]
    row_fill = CROSS_FILL if r['hulk_bank'] > 0 and r['balance'] <= 0 else fill
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        cell.fill = row_fill; cell.border = BORDER; cell.font = NORM_FONT
        if ci in NUM_COLS and isinstance(v, (int, float)):
            cell.alignment = RIGHT; cell.number_format = '#,##0'
        elif ci == 4:
            cell.alignment = CENTER; cell.number_format = 'DD-MMM-YYYY'
        else:
            cell.alignment = LEFT

def write_sheet(ws, label, sections):
    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
    ws['A1'] = label
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22

    row_n = 3
    seq = 1
    t_due = t_prev = t_cash = t_upi = t_tot = t_bal = 0

    for sec_label, rows_s, fill in sections:
        if not rows_s: continue
        write_section_header(ws, row_n, sec_label)
        row_n += 1
        for r in rows_s:
            write_data_row(ws, row_n, seq, r, fill)
            t_due += r['rent_due']; t_prev += r['prev_due']
            t_cash += r['cash']; t_upi += r['effective_upi']
            t_tot += r['true_total']; t_bal += r['balance']
            seq += 1; row_n += 1

    # totals row
    tot_vals = [None, '', 'TOTAL', None, None, t_due, t_prev, t_cash, None, None, None, t_upi, t_tot, t_bal, '', '']
    for ci, v in enumerate(tot_vals, 1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER
        if ci in NUM_COLS and isinstance(v, (int, float)):
            cell.alignment = RIGHT; cell.number_format = '#,##0'

    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'

# ── Build workbook ─────────────────────────────────────────────────────────────

wb_out = openpyxl.Workbook()
if wb_out.active: wb_out.remove(wb_out.active)

outstanding_rows = no_payment + partial

sections_main = [
    ('--- NO PAYMENT (no cash, no UPI, no bank match) ---', no_payment, NOPAY_FILL),
    ('--- PARTIAL PAYMENT (balance remaining) ---',          partial,    PARTIAL_FILL),
]
sections_all = sections_main + [
    ('--- CLEARED PER BANK / SHEET (no balance) ---',        cleared,    CLEAR_FILL),
]

ws1 = wb_out.create_sheet('THOR Outstanding', 0)
write_sheet(ws1, f'THOR — Outstanding Dues — May 2026  (as of 11 May 2026)', sections_main)

ws2 = wb_out.create_sheet('THOR All Tenants')
write_sheet(ws2, f'THOR — All Tenants — May 2026  (as of 11 May 2026)', sections_all)

os.makedirs('data/reports', exist_ok=True)
out_path = 'data/reports/THOR_Outstanding_May2026.xlsx'
wb_out.save(out_path)

print(f'\nSaved {out_path}')
print(f'  Outstanding: {len(outstanding_rows)} tenants, Rs.{total_outstanding:,.0f}')
print(f'  (blue rows = HULK cross-payment detected)')
print(f'\nNote: HULK April+May bank not yet available.')
print(f'      Re-run after Kiran uploads hulk bank statement april til now.')
