"""
Dry-run: parse 'pending dues (1).xlsx' Q column and compare against DB.

Usage:
    python scripts/_dry_run_pending_dues.py            # dry run (default)
    python scripts/_dry_run_pending_dues.py --write    # actually insert payments

For each row it prints one of:
    [SKIP]   — no action needed (already logged, no payment, or flagged)
    [OK]     — payment already in DB, matches comment
    [INSERT] — payment not in DB, will insert (or would insert in dry run)
    [FLAG]   — ambiguous, needs manual review
"""
import asyncio
import io
import os
import re
import sys
from datetime import date
from decimal import Decimal

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from dotenv import load_dotenv

load_dotenv()

WRITE = "--write" in sys.argv
EXCEL_FILE = "pending dues (1).xlsx"
MAY_PERIOD = date(2026, 5, 1)
APR_PERIOD = date(2026, 4, 1)
JUN_PERIOD = date(2026, 6, 1)

# Rows to always flag — needs human decision
MANUAL_FLAGS = {
    "Siddharth Dewangan": "No payment — 'without inform exit'. Needs checkout, not a payment.",
    "Prasad Vadlamani":   "Yet to be checkin — skip.",
    "V. Bhanu Prakash":   "Shared-phone unresolved pair (room 314). Manual fix required.",
    "Anshsinha":          "Comment mentions 'Adjusted in Anubav day stay rent 2400' — ambiguous split.",
}

# ── Comment parser ─────────────────────────────────────────────────────────────

def parse_comment(comment: str) -> list[dict]:
    """
    Extract payment actions from a free-text comment.
    Returns list of dicts: {amount, method, for_type, period, note}
    """
    if not comment:
        return []
    c = comment.strip()
    actions = []

    # Detect period overrides
    period = MAY_PERIOD
    if re.search(r'\bapr\b', c, re.I):
        period = APR_PERIOD
    # June carry-forward is a note, not a payment — skip
    jun_carry = re.search(r'(\d[\d,]+)\s*k?\s*(?:will\s+(?:be\s+)?collect|with\s+june)', c, re.I)
    jun_carry_amt = int(jun_carry.group(1).replace(',', '')) * (1000 if 'k' in jun_carry.group(0).lower() else 1) if jun_carry else 0

    # Find all amount+method pairs
    # Patterns: "cash 11k", "UPI 21000", "upi 5000 and cash 6500"
    pairs = re.findall(
        r'(cash|upi|neft|online|transfer)\s+(?:rs\.?\s*)?(\d[\d,]*)\s*(?:k\b)?|'
        r'(?:rs\.?\s*)?(\d[\d,]*)\s*(?:k\b)?\s+(cash|upi|neft|online|transfer)',
        c, re.I
    )

    for p in pairs:
        method_pre, amt_pre, amt_post, method_post = p
        method = (method_pre or method_post).upper()
        raw_amt = amt_pre or amt_post
        amt = int(raw_amt.replace(',', ''))
        # multiply by 1000 if followed by 'k' in original
        # check context around match
        if re.search(r'\b' + re.escape(raw_amt) + r'\s*k\b', c, re.I):
            amt *= 1000

        if amt == jun_carry_amt:
            continue  # this is a June carry note, not a payment now

        # Determine for_type
        for_type = "rent"
        ctx_start = max(0, c.lower().find(raw_amt) - 30)
        ctx = c[ctx_start:ctx_start + 60].lower()
        if "deposit" in ctx or "security" in ctx or "dep" in ctx:
            for_type = "deposit"

        if method in ("NEFT", "ONLINE", "TRANSFER"):
            method = "UPI"

        actions.append({
            "amount": amt,
            "method": method,
            "for_type": for_type,
            "period": period,
            "note": f"From pending dues sheet: {comment[:80]}",
        })

    return actions


# ── DB helpers ─────────────────────────────────────────────────────────────────

async def get_existing_payments(session, tenancy_id: int, period: date, for_type: str) -> Decimal:
    from sqlalchemy import select, func
    from src.database.models import Payment, PaymentFor, PaymentMode
    ft = PaymentFor(for_type)
    result = await session.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0))
        .where(
            Payment.tenancy_id == tenancy_id,
            Payment.is_void == False,
            Payment.for_type == ft,
            Payment.period_month == period,
        )
    )
    return Decimal(str(result or 0))


async def get_tenancy(session, room_number: str, building: str):
    from sqlalchemy import select
    from src.database.models import Tenancy, Tenant, Room, Property, TenancyStatus
    building_like = f"%{building}%"
    row = await session.execute(
        select(Tenancy, Tenant.name, Tenant.phone, Room.room_number)
        .join(Tenant, Tenant.id == Tenancy.tenant_id)
        .join(Room, Room.id == Tenancy.room_id)
        .join(Property, Property.id == Room.property_id)
        .where(
            Room.room_number == room_number,
            Tenancy.status.in_([TenancyStatus.active, TenancyStatus.no_show]),
            Property.name.ilike(building_like),
        )
        .order_by(Tenancy.id.desc())
        .limit(1)
    )
    return row.first()


# ── Main ───────────────────────────────────────────────────────────────────────

async def main():
    from src.database.db_manager import init_db, get_session
    from src.database.models import Payment, PaymentFor, PaymentMode, AuditLog

    db_url = os.environ["DATABASE_URL"]
    await init_db(db_url)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Row 3 onwards = data (row 1 = title, row 2 = headers)
    COL = {
        "num": 0, "building": 1, "name": 2, "room": 3, "phone": 4,
        "checkin": 5, "may_checkin": 6, "stay_type": 7,
        "rent": 8, "deposit": 9, "booking": 10,
        "may_due": 11, "may_paid": 12, "dep_owed": 13,
        "outstanding": 14, "pending_months": 15, "comment": 16,
    }

    total_inserts = 0
    total_skips = 0
    total_flags = 0

    print(f"\n{'='*70}")
    print(f"{'DRY RUN' if not WRITE else '*** WRITE MODE ***'} — pending dues (1).xlsx — {date.today()}")
    print(f"{'='*70}\n")

    async with get_session() as session:
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            num = row[COL["num"]]
            if not num or not isinstance(num, int):
                continue  # skip total row

            name     = str(row[COL["name"]] or "").strip()
            room     = str(row[COL["room"]] or "").strip()
            building = str(row[COL["building"]] or "").strip()
            comment  = str(row[COL["comment"]] or "").strip()
            outstanding = row[COL["outstanding"]] or 0

            prefix = f"[#{num:02d}] {name} (Room {room})"

            # ── Manual flags ──────────────────────────────────────────────
            if name in MANUAL_FLAGS:
                print(f"[FLAG]   {prefix}")
                print(f"         Reason: {MANUAL_FLAGS[name]}")
                print()
                total_flags += 1
                continue

            # ── No comment ────────────────────────────────────────────────
            if not comment or comment.lower() in ("", "none", "-"):
                print(f"[FLAG]   {prefix}")
                print(f"         No comment — outstanding ₹{int(outstanding):,}. Review manually.")
                print()
                total_flags += 1
                continue

            # ── Outstanding = 0, "all dues clear" ────────────────────────
            # Still parse — may have payments to log even if balance shows 0
            # (sheet balance may already reflect the payment)

            # ── Look up tenancy ───────────────────────────────────────────
            result = await get_tenancy(session, room, building)
            if not result:
                print(f"[FLAG]   {prefix}")
                print(f"         Tenancy not found in DB for room {room} / {building}.")
                print()
                total_flags += 1
                continue

            tenancy, db_name, db_phone, db_room = result

            # ── Parse comment ─────────────────────────────────────────────
            actions = parse_comment(comment)

            if not actions:
                print(f"[FLAG]   {prefix}")
                print(f"         Comment could not be parsed: '{comment}'")
                print()
                total_flags += 1
                continue

            # ── For each parsed payment, compare with DB ──────────────────
            for act in actions:
                already_paid = await get_existing_payments(
                    session, tenancy.id, act["period"], act["for_type"]
                )
                period_label = act["period"].strftime("%b %Y")
                amt = act["amount"]

                if already_paid >= Decimal(str(amt)):
                    print(f"[OK]     {prefix}")
                    print(f"         {act['for_type'].upper()} ₹{amt:,} {act['method']} {period_label} — already in DB (paid ₹{int(already_paid):,})")
                    print()
                    total_skips += 1
                    continue

                delta = amt - int(already_paid)
                print(f"[INSERT] {prefix}")
                print(f"         {act['for_type'].upper()} ₹{delta:,} {act['method']} {period_label}")
                print(f"         DB has ₹{int(already_paid):,} logged; comment says ₹{amt:,} → inserting delta ₹{delta:,}")
                print(f"         Comment: {comment[:80]}")

                if WRITE:
                    method_map = {"CASH": PaymentMode.cash, "UPI": PaymentMode.upi}
                    pmode = method_map.get(act["method"], PaymentMode.upi)
                    pfor  = PaymentFor(act["for_type"])
                    payment = Payment(
                        tenancy_id=tenancy.id,
                        amount=Decimal(str(delta)),
                        payment_mode=pmode,
                        for_type=pfor,
                        period_month=act["period"],
                        payment_date=date.today(),
                        notes=act["note"],
                        source="excel_import",
                        recorded_by="Kiran",
                        is_void=False,
                    )
                    session.add(payment)
                    await session.flush()
                    session.add(AuditLog(
                        changed_by="Kiran",
                        entity_type="payment",
                        entity_id=payment.id,
                        field="payment.log",
                        old_value="0",
                        new_value=str(delta),
                        room_number=room,
                        entity_name=db_name,
                        source="excel_import",
                        note=f"₹{delta:,} {act['method']} {act['for_type']} {period_label} — pending dues import",
                    ))
                    print(f"         ✓ Inserted payment id={payment.id}")

                print()
                total_inserts += 1

        if WRITE:
            await session.commit()

    print(f"{'='*70}")
    print(f"SUMMARY: {total_inserts} inserts {'(written)' if WRITE else '(dry run — use --write to apply)'}")
    print(f"         {total_skips} already OK in DB")
    print(f"         {total_flags} flagged for manual review")
    print(f"{'='*70}\n")


asyncio.run(main())
