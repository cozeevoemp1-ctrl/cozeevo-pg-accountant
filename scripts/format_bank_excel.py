"""
Re-format the bank statement extracted Excel into a polished multi-sheet workbook.

Usage:
    python scripts/format_bank_excel.py
    python scripts/format_bank_excel.py path/to/extracted.xlsx
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── colour palette ──────────────────────────────────────────────────────────
HEADER_BG   = "1F497D"
CREDIT_BG   = "E2EFDA"   # green  — deposits
DEBIT_BG    = "FCE4D6"   # orange — withdrawals
NEUTRAL_BG  = "F2F2F2"   # grey   — both/neither
TOTAL_BG    = "D9D9D9"

CAT_COLORS = {
    "salary":    "D6E4F0",
    "food":      "FFF2CC",
    "transport": "E2EFDA",
    "fuel":      "FCE4D6",
    "utilities": "EAD1DC",
    "shopping":  "D5E8D4",
    "banking":   "F4CCCC",
    "other":     "F3F3F3",
}

THIN  = Side(style="thin",   color="BBBBBB")
MED   = Side(style="medium", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOT_MED = Border(left=THIN, right=THIN, top=THIN, bottom=MED)


def _hdr_cell(ws, row, col, value):
    c = ws.cell(row, col, value)
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c


def _num(ws, row, col, value):
    c = ws.cell(row, col, value if value != "" else None)
    c.number_format = "₹#,##0.00"
    c.alignment     = Alignment(horizontal="right")
    c.border        = BORDER
    return c


def _set_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Sheet 1 : All Transactions ──────────────────────────────────────────────
def sheet_all(wb, df: pd.DataFrame):
    ws = wb.create_sheet("All Transactions")

    # ── title ─────────────────────────────────────────────────────────────────
    title_cols = 13
    ws.merge_cells(f"A1:{get_column_letter(title_cols)}1")
    t = ws["A1"]
    t.value     = "YES Bank Statement — LAKSHMI GORJALA  (Dec 2025 – Mar 2026)"
    t.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── headers ───────────────────────────────────────────────────────────────
    headers = [
        "#", "Txn Date", "Value Date", "Description", "Reference",
        "Withdrawals (₹)", "Deposits (₹)", "Balance (₹)",
        "Type", "UTR", "Payer UPI", "Payee UPI", "Category",
    ]
    col_keys = [
        None, "Transaction Date", "Value Date", "Description", "Reference",
        "Withdrawals", "Deposits", "Balance",
        "transaction_type", "utr", "payer_upi", "payee_upi", "category",
    ]
    for c, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, c, h)
    ws.row_dimensions[2].height = 30

    # ── data rows ─────────────────────────────────────────────────────────────
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = i + 2
        cat = str(row.get("category", "other")).lower()
        has_wd  = isinstance(row["Withdrawals"], float) and row["Withdrawals"] > 0
        has_dep = isinstance(row["Deposits"],    float) and row["Deposits"]    > 0

        if has_wd and not has_dep:
            bg = DEBIT_BG
        elif has_dep and not has_wd:
            bg = CREDIT_BG
        else:
            bg = CAT_COLORS.get(cat, NEUTRAL_BG)

        fill = PatternFill("solid", fgColor=bg)

        # row number
        c0 = ws.cell(r, 1, i)
        c0.fill = fill; c0.border = BORDER
        c0.alignment = Alignment(horizontal="center")
        c0.font = Font(name="Calibri", size=9, color="666666")

        # text columns
        for ci, key in enumerate(col_keys[1:], 2):
            val = row.get(key, "")
            if val != val:   # NaN
                val = ""
            c = ws.cell(r, ci, val)
            c.fill = fill; c.border = BORDER
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 4))

        # amount columns — right-aligned with ₹ format
        for ci, key in [(6, "Withdrawals"), (7, "Deposits"), (8, "Balance")]:
            val = row.get(key, "")
            cn = _num(ws, r, ci, val if isinstance(val, float) else None)
            cn.fill = fill
            cn.font = Font(name="Calibri", size=9,
                           color="CC0000" if key == "Withdrawals" else
                                 "006600" if key == "Deposits" else "000000")

    # ── totals ────────────────────────────────────────────────────────────────
    total_r = len(df) + 3
    ws.cell(total_r, 4, "TOTAL").font = Font(bold=True, name="Calibri")
    for ci, key in [(6, "Withdrawals"), (7, "Deposits")]:
        total = df[key].apply(lambda x: x if isinstance(x, float) else 0).sum()
        c = _num(ws, total_r, ci, total)
        c.font = Font(bold=True, name="Calibri",
                      color="CC0000" if key == "Withdrawals" else "006600")
    for col in range(1, title_cols + 1):
        cell = ws.cell(total_r, col)
        cell.fill   = PatternFill("solid", fgColor=TOTAL_BG)
        cell.border = BOT_MED

    _set_widths(ws, [4, 11, 11, 42, 20, 14, 14, 14, 7, 18, 26, 26, 11])
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"


# ─── Sheet 2 : Monthly Summary ───────────────────────────────────────────────
def sheet_monthly(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Monthly Summary")

    df2 = df.copy()
    df2["_date"] = pd.to_datetime(df2["Transaction Date"], dayfirst=True, errors="coerce")
    df2["Month"]  = df2["_date"].dt.strftime("%b %Y")

    months_order = list(dict.fromkeys(df2["Month"].dropna()))

    summary = df2.groupby("Month").agg(
        Transactions=("Transaction Date", "count"),
        Withdrawals=("Withdrawals",  lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
        Deposits=("Deposits",        lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
    ).reindex(months_order).reset_index()
    summary["Net"] = summary["Deposits"] - summary["Withdrawals"]

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Monthly Cash Flow Summary"
    t.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["Month", "Transactions", "Withdrawals (₹)", "Deposits (₹)", "Net (₹)", ""], 1):
        _hdr_cell(ws, 2, c, h)

    MONTH_COLORS = ["D6E4F0", "D5E8D4", "FFF2CC", "FCE4D6", "E1D5E7",
                    "F3F3F3", "EAD1DC", "D9EAD3"]

    for ri, (_, row) in enumerate(summary.iterrows(), 3):
        bg = MONTH_COLORS[ri % len(MONTH_COLORS)]
        fill = PatternFill("solid", fgColor=bg)

        c0 = ws.cell(ri, 1, row["Month"])
        c0.fill = fill; c0.border = BORDER
        c0.font = Font(name="Calibri", bold=True)

        c1 = ws.cell(ri, 2, int(row["Transactions"]))
        c1.fill = fill; c1.border = BORDER
        c1.alignment = Alignment(horizontal="center")

        for ci, key, color in [
            (3, "Withdrawals", "CC0000"),
            (4, "Deposits",    "006600"),
            (5, "Net",         "006600" if row["Net"] >= 0 else "CC0000"),
        ]:
            cn = _num(ws, ri, ci, row[key])
            cn.fill = fill
            cn.font = Font(name="Calibri", color=color)

        ws.cell(ri, 6).fill = fill; ws.cell(ri, 6).border = BORDER

    # totals
    total_r = len(summary) + 3
    ws.cell(total_r, 1, "TOTAL").font = Font(bold=True, name="Calibri")
    ws.cell(total_r, 2, int(summary["Transactions"].sum())).alignment = Alignment(horizontal="center")
    ws.cell(total_r, 2).font = Font(bold=True, name="Calibri")
    for ci, key, color in [(3, "Withdrawals", "CC0000"), (4, "Deposits", "006600")]:
        c = _num(ws, total_r, ci, summary[key].sum())
        c.font = Font(bold=True, name="Calibri", color=color)
    net = summary["Deposits"].sum() - summary["Withdrawals"].sum()
    c = _num(ws, total_r, 5, net)
    c.font = Font(bold=True, name="Calibri", color="006600" if net >= 0 else "CC0000")
    for col in range(1, 7):
        ws.cell(total_r, col).fill   = PatternFill("solid", fgColor=TOTAL_BG)
        ws.cell(total_r, col).border = BOT_MED

    _set_widths(ws, [13, 14, 18, 18, 18, 4])
    ws.freeze_panes = "A3"


# ─── Sheet 3 : Category Summary ──────────────────────────────────────────────
def sheet_category(wb, df: pd.DataFrame):
    ws = wb.create_sheet("By Category")

    cat_sum = df.groupby("category").agg(
        Transactions=("category", "count"),
        Withdrawals=("Withdrawals", lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
        Deposits=("Deposits",       lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
    ).reset_index().sort_values("Withdrawals", ascending=False).reset_index(drop=True)

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "Spending by Category"
    t.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["Category", "Transactions", "Withdrawals (₹)", "Deposits (₹)", ""], 1):
        _hdr_cell(ws, 2, c, h)

    for ri, (_, row) in enumerate(cat_sum.iterrows(), 3):
        cat = str(row["category"]).lower()
        bg  = CAT_COLORS.get(cat, NEUTRAL_BG)
        fill = PatternFill("solid", fgColor=bg)

        c0 = ws.cell(ri, 1, row["category"].title())
        c0.fill = fill; c0.border = BORDER
        c0.font = Font(name="Calibri", bold=True)

        c1 = ws.cell(ri, 2, int(row["Transactions"]))
        c1.fill = fill; c1.border = BORDER
        c1.alignment = Alignment(horizontal="center")

        for ci, key, color in [
            (3, "Withdrawals", "CC0000"),
            (4, "Deposits",    "006600"),
        ]:
            cn = _num(ws, ri, ci, row[key] if row[key] > 0 else None)
            cn.fill = fill
            cn.font = Font(name="Calibri", color=color)

        ws.cell(ri, 5).fill = fill; ws.cell(ri, 5).border = BORDER

    _set_widths(ws, [14, 14, 18, 18, 4])
    ws.freeze_panes = "A3"


# ─── Sheet 4 : UPI Contacts ──────────────────────────────────────────────────
def sheet_upi(wb, df: pd.DataFrame):
    ws = wb.create_sheet("UPI Contacts")

    upi_df = df[df["payer_upi"].str.len() > 0].copy()
    if upi_df.empty:
        ws["A1"] = "No UPI contacts found"
        return

    contacts = upi_df.groupby("payer_upi").agg(
        Transactions=("payer_upi", "count"),
        Total_Paid=("Withdrawals", lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
        Total_Received=("Deposits", lambda s: s.apply(lambda x: x if isinstance(x, float) else 0).sum()),
    ).reset_index().sort_values("Total_Paid", ascending=False).reset_index(drop=True)

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "UPI Contacts — Frequency & Volume"
    t.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["UPI ID", "Transactions", "Total Paid (₹)", "Total Received (₹)", ""], 1):
        _hdr_cell(ws, 2, c, h)

    for ri, (_, row) in enumerate(contacts.iterrows(), 3):
        bg = CREDIT_BG if ri % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        c0 = ws.cell(ri, 1, row["payer_upi"])
        c0.fill = fill; c0.border = BORDER
        c0.font = Font(name="Calibri", size=9)

        c1 = ws.cell(ri, 2, int(row["Transactions"]))
        c1.fill = fill; c1.border = BORDER; c1.alignment = Alignment(horizontal="center")

        for ci, key, color in [
            (3, "Total_Paid",     "CC0000"),
            (4, "Total_Received", "006600"),
        ]:
            cn = _num(ws, ri, ci, row[key] if row[key] > 0 else None)
            cn.fill = fill
            cn.font = Font(name="Calibri", size=9, color=color)
        ws.cell(ri, 5).fill = fill; ws.cell(ri, 5).border = BORDER

    _set_widths(ws, [36, 14, 18, 18, 4])
    ws.freeze_panes = "A3"


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "Statement-124563400000961-03-10-2026-20-15-08 (1)_extracted.xlsx"
    )
    if not src.exists():
        print(f"❌ File not found: {src}")
        return

    print(f"Reading: {src}")
    df = pd.read_excel(src)
    df["payer_upi"] = df["payer_upi"].fillna("").astype(str)
    df["category"]  = df["category"].fillna("other").astype(str)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    sheet_all(wb, df)
    sheet_monthly(wb, df)
    sheet_category(wb, df)
    sheet_upi(wb, df)

    out = src.stem.replace("_extracted", "") + "_formatted.xlsx"
    wb.save(out)

    total_wd  = df["Withdrawals"].apply(lambda x: x if isinstance(x, float) else 0).sum()
    total_dep = df["Deposits"].apply(lambda x: x if isinstance(x, float) else 0).sum()

    print("=" * 55)
    print("  Bank Statement — Formatted Excel")
    print("=" * 55)
    print(f"  File       : {out}")
    print(f"  Rows       : {len(df)}")
    print(f"  Withdrawals: Rs {total_wd:,.2f}")
    print(f"  Deposits   : Rs {total_dep:,.2f}")
    print(f"  Net        : Rs {total_dep - total_wd:,.2f}")
    print()
    print("  Sheets:")
    print("    1. All Transactions  — colour-coded debit/credit")
    print("    2. Monthly Summary   — cash flow by month")
    print("    3. By Category       — spending breakdown")
    print("    4. UPI Contacts      — who you transact with most")
    print("=" * 55)


if __name__ == "__main__":
    main()
