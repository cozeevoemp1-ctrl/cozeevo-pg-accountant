"""One-off: July 2026 classification-review workbook for Kiran.

Sheet 1 — every 'Other Expenses' row (classify: pick category + comment)
Sheet 2 — rows needing a yes/no confirmation (dupes, loans, bike)
Sheet 3 — open questions (Option A/B, cash count, etc.)
Categories sheet feeds the dropdown on Sheet 1.

Run:  venv/Scripts/python scripts/_export_july_review_workbook.py
Out:  data/reports/July_2026_Classification_Review.xlsx
Kiran fills the YOUR CATEGORY / ANSWER columns and hands it back; a follow-up
script applies the rulings + saves each as a permanent pnl_classify rule.
"""
from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from dotenv import load_dotenv
load_dotenv()

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text

from src.database.db_manager import get_session, init_engine
init_engine(os.environ["DATABASE_URL"])

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="332D29")
WRAP = Alignment(wrap_text=True, vertical="top")

CATEGORIES = [
    "Food & Groceries", "Staff & Labour", "Maintenance & Repairs", "Water",
    "Electricity", "Fuel & Diesel", "Furniture & Supplies", "Shopping & Supplies",
    "Cleaning Supplies", "Internet & WiFi", "IT & Software", "Waste Disposal",
    "Govt & Regulatory", "Property Rent", "Bank Charges",
    "Tenant Deposit Refund", "Non-Operating (loan/capital — not P&L)",
    "Other Expenses (leave as misc)",
]


def style_header(ws, row=1):
    for c in ws[row]:
        c.font = HDR
        c.fill = HDR_FILL
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


async def main() -> None:
    async with get_session() as s:
        other = (await s.execute(text(
            "SELECT txn_date, account_name, amount, description FROM bank_transactions "
            "WHERE txn_type='expense' AND category='Other Expenses' "
            "AND txn_date BETWEEN '2026-07-01' AND '2026-07-31' ORDER BY amount DESC"))).all()

    wb = Workbook()

    # ── Sheet 1: classify ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "1 - Classify these"
    ws.append(["Date", "Acct", "Amount", "Bank narration (payee + remark)",
               "YOUR CATEGORY (pick)", "YOUR COMMENT (who is this / what for)"])
    for d, acct, amt, desc in other:
        ws.append([str(d), acct, float(amt), desc, "", ""])
    for col, w in zip("ABCDEF", (11, 6, 10, 78, 34, 40)):
        ws.column_dimensions[col].width = w
    for r in ws.iter_rows(min_row=2):
        r[3].alignment = WRAP
    style_header(ws)

    # category dropdown
    cats = wb.create_sheet("Categories")
    for i, c in enumerate(CATEGORIES, 1):
        cats.cell(row=i, column=1, value=c)
    dv = DataValidation(type="list", formula1=f"=Categories!$A$1:$A${len(CATEGORIES)}",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{ws.max_row}")

    # ── Sheet 2: confirm ─────────────────────────────────────────────────────
    w2 = wb.create_sheet("2 - Confirm these")
    w2.append(["#", "Item", "Amount", "Current treatment", "Question", "YOUR ANSWER"])
    rows2 = [
        (1, "Jalluram paid TWICE Jul-11 (₹20,000 from HULK 'Housekeeping' + ₹20,000 from THOR)", 40000,
         "Both booked Staff & Labour", "Intentional (2 payments due) or double-paid by mistake?", ""),
        (2, "Ninjacart vegetables ₹8,025 TWICE Jul-15 to two different UPI handles (q283663049@ybl and rameshy1503@ybl)", 16050,
         "Both booked Food & Groceries", "Two real orders, or wrong-payee-then-repaid (one should be recovered/reversed)?", ""),
        (3, "chandrasekhar1996krish@oksbi — ₹90,000 OUT Jul-17 ('Hand Loan') / ₹79,900 IN Jul-29", 90000,
         "Both moved to Non-Operating loan (out/in), net ₹10,100 owed", "WHO is this — Chandra? Which loan-register account? (Same handle also got ₹20K Jan + ₹3K Feb labeled refunds in frozen months)", ""),
        (4, "G Ravikumar NEFT Jul-14 narration 'hand Loan'", 200000,
         "Non-Operating hand loan", "Confirm: new 5th loan-register account 'G Ravikumar' ₹2,00,000 outstanding?", ""),
        (5, "TVS 110cc Jupiter bike (+₹872 related payment)", 50872,
         "Furniture & Supplies (asset from PG account = opex, CAPEX abolished)", "OK as PG staff vehicle expense? Or personal (→ capital/drawings)?", ""),
        (6, "Inar Devi NEFT Jul-17", 7300,
         "Other Expenses / Unclassified", "Who is Inar Devi? (also pick category on Sheet 1)", ""),
    ]
    for r in rows2:
        w2.append(list(r))
    for col, w in zip("ABCDEF", (4, 52, 10, 34, 46, 30)):
        w2.column_dimensions[col].width = w
    for r in w2.iter_rows(min_row=2):
        for c in (r[1], r[3], r[4]):
            c.alignment = WRAP
    style_header(w2)

    # ── Sheet 3: open questions ──────────────────────────────────────────────
    w3 = wb.create_sheet("3 - Questions")
    w3.append(["#", "Question", "Options / context", "YOUR ANSWER"])
    rows3 = [
        (1, "Refund accounting — Option A or B?",
         "A = subtract deposit ONCE when collected; refunds become display-only; add 'Deposits forfeited' income line at exit (profit rises, cleaner books). "
         "B = keep current double subtraction (most conservative, understates profit).", ""),
        (2, "Cash line basis for dynamic months",
         "Approved rule A says cash income line = ALL cash received (rent+deposit+advance). Currently rent-only. "
         "July: rent-only 27,96,350 vs all-cash 28,31,250 (+ your offline 67,850 = your 28,99,100). Switch to all-cash + offline?", ""),
        (3, "July offline cash ₹67,850 (your 28,99,100 − app 28,31,250)",
         "Who paid this cash outside the app? (name/room/amount so it can be logged)", ""),
        (4, "Physical cash count on Jul-31 night",
         "Needed for the 'Cash in hand' line. Current stored figure 2,21,559 is from Jul-4.", ""),
        (5, "Chit cadence from Aug",
         "Balaji ₹3L/15mo + Boopalan + Tanvi ₹36K/5mo — monthly recurring from August? Modes?", ""),
        (6, "Internet accrual for May/Jun/Jul",
         "SOP books ₹15,514/mo (Airwire prepaid). Bank shows ~0 in dynamic months. Add the accrual line to May/Jun/Jul?", ""),
        (7, "July staff salaries look low (bank ₹1.42L vs ~₹1.9L normal)",
         "Were some July salaries paid in cash? If yes they belong in the cash-expense figure (currently only ₹2,000).", ""),
    ]
    for r in rows3:
        w3.append(list(r))
    for col, w in zip("ABCD", (4, 44, 80, 30)):
        w3.column_dimensions[col].width = w
    for r in w3.iter_rows(min_row=2):
        r[1].alignment = WRAP
        r[2].alignment = WRAP
    style_header(w3)

    out = Path("data/reports/July_2026_Classification_Review.xlsx")
    wb.save(out)
    print(f"saved {out} — {len(other)} rows to classify, {len(rows2)} confirmations, {len(rows3)} questions")


if __name__ == "__main__":
    asyncio.run(main())
