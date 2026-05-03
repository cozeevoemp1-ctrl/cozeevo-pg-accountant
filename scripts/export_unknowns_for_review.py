"""
scripts/export_unknowns_for_review.py
--------------------------------------
Export remaining "Other Expenses" transactions grouped by payee for
Kiran's yes/no classification review.

Sheet 1 — Summary: one row per unique payee (UPI handle / name),
          sorted by total amount desc. Includes "Is Deposit Refund?" column.
Sheet 2 — Detail: every transaction, sorted by payee then date.

Run:
  venv/Scripts/python scripts/export_unknowns_for_review.py
"""
import sys, os, glob as _glob, re
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

from src.parsers.yes_bank import parse_date, parse_amt, read_yes_bank_csv
from src.rules.pnl_classify import classify_txn
from src.utils.inr_format import INR_NUMBER_FORMAT

# ── Description parser (same as other_expenses_detail.py) ────────────────────
UPI_RE = re.compile(r"[a-zA-Z0-9.\-_+]+@[a-zA-Z]+")

def parse_desc(desc: str) -> dict:
    result = {"txn_type": "", "utr_ref": "", "payee_name": "", "payee_upi": "", "note": ""}
    if not desc or not isinstance(desc, str):
        return result
    d = desc.strip()
    dl = d.lower()
    if "upi" in dl:   result["txn_type"] = "UPI"
    elif "neft" in dl: result["txn_type"] = "NEFT"
    elif "imps" in dl: result["txn_type"] = "IMPS"
    elif "rtgs" in dl: result["txn_type"] = "RTGS"
    elif "atm" in dl or "cash" in dl: result["txn_type"] = "CASH"

    if result["txn_type"] == "UPI" and "/" in d:
        parts = [p.strip() for p in d.split("/")]
        if len(parts) >= 2 and re.match(r"^\d{10,16}$", parts[1]):
            result["utr_ref"] = parts[1]
            if len(parts) > 2: result["payee_name"] = parts[2]
            if len(parts) > 3 and "@" in parts[3]: result["payee_upi"] = parts[3]
            if len(parts) > 4: result["note"] = "/".join(parts[4:])
    elif result["txn_type"] == "NEFT":
        yib_m = re.match(r"(?:YIB-NEFT|NET-NEFT|YESOB)-(\w+)[-/](.+)$", d, re.IGNORECASE)
        if yib_m:
            result["utr_ref"] = yib_m.group(1)
            result["payee_name"] = yib_m.group(2).strip()

    if not result["payee_upi"]:
        upis = UPI_RE.findall(d)
        if upis:
            result["payee_upi"] = upis[-1]

    return result


def _payee_key(parsed: dict) -> str:
    """Canonical key for grouping: prefer UPI handle, fallback to payee name."""
    upi = (parsed["payee_upi"] or "").strip()
    if upi and upi.lower() not in ("", "null"):
        return upi.lower()
    name = (parsed["payee_name"] or "").strip()
    return name.lower() if name else "(no payee)"


# ── Load bank statements ──────────────────────────────────────────────────────
txns = []
csv_files = sorted(
    set(
        _glob.glob("Statement-*.csv")
        + _glob.glob("*statment*.csv")
        + _glob.glob("*statement*.csv")
        + _glob.glob("april month.csv")
        + _glob.glob("april*.csv")
    ),
    reverse=True,
)
for f in csv_files:
    t = read_yes_bank_csv(f)
    print(f"  {len(t):>4}  {f}")
    txns += t

for f in ["2025 statement.xlsx", "2026 statment.xlsx"]:
    if not os.path.exists(f):
        continue
    import openpyxl as _opx
    wb_s = _opx.load_workbook(f, data_only=True)
    ws_s = wb_s.active
    rows_s = list(ws_s.iter_rows(values_only=True))
    hrow = None
    for i, row in enumerate(rows_s):
        if row and any("transaction date" in str(c).lower() for c in row if c):
            hrow = i; break
    if hrow is not None:
        headers = [str(c).lower().strip() if c else "" for c in rows_s[hrow]]
        for row in rows_s[hrow + 1:]:
            if not any(row): continue
            d2  = dict(zip(headers, row))
            dt  = parse_date(d2.get("transaction date"))
            if not dt: continue
            wd  = parse_amt(d2.get("withdrawals", ""))
            dep = parse_amt(d2.get("deposits", ""))
            desc2 = str(d2.get("description") or "")
            if wd  > 0: txns.append((dt, desc2, "expense", wd))
            if dep > 0: txns.append((dt, desc2, "income",  dep))
    print(f"  Loaded {f}")

# Dedupe
seen = set(); deduped = []
for item in txns:
    dt, desc, typ, amt = item
    key = (dt.strftime("%Y-%m-%d"), round(float(amt), 2), (desc or "").strip().lower())
    if key not in seen:
        seen.add(key); deduped.append(item)
txns = deduped
print(f"\nTotal after dedupe: {len(txns)}")

# ── Filter to Other Expenses ──────────────────────────────────────────────────
other = []
for dt, desc, typ, amt in txns:
    if typ != "expense": continue
    cat, sub = classify_txn(desc, typ)
    if cat == "Other Expenses":
        p = parse_desc(desc)
        other.append({
            "date":        dt,
            "month":       dt.strftime("%b'%y"),
            "amount":      float(amt),
            "txn_type":    p["txn_type"],
            "payee_name":  p["payee_name"],
            "payee_upi":   p["payee_upi"],
            "note":        p["note"],
            "description": desc,
            "_key":        _payee_key(p),
        })

other.sort(key=lambda x: x["date"])
print(f"Other Expenses (after updated classifier): {len(other)} txns  Rs {sum(r['amount'] for r in other):,.0f}")

# ── Group by payee ────────────────────────────────────────────────────────────
groups: dict[str, list] = defaultdict(list)
for row in other:
    groups[row["_key"]].append(row)

summary = []
for key, rows in groups.items():
    dates  = sorted(r["date"] for r in rows)
    total  = sum(r["amount"] for r in rows)
    months = sorted({r["month"] for r in rows})
    sample = rows[0]
    summary.append({
        "payee_upi":   sample["payee_upi"] or key,
        "payee_name":  sample["payee_name"],
        "txn_count":   len(rows),
        "total_amt":   total,
        "first_date":  dates[0].strftime("%d %b %Y"),
        "last_date":   dates[-1].strftime("%d %b %Y"),
        "months":      ", ".join(months),
        "sample_note": sample["note"],
        "is_deposit_refund": "",   # blank — Kiran fills
        "category":    "",         # blank — Kiran fills if vendor
        "_rows":       rows,
    })

summary.sort(key=lambda x: -x["total_amt"])

print(f"\nUnique payees: {len(summary)}")
for s in summary:
    print(f"  Rs {s['total_amt']:>8,.0f}  x{s['txn_count']}  {s['payee_upi']:<35}  {s['payee_name']}")

# ── Excel styling helpers ─────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL  = PatternFill("solid", fgColor="EEF4FF")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
TOT_FILL  = PatternFill("solid", fgColor="D9D9D9")
GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")
YLW_FILL  = PatternFill("solid", fgColor="FFF2CC")
CTR       = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT     = Alignment(horizontal="right",  vertical="center")
NUM_FMT   = INR_NUMBER_FORMAT

def hdr(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    return c

def amt_cell(ws, row, col, val, fill=None):
    c = ws.cell(row, col, val)
    c.number_format = NUM_FMT
    c.font = Font(size=9, bold=False)
    c.alignment = RIGHT
    if fill: c.fill = fill
    return c

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)
ws1 = wb.create_sheet("Summary by Payee")

SUM_COLS = [
    ("#",                    4,  "center"),
    ("Payee UPI Handle",    30,  "left"),
    ("Payee Name",          25,  "left"),
    ("Txns",                 6,  "center"),
    ("Total Amount (Rs)",   16,  "right"),
    ("First Date",          13,  "center"),
    ("Last Date",           13,  "center"),
    ("Months",              22,  "left"),
    ("Sample Note",         30,  "left"),
    ("Is Deposit Refund? (Y/N)", 22, "center"),
    ("If vendor — Category",     24, "left"),
]

# Title
nc = len(SUM_COLS)
ws1.merge_cells(f"A1:{get_column_letter(nc)}1")
t = ws1["A1"]
t.value = f"Other Expenses — Payee Summary for Review ({len(summary)} payees  Rs {sum(s['total_amt'] for s in summary):,.0f})"
t.font = Font(bold=True, size=12, color="FFFFFF"); t.fill = HDR_FILL; t.alignment = CTR
ws1.row_dimensions[1].height = 22

# Headers
for ci, (h, w, _) in enumerate(SUM_COLS, 1):
    hdr(ws1, 2, ci, h)
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[2].height = 22

grand = 0.0
for ri, s in enumerate(summary, 1):
    r = ri + 2
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    vals = [ri, s["payee_upi"], s["payee_name"], s["txn_count"], s["total_amt"],
            s["first_date"], s["last_date"], s["months"], s["sample_note"],
            s["is_deposit_refund"], s["category"]]
    for ci, (val, (_, _, align)) in enumerate(zip(vals, SUM_COLS), 1):
        c = ws1.cell(r, ci, val)
        c.fill = fill
        if ci == 5:
            c.number_format = NUM_FMT
            c.font = Font(size=9, color="CC0000"); c.alignment = RIGHT
        elif ci == 10:  # Is Deposit Refund?
            c.fill = YLW_FILL
            c.font = Font(size=10, bold=True); c.alignment = CTR
        elif ci == 11:  # Category
            c.fill = GRN_FILL
            c.font = Font(size=9); c.alignment = LEFT
        else:
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal=align, vertical="center")
    grand += s["total_amt"]

# Total row
tr = len(summary) + 3
for ci in range(1, nc + 1):
    ws1.cell(tr, ci).fill = TOT_FILL
ws1.cell(tr, 1, "TOTAL").font = Font(bold=True)
ws1.cell(tr, 1).fill = TOT_FILL
c = ws1.cell(tr, 5, grand)
c.number_format = NUM_FMT; c.font = Font(bold=True, color="CC0000"); c.alignment = RIGHT; c.fill = TOT_FILL

ws1.freeze_panes = "A3"
ws1.auto_filter.ref = f"A2:{get_column_letter(nc)}2"

# ── Sheet 2: Detail ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("All Transactions (Detail)")

DET_COLS = [
    ("#",             4,  "center"),
    ("Payee UPI",    30,  "left"),
    ("Payee Name",   24,  "left"),
    ("Date",         12,  "center"),
    ("Month",         9,  "center"),
    ("Amount (Rs)",  14,  "right"),
    ("Txn Type",      8,  "center"),
    ("Note",         28,  "left"),
    ("Full Description", 70, "left"),
]

nc2 = len(DET_COLS)
ws2.merge_cells(f"A1:{get_column_letter(nc2)}1")
t2 = ws2["A1"]
t2.value = f"Other Expenses — All Transactions Detail ({len(other)} txns)"
t2.font = Font(bold=True, size=12, color="FFFFFF"); t2.fill = HDR_FILL; t2.alignment = CTR
ws2.row_dimensions[1].height = 22

for ci, (h, w, _) in enumerate(DET_COLS, 1):
    hdr(ws2, 2, ci, h)
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 22

# Sort detail by total_amt of group desc, then date
detail_sorted = []
for s in summary:
    detail_sorted.extend(sorted(s["_rows"], key=lambda x: x["date"]))

grand2 = 0.0
for ri, row in enumerate(detail_sorted, 1):
    r = ri + 2
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    vals = [ri, row["payee_upi"], row["payee_name"],
            row["date"].strftime("%d %b %Y"), row["month"], row["amount"],
            row["txn_type"], row["note"], row["description"]]
    for ci, (val, (_, _, align)) in enumerate(zip(vals, DET_COLS), 1):
        c = ws2.cell(r, ci, val)
        c.fill = fill
        if ci == 6:
            c.number_format = NUM_FMT; c.font = Font(size=9, color="CC0000"); c.alignment = RIGHT
        else:
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align=="left" and ci==9))
    grand2 += row["amount"]

tr2 = len(detail_sorted) + 3
for ci in range(1, nc2 + 1):
    ws2.cell(tr2, ci).fill = TOT_FILL
ws2.cell(tr2, 1, "TOTAL").font = Font(bold=True); ws2.cell(tr2, 1).fill = TOT_FILL
c2 = ws2.cell(tr2, 6, grand2)
c2.number_format = NUM_FMT; c2.font = Font(bold=True, color="CC0000"); c2.alignment = RIGHT; c2.fill = TOT_FILL

ws2.freeze_panes = "A3"
ws2.auto_filter.ref = f"A2:{get_column_letter(nc2)}2"

# ── Save ──────────────────────────────────────────────────────────────────────
OUT = os.path.join("data", "reports", "other_expenses_unknowns_review.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"  Summary sheet: {len(summary)} payees")
print(f"  Detail sheet:  {len(detail_sorted)} transactions  Rs {grand2:,.0f}")
