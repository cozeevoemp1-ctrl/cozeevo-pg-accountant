"""Clean reload: wipe all active/noshow tenancies, reload from Excel."""
import sys, os, re, asyncio
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime, date
from decimal import Decimal
from dotenv import load_dotenv
import openpyxl

from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

load_dotenv()

from src.database.models import Tenancy, Tenant, Room, TenancyStatus

DATABASE_URL = os.environ["DATABASE_URL"]
EXCEL_FILE = "April Month Collection.xlsx"


async def main():
    url = DATABASE_URL
    if url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+asyncpg://", 1)
    engine = create_async_engine(url, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    # Read Excel — active + noshow only
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Long term']

    excel_rows = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name or not str(name).strip():
            continue
        status_raw = str(ws.cell(row, 17).value or '').strip().upper()
        if status_raw == 'EXIT' or 'CANCEL' in status_raw:
            continue

        room = str(ws.cell(row, 1).value or '').strip()
        if room.endswith('.0'):
            room = room[:-2]
        sharing = str(ws.cell(row, 13).value or '').strip().lower()
        phone_raw = str(ws.cell(row, 4).value or '').strip()
        if phone_raw.endswith('.0'):
            phone_raw = phone_raw[:-2]
        gender = str(ws.cell(row, 3).value or '').strip().lower()
        checkin = ws.cell(row, 5).value

        rent_m = ws.cell(row, 10).value
        rent_f = ws.cell(row, 11).value
        rent_y = ws.cell(row, 12).value
        rm = float(rent_m or 0) if isinstance(rent_m, (int, float)) else 0
        rf = float(rent_f or 0) if isinstance(rent_f, (int, float)) else 0
        ry = float(rent_y or 0) if isinstance(rent_y, (int, float)) else 0
        rent = ry if ry > 0 else (rf if rf > 0 else rm)

        deposit = ws.cell(row, 7).value
        booking = ws.cell(row, 6).value
        maintenance = ws.cell(row, 8).value

        is_noshow = 'NO SHOW' in status_raw
        if sharing in ('premium', 'single ac', 'single a/c'):
            sharing_type = 'premium'
        elif sharing in ('single',):
            sharing_type = 'single'
        elif sharing in ('triple', '3'):
            sharing_type = 'triple'
        else:
            sharing_type = 'double'

        checkin_date = None
        if isinstance(checkin, datetime):
            checkin_date = checkin.date()
        elif isinstance(checkin, date):
            checkin_date = checkin

        excel_rows.append({
            'name': str(name).strip(),
            'room': room,
            'phone': phone_raw,
            'gender': 'female' if gender == 'female' else 'male',
            'sharing': sharing_type,
            'noshow': is_noshow,
            'checkin': checkin_date,
            'rent': rent,
            'deposit': float(deposit or 0) if isinstance(deposit, (int, float)) else 0,
            'booking': float(booking or 0) if isinstance(booking, (int, float)) else 0,
            'maintenance': float(maintenance or 0) if isinstance(maintenance, (int, float)) else 0,
        })

    print(f"Excel: {len(excel_rows)} active+noshow rows")

    async with Session() as session:
        # Step 1: Set ALL active/noshow to exited
        result = await session.execute(
            text("UPDATE tenancies SET status = 'exited' WHERE status IN ('active', 'no_show')")
        )
        print(f"Step 1: Set {result.rowcount} tenancies to exited")

        # Step 2: Room map
        rooms = {r.room_number: r for r in (await session.execute(select(Room))).scalars().all()}

        # Step 3: Process each Excel row
        matched = 0
        created_t = 0
        created_tn = 0
        errors = []

        for rec in excel_rows:
            room = rooms.get(rec['room'])
            if not room:
                errors.append(f"{rec['room']} {rec['name']} - room not found")
                continue

            # Find tenant by name
            tenant = await session.scalar(
                select(Tenant).where(Tenant.name.ilike(rec['name']))
            )
            if not tenant:
                digits = re.sub(r'\D', '', rec['phone'])
                if len(digits) >= 10:
                    phone_search = digits[-10:]
                    tenant = await session.scalar(
                        select(Tenant).where(Tenant.phone.like(f'%{phone_search}'))
                    )

            if not tenant:
                digits = re.sub(r'\D', '', rec['phone'])
                if digits.startswith('91') and len(digits) == 12:
                    digits = digits[2:]
                if len(digits) == 10 and digits[0] in '6789':
                    phone_db = f'+91{digits}'
                else:
                    safe = re.sub(r'[^a-zA-Z0-9]', '', rec['name'])[:12]
                    phone_db = f'NOPHONE_{rec["room"]}_{safe}'

                existing = await session.scalar(select(Tenant).where(Tenant.phone == phone_db))
                if existing:
                    safe = re.sub(r'[^a-zA-Z0-9]', '', rec['name'])[:12]
                    phone_db = f'NOPHONE_{rec["room"]}_{safe}'
                    existing2 = await session.scalar(select(Tenant).where(Tenant.phone == phone_db))
                    if existing2:
                        phone_db = f'{phone_db}_{int(datetime.now().timestamp()) % 10000}'

                tenant = Tenant(name=rec['name'], phone=phone_db, gender=rec['gender'])
                session.add(tenant)
                await session.flush()
                created_t += 1

            # Find tenancy for this tenant+room
            tenancy = await session.scalar(
                select(Tenancy).where(
                    Tenancy.tenant_id == tenant.id,
                    Tenancy.room_id == room.id,
                ).order_by(Tenancy.id.desc()).limit(1)
            )

            new_status = TenancyStatus.no_show if rec['noshow'] else TenancyStatus.active

            if tenancy:
                tenancy.status = new_status
                tenancy.sharing_type = rec['sharing']
                tenancy.agreed_rent = Decimal(str(rec['rent']))
                tenancy.security_deposit = Decimal(str(rec['deposit']))
                tenancy.booking_amount = Decimal(str(rec['booking']))
                tenancy.maintenance_fee = Decimal(str(rec['maintenance']))
                if rec['checkin']:
                    tenancy.checkin_date = rec['checkin']
                matched += 1
            else:
                tenancy = Tenancy(
                    tenant_id=tenant.id,
                    room_id=room.id,
                    status=new_status,
                    sharing_type=rec['sharing'],
                    checkin_date=rec['checkin'],
                    agreed_rent=Decimal(str(rec['rent'])),
                    security_deposit=Decimal(str(rec['deposit'])),
                    booking_amount=Decimal(str(rec['booking'])),
                    maintenance_fee=Decimal(str(rec['maintenance'])),
                    stay_type='monthly',
                )
                session.add(tenancy)
                created_tn += 1

            await session.flush()

        await session.commit()

        # Verify
        active = await session.scalar(select(func.count()).select_from(Tenancy).where(Tenancy.status == TenancyStatus.active)) or 0
        noshow = await session.scalar(select(func.count()).select_from(Tenancy).where(Tenancy.status == TenancyStatus.no_show)) or 0
        premium = await session.scalar(select(func.count()).select_from(Tenancy).where(
            Tenancy.status == TenancyStatus.active, Tenancy.sharing_type == 'premium')) or 0
        regular = active - premium
        beds = regular + (premium * 2) + noshow

        print(f"\nResults:")
        print(f"  Matched (reactivated): {matched}")
        print(f"  Created tenants: {created_t}")
        print(f"  Created tenancies: {created_tn}")
        if errors:
            print(f"  Errors ({len(errors)}):")
            for e in errors:
                print(f"    {e}")

        print(f"\n{'=' * 50}")
        print(f"DB:    {active} active ({regular} reg + {premium} prem) + {noshow} noshow")
        print(f"Beds:  {regular} + {premium*2} + {noshow} = {beds}")
        print(f"Vacant: 291 - {beds} = {291 - beds}")
        print(f"")
        print(f"Excel: 242 active (221 reg + 21 prem) + 13 noshow")
        print(f"Beds:  221 + 42 + 13 = 276")
        print(f"Vacant: 15")

    await engine.dispose()


asyncio.run(main())
