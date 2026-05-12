"""
One-off: export remaining Other Expenses + Volipi entries to Excel for Kiran to classify.
Run: venv/Scripts/python scripts/_export_other_expenses.py
"""
import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Raw data from DB ──────────────────────────────────────────────────────────

OTHER_EXPENSES = [
    (2617,"2025-12-02",50000.00,"CHANDRASEKHAR — Pg expenses","T2512021551086615341630","LAKSHMI_SBI"),
    (1174,"2026-04-08",49679.00,"UPI/300397072066/From:7358341775@ptyes/To:8951297583-3@ibl/Sent using Paytm UPI",None,"THOR"),
    (2412,"2025-12-08",24394.00,"UPI/372002673570/From:7358341775-2@ybl/To:arunphilip25@okicici/Payment from PhonePe",None,"THOR"),
    (1539,"2026-03-30",22000.00,"UPI/399840154458/From:7358341775@ptyes/To:9904388966-2@ybl/Sent using Paytm UPI",None,"THOR"),
    (2630,"2025-12-23",20000.00,"CHANDRASEKHAR — 1 Lakh spent by akhil for expenses","T2512231610387753187838","LAKSHMI_SBI"),
    (2320,"2025-12-21",16963.00,"UPI/155722169773/From:7358341775-2@ybl/To:M036TPQEK@ybl/Payment for 22dd8da925f54edb9f6c582f6fa58918",None,"THOR"),
    (2461,"2025-12-02",14950.00,"UPI/214515371994/From:7358341775-2@ybl/To:tpasha638@ybl/Payment from PhonePe",None,"THOR"),
    (1935,"2026-02-11",10000.00,"UPI/643083690694/From:7358341775-2@ybl/To:6202601070@ptsbi/Payment from PhonePe",None,"THOR"),
    (2470,"2025-12-02",9048.00,"UPI/303630701161/From:7358341775-2@ybl/To:tpasha638@ybl/Payment from PhonePe",None,"THOR"),
    (2481,"2025-12-01",9000.00,"UPI/682112034807/From:7358341775-2@ybl/To:bn895975@okicici/Payment from PhonePe",None,"THOR"),
    (1565,"2026-03-27",9000.00,"UPI/601673068816/From:7358341775@ptyes/To:paytmqr6wguqk@ptys/Verified Paytm Account",None,"THOR"),
    (2557,"2025-11-08",5000.00,"UPI/246566575161/From:7358341775-2@ybl/To:7667662128@ybl/Payment from PhonePe",None,"THOR"),
    (1943,"2026-02-10",4000.00,"UPI/424535655397/From:7358341775-2@ybl/To:sk5008786@ybl/Payment from PhonePe",None,"THOR"),
    (1934,"2026-02-11",4000.00,"UPI/334469752391/From:7358341775-2@ybl/To:6287677379@ybl/Payment from PhonePe",None,"THOR"),
    (2366,"2025-12-13",3200.00,"UPI/865529892235/From:7358341775-2@ybl/To:8073343903@kotak/Payment from PhonePe",None,"THOR"),
    (1146,"2026-04-11",2970.00,"UPI/602579149300/From:7358341775@ptyes/To:9148809732@ptaxis/Sent from Paytm",None,"THOR"),
    (1558,"2026-03-28",2100.00,"UPI/399726009762/From:7358341775@ptyes/To:9448259556@axl/Sent using Paytm UPI",None,"THOR"),
    (1670,"2026-03-13",2000.00,"UPI/398849152083/From:7358341775@ptyes/To:8310018565@ibl/Sent using Paytm UPI",None,"THOR"),
    (1037,"2026-04-25",2000.00,"UPI/301399176107/From:7358341775@ptyes/To:9989000250-2@axl/Sent using Paytm UPI",None,"THOR"),
    (1732,"2026-03-07",1925.00,"UPI/398500901601/From:7358341775@ptyes/To:9448259556@axl/Sent using Paytm UPI",None,"THOR"),
    (1710,"2026-03-10",1855.00,"UPI/201671032324/From:7358341775@ptyes/To:9989000250-6@ybl/Sent using Paytm UPI",None,"THOR"),
    (2374,"2025-12-13",1800.00,"UPI/836892507554/From:7358341775-2@ybl/To:8073343903@kotak/Payment from PhonePe",None,"THOR"),
    (2275,"2026-01-01",1800.00,"UPI/919998215483/From:7358341775-2@ybl/To:9146500827@ptyes/Payment from PhonePe",None,"THOR"),
    (1863,"2026-02-20",1464.00,"UPI/827763283504/From:7358341775-2@ybl/To:phegade08-2@okaxis/Payment from PhonePe",None,"THOR"),
    (2464,"2025-12-02",1400.00,"UPI/399139773744/From:7358341775-2@ybl/To:9148809732-2@axl/Payment from PhonePe",None,"THOR"),
    (1002,"2026-04-30",1350.00,"UPI/301696250278/From:7358341775@ptyes/To:9148809732@ptaxis/Sent from Paytm",None,"THOR"),
    (2321,"2025-12-20",1200.00,"UPI/070694910736/From:7358341775-2@ybl/To:9160224731@yescred/Payment from PhonePe",None,"THOR"),
    (1806,"2026-02-27",1200.00,"UPI/600032754033/From:7358341775@ptyes/To:9148809732@ptaxis/Sent from Paytm",None,"THOR"),
    (1671,"2026-03-13",1042.00,"UPI/600856521282/From:7358341775@ptyes/To:9902278720@axl/Sent using Paytm UPI",None,"THOR"),
    (1259,"2026-04-01",945.18,"UPI/203058174029/From:7358341775@ptyes/To:paybil3066@ptybl/Sent using Paytm UPI",None,"THOR"),
    (2075,"2026-01-26",940.00,"UPI/555114054085/From:7358341775-2@ybl/To:9663049651@ybl/Payment from PhonePe",None,"THOR"),
    (1214,"2026-04-05",900.00,"UPI/300208311624/From:7358341775@ptyes/To:9148809732@ptaxis/Sent from Paytm",None,"THOR"),
    (1996,"2026-02-03",830.18,"UPI/455844549167/From:7358341775-2@ybl/To:SV2512112238344230219611@ybl/Payment from PhonePe",None,"THOR"),
    (1705,"2026-03-11",827.18,"UPI/201688822041/From:7358341775@ptyes/To:paybil3066@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1566,"2026-03-27",826.00,"UPI/399658018512/From:7358341775@ptyes/To:kanjichoudhari170-3@okaxis/Sent using Paytm UPI",None,"THOR"),
    (1036,"2026-04-25",775.00,"UPI/204747995227/From:7358341775@ptyes/To:9989000250-2@axl/Sent using Paytm UPI",None,"THOR"),
    (1654,"2026-03-15",750.00,"UPI/398982112810/From:7358341775@ptyes/To:shahnawazlaskar362@okhdfcbank/Sent using Paytm UPI",None,"THOR"),
    (1031,"2026-04-25",700.00,"UPI/204762444153/From:7358341775@ptyes/To:9448259556@axl/Sent using Paytm UPI",None,"THOR"),
    (1639,"2026-03-16",680.00,"UPI/399059012084/From:7358341775@ptyes/To:9632460361@ybl/Sent using Paytm UPI",None,"THOR"),
    (1619,"2026-03-20",616.00,"UPI/202279273290/From:7358341775@ptyes/To:paytm-56505013@ptybl/Sent using Paytm UPI",None,"THOR"),
    (2194,"2026-01-09",562.00,"UPI/521173744075/From:7358341775-2@ybl/To:Q531107921@ybl/Transaction",None,"THOR"),
    (2312,"2025-12-22",560.00,"UPI/835734356071/From:7358341775-2@ybl/To:7829264915@ybl/Payment from PhonePe",None,"THOR"),
    (2297,"2025-12-25",500.00,"UPI/068274570824/From:7358341775-2@ybl/To:8310048369@ybl/Payment from PhonePe",None,"THOR"),
    (1818,"2026-02-26",500.00,"UPI/397972902534/From:7358341775@ptyes/To:8088826094@nyes/Sent using Paytm UPI",None,"THOR"),
    (1689,"2026-03-11",429.00,"UPI/201741711963/From:7358341775@ptyes/To:9108675353@ybl/Sent using Paytm UPI",None,"THOR"),
    (2096,"2026-01-22",407.00,"UPI/952198211108/From:7358341775-2@ybl/To:paytm-56505013@ptybl/Payment from PhonePe",None,"THOR"),
    (1127,"2026-04-13",400.00,"UPI/203939647448/From:7358341775@ptyes/To:9148809732-2@axl/Sent using Paytm UPI",None,"THOR"),
    (1918,"2026-02-12",389.20,"UPI/121496926025/From:7358341775-2@ybl/To:paytm-64646105@ptybl/Payment for UPI Autopay",None,"THOR"),
    (2154,"2026-01-14",377.00,"UPI/297045465747/From:7358341775-2@ybl/To:paytm-56505013@ptybl/Payment from PhonePe",None,"THOR"),
    (1231,"2026-04-04",330.00,"UPI/203276162215/From:7358341775@ptyes/To:8105769898@kotak811/Sent using Paytm UPI",None,"THOR"),
    (1586,"2026-03-25",326.00,"UPI/399532293305/From:7358341775@ptyes/To:paperandpie.63347339@hdfcbank/Sent using Paytm UPI",None,"THOR"),
    (1527,"2026-03-31",300.80,"UPI/203035154345/From:7358341775@ptyes/To:payair7673@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1164,"2026-04-10",300.80,"UPI/203676998976/From:7358341775@ptyes/To:payair7673@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1119,"2026-04-14",300.80,"UPI/204009001787/From:7358341775@ptyes/To:paybil3066@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1778,"2026-03-02",288.00,"UPI/201105316972/From:7358341775@ptyes/To:paytm-56505013@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1610,"2026-03-21",286.00,"UPI/202390382176/From:7358341775@ptyes/To:paytm-56505013@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1679,"2026-03-12",281.00,"UPI/398801549133/From:7358341775@ptyes/To:shahbaz80508637@oksbi/Sent using Paytm UPI",None,"THOR"),
    (2315,"2025-12-22",250.00,"UPI/509132190620/From:7358341775-2@ybl/To:premstealer1@ybl/Payment from PhonePe",None,"THOR"),
    (1217,"2026-04-05",250.00,"UPI/300193763539/From:7358341775@ptyes/To:akshyarathna168@okhdfcbank/Sent using Paytm UPI",None,"THOR"),
    (2173,"2026-01-11",247.00,"UPI/137250300149/From:7358341775-2@ybl/To:paytm-56505013@ptybl/Payment from PhonePe",None,"THOR"),
    (2164,"2026-01-12",231.00,"UPI/949841552044/From:7358341775-2@ybl/To:paytm-56505013@ptybl/Payment from PhonePe",None,"THOR"),
    (1881,"2026-02-17",231.00,"UPI/397480037136/From:7358341775@ptyes/To:yashumeena080@oksbi/Sent using Paytm UPI",None,"THOR"),
    (2425,"2025-12-06",204.00,"UPI/637380742262/From:7358341775-2@ybl/To:Q424363623@ybl/non plu",None,"THOR"),
    (1984,"2026-02-04",200.00,"UPI/603536803354/From:917358341775@waaxis/To:paytmqr2810050501010j1n5l1atjq0@paytm/UPI",None,"THOR"),
    (1852,"2026-02-21",200.00,"UPI/786037856686/From:7358341775-2@ybl/To:6909620430@ybl/Payment from PhonePe",None,"THOR"),
    (1042,"2026-04-25",175.00,"UPI/301385799720/From:7358341775@ptyes/To:7981501263@ybl/Sent using Paytm UPI",None,"THOR"),
    (1683,"2026-03-12",150.00,"UPI/201763433264/From:7358341775@ptyes/To:8795514149@ibl/Sent using Paytm UPI",None,"THOR"),
    (1872,"2026-02-18",144.00,"UPI/200423124044/From:7358341775@ptyes/To:paytm-56505013@ptybl/Sent using Paytm UPI",None,"THOR"),
    (1932,"2026-02-11",100.00,"UPI/091835554372/From:7358341775-2@ybl/To:8197867322-3@ybl/Payment from PhonePe",None,"THOR"),
    (1722,"2026-03-08",100.00,"UPI/398557285867/From:7358341775@ptyes/To:sujanmallik8348@oksbi/Sent using Paytm UPI",None,"THOR"),
    (1708,"2026-03-11",100.00,"UPI/398706438372/From:7358341775@ptyes/To:anudeepbishtab@okicici/Sent using Paytm UPI",None,"THOR"),
    (1216,"2026-04-05",100.00,"UPI/203349936890/From:7358341775@ptyes/To:7061052050@axisbank/Sent using Paytm UPI",None,"THOR"),
    (2418,"2025-12-08",80.00,"UPI/565151803977/From:7358341775-2@ybl/To:jaydevjena73@oksbi/Payment from PhonePe",None,"THOR"),
    (1601,"2026-03-22",60.00,"UPI/399387956981/From:7358341775@ptyes/To:8816019354@ptsbi/Sent from Paytm",None,"THOR"),
    (1765,"2026-03-03",50.00,"UPI/398258703330/From:7358341775@ptyes/To:premstealer1@ybl/Sent using Paytm UPI",None,"THOR"),
    (1110,"2026-04-14",20.00,"UPI/204021021978/From:7358341775@ptyes/To:q937051021@ybl/water",None,"THOR"),
    (2417,"2025-12-08",7.00,"UPI/123557107818/From:7358341775-2@ybl/To:jaydevjena73@oksbi/Payment from PhonePe",None,"THOR"),
]

VOLIPI = [
    (1685,"2026-03-12",13000.00,"UPI/847812545326/From:7358341775-2@ybl/To:volipi.l@ptyes/Payment from PhonePe","THOR","Operational Expenses"),
    (1686,"2026-03-12",16000.00,"UPI/167481104453/From:7358341775-2@ybl/To:volipi.l@ptyes/Payment from PhonePe","THOR","Operational Expenses"),
    (1644,"2026-03-16",444.00,"UPI/284654207742/From:7358341775-2@ybl/To:volipi.l@ptyes/Payment from PhonePe","THOR","Operational Expenses"),
    (1143,"2026-04-11",340.00,"UPI/203805822962/From:7358341775@ptyes/To:volipi.l@ptyes/kastig soda","THOR","Operational Expenses"),
    (1109,"2026-04-14",326.00,"UPI/204022094584/From:7358341775@ptyes/To:volipi.l@ptyes/bus ticket","THOR","Operational Expenses"),
    (1082,"2026-04-17",400.00,"UPI/204218176152/From:7358341775@ptyes/To:volipi.l@ptyes/mirrors porte","THOR","Operational Expenses"),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_upi(desc):
    """Extract (from_upi, to_upi, memo) from UPI description string."""
    if not desc.startswith("UPI/"):
        return ("—", "—", desc)
    m_from = re.search(r'From[:\s]+([^\s/]+)', desc, re.IGNORECASE)
    m_to   = re.search(r'To[:\s]+([^\s/]+)', desc, re.IGNORECASE)
    # memo = everything after the To:xxx/ part
    m_memo = re.search(r'To[:\s]+[^\s/]+[/\s]+(.*)', desc, re.IGNORECASE)
    from_upi = m_from.group(1) if m_from else "—"
    to_upi   = m_to.group(1)   if m_to   else "—"
    memo     = m_memo.group(1).strip() if m_memo else "—"
    # clean generic memos
    if memo.lower() in ("sent using paytm upi","payment from phonepe","sent from paytm","upi","transaction",""):
        memo = "—"
    return from_upi, to_upi, memo


GROUP_MAP = {
    # (to_upi_contains, group_label)
    "chandrasekhar":     "A — Chandrasekhar (PG Expenses)",
    "8951297583":        "B — 8951297583 (UNKNOWN)",
    "arunphilip25":      "C — arunphilip25 (UNKNOWN)",
    "9904388966":        "D — 9904388966 (UNKNOWN)",
    "tpasha638":         "E — tpasha638 (UNKNOWN)",
    "m036tpqek":         "F — M036TPQEK (Merchant QR)",
    "6202601070":        "G — 6202601070 (UNKNOWN)",
    "bn895975":          "H — bn895975 (UNKNOWN)",
    "paytmqr6wguqk":     "I — paytmqr6wguqk (Merchant QR)",
    "8073343903":        "J — 8073343903 (UNKNOWN)",
    "7667662128":        "K — 7667662128 (UNKNOWN)",
    "sk5008786":         "L1 — sk5008786 (UNKNOWN)",
    "6287677379":        "L2 — 6287677379 (UNKNOWN)",
    "9148809732":        "M — 9148809732 (Recurring UNKNOWN)",
    "9448259556":        "N — 9448259556 (Recurring UNKNOWN)",
    "9989000250":        "O — 9989000250 (Recurring UNKNOWN)",
    "8310018565":        "P — 8310018565 (UNKNOWN)",
    "9146500827":        "Q — 9146500827 (UNKNOWN)",
    "phegade08":         "R — phegade08 (UNKNOWN)",
    "9160224731":        "S — 9160224731 (Yes Credit?)",
    "9902278720":        "T — 9902278720 (UNKNOWN)",
    "paybil3066":        "U — paybil3066 (Bill Pay service)",
    "9663049651":        "V — 9663049651 (UNKNOWN)",
    "sv2512112238":      "W — SV vendor (UNKNOWN)",
    "kanjichoudhari":    "X — kanjichoudhari (UNKNOWN)",
    "shahnawazlaskar":   "Y — shahnawazlaskar (UNKNOWN)",
    "9632460361":        "Z — 9632460361 (UNKNOWN)",
    "paytm-56505013":    "Z1 — paytm-56505013 (Paytm bill/misc)",
    "payair7673":        "Z2 — payair7673 (UNKNOWN service)",
    "paytm-64646105":    "Z3 — paytm-64646105 (UPI Autopay)",
    "8105769898":        "Z4 — 8105769898 (UNKNOWN)",
    "paperandpie":       "Z5 — paperandpie (Stationery?)",
    "7829264915":        "Z6 — 7829264915 (UNKNOWN)",
    "8310048369":        "Z7 — 8310048369 (UNKNOWN)",
    "8088826094":        "Z8 — 8088826094 (UNKNOWN)",
    "9108675353":        "Z9 — 9108675353 (UNKNOWN)",
    "q531107921":        "Z10 — Q531107921 (Merchant QR)",
    "premstealer1":      "Z11 — premstealer1 (gaming/unknown)",
    "akshyarathna168":   "Z12 — akshyarathna168 (UNKNOWN)",
    "yashumeena080":     "Z13 — yashumeena080 (UNKNOWN)",
    "q424363623":        "Z14 — Q424363623 (non plu merchant)",
    "paytmqr2810050501": "Z15 — Paytm QR merchant",
    "6909620430":        "Z16 — 6909620430 (UNKNOWN)",
    "7981501263":        "Z17 — 7981501263 (UNKNOWN)",
    "8795514149":        "Z18 — 8795514149 (UNKNOWN)",
    "8197867322":        "Z19 — 8197867322 (UNKNOWN)",
    "sujanmallik8348":   "Z20 — sujanmallik (UNKNOWN)",
    "anudeepbishtab":    "Z21 — anudeepbishtab (UNKNOWN)",
    "7061052050":        "Z22 — 7061052050 (UNKNOWN)",
    "jaydevjena73":      "Z23 — jaydevjena73 (UNKNOWN)",
    "8816019354":        "Z24 — 8816019354 (UNKNOWN)",
    "q937051021":        "Z25 — water (small)",
    "shahbaz80508637":   "Z26 — shahbaz (UNKNOWN)",
}

def get_group(to_upi, desc):
    key = (to_upi + " " + desc).lower()
    for pattern, label in GROUP_MAP.items():
        if pattern in key:
            return label
    return "ZZ — Other"


# ── Build rows ────────────────────────────────────────────────────────────────

rows = []
for rec in OTHER_EXPENSES:
    db_id, date, amount, desc, upi_ref, account = rec
    from_upi, to_upi, memo = parse_upi(desc)
    group = get_group(to_upi, desc)
    rows.append({
        "Group":         group,
        "Date":          date,
        "Amount (Rs.)":  amount,
        "To / Payee":    to_upi,
        "Memo":          memo,
        "Account":       account,
        "DB ID":         db_id,
        "YOUR CATEGORY": "",
    })

rows.sort(key=lambda r: (r["Group"], -r["Amount (Rs.)"]))

volipi_rows = []
for rec in VOLIPI:
    db_id, date, amount, desc, account, current_cat = rec
    _, to_upi, memo = parse_upi(desc)
    volipi_rows.append({
        "DB ID":             db_id,
        "Date":              date,
        "Amount (₹)":        amount,
        "Account":           account,
        "To / Payee":        to_upi,
        "Memo":              memo,
        "Current Category":  current_cat,
        "Confirm / Change":  "",
        "Notes":             "",
    })


# ── Write Excel ───────────────────────────────────────────────────────────────

out_path = "data/reports/other_expenses_classify.xlsx"

HEADER_FILL   = PatternFill("solid", fgColor="1F2D3D")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL    = PatternFill("solid", fgColor="EEF2F7")
GROUP_FONT    = Font(bold=True, color="1F2D3D", size=9)
AMOUNT_FILL   = PatternFill("solid", fgColor="FFF9E6")
INPUT_FILL    = PatternFill("solid", fgColor="E8F5E9")
BORDER_THIN   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

    # ── Sheet 1: Unclassified ─────────────────────────────────────────────────
    df1 = pd.DataFrame(rows)
    df1.to_excel(writer, sheet_name="Classify These (77 rows)", index=False)
    ws1 = writer.sheets["Classify These (77 rows)"]

    # Columns: Group | Date | Amount | To/Payee | Memo | Account | DB ID | YOUR CATEGORY
    col_widths = {
        "A": 38,  # Group
        "B": 12,  # Date
        "C": 13,  # Amount
        "D": 30,  # To/Payee
        "E": 22,  # Memo
        "F": 14,  # Account
        "G": 8,   # DB ID
        "H": 28,  # YOUR CATEGORY
    }
    for col, w in col_widths.items():
        ws1.column_dimensions[col].width = w

    # Header row styling
    for cell in ws1[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER_THIN
    ws1.row_dimensions[1].height = 28

    # Rows — alternate fill per group
    prev_group = None
    alt = False
    for row_idx, row_data in enumerate(rows, start=2):
        grp = row_data["Group"]
        if grp != prev_group:
            alt = not alt
            prev_group = grp

        row_fill = PatternFill("solid", fgColor="F0F4FA") if alt else PatternFill("solid", fgColor="FFFFFF")

        for col_idx, col_key in enumerate(
            ["Group","Date","Amount (Rs.)","To / Payee","Memo","Account","DB ID","YOUR CATEGORY"],
            start=1
        ):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border    = BORDER_THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font      = Font(size=9)

            if col_key == "Amount (Rs.)":
                cell.number_format = '#,##0.00'
                amt = row_data["Amount (Rs.)"]
                if amt >= 5000:
                    cell.fill = PatternFill("solid", fgColor="FFE0B2")
                    cell.font = Font(bold=True, size=9, color="BF360C")
                elif amt >= 1000:
                    cell.fill = PatternFill("solid", fgColor="FFF9E6")
                    cell.font = Font(bold=True, size=9)
                else:
                    cell.fill = row_fill
            elif col_key == "YOUR CATEGORY":
                cell.fill = INPUT_FILL
                cell.font = Font(size=9, color="1B5E20", bold=True)
            elif col_key == "Group":
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
            else:
                cell.fill = row_fill

        ws1.row_dimensions[row_idx].height = 16

    # Total row
    total_row = len(rows) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=10)
    total_cell = ws1.cell(row=total_row, column=3,
                          value=sum(r["Amount (Rs.)"] for r in rows))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True, size=10, color="BF360C")
    total_cell.fill = PatternFill("solid", fgColor="FFE0B2")

    ws1.freeze_panes = "B2"

    # ── Sheet 2: Volipi ───────────────────────────────────────────────────────
    df2 = pd.DataFrame(volipi_rows)
    df2.to_excel(writer, sheet_name="Volipi (confirm sub-cat)", index=False)
    ws2 = writer.sheets["Volipi (confirm sub-cat)"]

    col_widths2 = {"A":8,"B":12,"C":12,"D":12,"E":28,"F":22,"G":28,"H":28,"I":22}
    for col, w in col_widths2.items():
        ws2.column_dimensions[col].width = w

    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    for row_idx in range(2, len(volipi_rows) + 2):
        for col_idx in range(1, 10):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            if col_idx == 3:
                cell.number_format = '#,##0.00'
            if col_idx == 8:
                cell.fill = INPUT_FILL
                cell.font = Font(size=9, color="1B5E20", bold=True)

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    summary_data = []
    from collections import defaultdict
    group_totals = defaultdict(float)
    group_counts = defaultdict(int)
    for r in rows:
        group_totals[r["Group"]] += r["Amount (Rs.)"]
        group_counts[r["Group"]] += 1

    for g in sorted(group_totals):
        summary_data.append({"Group": g, "# Txns": group_counts[g], "Total (₹)": group_totals[g]})

    df3 = pd.DataFrame(summary_data)
    df3.to_excel(writer, sheet_name="Summary by Group", index=False)
    ws3 = writer.sheets["Summary by Group"]
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 16

    for cell in ws3[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, len(summary_data) + 2):
        ws3.cell(row=row_idx, column=1).font = Font(size=9)
        ws3.cell(row=row_idx, column=2).font = Font(size=9)
        amt_cell = ws3.cell(row=row_idx, column=3)
        amt_cell.number_format = '#,##0.00'
        amt_cell.font = Font(size=9, bold=True if group_totals[summary_data[row_idx-2]["Group"]] >= 5000 else False)

    grand = ws3.cell(row=len(summary_data)+2, column=3, value=sum(r["Total (₹)"] for r in summary_data))
    grand.number_format = '#,##0.00'
    grand.font = Font(bold=True, size=10, color="BF360C")
    ws3.cell(row=len(summary_data)+2, column=1, value="GRAND TOTAL").font = Font(bold=True, size=10)


print(f"Done -> {out_path}")
total = sum(r["Amount (Rs.)"] for r in rows)
print(f"  Sheet 1: {len(rows)} unclassified rows  |  Total Rs.{total:,.0f}")
print(f"  Sheet 2: {len(volipi_rows)} volipi rows to confirm")
print(f"  Sheet 3: {len(summary_data)} groups")
