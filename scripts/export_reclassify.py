"""
scripts/export_reclassify.py
----------------------------
Exports three P&L categories to Excel for manual reclassification:
  Sheet 1 — Operational Expenses (all accounts)
  Sheet 2 — Other Expenses (all accounts)
  Sheet 3 — Partner Reimbursable SBI 0167 (all PERSONAL_SBI_0167 txns)

Each sheet has a blank "New Category" column with dropdown validation.
Run: python scripts/export_reclassify.py
Output: data/reports/reclassify_review.xlsx
"""
from datetime import date
from pathlib import Path

import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

DB = "postgresql://postgres:Anchorstrong123!@db.oxiqomoilqwfxjauxhzp.supabase.co:5432/postgres"
OUTPUT_PATH = Path("data/reports/reclassify_review.xlsx")

VALID_CATEGORIES = [
    "Bank Charges", "Capital Investment", "Cleaning Supplies", "Electricity",
    "Food & Groceries", "Fuel & Diesel", "Furniture & Fittings", "Govt & Regulatory",
    "Internet & WiFi", "IT & Software", "Maintenance & Repairs", "Marketing",
    "Non-Operating", "Operational Expenses", "Other Expenses", "Other Income",
    "Partner Capital (Whitefield)", "Property Rent", "Rent Income",
    "Shopping & Supplies", "Staff & Labour", "Tenant Deposit Refund",
    "Waste Disposal", "Water",
]

HEADERS = ["Date", "Description", "Amount", "Current Category", "Account", "Month", "New Category"]

HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")
NEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
INR_FMT  = '#,##0.00'


def fetch(cur, category: str | None = None, account: str | None = None) -> list[dict]:
    conditions = ["txn_date BETWEEN '2025-10-01' AND '2026-04-30'"]
    params: list = []
    if category:
        conditions.append("category = %s")
        params.append(category)
    if account:
        conditions.append("account_name = %s")
        params.append(account)
    sql = (
        "SELECT txn_date, description, amount, category, account_name "
        f"FROM bank_transactions WHERE {' AND '.join(conditions)} ORDER BY txn_date"
    )
    cur.execute(sql, params)
    return cur.fetchall()


def month_label(d) -> str:
    if isinstance(d, str):
        d = date.fromisoformat(d)
    return d.strftime("%b'%y")


def write_sheet(ws, rows: list[dict]):
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(1, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 2):
        alt = (i % 2 == 0)
        date_cell = ws.cell(i, 1, r["txn_date"])
        date_cell.number_format = "DD-MMM-YY"
        ws.cell(i, 2, r["description"])
        amt = ws.cell(i, 3, float(r["amount"]))
        amt.number_format = INR_FMT
        ws.cell(i, 4, r["category"])
        ws.cell(i, 5, r["account_name"])
        ws.cell(i, 6, month_label(r["txn_date"]))
        new_cat = ws.cell(i, 7, "")
        new_cat.fill = NEW_FILL
        if alt:
            for col in range(1, 7):
                ws.cell(i, col).fill = ALT_FILL

    # Dropdown validation on New Category column
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_CATEGORIES) + '"',
        allow_blank=True,
    )
    dv.sqref = f"G2:G{max(len(rows) + 1, 2)}"
    ws.add_data_validation(dv)

    # Totals row
    total_row = len(rows) + 2
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    total_cell = ws.cell(total_row, 3, f"=SUM(C2:C{len(rows)+1})")
    total_cell.number_format = INR_FMT
    total_cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 24
    ws.freeze_panes = "A2"


def main():
    conn = psycopg2.connect(DB)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    op_exp    = fetch(cur, category="Operational Expenses")
    other_exp = fetch(cur, category="Other Expenses")
    sbi_0167  = fetch(cur, account="PERSONAL_SBI_0167")

    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    del wb["Sheet"]  # remove default empty sheet

    ws1 = wb.create_sheet("Operational Expenses")
    write_sheet(ws1, op_exp)

    ws2 = wb.create_sheet("Other Expenses")
    write_sheet(ws2, other_exp)

    ws3 = wb.create_sheet("Partner Reimbursable SBI0167")
    write_sheet(ws3, sbi_0167)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved -> {OUTPUT_PATH}")
    print(f"  Operational Expenses : {len(op_exp)} rows")
    print(f"  Other Expenses       : {len(other_exp)} rows")
    print(f"  Partner SBI 0167     : {len(sbi_0167)} rows")


if __name__ == "__main__":
    main()
