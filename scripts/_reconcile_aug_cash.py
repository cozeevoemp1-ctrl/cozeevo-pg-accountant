"""
One-off: three-way August-2026 CASH reconciliation.

  RECEIPT  Kiran's physical receipt book, transcribed
           data/uploads/csv/Cash -Kiran.xlsx -> 'CASH COLLECTION August-26'
  APP      DB payments, payment_mode=cash, is_void=false, payment_date in the
           window the receipt book actually covers (21-Jul .. 31-Aug-2026).
           NOTE: the book's first 6 rows are dated 21-31 Jul but are August rent,
           so a plain "August only" window loses ~Rs 1.3L and shows false gaps.
  GSHEET   data/uploads/csv/google sheet local excel.xlsx -> col 'Aug Cash'

Output: data/reports/Aug2026_Cash_3Way_Reconciliation.xlsx
  Tab 1 'Reconciliation' — one row per tenant, three amount columns side by side
  Tab 2 'Summary'        — totals, variances, breakdown by status
  Tab 3 'Receipt Book'   — every receipt row + which tenant it matched

Matching: room number, then fuzzy name inside the room, then a global fuzzy pass
(handwritten room numbers are sometimes wrong). Joint receipts written as
"Lenin/Madhu" are split across both tenants in proportion to their app/sheet
amounts. Every non-exact link is flagged in the Notes column.
"""
from __future__ import annotations

import asyncio
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
load_dotenv()

from src.database.db_manager import get_session, init_engine  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
RECEIPT_XLSX = ROOT / "data/uploads/csv/Cash -Kiran.xlsx"
GSHEET_XLSX = ROOT / "data/uploads/csv/google sheet local excel.xlsx"
OUT_XLSX = ROOT / "data/reports/Aug2026_Cash_3Way_Reconciliation.xlsx"

RECEIPT_SHEET = "CASH COLLECTION August-26"
RECEIPT_FIRST_ROW, RECEIPT_LAST_ROW = 5, 176  # row 177 = totals

APP_FROM, APP_TO = "2026-07-21", "2026-08-31"

# ---------------------------------------------------------------- normalisers

_NOISE = re.compile(r"[^A-Z ]")
_STOP = {
    "MR", "MS", "MRS", "KUMAR", "SINGH", "REDDY", "BABU", "ALIEAS",
    "S", "K", "M", "P", "N", "J", "B", "D", "R", "A", "V", "T", "G", "C",
}


def norm_room(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return re.sub(r"\.0$", "", str(v).strip().upper())


def norm_name(v) -> str:
    if not v:
        return ""
    s = str(v).upper().replace("/", " ").replace(".", " ").replace(",", " ")
    return " ".join(_NOISE.sub(" ", s).split())


def sig_tokens(name: str) -> set[str]:
    return {t for t in name.split() if len(t) >= 4 and t not in _STOP}


def flat(name: str) -> str:
    return name.replace(" ", "")


def _ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def similar(a: str, b: str) -> bool:
    """True if two normalised names plausibly refer to the same person.

    Tuned against the real Aug-26 data, which contains pairs like:
      SAMSUDHI/SAMRUDDHI THANWAR   NAVEENETH P J/NAVANEETH P J
      AVANISH/AVINASH YADAV        ANSH/ANSHSINHA
      SHREYA/SHREYAKARPE           SWRAP/SWARUP RAVINDRA FUTANE
      SATHYA/SAHITHYA V            PRAKASHSITA/PRAKASHITA
      NARENDHRA S/S NARENDH        HARISH PRAJAPATHI/PRAJAPATI HARESHBHAI
    """
    if not a or not b:
        return False
    if a == b:
        return True

    fa, fb = flat(a), flat(b)
    # one name fully contained in the other  (ANSH -> ANSHSINHA)
    if len(fa) >= 4 and len(fb) >= 4 and (fa in fb or fb in fa):
        return True
    if _ratio(fa, fb) >= 0.80:
        return True

    ta, tb = sig_tokens(a), sig_tokens(b)
    if ta & tb:
        return True
    # any significant token pair close enough  (NAVEENETH ~ NAVANEETH)
    for x in ta:
        for y in tb:
            if _ratio(x, y) >= 0.78:
                return True
            if len(x) >= 5 and len(y) >= 5 and (x in y or y in x):
                return True
    return False


def loose_similar(a: str, b: str) -> bool:
    """Deliberately permissive — only used on the small residual set, where the
    alternative is a human eyeballing 'SWRAP' against 'SWARUP RAVINDRA FUTANE'."""
    if similar(a, b):
        return True
    fa, fb = flat(a), flat(b)
    if _ratio(fa, fb) >= 0.62:
        return True
    if fa[:4] and fa[:4] == fb[:4]:
        return True
    for x in a.split():
        for y in b.split():
            if len(x) >= 5 and len(y) >= 5 and _ratio(x, y) >= 0.65:
                return True
    return False


def money(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


# ------------------------------------------------------------------ loaders


def load_receipt() -> list[dict]:
    ws = openpyxl.load_workbook(RECEIPT_XLSX, data_only=True)[RECEIPT_SHEET]
    out, last_date = [], None
    for r in range(RECEIPT_FIRST_ROW, RECEIPT_LAST_ROW + 1):
        name, amt = ws.cell(r, 5).value, money(ws.cell(r, 6).value)
        if not name or amt <= 0:
            continue
        d = as_date(ws.cell(r, 3).value)
        if d:
            last_date = d
        raw = str(name).strip()
        out.append(
            {
                "row": r,
                "date": d,
                "eff_date": d or last_date,       # book stops writing dates part-way
                "date_carried": d is None,
                "room": norm_room(ws.cell(r, 4).value),
                "raw_name": raw,
                "name": norm_name(raw),
                "parts": [norm_name(p) for p in re.split(r"[/&]", raw) if norm_name(p)],
                "amount": amt,
                "collector": str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "",
            }
        )
    return out


def load_gsheet() -> list[dict]:
    ws = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)["Sheet1"]
    hdr = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            hdr.setdefault(str(h).strip(), c)
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, hdr["Name"]).value
        if not name:
            continue
        out.append(
            {
                "room": norm_room(ws.cell(r, hdr["Room No"]).value),
                "raw_name": str(name).strip(),
                "name": norm_name(name),
                "amount": money(ws.cell(r, hdr["Aug Cash"]).value),
                "status": ws.cell(r, hdr["Checkin/out"]).value,
            }
        )
    return out


async def load_app() -> list[dict]:
    init_engine(os.environ["DATABASE_URL"])
    async with get_session() as s:
        res = await s.execute(
            text(
                """
                SELECT r.room_number, t.name, p.payment_date, p.amount,
                       p.for_type::text AS for_type, tc.status::text AS tstatus
                FROM payments p
                JOIN tenancies tc ON tc.id = p.tenancy_id
                JOIN tenants   t  ON t.id  = tc.tenant_id
                LEFT JOIN rooms r ON r.id  = tc.room_id
                WHERE p.is_void = false
                  AND p.payment_mode = 'cash'
                  AND p.payment_date <= :d2
                  AND (
                        p.payment_date >= DATE '2026-08-01'
                        -- late-July collections of August rent: the book's first
                        -- six rows. Exclude July-period money collected in the
                        -- same days, which belongs to the July book.
                     OR (p.payment_date >= :d1
                         AND (p.period_month = DATE '2026-08-01'
                              OR p.for_type <> 'rent'))
                  )
                ORDER BY r.room_number, t.name
                """
            ),
            {"d1": date.fromisoformat(APP_FROM), "d2": date.fromisoformat(APP_TO)},
        )
        return [
            {
                "room": norm_room(m.room_number),
                "raw_name": (m.name or "").strip(),
                "name": norm_name(m.name),
                "amount": float(m.amount),
                "date": m.payment_date,
                "for_type": m.for_type,
                "tstatus": m.tstatus,
            }
            for m in res.mappings()
        ]


# ------------------------------------------------------------------ model


class Tenant:
    __slots__ = ("room", "names", "receipt", "app", "gsheet", "notes",
                 "app_dates", "receipt_dates", "tstatus", "orphan", "merged")

    def __init__(self, room: str, name: str):
        self.room = room
        self.names = [name]
        self.receipt = self.app = self.gsheet = 0.0
        self.notes: list[str] = []
        self.app_dates: list[date] = []
        self.receipt_dates: list[date] = []
        self.tstatus = ""
        self.orphan = False   # created solely from an unmatched receipt row
        self.merged = False   # absorbed into another tenant by the residual pass

    @property
    def display(self) -> str:
        return max(self.names, key=len)

    def matches(self, name: str) -> bool:
        return any(similar(n, name) for n in self.names)

    def add_name(self, name: str):
        if name not in self.names:
            self.names.append(name)

    def note(self, msg: str):
        if msg not in self.notes:
            self.notes.append(msg)


def build(receipt, app, gsheet) -> tuple[list[Tenant], list[dict]]:
    by_room: dict[str, list[Tenant]] = defaultdict(list)
    everyone: list[Tenant] = []

    def find(room: str, name: str) -> Tenant | None:
        for t in by_room[room]:
            if t.matches(name):
                return t
        return None

    def get(room: str, name: str) -> Tenant:
        t = find(room, name)
        if t is None:
            t = Tenant(room, name)
            by_room[room].append(t)
            everyone.append(t)
        else:
            t.add_name(name)
        return t

    def find_anywhere(name: str, exclude_room: str = "") -> list[Tenant]:
        return [t for t in everyone if t.room != exclude_room and t.matches(name)]

    # 1. master list = GSHEET (tenant register) + APP
    for g in gsheet:
        t = get(g["room"], g["name"])
        t.gsheet += g["amount"]
        if g["status"]:
            t.tstatus = str(g["status"])
    for a in app:
        t = get(a["room"], a["name"])
        t.app += a["amount"]
        t.app_dates.append(a["date"])
        if a["for_type"] != "rent":
            t.note(f"app row is {a['for_type']} Rs {int(a['amount']):,}")
        if a["tstatus"] not in ("active", "", None):
            t.tstatus = a["tstatus"]

    # 2. overlay receipt book
    trace = []

    def pick(hits: list[Tenant], amount: float) -> Tenant | None:
        """Break a multi-name tie using the amount — only if exactly one fits."""
        exact = [t for t in hits if round(t.app) == round(amount) or round(t.gsheet) == round(amount)]
        return exact[0] if len(exact) == 1 else None

    def resolve(name: str, room: str, amount: float) -> tuple[Tenant | None, str]:
        """-> (tenant, how)"""
        if room:
            t = find(room, name)
            if t:
                return t, "room+name"
            hits = find_anywhere(name, exclude_room=room)
            if len(hits) == 1:
                hits[0].note(f"book says room {room}, register says {hits[0].room}")
                return hits[0], "name only (room differs)"
            if len(hits) > 1:
                t = pick(hits, amount)
                if t:
                    t.note(f"book says room {room}, register says {t.room}")
                    return t, "name + amount (room differs)"
                return None, f"ambiguous - {len(hits)} name matches"
            return None, "no match"

        hits = find_anywhere(name)
        if len(hits) == 1:
            hits[0].note("receipt row had no room number")
            return hits[0], "name only (no room in book)"
        if len(hits) > 1:
            t = pick(hits, amount)
            if t:
                t.note(f"receipt row had no room number; matched on amount to {t.room}")
                return t, "name + amount (no room in book)"
            return None, f"ambiguous - {len(hits)} name matches"
        return None, "no match"

    for rec in receipt:
        parts = rec["parts"] if len(rec["parts"]) > 1 else [rec["name"]]
        share_hint = rec["amount"] / len(parts)
        resolved = [(p, *resolve(p, rec["room"], share_hint if len(parts) > 1 else rec["amount"]))
                    for p in parts]
        hit = [(p, t, how) for p, t, how in resolved if t is not None]

        if not hit:
            t = get(rec["room"] or "(no room)", rec["name"])
            t.receipt += rec["amount"]
            t.receipt_dates.append(rec["eff_date"])
            t.note(f"receipt only - {resolved[0][2]}")
            t.orphan = True
            trace.append({**rec, "matched": "", "how": resolved[0][2], "split": rec["amount"]})
            continue

        if len(hit) == 1:
            _, t, how = hit[0]
            t.receipt += rec["amount"]
            t.receipt_dates.append(rec["eff_date"])
            if how != "room+name":
                t.note(f"matched by {how}")
            if len(parts) > 1:
                t.note(f"joint receipt '{rec['raw_name']}' — only this name matched")
            trace.append({**rec, "matched": t.display, "how": how, "split": rec["amount"]})
            continue

        # joint receipt across 2+ tenants — split by their app/sheet amounts
        weights = [max(t.app, t.gsheet) for _, t, _ in hit]
        if sum(weights) <= 0:
            weights = [1.0] * len(hit)
        total_w = sum(weights)
        for (p, t, how), w in zip(hit, weights):
            share = round(rec["amount"] * w / total_w, 2)
            t.receipt += share
            t.receipt_dates.append(rec["eff_date"])
            t.note(f"joint receipt '{rec['raw_name']}' Rs {int(rec['amount']):,} split")
            trace.append({**rec, "matched": t.display, "how": f"joint / {how}", "split": share})

    # 3. residual pass — pair leftover receipt-only rows against tenants that
    #    have app/sheet money but no receipt. Both sets are small, so a looser
    #    threshold is safe here and every merge is flagged in Notes.
    orphans = [t for t in everyone if t.orphan and not t.merged]
    no_receipt = [t for t in everyone if not t.orphan and t.receipt == 0 and (t.app or t.gsheet)]

    for o in orphans:
        cands = [t for t in no_receipt if not t.merged and loose_similar(o.display, t.display)]
        if not cands:
            continue
        if len(cands) > 1:
            exact = [t for t in cands if round(t.app) == round(o.receipt)
                     or round(t.gsheet) == round(o.receipt)]
            if len(exact) != 1:
                o.note(f"possible match to {len(cands)} tenants - resolve manually")
                continue
            cands = exact
        target = cands[0]
        target.receipt += o.receipt
        target.receipt_dates += o.receipt_dates
        target.note(
            f"book wrote '{o.display.title()}'"
            + (f" room {o.room}" if o.room and o.room != "(no room)" else " with no room")
            + f" - treated as {target.display.title()} room {target.room}"
        )
        o.merged = True
        for x in trace:
            if not x["matched"] and norm_name(x["raw_name"]) in o.names:
                x["matched"] = target.display
                x["how"] = "residual pass (loose name match)"

    return [t for t in everyone if not t.merged], trace


def room_key(room: str):
    m = re.match(r"^(\d+)$", room)
    return (0, int(m.group(1)), "") if m else (1, 0, room)


def classify(t: Tenant) -> str:
    present = [k for k, v in (("receipt", t.receipt), ("app", t.app), ("gsheet", t.gsheet)) if v > 0]
    if not present:
        return ""
    if len(present) == 3 and round(t.receipt) == round(t.app) == round(t.gsheet):
        return "MATCH"
    if len(present) == 1:
        return f"ONLY IN {present[0].upper()}"
    if len(present) == 2:
        return f"MISSING IN {({'receipt', 'app', 'gsheet'} - set(present)).pop().upper()}"
    return "AMOUNT MISMATCH"


# ------------------------------------------------------------------ writer

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = PatternFill("solid", fgColor="1F3864")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREEN = PatternFill("solid", fgColor="C6EFCE")
BAND = PatternFill("solid", fgColor="DDEBF7")
INR = '#,##0;[Red]-#,##0;"-"'

HEADERS = ["Room", "Tenant", "Receipt Book", "App (PWA)", "Google Sheet",
           "Receipt - App", "Receipt - GSheet", "App - GSheet",
           "Status", "Tenancy", "App Dates", "Notes"]


def style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write(tenants: list[Tenant], trace: list[dict], totals: dict):
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ---------------------------------------------------- Tab 1: reconciliation
    ws = wb.active
    ws.title = "Reconciliation"
    rows = sorted([t for t in tenants if (t.receipt or t.app or t.gsheet)],
                  key=lambda t: (room_key(t.room), t.display))

    ws.append(["AUGUST 2026 CASH — RECEIPT BOOK vs APP vs GOOGLE SHEET"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"App window: cash payments {APP_FROM} to {APP_TO} (the range the receipt book covers)"])
    ws["A2"].font = Font(italic=True, color="808080")
    ws.append(HEADERS)
    style_header(ws, 3, len(HEADERS))

    first = 4
    for t in rows:
        st = classify(t)
        dates = ", ".join(sorted({d.strftime("%d-%b") for d in t.app_dates if d}))
        ws.append([t.room, t.display.title(), t.receipt or None, t.app or None,
                   t.gsheet or None, t.receipt - t.app, t.receipt - t.gsheet,
                   t.app - t.gsheet, st, t.tstatus or "", dates, "; ".join(t.notes)])
        r = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).border = BORDER
        for c in range(3, 9):
            ws.cell(r, c).number_format = INR
        sc = ws.cell(r, 9)
        if st == "MATCH":
            sc.fill = GREEN
        elif st == "AMOUNT MISMATCH":
            sc.fill = AMBER
            sc.font = Font(bold=True)
        elif st:
            sc.fill = RED
            sc.font = Font(bold=True)

    last = ws.max_row
    ws.append([])
    ws.append(["", "TOTAL"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{last})"
                               for c in range(3, 9)] + ["", "", "", ""])
    tr = ws.max_row
    for c in range(1, len(HEADERS) + 1):
        ws.cell(tr, c).font = Font(bold=True)
        ws.cell(tr, c).fill = BAND
        ws.cell(tr, c).border = BORDER
    for c in range(3, 9):
        ws.cell(tr, c).number_format = INR

    for i, w in enumerate([9, 30, 14, 13, 14, 14, 15, 13, 19, 11, 18, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:L{last}"

    # ---------------------------------------------------- Tab 2: summary
    s = wb.create_sheet("Summary")
    s.append(["AUGUST 2026 CASH — SUMMARY"])
    s["A1"].font = Font(bold=True, size=14)
    s.append([])
    s.append(["Source", "Entries", "Total (Rs)"])
    style_header(s, 3, 3)
    for label, k in [("Receipt book (Cash -Kiran.xlsx)", "receipt"),
                     ("App / PWA (DB cash payments)", "app"),
                     ("Google sheet local excel ('Aug Cash')", "gsheet")]:
        s.append([label, totals[k]["n"], totals[k]["sum"]])
        s.cell(s.max_row, 3).number_format = INR
    s.append([])
    s.append(["VARIANCE"])
    s.cell(s.max_row, 1).font = Font(bold=True)
    for label, v in [("Receipt book - App", totals["receipt"]["sum"] - totals["app"]["sum"]),
                     ("Receipt book - Google Sheet", totals["receipt"]["sum"] - totals["gsheet"]["sum"]),
                     ("App - Google Sheet", totals["app"]["sum"] - totals["gsheet"]["sum"])]:
        s.append([label, "", v])
        s.cell(s.max_row, 3).number_format = INR
        s.cell(s.max_row, 3).font = Font(bold=True)
    s.append([])
    s.append(["BREAKDOWN BY STATUS", "Tenants", "Receipt Rs", "App Rs", "GSheet Rs"])
    style_header(s, s.max_row, 5)
    agg = defaultdict(lambda: [0, 0.0, 0.0, 0.0])
    for t in rows:
        a = agg[classify(t)]
        a[0] += 1
        a[1] += t.receipt
        a[2] += t.app
        a[3] += t.gsheet
    for st in sorted(agg, key=lambda k: -agg[k][0]):
        n, rc, ap, gs = agg[st]
        s.append([st, n, rc, ap, gs])
        for c in (3, 4, 5):
            s.cell(s.max_row, c).number_format = INR
        if st != "MATCH":
            s.cell(s.max_row, 1).fill = RED if not st.startswith("AMOUNT") else AMBER
    for i, w in enumerate([40, 12, 16, 16, 16], start=1):
        s.column_dimensions[get_column_letter(i)].width = w

    # ---------------------------------------------------- Tab 3: receipt trace
    t3 = wb.create_sheet("Receipt Book")
    t3.append(["EVERY ROW OF THE RECEIPT BOOK AND HOW IT WAS MATCHED"])
    t3["A1"].font = Font(bold=True, size=14)
    t3.append([])
    cols = ["Book Row", "Date", "Date Carried", "Room", "Name in Book",
            "Amount", "Matched Tenant", "Match Method", "Amount Applied"]
    t3.append(cols)
    style_header(t3, 3, len(cols))
    for x in sorted(trace, key=lambda z: z["row"]):
        t3.append([x["row"], x["eff_date"], "yes" if x["date_carried"] else "",
                   x["room"], x["raw_name"], x["amount"],
                   x["matched"], x["how"], x["split"]])
        r = t3.max_row
        for c in range(1, len(cols) + 1):
            t3.cell(r, c).border = BORDER
        t3.cell(r, 2).number_format = "dd-mmm"
        t3.cell(r, 6).number_format = INR
        t3.cell(r, 9).number_format = INR
        if not x["matched"]:
            t3.cell(r, 7).fill = RED
            t3.cell(r, 8).fill = RED
        elif x["how"] != "room+name":
            t3.cell(r, 8).fill = AMBER
    for i, w in enumerate([10, 11, 13, 9, 28, 13, 28, 30, 15], start=1):
        t3.column_dimensions[get_column_letter(i)].width = w
    t3.freeze_panes = "A4"
    t3.auto_filter.ref = f"A3:I{t3.max_row}"

    wb.save(OUT_XLSX)
    return rows, agg


# ------------------------------------------------------------------ main


async def main():
    receipt, gsheet = load_receipt(), load_gsheet()
    app = await load_app()

    totals = {
        "receipt": {"n": len(receipt), "sum": sum(r["amount"] for r in receipt)},
        "app": {"n": len(app), "sum": sum(a["amount"] for a in app)},
        "gsheet": {"n": sum(1 for g in gsheet if g["amount"]),
                   "sum": sum(g["amount"] for g in gsheet)},
    }

    tenants, trace = build(receipt, app, gsheet)
    rows, agg = write(tenants, trace, totals)

    w = 14
    print(f"receipt book : {totals['receipt']['n']:>4} rows   Rs {totals['receipt']['sum']:>{w},.0f}")
    print(f"app (cash)   : {totals['app']['n']:>4} rows   Rs {totals['app']['sum']:>{w},.0f}")
    print(f"google sheet : {totals['gsheet']['n']:>4} rows   Rs {totals['gsheet']['sum']:>{w},.0f}")
    print()
    print(f"receipt - app    : Rs {totals['receipt']['sum'] - totals['app']['sum']:>{w},.0f}")
    print(f"receipt - gsheet : Rs {totals['receipt']['sum'] - totals['gsheet']['sum']:>{w},.0f}")
    print(f"app - gsheet     : Rs {totals['app']['sum'] - totals['gsheet']['sum']:>{w},.0f}")
    print()
    for st in sorted(agg, key=lambda k: -agg[k][0]):
        n, rc, ap, gs = agg[st]
        print(f"  {st:<20} {n:>4} tenants   R {rc:>11,.0f}  A {ap:>11,.0f}  G {gs:>11,.0f}")
    unmatched = [x for x in trace if not x["matched"]]
    print(f"\nreceipt rows with no tenant match: {len(unmatched)}")
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    asyncio.run(main())
