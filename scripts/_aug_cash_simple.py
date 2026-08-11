"""
Simple August-2026 cash comparison — one tab, one row per tenant.

Columns: Room | Name | Date | Receipt Book | App | Google Sheet | Match?

Reuses the matching logic in _reconcile_aug_cash.py — do not duplicate it here.
Output: data/reports/Aug2026_Cash_Simple.xlsx
"""
from __future__ import annotations

import asyncio
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from _reconcile_aug_cash import build, load_app, load_gsheet, load_receipt, room_key

OUT = Path(__file__).resolve().parents[1] / "data/reports/Aug2026_Cash_Simple.xlsx"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = PatternFill("solid", fgColor="1F3864")
RED = PatternFill("solid", fgColor="FFC7CE")
BAND = PatternFill("solid", fgColor="DDEBF7")
INR = '#,##0;[Red]-#,##0;"-"'

COLS = ["Room", "Name", "Date", "Receipt Book", "App", "Google Sheet", "Match?"]


def date_str(tenant) -> str:
    """Prefer the receipt-book date; fall back to the app payment date."""
    src = tenant.receipt_dates or tenant.app_dates
    return ", ".join(sorted({d.strftime("%d-%b") for d in src if d}))


async def main():
    receipt, gsheet = load_receipt(), load_gsheet()
    app = await load_app()
    tenants, _ = build(receipt, app, gsheet)

    rows = sorted(
        [t for t in tenants if (t.receipt or t.app or t.gsheet)],
        key=lambda t: (room_key(t.room), t.display),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aug 2026 Cash"

    ws.append(["AUGUST 2026 — CASH COLLECTED PER TENANT"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(3, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in rows:
        ok = round(t.receipt) == round(t.app) == round(t.gsheet)
        ws.append([
            t.room,
            t.display.title(),
            date_str(t),
            t.receipt or None,
            t.app or None,
            t.gsheet or None,
            "" if ok else "CHECK",
        ])
        r = ws.max_row
        for c in range(1, len(COLS) + 1):
            ws.cell(r, c).border = BORDER
        for c in (4, 5, 6):
            ws.cell(r, c).number_format = INR
        if not ok:
            for c in range(1, len(COLS) + 1):
                ws.cell(r, c).fill = RED

    last = ws.max_row
    ws.append([])
    ws.append(["", "TOTAL", ""] + [f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{last})"
                                   for c in (4, 5, 6)] + [""])
    tr = ws.max_row
    for c in range(1, len(COLS) + 1):
        ws.cell(tr, c).font = Font(bold=True)
        ws.cell(tr, c).fill = BAND
        ws.cell(tr, c).border = BORDER
    for c in (4, 5, 6):
        ws.cell(tr, c).number_format = INR

    for i, w in enumerate([9, 32, 22, 15, 13, 15, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{last}"

    wb.save(OUT)
    n_check = sum(1 for t in rows if not (round(t.receipt) == round(t.app) == round(t.gsheet)))
    print(f"{len(rows)} tenants, {n_check} marked CHECK")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
