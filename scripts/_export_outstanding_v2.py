"""
Outstanding dues report — reads directly from Google Sheets APRIL/MAY 2026 tabs.
April: uses Sheet cash + UPI as-is (receptionist entries, backed by Lakshmi app).
May: uses Sheet cash + UPI, then cross-checks against May bank statement files.
      Where bank > Sheet UPI, uses bank amount (real payment not yet entered in Sheet).

Output: data/reports/April_Outstanding_2026.xlsx + May_Outstanding_2026.xlsx
Each file: ALL Buildings tab (first) + HULK tab + THOR tab.
Columns: #, Building, Name, Room, Check-in, Rent, Rent Due, Basis, Cash, UPI, Total Paid, Balance, Status
"""
import os, csv, re, openpyxl
import gspread
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ── Google Sheets ─────────────────────────────────────────────────────────────
creds = Credentials.from_service_account_file(
    'credentials/gsheets_service_account.json',
    scopes=['https://www.googleapis.com/auth/spreadsheets']
)
gc = gspread.authorize(creds)
sh = gc.open_by_key('1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw')

def safe_float(v):
    try: return float(str(v).replace(',', '').replace('Rs.', '').strip() or 0)
    except: return 0

def read_month_sheet(ws_name):
    ws = sh.worksheet(ws_name)
    rows = ws.get_all_values()
    H = {h: i for i, h in enumerate(rows[7])}
    tenants = []
    for row in rows[8:]:
        if len(row) <= H['UPI']: continue
        room = row[H['Room']].strip()
        name = row[H['Name']].strip()
        if not room or not name: continue
        rent_due = safe_float(row[H['Rent Due']])
        if rent_due <= 0: continue
        prev_due = safe_float(row[H['Prev Due']]) if 'Prev Due' in H and len(row) > H['Prev Due'] else 0
        building = row[H['Building']].strip()
        bname = 'HULK' if 'HULK' in building.upper() else 'THOR'
        checkin_raw = row[H['Check-in']].strip() if 'Check-in' in H and len(row) > H['Check-in'] else ''
        # Parse check-in date for display
        checkin = None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                from datetime import datetime
                checkin = datetime.strptime(checkin_raw, fmt).date()
                break
            except: pass
        rent_str = row[H['Rent']].strip() if 'Rent' in H and len(row) > H['Rent'] else '0'
        rent = safe_float(rent_str)
        tenants.append({
            'room': room, 'name': name, 'phone': row[H['Phone']].strip(),
            'building': bname, 'rent': rent,
            'rent_due': rent_due, 'prev_due': prev_due,
            'cash': safe_float(row[H['Cash']]),
            'upi': safe_float(row[H['UPI']]),
            'bal': safe_float(row[H['Balance']]),
            'status': row[H['Status']].strip(),
            'checkin': checkin,
        })
    return tenants

# ── May bank statement (for cross-check) ─────────────────────────────────────
def extract_phone(upi_id):
    m = re.match(r'^(\d{10})(?:-\d+)?@', str(upi_id))
    return m.group(1) if m else None

def normalize_phone(ph):
    if not ph: return None
    ph = re.sub(r'\D', '', str(ph))
    if ph.startswith('91') and len(ph) == 12: ph = ph[2:]
    return ph if len(ph) == 10 else None

def name_tokens(s):
    return [p for p in re.sub(r'[^A-Z0-9 ]', ' ', s.upper()).split() if len(p) > 1]

bank_txns = []
wb = openpyxl.load_workbook('Hulk may 11th bank statement.xlsx')
for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
    if row[11] == 'SUCCESS' and row[3]:
        bank_txns.append({'amt': float(row[3]), 'phone': extract_phone(row[7]),
                           'name': str(row[8]).strip().upper(), 'used': False})
with open('thor may 11th.csv', newline='') as f:
    for row in csv.DictReader(f):
        if row['Settlement_Status'] == 'SUCCESS' and row['TXN_AMOUNT']:
            bank_txns.append({'amt': float(row['TXN_AMOUNT']), 'phone': extract_phone(row['Payer_VPA']),
                               'name': row['Payer_Name'].strip().upper(), 'used': False})

def find_bank_total(name, phone, txns):
    tphone = normalize_phone(phone)
    name_up = name.strip().upper()
    tparts = name_tokens(name_up)
    total = 0
    if tphone:
        for t in txns:
            if normalize_phone(t['phone']) == tphone and not t['used']:
                total += t['amt']; t['used'] = True
    if total > 0: return total
    for t in txns:
        if not t['used'] and t['name'] == name_up:
            t['used'] = True; return t['amt']
    if len(tparts) >= 2:
        best, bi = 0, None
        for i, t in enumerate(txns):
            if t['used']: continue
            bparts = name_tokens(t['name'])
            hits = sum(1 for tp in tparts if any(tp in bp or bp in tp for bp in bparts))
            score = hits / len(tparts)
            if hits >= 2 and score >= 0.6 and score > best:
                best, bi = score, i
        if bi is not None:
            txns[bi]['used'] = True; return txns[bi]['amt']
    return 0

# ── Read both months ──────────────────────────────────────────────────────────
print('Reading April 2026 sheet...')
apr_tenants = read_month_sheet('APRIL 2026')
print(f'  {len(apr_tenants)} tenants with dues')

print('Reading May 2026 sheet...')
may_tenants_raw = read_month_sheet('MAY 2026')

# For May: check bank for each tenant; use bank_upi if bank > sheet_upi
may_tenants = []
bank_corrections = []
for t in may_tenants_raw:
    bank_upi = find_bank_total(t['name'], t['phone'], bank_txns)
    effective_upi = t['upi']
    correction = 0
    if bank_upi > t['upi'] + 100:
        correction = bank_upi - t['upi']
        effective_upi = bank_upi
    true_total = effective_upi + t['cash']
    true_bal = t['rent_due'] + t['prev_due'] - true_total
    if correction > 0:
        bank_corrections.append({'room': t['room'], 'name': t['name'],
                                  'sheet_upi': t['upi'], 'bank_upi': bank_upi,
                                  'correction': correction, 'new_bal': true_bal})
    may_tenants.append({**t, 'upi': effective_upi, 'total': true_total,
                        'bal': true_bal, 'bank_corrected': correction > 0})

print(f'  {len(may_tenants)} tenants with dues')
print(f'  {len(bank_corrections)} tenants had UPI corrected from bank statement:')
for c in sorted(bank_corrections, key=lambda x: -x['correction']):
    print(f'    Room {c["room"]:>5}  {c["name"]:<28}  Sheet UPI {c["sheet_upi"]:>7,.0f}  Bank {c["bank_upi"]:>7,.0f}  +{c["correction"]:,.0f}')

# Unmatched bank entries
unmatched = [(t['name'], t['amt']) for t in bank_txns if not t['used'] and t['amt'] >= 5000]
unmatched.sort(key=lambda x: -x[1])
if unmatched:
    print(f'\n  {len(unmatched)} bank entries >=5000 not matched to any Sheet tenant:')
    for name, amt in unmatched:
        print(f'    Rs.{amt:>8,.0f}  {name}')

# ── Excel styles ──────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill('solid', fgColor='1F3864')
HDR_FONT    = Font(bold=True, color='FFFFFF', size=10)
PARTIAL_FILL= PatternFill('solid', fgColor='FFF2CC')
NOPAY_FILL  = PatternFill('solid', fgColor='FCE4D6')
CORR_FILL   = PatternFill('solid', fgColor='E2EFDA')  # green for bank-corrected
TOTAL_FILL  = PatternFill('solid', fgColor='BDD7EE')
TOTAL_FONT  = Font(bold=True, size=10)
NORM_FONT   = Font(size=10)
thin        = Side(style='thin', color='BBBBBB')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT       = Alignment(horizontal='right', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')

COL_HEADERS = ['#', 'Building', 'Name', 'Room', 'Check-in', 'Rent', 'Rent Due', 'Prev Due',
               'Cash', 'UPI', 'Total Paid', 'Balance', 'Status']
COL_WIDTHS  = [4, 8, 28, 6, 12, 10, 10, 10, 10, 10, 10, 10, 20]
NUM_COLS    = {6, 7, 8, 9, 10, 11, 12}  # 1-indexed: Rent, Rent Due, Prev Due, Cash, UPI, Total, Balance

def write_sheet(ws, label, rows_out, is_may=False):
    ws.merge_cells(f'A1:{get_column_letter(len(COL_HEADERS))}1')
    ws['A1'] = label
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22

    row_n = 3
    t_due = t_prev = t_cash = t_upi = t_tot = t_bal = 0
    partial = [r for r in rows_out if r['total'] > 0]
    nopay   = [r for r in rows_out if r['total'] == 0]

    for section, rows_s, base_fill in [
        ('— Partial / Carry-over (sorted by balance) —', partial, PARTIAL_FILL),
        ('— No Payment —', nopay, NOPAY_FILL)
    ]:
        if not rows_s: continue
        ws.merge_cells(f'A{row_n}:{get_column_letter(len(COL_HEADERS))}{row_n}')
        ws.cell(row=row_n, column=1, value=section).font = Font(bold=True, italic=True, size=10, color='595959')
        ws.cell(row=row_n, column=1).alignment = LEFT
        row_n += 1
        for r in rows_s:
            fill = CORR_FILL if is_may and r.get('bank_corrected') else base_fill
            vals = [r['no'], r['building'], r['name'], r['room'],
                    r['checkin'], r['rent'], r['rent_due'], r['prev_due'],
                    r['cash'], r['upi'], r['total'], r['bal'], r['status']]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=ci, value=v)
                cell.fill = fill; cell.border = BORDER; cell.font = NORM_FONT
                if ci in NUM_COLS:
                    cell.alignment = RIGHT; cell.number_format = '#,##0'
                elif ci == 5:
                    cell.alignment = CENTER; cell.number_format = 'DD-MMM-YYYY'
                else:
                    cell.alignment = LEFT
            row_n += 1
            t_due += r['rent_due']; t_prev += r['prev_due']
            t_cash += r['cash']; t_upi += r['upi']
            t_tot += r['total']; t_bal += r['bal']

    totals = [None, '', 'TOTAL', '', None, None, t_due, t_prev, t_cash, t_upi, t_tot, t_bal, '']
    for ci, v in enumerate(totals, 1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER
        if ci in NUM_COLS and isinstance(v, (int, float)):
            cell.alignment = RIGHT; cell.number_format = '#,##0'

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'
    return t_bal

# ── Build per-month outputs ───────────────────────────────────────────────────
def build_rows(tenants):
    """Filter to outstanding only and format for sheet writing."""
    rows = []
    for t in tenants:
        bal = t['bal']
        if bal <= 0: continue
        total = t.get('total', t['upi'] + t['cash'])
        rows.append({
            'building': t['building'], 'name': t['name'], 'room': t['room'],
            'checkin': t.get('checkin'), 'rent': t['rent'],
            'rent_due': t['rent_due'], 'prev_due': t.get('prev_due', 0),
            'cash': t['cash'], 'upi': t['upi'], 'total': total, 'bal': bal,
            'status': 'No Payment' if total == 0 else f'{int(bal):,} outstanding',
            'bank_corrected': t.get('bank_corrected', False),
            'no': 0,
        })
    rows.sort(key=lambda x: -x['bal'])
    for i, r in enumerate(rows, 1): r['no'] = i
    return rows

MONTHS = [
    ('April', apr_tenants, False,
     'data/reports/April_Outstanding_2026.xlsx',
     'HULK + THOR — Outstanding Dues — April 2026  (as of 11 May 2026)'),
    ('May', may_tenants, True,
     'data/reports/May_Outstanding_2026.xlsx',
     'HULK + THOR — Outstanding Dues — May 2026  (as of 11 May 2026)'),
]

for month_name, tenants, is_may, filepath, all_label in MONTHS:
    wb_out = openpyxl.Workbook()
    if wb_out.active: wb_out.remove(wb_out.active)

    all_rows = build_rows(tenants)
    all_rows_combined = list(all_rows)
    for i, r in enumerate(all_rows_combined, 1): r['no'] = i

    ws_all = wb_out.create_sheet('ALL Buildings', 0)
    total_bal = write_sheet(ws_all, all_label, all_rows_combined, is_may)

    for bname in ['HULK', 'THOR']:
        brows = [r for r in all_rows if r['building'] == bname]
        for i, r in enumerate(brows, 1): r['no'] = i
        ws = wb_out.create_sheet(bname)
        lbl = f'{bname} — Outstanding Dues — {month_name} 2026  (as of 11 May 2026)'
        write_sheet(ws, lbl, brows, is_may)

    os.makedirs('data/reports', exist_ok=True)
    wb_out.save(filepath)
    note = ' (green = UPI corrected from bank)' if is_may else ''
    print(f'\nSaved {filepath}')
    print(f'  {len(all_rows_combined)} outstanding tenants, total balance Rs.{total_bal:,.0f}{note}')

print('\nDone.')
