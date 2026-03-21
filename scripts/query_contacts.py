"""
Query pg_contacts from Supabase.

Usage:
  python scripts/query_contacts.py                   # all contacts
  python scripts/query_contacts.py plumber            # filter by category
  python scripts/query_contacts.py electrician        # filter by category
  python scripts/query_contacts.py "Ibrahim"          # search by name
  python scripts/query_contacts.py --export           # export to Excel
"""
import asyncio
import sys
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import text

DB_URL = "postgresql+asyncpg://postgres:Anchorstrong123!@db.oxiqomoilqwfxjauxhzp.supabase.co:5432/postgres"


async def main():
    engine = create_async_engine(DB_URL)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    query = sys.argv[1] if len(sys.argv) > 1 else None
    export = query == "--export"
    if export:
        query = None

    async with async_session() as session:
        if query:
            # Search by category OR name (case-insensitive)
            res = await session.execute(
                text("""
                    SELECT id, name, phone, category, contact_for, amount_paid, remaining, comments
                    FROM pg_contacts
                    WHERE LOWER(category) LIKE :q
                       OR LOWER(name) LIKE :q
                       OR LOWER(contact_for) LIKE :q
                    ORDER BY category, name
                """),
                {"q": f"%{query.lower()}%"},
            )
        else:
            res = await session.execute(
                text("""
                    SELECT id, name, phone, category, contact_for, amount_paid, remaining, comments
                    FROM pg_contacts
                    ORDER BY category, name
                """)
            )

        rows = res.fetchall()

        if export:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            except ImportError:
                print("pip install openpyxl first")
                await engine.dispose()
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PG Contacts"
            headers = ["ID", "Name", "Phone", "Category", "Contact For", "Amount Paid", "Remaining", "Comments"]
            hfont = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
            widths = [5, 30, 18, 15, 45, 12, 20, 35]
            for i, (h, w) in enumerate(zip(headers, widths), 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = hfont
                c.fill = hfill
                c.border = border
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            for ri, row in enumerate(rows, 2):
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=str(val) if val else "")
                    c.border = border
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            outpath = "data/reports/pg_contacts_export.xlsx"
            wb.save(outpath)
            print(f"Exported {len(rows)} contacts to {outpath}")
        else:
            if not rows:
                print(f"No contacts found{f' matching: {query}' if query else ''}.")
                await engine.dispose()
                return

            current_cat = None
            for row in rows:
                _id, name, phone, cat, contact_for, paid, remaining, comments = row
                cat_display = (cat or "other").replace("_", " ").title()
                if cat_display != current_cat:
                    current_cat = cat_display
                    print(f"\n{'='*60}")
                    print(f"  {current_cat}")
                    print(f"{'='*60}")
                phone_str = phone or "-"
                print(f"  {name or 'Unknown':<30} {phone_str:<15} {contact_for or ''}")
                if paid:
                    print(f"    Paid: {paid}  |  Remaining: {remaining or 'N/A'}")

            print(f"\nTotal: {len(rows)} contacts")

    await engine.dispose()


asyncio.run(main())
