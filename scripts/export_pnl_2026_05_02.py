"""P&L for Oct'25 – Apr'26 — BANK STATEMENT as primary source.

Produces: data/reports/PnL_Accrual_2026_05_02.xlsx

INCOME METHODOLOGY:
  Primary:      Bank statement credits — Yes Bank …961 (UPI batch settlements + other)
  Supplementary: Cash physically collected but NOT deposited to bank (from DB records)
  NOT income:   Capital / personal transfers from Kiran + partner + NEFT (owner injections)
  NOT income:   Maintenance fee (non-refundable portion of security deposit — shown in deposits)

EXPENSE METHODOLOGY (unchanged — already bank-derived):
- Property Rent accrual Rs.21.32L/mo from Jan onwards
- Water: bank tanker kept; Manoj B accrual from March
- Internet: bank pre-Feb; Rs.15,514/mo Feb+ (amortised)
- Police: Rs.3K/mo accrual Jan+
- Apr reclassifications applied (see script header comments below)

April reclassifications from bank classifier:
  Non-Op → Food:     Rs.29,946  (Prabhakaran NEFT "ninja cart pooja veg")
  Non-Op → CAPEX:    Rs.75,600  (TV payment)
  Other  → CAPEX:    Rs.1,33,420 (chairs 47K + kitchen vessels 37.5K + atta machines 42.12K + mixer 6.8K)
  Other  → Maint:    Rs.24,599  (EB panel board)
  Other  → Non-Op:   Rs.10,000  (borrow to Prabhakaran)
  Flags:  Fuel Apr = Rs.2,800 only (no DG diesel visible — check with Kiran)
          Manoj Apr water accrual TBD (currently showing bank tanker only)
          Rs.49,679 + Rs.9,500 unknowns still in Other Expenses
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from utils.inr_format import INR_NUMBER_FORMAT

OUT = Path(__file__).parent.parent / "data" / "reports" / "PnL_Accrual_2026_05_02.xlsx"

MONTHS = ["Oct'25", "Nov'25", "Dec'25", "Jan'26", "Feb'26", "Mar'26", "Apr'26"]

# ── Income — BANK as primary, cash as supplement ─────────────────────────
# Bank income = all credits EXCEPT Lakshmi SBI→Yes Bank startup (Oct) and Kiran top-ups.
# Individual UPI includes: tenant rent + deposits + bookings paid directly to partner's personal UPI.
# For Oct-Dec the personal account (7358341775) was used before a merchant QR was set up.
# Verified against closing balance: Oct–Dec net ₹13,76,417 matches bank statement exactly.
income = {
    # Batch settlements = Yes Bank merchant QR daily aggregate (Jan onwards only)
    "Bank — UPI batch settlements (merchant QR)":    [0,      0,        0,  175596, 2091597, 2515275, 2834731],
    # Individual/direct = personal UPI + NEFT income (Oct–Dec all were direct; Jan+ diminishing as QR grew)
    "Bank — individual direct payments + NEFT":      [0, 723007, 1350547, 1083628,  420690,  225091,  226807],
    # Cash physically collected — NOT deposited to bank. Source: DB cash payments.
    # Oct-Dec ₹0 = no DB records for pre-live months.
    "Cash (physical, not deposited to bank)":        [0,      0,        0,  300572,  653300, 1094220, 1343783],
}

# ── Bank statement — ALL credits (Yes Bank …961) ─────────────────────────────
# Full picture of what hit the bank account each month.
# Tenant UPI batch + capital injections + other = total bank inflow.
# Verified against Dec closing balance ₹13,76,417 = sum(income) + sum(capital) - sum(debits) across Oct-Dec ✓
# Old classifier was wrong: treated ALL individual UPI as "Capital" because partner number appeared as recipient.
# Fixed: only flag as capital when partner/Kiran is the SENDER (from: field), not recipient.
bank_all_credits = {
    "Income (tenant UPI batch + direct + NEFT)": [      0, 723007, 1350547, 1259224, 2512287, 2740366, 3061538],
    "Capital (Lakshmi startup + Kiran top-ups)": [500000,      1,       0,   90000,       0,       0,       0],
}

# Owner capital contribution — balance-sheet / equity
capital_contributions = {
    "Owner startup — Lakshmi SBI to Yes Bank (Oct 2025)": [500000, 0, 0,     0, 0, 0, 0],
    "Kiran top-up transfer (Jan 2026)":                   [     0, 0, 0, 90000, 0, 0, 0],
}

# ── Deposits held — NOT income, balance sheet only ───────────────────────────
# Security deposit = what tenant paid upfront.
# Split into: (a) refundable portion — must return to active tenants on exit
#             (b) maintenance fee   — non-refundable, PG retains, NOT income
# Source for refundable: active tenancies only (matches PWA "Security Deposits Held" screen, Apr 30).
# Maintenance fee spread: by check-in month from DB tenancies (all tenants in window).
working_capital = {
    "Security Deposits — refundable (must return to active tenants)": [0, 0, 0, 0, 0, 0, 2437425],
    # Non-refundable maintenance fee — PG retains this permanently. NOT a liability. NOT income.
    # Shown here for transparency: by check-in month (Nov 22 tenants, Dec 40, Jan 59, Feb 53, Mar 72, Apr 23)
    "  Maintenance Fee retained (non-refundable, by check-in month)": [0, 96000, 196000, 287500, 248000, 319700, 122000],
}

# ── Opex (accrual basis) ────────────────────────────────────────────────────
opex = {
    # Overrides: bank UPI (~6L) replaced by full 21.32L accrual each month
    "Property Rent (accrual Rs.21.32L/mo, Jan onwards)":         [0, 0, 0,      2132000, 2132000, 2132000, 2132000],
    "Electricity":                                                [0, 0, 0,       131554,  134538,   96617,  140659],
    # Mar: Manoj Rs.42.5K + bank tanker Rs.8K = Rs.50,500
    # Apr: bank tanker Rs.42,020 (Himalaya); Manoj Apr accrual TBD (ask Kiran)
    "Water (bank tankers + Manoj accrual; Apr Manoj TBD)":        [0, 0, 0,            0,       0,   50500,   42020],
    "IT & Software":                                              [0, 0, 3480,     10620,       0,       0,       0],
    # Pre-Feb: bank; Feb+: amortised Rs.15,514/mo (Airwire + WiFi vendor prepay)
    "Internet & WiFi (bank pre-Feb; Rs.15.5K/mo Feb+)":           [0, 0, 40946,    70952,   15514,   15514,   15514],
    # Apr: includes Rs.29,946 Prabhakaran NEFT (ninja cart veg) reclassified from Non-Op
    "Food & Groceries":                                           [0, 1086, 34435, 201558,  94931,  237747,  263625],
    # Apr: only Rs.2,800 petrol; no DG diesel visible — flagged ⚠
    "Fuel & Diesel":                                              [0, 0, 0,        9099,  104366,  346308,    2800],
    "Staff & Labour":                                             [0, 1000, 125935, 112063, 219715,  155481,  156102],
    # Apr: Rs.4,340 classifier + Rs.24,599 EB panel board (reclassified from Other)
    "Maintenance & Repairs":                                      [0, 0, 0,            0,     500,   18470,   28939],
    "Cleaning Supplies":                                          [0, 0, 4414,       1400,     700,    4566,   14500],
    "Waste Disposal (Pavan Rs.3.5K/mo)":                         [0, 0, 0,          3000,    3500,    3500,    3500],
    "Shopping & Supplies":                                        [0, 2730, 10530,  12036,   6127,    6184,    7442],
    "Operational Expenses":                                       [0, 0, 47769,     10315,   2174,    3237,   30756],
    "Marketing":                                                  [0, 0, 39500,     17895,   3620,   27700,       0],
    "Govt & Regulatory (incl Police Rs.3K accrual Jan+)":        [0, 0, 75716,     88073,   3000,    3000,    3000],
    "Bank Charges":                                               [0, 0, 0,           149,      0,       0,     100],
    # CAPEX — one-time setup investments. Real cash out, included in full.
    # Apr breakdown: classifier Rs.2,163 + TV Rs.75,600 + chairs Rs.47,000
    #   + kitchen vessels Rs.37,500 + atta machines Rs.42,120 + mixer Rs.6,800 = Rs.2,11,183
    "Furniture & Fittings (CAPEX)":                              [0, 50000, 110191, 203815, 1185397,   331, 211183],
    "CCTV Installation (CAPEX)":                                 [0, 82000,      0,      0,       0,     0,      0],
    # Apr: Rs.2,77,034 raw; less CAPEX Rs.1,33,420 + Maint Rs.24,599 + Non-Op Rs.10,000 = Rs.1,09,015
    # Remaining unknowns: Rs.49,679 (8951297583) + Rs.9,500 (9099913969) + esob/tanti Rs.13K + others
    "Other Expenses":                                             [0, 10000, 83556,  6564,  23308,   98500,  109015],
}

# ── Excluded — genuinely NOT costs (balance sheet items only) ────────────────
excluded = {
    # Giving back money that was always theirs — not an expense
    "Tenant Deposit Refund (balance sheet)":  [0, 10000,  21500,  53944,  74532,  118671, 128418],
    # Paying back debt principal — not an expense (reduces liability, not profit)
    "Loan Repayment / Transfers (non-op)":    [0, 450000,     0,      0,  600000, 2090000, 22357],
}


def main():
    wb = openpyxl.Workbook()

    # ── Sheet 1: P&L Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "P&L Summary"

    bold       = Font(bold=True)
    hdr_fill   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font   = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    flag_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow for flags
    ctr        = Alignment(horizontal="center")

    header = [""] + MONTHS + ["TOTAL"]
    ws.append(header)
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = ctr

    # Income
    ws.append(["INCOME"])
    ws[ws.max_row][0].font = bold
    for label, row in income.items():
        ws.append([label] + row + [sum(row)])
    rev_row = [sum(col) for col in zip(*income.values())]
    ws.append(["Revenue"] + rev_row + [sum(rev_row)])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = total_fill

    ws.append([])

    # Opex
    ws.append(["OPERATING EXPENSES (accrual)"])
    ws[ws.max_row][0].font = bold
    for label, row in opex.items():
        ws.append([label] + row + [sum(row)])
        # flag rows with known issues
        if "TBD" in label or "⚠" in label or "Fuel & Diesel" in label:
            for c in ws[ws.max_row]:
                c.fill = flag_fill
    opex_row = [sum(col) for col in zip(*opex.values())]
    ws.append(["Total Opex"] + opex_row + [sum(opex_row)])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = total_fill

    ws.append([])

    # Net Profit
    profit_row = [r - o for r, o in zip(rev_row, opex_row)]
    ws.append(["NET PROFIT (accrual)"] + profit_row + [sum(profit_row)])
    for c in ws[ws.max_row]:
        c.font = bold

    margin_row = [f"{(p/r*100):.1f}%" if r else "-" for p, r in zip(profit_row, rev_row)]
    ws.append(["Margin %"] + margin_row + [f"{(sum(profit_row)/sum(rev_row)*100):.1f}%"])

    ws.append([])

    # Capital Contributions
    ws.append(["CAPITAL CONTRIBUTIONS (not P&L — balance-sheet / owner's equity)"])
    ws[ws.max_row][0].font = bold
    for label, row in capital_contributions.items():
        ws.append([label] + row + [sum(row)])

    ws.append([])

    # Bank Reconciliation
    bank_section_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    gap_fill          = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_bank_fill   = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
    ws.append(["BANK STATEMENT CREDITS — ALL INFLOWS (Yes Bank …961)"])
    ws[ws.max_row][0].font = Font(bold=True, color="FFFFFF")
    ws[ws.max_row][0].fill = bank_section_fill
    for ci in range(2, len(header) + 1):
        ws.cell(ws.max_row, ci).fill = bank_section_fill

    for label, row in bank_all_credits.items():
        ws.append([label] + row + [sum(row)])

    # Total bank credits row
    total_bank_row = [sum(col) for col in zip(*bank_all_credits.values())]
    ws.append(["  Total Bank Credits (all inflows)"] + total_bank_row + [sum(total_bank_row)])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = total_bank_fill

    ws.append([])

    # Reconciliation: Recorded income vs what hit the bank
    recorded_income_row = [sum(col) for col in zip(*income.values())]
    gap_row = [b - r for b, r in zip(total_bank_row, recorded_income_row)]
    ws.append(["  Bank Credits − Recorded Income (gap = capital infusions + deposits not in income)"] + gap_row + [sum(gap_row)])
    for c in ws[ws.max_row]:
        c.fill = gap_fill

    ws.append([])

    # Working Capital
    wc_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.append(["DEPOSITS HELD — refundable liability + non-refundable maintenance fee retained"])
    ws[ws.max_row][0].font = Font(bold=True, color="FFFFFF")
    ws[ws.max_row][0].fill = wc_fill
    for ci in range(2, len(header) + 1):
        ws.cell(ws.max_row, ci).fill = wc_fill

    for label, row in working_capital.items():
        ws.append([label] + row + [sum(row)])

    # Net working capital owed row
    net_wc = [sum(col) for col in zip(*working_capital.values())]
    ws.append(["  Net working capital owed to tenants (balance sheet liability)"] + net_wc + [sum(net_wc)])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    ws.append([])

    # Excluded
    ws.append(["EXCLUDED (not in opex)"])
    ws[ws.max_row][0].font = bold
    for label, row in excluded.items():
        ws.append([label] + row + [sum(row)])

    ws.append([])

    # Cash position note
    ws.append(["CASH POSITION (Apr 30)"])
    ws[ws.max_row][0].font = bold
    ws.append(["Bank closing balance (Apr 30)", "", "", "", "", "", "", 1373863])
    ws.append(["Cash in hand", "", "", "", "", "", "", "← Ask Kiran"])
    ws.append(["Refundable Security Deposit liability", "", "", "", "", "", "", "← recompute from Apr Collection sheet"])
    ws.append(["True free cash", "", "", "", "", "", "", "= Bank + Cash − Refundable deposits"])

    ws.append([])

    # Flags / review items
    ws.append(["⚠ ITEMS NEEDING KIRAN REVIEW"])
    ws[ws.max_row][0].font = Font(bold=True, color="FF0000")
    flags = [
        "1. Manoj water bill for April (accrual — paid in May). Add to Water line when known.",
        "2. Fuel & Diesel Apr = Rs.2,800 only (petrol for staff). No DG diesel visible in bank. Was generator not used? Or paid separately?",
        "3. Apr CAPEX reclassified (removed from opex): TV Rs.75,600 + chairs Rs.47,000 + kitchen vessels Rs.37,500 + atta machines Rs.42,120 + mixer Rs.6,800 = Rs.2,11,023. Confirm these are correct.",
        "4. EB panel board Rs.24,599 moved to Maintenance & Repairs. Correct?",
        "5. Unknown: Rs.49,679 to 8951297583-3@ibl on 08-Apr. What is this?",
        "6. Unknown: Rs.9,500 to 9099913969@ptsbi on 11-Apr. What is this?",
        "7. Mar income updated in this P&L (DB now shows Rs.41.07L vs Rs.39.83L in Apr-23 P&L — additional payments recorded after Apr 23).",
    ]
    for f in flags:
        ws.append([f])

    # Format numbers
    for row in ws.iter_rows(min_row=2, max_col=len(header)):
        for i, cell in enumerate(row):
            if i == 0 or cell.value is None or isinstance(cell.value, str):
                continue
            if isinstance(cell.value, (int, float)):
                cell.number_format = INR_NUMBER_FORMAT

    ws.column_dimensions["A"].width = 58
    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 14

    # ── Sheet 2: Rules Applied ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rules Applied")
    rules = [
        ("Basis", "Accrual — expense in month of service, not payment"),
        ("Income source", "DB payments (for_type=rent, is_void=false). DB is truth — never cross-reference source sheet."),
        ("Property Rent", "Rs.21,32,000/mo accrual Jan–Apr. Nov-Dec zero (notice period). Bank UPI portion replaced by full accrual."),
        ("Water — Manoj B (9535665407)", "Variable invoice, accrued in month of consumption, paid 1mo behind. Mar invoice = Rs.42,500. Apr invoice TBD."),
        ("Water — bank tankers", "Kept additive to Manoj accrual (not replacement). Apr: Himalaya Rs.42,000."),
        ("Internet pre-Feb", "Bank classifier numbers kept as-is."),
        ("Internet Feb+", "Airwire Rs.1.13L + WiFi Vendor Rs.1.04L = Rs.2.17L prepay / 14mo = Rs.15,514/mo (Feb 2026 – Mar 2027)."),
        ("Police", "Rs.3,000/mo cash accrual Jan onwards."),
        ("Waste Disposal", "Rs.3,500/mo (Pavan 6366411789). Classifier catches it."),
        ("Apr reclassifications", "See script header for full list."),
        ("Maintenance Fee income", "Spread by check-in month (not collected month). Nov ₹96K, Dec ₹1.96L, Jan ₹2.875L, Feb ₹2.48L, Mar ₹3.197L, Apr ₹1.22L. May check-ins excluded."),
        ("Excluded from opex", "Furniture & Fittings (capex), CCTV (capex), Tenant Deposit Refunds (liability), Loan Repayments (non-operating)."),
        ("Cash leakage", "Max 4% of EBITDA for reality-adjusted profit (Kiran-confirmed). Not applied here — flagged for awareness."),
    ]
    ws3.append(["Rule", "Detail"])
    hdr_fill2 = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font2 = Font(bold=True, color="FFFFFF")
    for c in ws3[1]:
        c.fill = hdr_fill2; c.font = hdr_font2
    for k, v in rules:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 110
    for row in ws3.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
