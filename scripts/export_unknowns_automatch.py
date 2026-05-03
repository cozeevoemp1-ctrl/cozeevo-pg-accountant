"""
scripts/export_unknowns_automatch.py
--------------------------------------
For every remaining "Other Expenses" transaction:
  1. Extract phone number from UPI handle (e.g. 9904388966-2@ybl → 9904388966)
  2. Cross-reference against ALL tenants in DB (active + exited)
  3. Cross-reference against known vendor keywords in pnl_classify.py
  4. Output Excel with match column — confirmed unknowns stay blank for Kiran

Run:
  venv/Scripts/python scripts/export_unknowns_automatch.py
"""
import sys, os, glob as _glob, re, asyncio
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

from src.parsers.yes_bank import parse_date, parse_amt, read_yes_bank_csv
from src.rules.pnl_classify import classify_txn, EXPENSE_RULES
from src.utils.inr_format import INR_NUMBER_FORMAT

# ── Extract phone number from UPI handle ──────────────────────────────────────
PHONE_RE = re.compile(r"\b([6-9]\d{9})\b")
UPI_RE   = re.compile(r"[a-zA-Z0-9.\-_+]+@[a-zA-Z]+")

def extract_phones(upi_handle: str) -> list[str]:
    """Return all 10-digit mobile numbers found in a UPI handle."""
    return PHONE_RE.findall(upi_handle.replace("-", "").replace("_", ""))

def parse_desc(desc: str) -> dict:
    result = {"txn_type": "", "payee_name": "", "payee_upi": "", "note": ""}
    if not desc or not isinstance(desc, str):
        return result
    d = desc.strip(); dl = d.lower()
    if "upi" in dl:    result["txn_type"] = "UPI"
    elif "neft" in dl: result["txn_type"] = "NEFT"
    elif "imps" in dl: result["txn_type"] = "IMPS"

    if result["txn_type"] == "UPI" and "/" in d:
        parts = [p.strip() for p in d.split("/")]
        if len(parts) >= 2 and re.match(r"^\d{10,16}$", parts[1]):
            if len(parts) > 2: result["payee_name"] = parts[2]
            if len(parts) > 3 and "@" in parts[3]: result["payee_upi"] = parts[3]
            if len(parts) > 4: result["note"] = "/".join(parts[4:])
    if not result["payee_upi"]:
        upis = UPI_RE.findall(d)
        if upis: result["payee_upi"] = upis[-1]
    return result

def _payee_key(p: dict) -> str:
    upi = (p["payee_upi"] or "").strip().lower()
    return upi if upi else (p["payee_name"] or "").strip().lower()

# ── Build vendor keyword set from classifier (for auto-tagging) ───────────────
KNOWN_VENDOR_KEYWORDS: set[str] = set()
for cat, sub, kws in EXPENSE_RULES:
    for kw in kws:
        KNOWN_VENDOR_KEYWORDS.add(kw.lower())

# ── Load all tenants from DB ──────────────────────────────────────────────────
import asyncpg
from dotenv import load_dotenv
load_dotenv()

DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")

async def load_tenants():
    conn = await asyncpg.connect(DB_URL)
    try:
        return await conn.fetch("""
            SELECT t.name, t.phone, tn.stay_type,
                   tn.status, tn.security_deposit
            FROM tenants t
            JOIN tenancies tn ON tn.tenant_id = t.id
            ORDER BY t.name
        """)
    finally:
        await conn.close()

tenants = asyncio.run(load_tenants())
print(f"Loaded {len(tenants)} tenants from DB")

# Build phone → tenant lookup
phone_to_tenant: dict[str, str] = {}
for row in tenants:
    name = row["name"] or ""
    ph_field = row["phone"]
    if ph_field:
        ph = re.sub(r"\D", "", str(ph_field))[-10:]
        if len(ph) == 10:
            phone_to_tenant[ph] = f"{name} ({row['status']} / dep={row['security_deposit']})"

print(f"Phone lookup: {len(phone_to_tenant)} entries")

# ── Load bank statements ──────────────────────────────────────────────────────
txns = []
csv_files = sorted(set(
    _glob.glob("Statement-*.csv") + _glob.glob("*statment*.csv") +
    _glob.glob("*statement*.csv") + _glob.glob("april month.csv") + _glob.glob("april*.csv")
), reverse=True)
for f in csv_files:
    txns += read_yes_bank_csv(f)

for f in ["2025 statement.xlsx", "2026 statment.xlsx"]:
    if not os.path.exists(f): continue
    import openpyxl as _opx
    wb_s = _opx.load_workbook(f, data_only=True); ws_s = wb_s.active
    rows_s = list(ws_s.iter_rows(values_only=True))
    hrow = next((i for i, r in enumerate(rows_s) if r and any("transaction date" in str(c).lower() for c in r if c)), None)
    if hrow is not None:
        headers = [str(c).lower().strip() if c else "" for c in rows_s[hrow]]
        for row in rows_s[hrow + 1:]:
            if not any(row): continue
            d2 = dict(zip(headers, row))
            dt = parse_date(d2.get("transaction date"))
            if not dt: continue
            wd = parse_amt(d2.get("withdrawals", ""))
            dep = parse_amt(d2.get("deposits", ""))
            desc2 = str(d2.get("description") or "")
            if wd > 0:  txns.append((dt, desc2, "expense", wd))
            if dep > 0: txns.append((dt, desc2, "income", dep))

seen = set(); deduped = []
for item in txns:
    dt, desc, typ, amt = item
    key = (dt.strftime("%Y-%m-%d"), round(float(amt), 2), (desc or "").strip().lower())
    if key not in seen: seen.add(key); deduped.append(item)
txns = deduped

# ── Filter to Other Expenses + auto-match ────────────────────────────────────
rows_out = []
for dt, desc, typ, amt in txns:
    if typ != "expense": continue
    cat, sub = classify_txn(desc, typ)
    if cat != "Other Expenses": continue
    p = parse_desc(desc)
    upi_handle = p["payee_upi"] or ""
    payee_name = p["payee_name"] or ""

    # Try to match phone in UPI handle against tenant DB
    phones = extract_phones(upi_handle)
    tenant_match = ""
    for ph in phones:
        if ph in phone_to_tenant:
            tenant_match = phone_to_tenant[ph]
            break

    # Try name-based match (first word of payee name against tenant names)
    name_match = ""
    if not tenant_match and payee_name:
        name_lc = payee_name.lower()
        for row in tenants:
            t_name = (row["name"] or "").lower()
            if t_name and (name_lc in t_name or t_name in name_lc):
                name_match = f"{row['name']} ({row['status']})"
                break

    auto_tag = tenant_match or (f"Name match: {name_match}" if name_match else "")
    suggestion = ""
    if tenant_match:
        suggestion = "POSSIBLE DEPOSIT REFUND"
    elif name_match:
        suggestion = "CHECK - name match"

    rows_out.append({
        "date":         dt,
        "month":        dt.strftime("%b'%y"),
        "amount":       float(amt),
        "payee_upi":    upi_handle,
        "payee_name":   payee_name,
        "note":         p["note"],
        "auto_match":   auto_tag,
        "suggestion":   suggestion,
        "description":  desc,
    })

rows_out.sort(key=lambda x: x["date"])
total = sum(r["amount"] for r in rows_out)
matched = [r for r in rows_out if r["auto_match"]]
print(f"\nRemaining Other Expenses: {len(rows_out)} txns  Rs {total:,.0f}")
print(f"Auto-matched to tenants:  {len(matched)}")
for r in matched:
    print(f"  {r['date'].strftime('%d %b')}  Rs {r['amount']:>8,.0f}  {r['payee_upi']:<35}  {r['auto_match']}")

# ── Group summary ─────────────────────────────────────────────────────────────
groups: dict[str, list] = defaultdict(list)
for r in rows_out:
    key = r["payee_upi"].lower() or r["payee_name"].lower() or "(no payee)"
    groups[key].append(r)

summary = []
for key, grp in groups.items():
    dates = sorted(r["date"] for r in grp)
    total_g = sum(r["amount"] for r in grp)
    sample = grp[0]
    auto_match = next((r["auto_match"] for r in grp if r["auto_match"]), "")
    suggestion = next((r["suggestion"] for r in grp if r["suggestion"]), "")
    summary.append({
        "payee_upi":   sample["payee_upi"] or key,
        "payee_name":  sample["payee_name"],
        "txn_count":   len(grp),
        "total_amt":   total_g,
        "first_date":  dates[0].strftime("%d %b %Y"),
        "last_date":   dates[-1].strftime("%d %b %Y"),
        "months":      ", ".join(sorted({r["month"] for r in grp})),
        "auto_match":  auto_match,
        "suggestion":  suggestion,
        "note":        sample["note"],
    })

summary.sort(key=lambda x: -x["total_amt"])

# ── Excel ─────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="EEF4FF")
WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="D9D9D9")
YLW_FILL = PatternFill("solid", fgColor="FFF2CC")
ORG_FILL = PatternFill("solid", fgColor="FCE4D6")  # orange = possible deposit
NUM_FMT  = INR_NUMBER_FORMAT
CTR      = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left", vertical="center")
RIGHT    = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Summary by Payee")
SUM_COLS = [
    ("#",                      4, "center"),
    ("Payee UPI Handle",      32, "left"),
    ("Payee Name",            22, "left"),
    ("Txns",                   6, "center"),
    ("Total Amount (Rs)",     16, "right"),
    ("First Date",            13, "center"),
    ("Last Date",             13, "center"),
    ("Months",                22, "left"),
    ("Auto-Match (Tenant?)",  30, "left"),
    ("Suggestion",            24, "left"),
    ("Is Deposit Refund? Y/N",22, "center"),
    ("If vendor — Category",  26, "left"),
]
nc = len(SUM_COLS)
ws1.merge_cells(f"A1:{get_column_letter(nc)}1")
t = ws1["A1"]
t.value = f"Other Expenses — Review after auto-match ({len(summary)} payees  Rs {sum(s['total_amt'] for s in summary):,.0f})"
t.font = Font(bold=True, size=12, color="FFFFFF"); t.fill = HDR_FILL; t.alignment = CTR
ws1.row_dimensions[1].height = 22

for ci, (h, w, _) in enumerate(SUM_COLS, 1):
    c = ws1.cell(2, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[2].height = 22

grand = 0.0
for ri, s in enumerate(summary, 1):
    r = ri + 2
    is_match = bool(s["auto_match"])
    row_fill = ORG_FILL if is_match else (ALT_FILL if ri % 2 == 0 else WHT_FILL)
    vals = [ri, s["payee_upi"], s["payee_name"], s["txn_count"], s["total_amt"],
            s["first_date"], s["last_date"], s["months"],
            s["auto_match"], s["suggestion"], "", ""]
    for ci, (val, (_, _, align)) in enumerate(zip(vals, SUM_COLS), 1):
        c = ws1.cell(r, ci, val)
        c.fill = row_fill
        if ci == 5:
            c.number_format = NUM_FMT; c.font = Font(size=9, color="CC0000"); c.alignment = RIGHT
        elif ci == 10:  # Suggestion
            c.font = Font(size=9, bold=is_match, color="7B0000" if is_match else "000000")
            c.alignment = LEFT
        elif ci == 11:  # Y/N column
            c.fill = YLW_FILL; c.font = Font(size=10, bold=True); c.alignment = CTR
        else:
            c.font = Font(size=9); c.alignment = Alignment(horizontal=align, vertical="center")
    grand += s["total_amt"]

tr = len(summary) + 3
for ci in range(1, nc + 1): ws1.cell(tr, ci).fill = TOT_FILL
ws1.cell(tr, 1, "TOTAL").font = Font(bold=True); ws1.cell(tr, 1).fill = TOT_FILL
c = ws1.cell(tr, 5, grand)
c.number_format = NUM_FMT; c.font = Font(bold=True, color="CC0000"); c.alignment = RIGHT; c.fill = TOT_FILL
ws1.freeze_panes = "A3"; ws1.auto_filter.ref = f"A2:{get_column_letter(nc)}2"

# ── Sheet 2: Detail ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("All Transactions")
DET_COLS = [
    ("#",               4, "center"),
    ("Payee UPI",      32, "left"),
    ("Payee Name",     22, "left"),
    ("Date",           12, "center"),
    ("Month",           9, "center"),
    ("Amount (Rs)",    14, "right"),
    ("Auto-Match",     30, "left"),
    ("Suggestion",     22, "left"),
    ("Note",           28, "left"),
    ("Full Description", 70, "left"),
]
nc2 = len(DET_COLS)
ws2.merge_cells(f"A1:{get_column_letter(nc2)}1")
t2 = ws2["A1"]
t2.value = f"Other Expenses — Detail ({len(rows_out)} txns  Rs {total:,.0f})"
t2.font = Font(bold=True, size=12, color="FFFFFF"); t2.fill = HDR_FILL; t2.alignment = CTR
ws2.row_dimensions[1].height = 22
for ci, (h, w, _) in enumerate(DET_COLS, 1):
    c = ws2.cell(2, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 22

grand2 = 0.0
for ri, row in enumerate(rows_out, 1):
    r = ri + 2
    is_match = bool(row["auto_match"])
    fill = ORG_FILL if is_match else (ALT_FILL if ri % 2 == 0 else WHT_FILL)
    vals = [ri, row["payee_upi"], row["payee_name"],
            row["date"].strftime("%d %b %Y"), row["month"], row["amount"],
            row["auto_match"], row["suggestion"], row["note"], row["description"]]
    for ci, (val, (_, _, align)) in enumerate(zip(vals, DET_COLS), 1):
        c = ws2.cell(r, ci, val); c.fill = fill
        if ci == 6:
            c.number_format = NUM_FMT; c.font = Font(size=9, color="CC0000"); c.alignment = RIGHT
        else:
            c.font = Font(size=9, bold=(is_match and ci in (7, 8)))
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align=="left" and ci==10))
    grand2 += row["amount"]

tr2 = len(rows_out) + 3
for ci in range(1, nc2 + 1): ws2.cell(tr2, ci).fill = TOT_FILL
ws2.cell(tr2, 1, "TOTAL").font = Font(bold=True); ws2.cell(tr2, 1).fill = TOT_FILL
c2 = ws2.cell(tr2, 6, grand2)
c2.number_format = NUM_FMT; c2.font = Font(bold=True, color="CC0000"); c2.alignment = RIGHT; c2.fill = TOT_FILL
ws2.freeze_panes = "A3"; ws2.auto_filter.ref = f"A2:{get_column_letter(nc2)}2"

OUT = os.path.join("data", "reports", "other_expenses_automatch.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Summary: {len(summary)} payees  |  Detail: {len(rows_out)} txns  Rs {grand2:,.0f}")
