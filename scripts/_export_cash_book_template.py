"""Monthly CASH BOOK template (Kiran 2026-08-08) — one sheet per month.

Formulas compute expected closing + variance automatically, so a mismatch
like July's +64,841 shows up in red the moment numbers go in.
Sheet 1 = blank AUG 2026 ready to fill; Sheet 2 = JUL 2026 worked example
with Kiran's real numbers (variance visible).

Run:  venv/Scripts/python scripts/_export_cash_book_template.py
Out:  data/reports/Cash_Book_2026.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="332D29")
SEC_FILL = PatternFill("solid", fgColor="EAE3DC")
OK_FILL = PatternFill("solid", fgColor="E2F1E8")
WARN_FILL = PatternFill("solid", fgColor="F9E6E1")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))

ROWS = [
    # (label, kind) kind: sec=section header, in=cash in, out=cash out, calc/count/var
    ("A. OPENING — cash in hand on the 1st (= last month's counted closing)", "sec"),
    ("Opening cash in hand", "open"),
    ("B. CASH IN (every rupee that entered the box)", "sec"),
    ("Rent / dues collected — recorded in app", "in"),
    ("Deposits + booking advances collected in cash (app)", "in"),
    ("Cash collected but NOT in app (name each in Notes!)", "in"),
    ("Loan / chit repayments received in cash", "in"),
    ("Cash withdrawn from bank into the box", "in"),
    ("C. CASH OUT (every rupee that left the box)", "sec"),
    ("Property rent paid to landlords in cash", "out"),
    ("Staff salaries / advances paid in cash", "out"),
    ("Operating expenses in cash (attach the day list)", "out"),
    ("Tenant deposit REFUNDS paid in cash (tenant + room each)", "out"),
    ("Loans given / chit installments paid in cash (name each)", "out"),
    ("Cash deposited INTO the bank", "out"),
    ("Owner drawings / personal (name who)", "out"),
    ("D. CLOSE", "sec"),
    ("EXPECTED closing  (A + all B − all C)", "calc"),
    ("PHYSICAL COUNT on last night (who counted, date)", "count"),
    ("VARIANCE (count − expected) — must be ~0, explain if not", "var"),
]

JULY = {  # worked example — Kiran's real July numbers
    "Opening cash in hand": 235059,
    "Rent / dues collected — recorded in app": 2831250,
    "Cash collected but NOT in app (name each in Notes!)": 67850,
    "Property rent paid to landlords in cash": 1532000,
    "Staff salaries / advances paid in cash": 2000,
    "Loans given / chit installments paid in cash (name each)": 1600000,
    "PHYSICAL COUNT on last night (who counted, date)": 65000,
}
JULY_NOTES = {
    "Cash collected but NOT in app (name each in Notes!)": "app entries missed — WHO? (open)",
    "Loans given / chit installments paid in cash (name each)":
        "Chit Belandur 5,00,000 · Chit Boobalan 3,50,000 · Chandra 50,000 · Loan to Mama 7,00,000",
    "VARIANCE (count − expected) — must be ~0, explain if not":
        "+64,841 UNEXPLAINED — a spend was written bigger than what left the box, or collection understated",
}


def build_sheet(ws, title: str, values: dict, notes: dict):
    ws.append([title, "Amount (₹)", "Notes / breakdown"])
    for c in ws[1]:
        c.font = HDR
        c.fill = HDR_FILL
    in_rows, out_rows = [], []
    open_row = None
    for label, kind in ROWS:
        ws.append([label, values.get(label, None), notes.get(label, "")])
        r = ws.max_row
        cell_a, cell_b = ws.cell(r, 1), ws.cell(r, 2)
        cell_b.number_format = "#,##0"
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
        cell_a.alignment = Alignment(wrap_text=True, vertical="top")
        for c in (cell_a, cell_b, ws.cell(r, 3)):
            c.border = THIN
        if kind == "sec":
            cell_a.font = BOLD
            for c in (cell_a, cell_b, ws.cell(r, 3)):
                c.fill = SEC_FILL
        elif kind == "open":
            open_row = r
        elif kind == "in":
            in_rows.append(r)
        elif kind == "out":
            out_rows.append(r)
        elif kind == "calc":
            cell_a.font = BOLD
            cell_b.font = BOLD
            cell_b.value = (f"=B{open_row}"
                            + "".join(f"+B{x}" for x in in_rows)
                            + "".join(f"-B{x}" for x in out_rows))
            calc_row = r
        elif kind == "count":
            cell_a.font = BOLD
            count_row = r
        elif kind == "var":
            cell_a.font = BOLD
            cell_b.font = BOLD
            cell_b.value = f"=B{count_row}-B{calc_row}"
            cell_b.fill = WARN_FILL
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"


wb = Workbook()
build_sheet(wb.active, "CASH BOOK — AUG 2026", {}, {})
wb.active.title = "AUG 2026"
ex = wb.create_sheet("JUL 2026 (example)")
build_sheet(ex, "CASH BOOK — JUL 2026 (worked example)", JULY, JULY_NOTES)

out = Path("data/reports/Cash_Book_2026.xlsx")
wb.save(out)
print(f"saved {out}")
