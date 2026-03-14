"""
Parse tenant Comments from Excel → structured agreed terms.
Stores clean structured text in tenancy.notes (internal only, never shown to tenants).
Also updates lock_in_months where identifiable.

Run (PREVIEW — no DB writes):
    PYTHONPATH=. PYTHONUTF8=1 venv/Scripts/python scripts/parse_tenant_agreements.py

Run (UPDATE DB):
    PYTHONPATH=. PYTHONUTF8=1 venv/Scripts/python scripts/parse_tenant_agreements.py --update

Output Excel (always produced):
    tenant_agreements_parsed_<date>.xlsx
"""
import re, sys, asyncio, os, openpyxl
from datetime import date
from dotenv import load_dotenv
load_dotenv()

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from sqlalchemy.ext.asyncio import create_async_engine
from sqlalchemy import text
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATABASE_URL = os.environ["DATABASE_URL"]
EXCEL_PATH   = "Cozeevo Monthly stay (3).xlsx"
TODAY        = date.today()

ACTIVE_STATUSES = {"CHECKIN", "NO SHOW"}
SKIP_COMMENTS   = {"no due", "no dues", "-", "", "none", "exit", "n/a", "no  due", "nodue"}


# ── Parser ─────────────────────────────────────────────────────────────────────

def parse_comment(raw: str) -> dict:
    """
    Returns a dict with keys:
      cash_rent      int | None
      upi_rent       int | None
      payment_pref   'cash' | 'upi' | None
      lock_in_months int | None
      escalation     list of str  e.g. ["Feb-Mar: 27,500", "Apr+: 29,000"]
      collect_note   str | None   pending amount or collection instruction
      other_note     str | None   remaining useful context
    """
    c   = raw.strip()
    cl  = c.lower()
    out = {
        "cash_rent":      None,
        "upi_rent":       None,
        "payment_pref":   None,
        "lock_in_months": None,
        "escalation":     [],
        "collect_note":   None,
        "other_note":     None,
    }

    # ── Cash vs UPI amounts ───────────────────────────────────────────────────
    # Patterns: "if cash 15500", "cash 21000 if UPI 22000", "if Cash 14000 otherwise 14500"
    # "If cash then 15500 otherwise collect monthly rent 16000 if UPI"
    cash_m = re.search(
        r'(?:if\s+)?cash\s*(?:then\s+)?(?:rs\.?\s*)?(\d[\d,]+)',
        c, re.I
    )
    upi_m = re.search(
        r'(?:if\s+)?(?:upi|otherwise)\s*(?:then\s+)?(?:rs\.?\s*)?(\d[\d,]+)'  # "if UPI 16000"
        r'|(\d[\d,]+)\s+if\s+(?:upi|online)',                                   # "16000 if UPI"
        c, re.I
    )
    # "always by cash 21000" pattern
    cash_only_m = re.search(r'(?:always\s+by\s+cash|cash\s+only)\s*(?:rs\.?\s*)?(\d[\d,]+)', c, re.I)

    if cash_m:
        out["cash_rent"] = int(cash_m.group(1).replace(",", ""))
    elif cash_only_m:
        out["cash_rent"] = int(cash_only_m.group(1).replace(",", ""))

    if upi_m:
        upi_val = upi_m.group(1) or upi_m.group(2)
        out["upi_rent"] = int(upi_val.replace(",", ""))

    # ── Payment preference ────────────────────────────────────────────────────
    if re.search(r'always\s+(?:by\s+)?cash|ask\s+for\s+cash|cash\s+only|prefer\s+cash'
                 r'|forgot\s+to\s+ask.*cash|cash\s+comes|cash\s+preferred', cl):
        out["payment_pref"] = "cash"

    # ── Lock-in ───────────────────────────────────────────────────────────────
    # Convert "one/two/three" words → digits
    _WORD_NUM = {"one": "1", "two": "2", "three": "3", "four": "4", "six": "6"}
    c_norm = c
    for w, d in _WORD_NUM.items():
        c_norm = re.sub(rf'\b{w}\b', d, c_norm, flags=re.I)

    lockin_m = re.search(
        r'(\d+)\s*mon(?:tt?h[s]?|h[s]?|hs?)?\s*lock(?:[\-\s]?in)?'  # "3 months/montths/monhs lockin"
        r'|lock(?:[\-\s]?in)?\s+(?:is\s+only\s+for\s+)?(\d+)\s*month'   # "lockin 1 month"
        r'|(\d+)\s*month[s]?\s+lock',                                      # "1 month lock"
        c_norm, re.I
    )
    if lockin_m:
        val = next(v for v in lockin_m.groups() if v)
        out["lock_in_months"] = int(val)

    # ── Rent escalation ───────────────────────────────────────────────────────
    # Patterns like: "April - 15000 may 1st- 15500"
    #                "until jan 31st-12000, until mar31st-14000, from apr 1st-14500"
    #                "Feb&Mar 27500 and after that 29000"
    #                "12000 until april 30 then 13000"
    #                "9900 from Feb till april 31st"
    #                "mar & April 10000 and From First may 1st 12k"

    esc_parts = []
    MO = r'(jan\w*|feb\w*|mar\w*|apr\w*|may|jun\w*|jul\w*|aug\w*|sep\w*|oct\w*|nov\w*|dec\w*)'

    def mo_cap(s):
        return s.strip().capitalize()[:3] if s else s

    def big(n):  # only treat as rent amount if > 500 (avoids reading dates like "30" or "31")
        return n > 500

    # Pattern A: "[month] [date?] - [amount]"  e.g. "April - 15000", "may 1st - 15500"
    for m in re.finditer(
        MO + r'(?:\s+\d+(?:st|nd|rd|th)?)?\s*[-:→]\s*(?:rs\.?\s*)?(\d[\d,]+)',
        c, re.I
    ):
        amt = int(m.group(2).replace(",", ""))
        if big(amt):
            esc_parts.append(f"{mo_cap(m.group(1))}: Rs {amt:,}")

    # Pattern B: "until/from/till [month] [date] - [amount]"  — REQUIRES dash/colon before amount
    for m in re.finditer(
        r'(?:until|from|till)\s+(?:first\s+)?' + MO +
        r'(?:\s+\d+(?:st|nd|rd|th)?)?\s*[-:→]\s*(?:rs\.?\s*)?(\d[\d,]+)',
        c, re.I
    ):
        amt = int(m.group(2).replace(",", ""))
        if big(amt):
            label = re.sub(r'[-:→\s]+$', '', m.group(0).rsplit(m.group(2), 1)[0]).strip()
            if label not in [p.split(":")[0] for p in esc_parts]:
                esc_parts.append(f"{label}: Rs {amt:,}")

    # Pattern C: "[amount] until [month] [date?] then [amount]"
    # e.g. "12000 until april 30 then 13000"
    c_then = re.search(
        r'(\d[\d,]+)\s+until\s+\w+(?:\s+\d+)?\s+then\s+(\d[\d,]+)', c, re.I
    )
    if c_then:
        a1, a2 = int(c_then.group(1).replace(",", "")), int(c_then.group(2).replace(",", ""))
        if big(a1) and big(a2):
            esc_parts.append(f"Current: Rs {a1:,} → After: Rs {a2:,}")

    # Pattern D: "[months] [amount] and after that [amount]"
    # e.g. "Feb&Mar 27500 and after that 29000"
    after_that = re.search(
        r'(\d[\d,]+)\s+(?:and\s+)?after\s+that\s+(\d[\d,]+)', c, re.I
    )
    if after_that:
        a1, a2 = int(after_that.group(1).replace(",", "")), int(after_that.group(2).replace(",", ""))
        if big(a1) and big(a2):
            mo_m = re.search(r'(feb\w*&\w*mar\w*|jan\w*&\w*feb\w*|' + MO[1:-1] + r')', c[:after_that.start()], re.I)
            prefix = (mo_m.group(1).capitalize() + ": ") if mo_m else ""
            esc_parts.append(f"{prefix}Rs {a1:,} → After: Rs {a2:,}")

    # Pattern E: "[amount] from [month] (till [month])?" — same rate over a range
    # e.g. "9900 from Feb till april", "9900 till april 31st" — just note the rate
    # Guard: skip if the matched text is surrounded by deposit/installment language
    if not esc_parts:
        rate_range = re.search(
            r'(\d[\d,]+)\s+(?:from\s+)?' + MO + r'(?:\s+till\s+(' + MO[1:-1] + r'))?',
            c, re.I
        )
        if rate_range:
            amt = int(rate_range.group(1).replace(",", ""))
            ctx_start = max(0, rate_range.start() - 40)
            ctx_end   = min(len(c), rate_range.end() + 40)
            ctx = c[ctx_start:ctx_end].lower()
            near_deposit = bool(re.search(r'deposit|installment|balance|security', ctx))
            if big(amt) and not near_deposit:
                m1 = mo_cap(rate_range.group(2))
                m2 = mo_cap(rate_range.group(3)) if rate_range.group(3) else ""
                label = f"{m1}–{m2}" if m2 else f"from {m1}"
                esc_parts.append(f"Rent Rs {amt:,} ({label})")

    # Pattern F: "[month1]&[month2] [amount] and from first [month3] [amount2]"
    # e.g. "mar & April 10000 and From First may 1st 12k"
    mar_may = re.search(
        MO + r'\s*[&,]\s*' + MO + r'\s+(\d[\d,]+)'
        r'.*?(?:from\s+first\s+|from\s+)' + MO +
        r'(?:\s+\d+(?:st|nd|rd|th)?)?\s+(\d[\d,k]+)',
        c, re.I
    )
    if mar_may:
        a1 = int(mar_may.group(3).replace(",", ""))
        a2 = _parse_amount(mar_may.group(5))
        if big(a1) and big(a2):
            esc_parts.append(f"{mo_cap(mar_may.group(1))}/{mo_cap(mar_may.group(2))}: Rs {a1:,} → {mo_cap(mar_may.group(4))}+: Rs {a2:,}")

    # Pattern G: "From [month].Xst price increased [amount]"
    # e.g. "From Feb.1st price increased 16000"
    price_incr = re.search(
        r'(?:from\s+)?' + MO + r'\.?\w*\s+(?:price|rent)\s+increas\w+\s+(?:rs\.?\s*)?(\d[\d,]+)',
        c, re.I
    )
    if price_incr:
        amt = int(price_incr.group(2).replace(",", ""))
        if big(amt):
            esc_parts.append(f"New rent from {mo_cap(price_incr.group(1))}: Rs {amt:,}")

    if esc_parts:
        # Deduplicate: prefer longer/more-descriptive label when same amount appears twice
        # e.g. "until jan 31st: Rs 12,000" vs "Jan: Rs 12,000" — keep only the longer one
        amt_seen: dict = {}
        for p in esc_parts:
            amt_m = re.search(r'Rs ([\d,]+)', p)
            if amt_m:
                key = amt_m.group(1)
                if key not in amt_seen or len(p) > len(amt_seen[key]):
                    amt_seen[key] = p
            else:
                amt_seen[p] = p  # no amount — keep as-is keyed by full string
        # Restore original order
        deduped = []
        used = set()
        for p in esc_parts:
            amt_m = re.search(r'Rs ([\d,]+)', p)
            key = amt_m.group(1) if amt_m else p
            if key not in used:
                used.add(key)
                deduped.append(amt_seen[key])
        out["escalation"] = deduped

    # ── Collect notes (pending amounts, balance due) ──────────────────────────
    collect_parts = []
    # "collect XXXX on <date>"
    col_m = re.search(r'collect\s+(?:rs\.?\s*)?(\d[\d,]+)\s+(?:on\s+)?(.{0,30}?)(?:\.|$)', c, re.I)
    if col_m:
        amt  = int(col_m.group(1).replace(",", ""))
        when = col_m.group(2).strip()
        collect_parts.append(f"Collect Rs {amt:,} on {when}" if when else f"Collect Rs {amt:,}")

    # "pending XXXX" or "balance XXXX"
    for m in re.finditer(r'(?:pending|balance(?:\s+amount)?|remaining\s+amount)\s+(?:(?:is|of|rs\.?)\s+)?(?:rs\.?\s*)?(\d[\d,]+)', c, re.I):
        amt = int(m.group(1).replace(",", ""))
        collect_parts.append(f"Balance: Rs {amt:,}")

    # "XXXX BALANCE AMOUNT ON DATE"
    bal_on_m = re.search(r'(?:rs\.?\s*)?(\d[\d,]+)\s+balance\s+amount\s+on\s+(.{0,25})', c, re.I)
    if bal_on_m:
        amt  = int(bal_on_m.group(1).replace(",", ""))
        when = bal_on_m.group(2).strip().rstrip(".")
        collect_parts.append(f"Collect Rs {amt:,} on {when}")

    # "deposit in 2 installments" / "half deposit at april"
    if re.search(r'deposit\s+in\s+\d+\s+install|half\s+deposit', c, re.I):
        collect_parts.append("Deposit: installments agreed")

    # "collect on Xth around Xpm"
    time_m = re.search(r'(\d+)(?:st|nd|rd|th)\s+around\s+(.{0,20}?)\s+(?:cash|upi|$)', c, re.I)
    if time_m:
        collect_parts.append(f"Collect on {time_m.group(1)}{['st','nd','rd','th'][min(int(time_m.group(1))-1,3)]} ~{time_m.group(2).strip()}")

    # "25th around 2 pm cash only"
    time2_m = re.search(r'(?:cash|payment)\s+comes\s+around\s+(.+?)(?:on|$)', c, re.I)
    if time2_m:
        collect_parts.append(f"Arrives: {time2_m.group(1).strip()}")

    # GST
    gst_m = re.search(r'(\d+)\s*%\s*gst', c, re.I)
    if gst_m:
        collect_parts.append(f"GST: +{gst_m.group(1)}% on rent")

    if collect_parts:
        out["collect_note"] = " | ".join(collect_parts)

    # ── Other notes ───────────────────────────────────────────────────────────
    other = []
    if re.search(r'no\s+deposit\s+refund|deposit\s+(?:not|won.t\s+be)\s+refund', c, re.I):
        other.append("No deposit refund if exits early")
    if re.search(r'only\s+for\s+(\d+)\s+months?|stay\s+only\s+(\d+)\s+months?|plans?\s+to\s+stay.*?(\d+)\s+months?', c, re.I):
        m = re.search(r'(\d+)\s*month', c, re.I)
        if m:
            other.append(f"Short stay: {m.group(1)} months")
    if re.search(r'vacation|luggage|leave.*safe', c, re.I):
        other.append("On vacation — keeping room at reduced rate")
    if re.search(r'family.*visit|family.*stay', c, re.I):
        other.append("Family visits allowed")
    if re.search(r'want\s+to\s+move|preferred\s+room|sattva\s+side|building\s+side', c, re.I):
        other.append("Room transfer requested")
    if re.search(r'upi\s+name\s*:', c, re.I):
        upi_name_m = re.search(r'upi\s+name\s*:\s*(.+?)(?:,|$)', c, re.I)
        if upi_name_m:
            other.append(f"UPI sender name: {upi_name_m.group(1).strip()}")

    if other:
        out["other_note"] = " | ".join(other)

    return out


def _parse_amount(s: str) -> int:
    """Convert '12k' → 12000, '29,000' → 29000."""
    s = s.strip().lower().replace(",", "")
    if s.endswith("k"):
        return int(float(s[:-1]) * 1000)
    try:
        return int(s)
    except ValueError:
        return 0


def format_notes(parsed: dict, original: str) -> str:
    """Build the clean structured notes string for DB storage."""
    parts = []

    if parsed["cash_rent"] and parsed["upi_rent"]:
        parts.append(f"Cash: Rs {parsed['cash_rent']:,} | UPI: Rs {parsed['upi_rent']:,}")
    elif parsed["cash_rent"]:
        parts.append(f"Cash rent: Rs {parsed['cash_rent']:,}")
    elif parsed["upi_rent"]:
        parts.append(f"UPI rent: Rs {parsed['upi_rent']:,}")

    if parsed["payment_pref"] == "cash":
        parts.append("Payment: cash preferred")

    if parsed["lock_in_months"]:
        parts.append(f"Lock-in: {parsed['lock_in_months']} month(s)")

    if parsed["escalation"]:
        parts.append("Escalation: " + " → ".join(parsed["escalation"]))

    if parsed["collect_note"]:
        parts.append(parsed["collect_note"])

    if parsed["other_note"]:
        parts.append(parsed["other_note"])

    if not parts:
        # Keep original as fallback for unstructured notes
        return f"Note: {original.strip()}"

    return " | ".join(parts)


# ── Load Excel ─────────────────────────────────────────────────────────────────

def load_active_tenants():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["History"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[idx["IN/OUT"]]
        if status not in ACTIVE_STATUSES:
            continue
        comment = row[idx["Comments"]]
        if not comment:
            continue
        cs = str(comment).strip()
        if cs.lower().replace(" ", "") in SKIP_COMMENTS:
            continue
        mobile_raw = row[idx["Mobile Number"]]
        mobile = None
        if mobile_raw:
            m = re.sub(r"[^\d]", "", str(mobile_raw))
            if len(m) >= 10:
                mobile = m[-10:]   # last 10 digits

        rows.append({
            "room":    str(row[idx["Room No"]]).replace(".0", ""),
            "name":    str(row[idx["Name"]]).strip() if row[idx["Name"]] else "",
            "mobile":  mobile,
            "status":  status,
            "comment": cs,
        })
    return rows


# ── DB lookup ──────────────────────────────────────────────────────────────────

async def fetch_tenancies():
    """Return dict keyed by (mobile, name_lower) → tenancy_id."""
    engine = create_async_engine(DATABASE_URL, echo=False)
    rows = []
    async with engine.connect() as conn:
        result = await conn.execute(text("""
            SELECT tn.id, t.phone, t.name, tn.notes, tn.lock_in_months
            FROM tenancies tn
            JOIN tenants t ON t.id = tn.tenant_id
            WHERE tn.status IN ('active', 'no_show')
        """))
        rows = result.fetchall()
    await engine.dispose()
    # Build lookup: mobile → tenancy_id, name_lower → tenancy_id
    by_phone = {}
    by_name  = {}
    for r in rows:
        phone = (r[1] or "").strip()
        name  = (r[2] or "").strip().lower()
        if phone:
            by_phone[phone[-10:]] = {"id": r[0], "notes": r[3], "lock_in": r[4]}
        if name:
            by_name[name] = {"id": r[0], "notes": r[3], "lock_in": r[4]}
    return by_phone, by_name


async def update_db(updates: list):
    """updates = list of (tenancy_id, new_notes, lock_in_months_or_None)"""
    engine = create_async_engine(DATABASE_URL, echo=False)
    async with engine.connect() as conn:
        for tenancy_id, new_notes, lock_in in updates:
            if lock_in is not None:
                await conn.execute(text("""
                    UPDATE tenancies SET notes = :n, lock_in_months = :l
                    WHERE id = :id
                """), {"n": new_notes, "l": lock_in, "id": tenancy_id})
            else:
                await conn.execute(text("""
                    UPDATE tenancies SET notes = :n WHERE id = :id
                """), {"n": new_notes, "id": tenancy_id})
        await conn.commit()
    await engine.dispose()


# ── Main ───────────────────────────────────────────────────────────────────────

async def main():
    do_update = "--update" in sys.argv

    tenants = load_active_tenants()
    by_phone, by_name = await fetch_tenancies()

    results = []
    for t in tenants:
        parsed   = parse_comment(t["comment"])
        new_note = format_notes(parsed, t["comment"])

        # Match to DB
        db_rec = None
        if t["mobile"] and t["mobile"] in by_phone:
            db_rec = by_phone[t["mobile"]]
        elif t["name"].lower() in by_name:
            db_rec = by_name[t["name"].lower()]

        results.append({
            "room":       t["room"],
            "name":       t["name"],
            "status":     t["status"],
            "mobile":     t["mobile"] or "",
            "original":   t["comment"],
            "structured": new_note,
            "lock_in":    parsed["lock_in_months"],
            "tenancy_id": db_rec["id"] if db_rec else None,
            "matched":    db_rec is not None,
        })

    # ── Output Excel ──────────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Agreed Terms Parsed"

    COLS = [
        ("Room",         9),
        ("Name",         22),
        ("Status",       10),
        ("Original Comment",   38),
        ("Structured Terms",   55),
        ("Lock-in (m)",  12),
        ("DB Matched",   12),
    ]

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, r in enumerate(results, 2):
        vals = [
            r["room"], r["name"], r["status"],
            r["original"], r["structured"],
            r["lock_in"] or "",
            "✓" if r["matched"] else "✗ NOT FOUND",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in (4, 5)))
            if r["status"] == "NO SHOW":
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            if not r["matched"] and ci == 7:
                cell.fill = PatternFill("solid", fgColor="FFCCCC")
        # Structured terms = yellow
        ws.cell(row=ri, column=5).fill = PatternFill("solid", fgColor="FFF2CC")

    out_path = f"tenant_agreements_parsed_{TODAY}.xlsx"
    wb.save(out_path)
    print(f"  Excel saved: {out_path}  ({len(results)} rows)")

    matched   = sum(1 for r in results if r["matched"])
    unmatched = [r for r in results if not r["matched"]]
    with_lock = sum(1 for r in results if r["lock_in"])

    print(f"  DB matched:  {matched}/{len(results)}")
    print(f"  Lock-in parsed: {with_lock}")
    print()

    if unmatched:
        print(f"  ⚠  {len(unmatched)} tenant(s) NOT matched in DB (mobile/name mismatch):")
        for r in unmatched:
            print(f"     Room {r['room']} | {r['name']} | mobile: {r['mobile'] or 'none'}")
        print()

    # ── Preview structured output ─────────────────────────────────────────────
    print("  Sample structured terms:")
    for r in results[:15]:
        print(f"  [{r['room']}] {r['name']}")
        print(f"    → {r['structured']}")
    print()

    if do_update:
        updates = [
            (r["tenancy_id"], r["structured"], r["lock_in"])
            for r in results if r["matched"]
        ]
        await update_db(updates)
        print(f"  ✓ Updated {len(updates)} tenancy.notes in DB.")
    else:
        print("  Run with --update to write to DB.")


if __name__ == "__main__":
    asyncio.run(main())
