"""Full monthly report from Excel with correct cleaning logic."""
import sys, re, os
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from collections import Counter
import openpyxl

EXCEL = "Cozeevo Monthly stay (4).xlsx"
TOTAL_BEDS = 291

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def pd(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ('%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: continue
    return None

def cst(v):
    s = str(v).strip().upper() if v else ''
    if s in ('CHECKIN', 'CHECK IN', ''): return 'Active'
    if 'EXIT' in s: return 'Exited'
    if 'NO SHOW' in s: return 'No-show'
    if 'CANCEL' in s: return 'Cancelled'
    return 'Active'

def rst(v):
    s = str(v).strip().upper() if v else ''
    if not s: return ''
    if 'PAID' in s and 'NOT' not in s and 'PARTIAL' not in s: return 'PAID'
    if 'PARTIAL' in s: return 'PARTIAL'
    if 'NOT PAID' in s or 'UNPAID' in s: return 'UNPAID'
    if 'EXIT' in s: return 'EXIT'
    if 'NO SHOW' in s: return 'NO SHOW'
    if 'ADVANCE' in s: return 'ADVANCE'
    if 'CANCEL' in s: return 'CANCELLED'
    return s

CHANDRA_RE = re.compile(r'received\s+by\s+chandra', re.I)
EXIT_PATS = [
    re.compile(r'(?:march|feb|jan|apr)\s*\d+\s*(?:st|nd|rd|th)?\s*exit', re.I),
    re.compile(r'exit\s+on\s+', re.I),
]

def clean_num(raw, is_balance=False, cash_num=0, upi_num=0, rent=0):
    """Clean a cell. Returns (number, note, chandra_amt, lakshmi_amt)."""
    if raw is None: return 0, '', 0, 0
    s = str(raw).strip()
    if not s: return 0, '', 0, 0
    try: return float(s), '', 0, 0
    except: pass

    chandra = 0
    lakshmi = 0

    # Chandra in balance col
    if is_balance and CHANDRA_RE.search(s):
        amt = cash_num if cash_num > 0 else (upi_num if upi_num > 0 else rent)
        return 0, s, amt, 0

    # Chandra in cash col: return 0 for cash, track separately
    if not is_balance and CHANDRA_RE.search(s):
        # Split: "13000 Received by Chandra /5000 Lakshmi gorjala"
        sp = re.match(r'(\d[\d,]*)\s*Received.*?/\s*(\d[\d,]*)\s*(.*)', s, re.I)
        if sp:
            a1 = float(sp.group(1).replace(',', ''))
            a2 = float(sp.group(2).replace(',', ''))
            return 0, s, a1, a2
        # Plain "Received by Chandra anna"
        amt = upi_num if upi_num > 0 else rent
        return 0, s, amt, 0

    # Exit dates = 0
    for p in EXIT_PATS:
        if p.search(s): return 0, s, 0, 0

    # Return = refund
    if re.match(r'return\s+\d+', s, re.I): return 0, s, 0, 0
    # paid in X = 0
    if re.match(r'paid\s+in\s+', s, re.I): return 0, s, 0, 0
    # hitachi = 0
    if re.match(r'hitachi', s, re.I): return 0, s, 0, 0
    # dash = 0
    if s == '-': return 0, '-', 0, 0

    # Expression: 516*16=8256
    eq = re.search(r'=\s*([\d,.]+)\s*$', s)
    if eq: return float(eq.group(1).replace(',', '')), s, 0, 0

    # Number + text: "5500 on april 1st", "9000 deposit"
    m = re.search(r'([\d,]+)', s)
    if m: return float(m.group(1).replace(',', '')), s, 0, 0

    return 0, s, 0, 0


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb['History']

    tenants = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name or not str(name).strip(): continue
        rm = sn(ws.cell(row, 10).value)
        rf = sn(ws.cell(row, 11).value)
        ry = sn(ws.cell(row, 12).value)
        rent = ry if ry > 0 else (rf if rf > 0 else rm)

        tenants.append({
            'row': row, 'name': str(name).strip(),
            'room': str(ws.cell(row, 1).value or '').strip().replace('.0', ''),
            'block': str(ws.cell(row, 18).value or '').strip(),
            'sharing': str(ws.cell(row, 13).value or '').strip().lower(),
            'status': cst(ws.cell(row, 17).value),
            'checkin': pd(ws.cell(row, 5).value),
            'rent': rent,
            'dec_st': rst(ws.cell(row, 21).value),
            'jan_st': rst(ws.cell(row, 22).value),
            'jan_bal': ws.cell(row, 23).value, 'jan_cash': ws.cell(row, 24).value, 'jan_upi': ws.cell(row, 25).value,
            'feb_st': rst(ws.cell(row, 26).value),
            'feb_bal': ws.cell(row, 28).value, 'feb_cash': ws.cell(row, 29).value, 'feb_upi': ws.cell(row, 30).value,
            'mar_st': rst(ws.cell(row, 27).value),
            'mar_bal': ws.cell(row, 31).value, 'mar_cash': ws.cell(row, 32).value, 'mar_upi': ws.cell(row, 33).value,
        })

    print(f"Parsed {len(tenants)} tenants\n")

    months = [
        ('DECEMBER 2025', date(2025,12,1), date(2025,12,31), 'dec_st', None, None, None),
        ('JANUARY 2026',  date(2026,1,1),  date(2026,1,31),  'jan_st', 'jan_cash', 'jan_upi', 'jan_bal'),
        ('FEBRUARY 2026', date(2026,2,1),  date(2026,2,28),  'feb_st', 'feb_cash', 'feb_upi', 'feb_bal'),
        ('MARCH 2026',    date(2026,3,1),  date(2026,3,31),  'mar_st', 'mar_cash', 'mar_upi', 'mar_bal'),
    ]

    for label, ms, me, stk, ck, uk, bk in months:
        print(f"{'='*70}")
        print(f"  {label}")
        print(f"{'='*70}")

        # Clean payments
        cash_total = 0
        upi_total = 0
        bal_total = 0
        chandra_total = 0
        lakshmi_total = 0

        if ck:
            for t in tenants:
                cash_num = sn(t[ck]) if isinstance(t[ck], (int, float)) else 0
                upi_num = sn(t[uk]) if isinstance(t[uk], (int, float)) else 0

                cash, _, cc, cl = clean_num(t[ck], False, 0, upi_num, t['rent'])
                upi, _, _, _ = clean_num(t[uk])
                bal, _, bc, bl = clean_num(t[bk], True, cash_num, upi_num, t['rent'])

                cash_total += cash
                upi_total += upi
                bal_total += bal
                chandra_total += cc + bc
                lakshmi_total += cl + bl

        # Occupancy
        # Active: checked in by month end
        active = [t for t in tenants if t['status'] == 'Active' and t['checkin'] and t['checkin'] <= me]
        # No-show: ALL no-shows, NO date filter (same count every month)
        noshow = [t for t in tenants if t['status'] == 'No-show']

        prem_a = sum(1 for t in active if t['sharing'] == 'premium')
        reg_a = len(active) - prem_a
        a_beds = reg_a + prem_a * 2

        prem_n = sum(1 for t in noshow if t['sharing'] == 'premium')
        reg_n = len(noshow) - prem_n
        n_beds = reg_n + prem_n * 2

        vacant = TOTAL_BEDS - a_beds - n_beds

        # THOR / HULK
        thor = [t for t in active if t['block'] == 'THOR']
        hulk = [t for t in active if t['block'] == 'HULK']
        tp = sum(1 for t in thor if t['sharing'] == 'premium')
        hp = sum(1 for t in hulk if t['sharing'] == 'premium')
        tb = (len(thor) - tp) + tp * 2
        hb = (len(hulk) - hp) + hp * 2

        # Status
        stc = Counter(t[stk] for t in tenants if t[stk])

        tenant_dues = bal_total  # Balance column = tenant dues (Chandra tracked separately)

        print(f"  OCCUPANCY")
        print(f"    Checked-in:  {a_beds} beds ({reg_a} reg + {prem_a} prem x2)")
        print(f"    No-show:     {n_beds} beds ({reg_n} reg + {prem_n} prem x2)")
        print(f"    Vacant:      {vacant}")
        print(f"    Occupancy:   {a_beds/TOTAL_BEDS*100:.1f}%")
        print(f"    THOR: {tb} beds ({len(thor)}t, {tp}P)   HULK: {hb} beds ({len(hulk)}t, {hp}P)")

        if ck:
            print(f"  COLLECTIONS")
            print(f"    Cash:        Rs.{cash_total:>10,.0f}")
            print(f"    UPI:         Rs.{upi_total:>10,.0f}")
            print(f"    Total:       Rs.{cash_total + upi_total:>10,.0f}")
            print(f"  BALANCE / DUES")
            print(f"    Tenant dues: Rs.{tenant_dues:>10,.0f}")
            print(f"    Chandra:     Rs.{chandra_total:>10,.0f}")
            if lakshmi_total:
                print(f"    Lakshmi:     Rs.{lakshmi_total:>10,.0f}")
            print(f"    Total:       Rs.{bal_total:>10,.0f}")

        print(f"  STATUS COUNTS")
        for s in ['PAID', 'PARTIAL', 'UNPAID', 'NO SHOW', 'EXIT', 'ADVANCE', 'CANCELLED']:
            if stc.get(s, 0) > 0:
                print(f"    {s:12s}  {stc[s]}")
        print()


if __name__ == "__main__":
    main()
