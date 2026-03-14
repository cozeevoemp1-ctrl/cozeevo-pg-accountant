"""
Export staff salary records to a formatted Excel file.
Reads directly from the SALARY_ROWS in import_staff_salaries.py.

Usage:
    python scripts/export_salary_to_excel.py
"""
import sys
import os
from datetime import date
from pathlib import Path

# ── pull SALARY_ROWS from the import script ───────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent.parent))
from scripts.import_staff_salaries import SALARY_ROWS

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Category labels ───────────────────────────────────────────────────────────
CAT_LABELS = {5: "Salary", 6: "Bonus", 1: "Electricity", 7: "Maintenance"}

MONTH_COLORS = {
    "Nov 2025": "D6E4F0",
    "Dec 2025": "D5E8D4",
    "Jan 2026": "FFF2CC",
    "Feb 2026": "FCE4D6",
    "Mar 2026": "E1D5E7",
}

FLAG_COLOR   = "FF0000"
HEADER_COLOR = "2F4F8F"
TOTAL_COLOR  = "F0F0F0"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def month_label(d: date) -> str:
    return d.strftime("%b %Y")


def build_df() -> pd.DataFrame:
    rows = []
    for i, (dt, name, amt, cat_id, desc, flagged) in enumerate(SALARY_ROWS, 1):
        rows.append({
            "#":           i,
            "Date":        dt,
            "Month":       month_label(dt),
            "Person":      name,
            "Amount (₹)":  amt,
            "Category":    CAT_LABELS.get(cat_id, str(cat_id)),
            "Description": desc,
            "Flagged":     "⚠" if flagged else "",
        })
    return pd.DataFrame(rows)


def write_excel(df: pd.DataFrame, out_path: str) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Sheet 1: All Transactions ─────────────────────────────────────────
        df_out = df.drop(columns=["Month"])
        df_out.to_excel(writer, sheet_name="All Transactions", index=False, startrow=1)

        ws = writer.sheets["All Transactions"]

        # Title row
        ws["A1"] = "Staff Salary Ledger — THOR PG  (Nov 2025 – Mar 2026)"
        ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A1:{get_column_letter(len(df_out.columns))}1")

        # Header row (row 2)
        header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        # Data rows — colour by month, flag in red
        months_seen = list(dict.fromkeys(df["Month"]))
        month_idx = {m: i % 2 for i, m in enumerate(months_seen)}  # alternating

        for row_num, (_, series) in enumerate(df.iterrows(), start=3):
            month = series["Month"]
            bg = MONTH_COLORS.get(month, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)
            flagged = series["Flagged"] == "⚠"

            for col_num, cell in enumerate(ws[row_num], start=1):
                cell.fill = fill
                cell.border = BORDER
                cell.font = Font(name="Calibri", size=9,
                                 color=FLAG_COLOR if flagged else "000000",
                                 bold=flagged)
                cell.alignment = Alignment(vertical="center")

            # Right-align Amount column (col 5)
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=row_num, column=5).number_format = '#,##0'

        # Column widths
        col_widths = [4, 12, 22, 12, 12, 18, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A3"

        # Totals row
        total_row = len(df) + 3
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name="Calibri")
        ws.cell(total_row, 5, df["Amount (₹)"].sum()).number_format = '#,##0'
        ws.cell(total_row, 5).font = Font(bold=True, name="Calibri")
        ws.cell(total_row, 5).alignment = Alignment(horizontal="right")
        for c in range(1, len(df_out.columns) + 1):
            cell = ws.cell(total_row, c)
            cell.fill = PatternFill("solid", fgColor=TOTAL_COLOR)
            cell.border = BORDER

        # ── Sheet 2: Monthly Summary ──────────────────────────────────────────
        summary = (
            df.groupby(["Month", "Category"], sort=False)["Amount (₹)"]
            .sum()
            .reset_index()
        )
        pivot = summary.pivot_table(
            index="Month", columns="Category", values="Amount (₹)",
            aggfunc="sum", fill_value=0
        ).reset_index()

        # Preserve month order
        month_order = list(dict.fromkeys(df["Month"]))
        pivot["_order"] = pivot["Month"].map({m: i for i, m in enumerate(month_order)})
        pivot = pivot.sort_values("_order").drop(columns=["_order"])

        pivot["Grand Total"] = pivot.select_dtypes("number").sum(axis=1)
        pivot.to_excel(writer, sheet_name="Monthly Summary", index=False, startrow=1)

        ws2 = writer.sheets["Monthly Summary"]
        ws2["A1"] = "Monthly Summary by Category"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws2["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.merge_cells(f"A1:{get_column_letter(len(pivot.columns))}1")

        for cell in ws2[2]:
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        for row_num, (_, series) in enumerate(pivot.iterrows(), start=3):
            month = series["Month"]
            bg = MONTH_COLORS.get(month, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)
            for col_num, cell in enumerate(ws2[row_num], start=1):
                cell.fill = fill
                cell.border = BORDER
                cell.font = Font(name="Calibri", size=10)
            # Bold Grand Total column
            ws2.cell(row_num, len(pivot.columns)).font = Font(bold=True, name="Calibri")

        # Format numbers in summary
        for row in ws2.iter_rows(min_row=3, max_row=2 + len(pivot),
                                  min_col=2, max_col=len(pivot.columns)):
            for cell in row:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

        for i in range(1, len(pivot.columns) + 1):
            ws2.column_dimensions[get_column_letter(i)].width = 16
        ws2.column_dimensions["A"].width = 12

        # Grand total row
        total_row2 = len(pivot) + 3
        ws2.cell(total_row2, 1, "TOTAL").font = Font(bold=True, name="Calibri")
        num_cols = pivot.select_dtypes("number").columns
        for col_idx, col_name in enumerate(pivot.columns[1:], start=2):
            val = pivot[col_name].sum() if col_name in num_cols else ""
            cell = ws2.cell(total_row2, col_idx, val if val else "")
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=TOTAL_COLOR)
            cell.border = BORDER
        ws2.cell(total_row2, 1).fill = PatternFill("solid", fgColor=TOTAL_COLOR)
        ws2.cell(total_row2, 1).border = BORDER

        ws2.freeze_panes = "A3"

        # ── Sheet 3: Per-Person Summary ───────────────────────────────────────
        person_summary = (
            df.groupby("Person")["Amount (₹)"]
            .sum()
            .reset_index()
            .sort_values("Amount (₹)", ascending=False)
            .reset_index(drop=True)
        )
        person_summary.index += 1
        person_summary.index.name = "#"
        person_summary.to_excel(writer, sheet_name="Per Person", startrow=1)

        ws3 = writer.sheets["Per Person"]
        ws3["A1"] = "Total Paid Per Person"
        ws3["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws3["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
        ws3["A1"].alignment = Alignment(horizontal="center")
        ws3.merge_cells("A1:C1")

        for cell in ws3[2]:
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        for row in ws3.iter_rows(min_row=3, max_row=2 + len(person_summary)):
            for cell in row:
                cell.border = BORDER
                cell.font = Font(name="Calibri", size=10)
        for row in ws3.iter_rows(min_row=3, max_row=2 + len(person_summary),
                                  min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

        ws3.column_dimensions["A"].width = 5
        ws3.column_dimensions["B"].width = 24
        ws3.column_dimensions["C"].width = 14

        # Person total
        total_row3 = len(person_summary) + 3
        ws3.cell(total_row3, 2, "TOTAL").font = Font(bold=True, name="Calibri")
        ws3.cell(total_row3, 3, person_summary["Amount (₹)"].sum()).number_format = '#,##0'
        ws3.cell(total_row3, 3).font = Font(bold=True, name="Calibri")
        ws3.cell(total_row3, 3).alignment = Alignment(horizontal="right")
        for c in [1, 2, 3]:
            ws3.cell(total_row3, c).fill = PatternFill("solid", fgColor=TOTAL_COLOR)
            ws3.cell(total_row3, c).border = BORDER

        ws3.freeze_panes = "A3"


def main():
    df = build_df()
    out = "staff_salary_ledger.xlsx"
    write_excel(df, out)

    total = df["Amount (₹)"].sum()
    salary_total = df[df["Category"] == "Salary"]["Amount (₹)"].sum()
    bonus_total  = df[df["Category"] == "Bonus"]["Amount (₹)"].sum()
    other_total  = df[df["Category"].isin(["Electricity", "Maintenance"])]["Amount (₹)"].sum()

    print("=" * 55)
    print("  Staff Salary Ledger — Excel Export")
    print("=" * 55)
    print(f"  File       : {out}")
    print(f"  Rows       : {len(df)}")
    print(f"  Salary     : ₹{salary_total:,.0f}")
    print(f"  Bonus      : ₹{bonus_total:,.0f}")
    print(f"  Other      : ₹{other_total:,.0f}")
    print(f"  Grand Total: ₹{total:,.0f}")
    print()
    print("  Sheets:")
    print("    1. All Transactions  — colour-coded by month")
    print("    2. Monthly Summary   — pivot by category")
    print("    3. Per Person        — total per staff member")
    print("=" * 55)


if __name__ == "__main__":
    main()
