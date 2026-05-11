"""
Export outstanding dues — 2 files: April_Outstanding_2026.xlsx + May_Outstanding_2026.xlsx
Each file: both HULK + THOR combined, one sheet per building tab + one combined tab
Columns: #, Building, Name, Room, Check-in, Rent, Amount Due, Basis, UPI Paid, Cash Paid, Total Paid, Balance, Status
Sorted by balance descending. Partial payment (yellow) + No payment (red) sections.
"""
import os, csv, datetime, re
from dotenv import load_dotenv
load_dotenv()
db_url = os.environ['DATABASE_URL'].replace('postgresql+asyncpg://', 'postgresql://')
import psycopg2, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APR_DAYS = 30
MAY_DAYS = 31

def extract_phone(upi_id):
    m = re.match(r'^(\d{10})(?:-\d+)?@', str(upi_id))
    return m.group(1) if m else None

hulk_entries, thor_entries = [], []
wb = openpyxl.load_workbook('Hulk may 11th bank statement.xlsx')
for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
    if row[11] == 'SUCCESS' and row[3]:
        hulk_entries.append({'amt': float(row[3]), 'phone': extract_phone(row[7]),
                              'name': str(row[8]).strip().upper(), 'matched': False})
with open('thor may 11th.csv', newline='') as f:
    for row in csv.DictReader(f):
        if row['Settlement_Status'] == 'SUCCESS' and row['TXN_AMOUNT']:
            thor_entries.append({'amt': float(row['TXN_AMOUNT']), 'phone': extract_phone(row['Payer_VPA']),
                                  'name': row['Payer_Name'].strip().upper(), 'matched': False})
may_entries = hulk_entries + thor_entries

conn = psycopg2.connect(db_url)
cur = conn.cursor()
cur.execute("""
SELECT t.name, t.phone, p.name, tn.agreed_rent, tn.security_deposit,
       tn.checkin_date, tn.checkout_date, r.room_number, tn.status, tn.id
FROM tenancies tn
JOIN tenants t ON t.id = tn.tenant_id
JOIN rooms r ON r.id = tn.room_id
JOIN properties p ON p.id = r.property_id
WHERE tn.status = 'active'
   OR (tn.checkout_date >= '2026-04-01' AND tn.checkout_date <= '2026-05-11' AND tn.status = 'exited')
ORDER BY p.name, t.name
""")
tenants = cur.fetchall()
cur.execute("""
SELECT tenancy_id, payment_mode, SUM(amount)
FROM payments WHERE is_void=false AND period_month='2026-04-01'
GROUP BY tenancy_id, payment_mode
""")
apr_pay = {}
for r in cur.fetchall():
    apr_pay.setdefault(r[0], {'upi': 0, 'cash': 0})
    apr_pay[r[0]]['upi' if r[1] in ('upi','bank') else 'cash'] += float(r[2])
cur.execute("""
SELECT tenancy_id, SUM(amount) FROM payments
WHERE is_void=false AND payment_mode='cash' AND period_month='2026-05-01'
GROUP BY tenancy_id
""")
may_cash = {r[0]: float(r[1]) for r in cur.fetchall()}
cur.close(); conn.close()

def normalize_phone(ph):
    if not ph: return None
    ph = re.sub(r'\D', '', str(ph))
    if ph.startswith('91') and len(ph)==12: ph = ph[2:]
    return ph if len(ph)==10 else None

def name_tokens(s):
    return [p for p in re.sub(r'[^A-Z0-9 ]',' ',s.upper()).split() if len(p)>1]

def find_may_bank(name, phone, entries):
    tphone = normalize_phone(phone)
    name_up = name.strip().upper()
    tparts = name_tokens(name_up)
    if tphone:
        total = sum(e['amt'] for e in entries if e['phone']==tphone)
        if total > 0:
            for e in entries:
                if e['phone']==tphone: e['matched']=True
            return total
    for e in entries:
        if not e['matched'] and e['name']==name_up:
            e['matched']=True; return e['amt']
    if len(tparts)>=2:
        best,bi = 0,None
        for i,e in enumerate(entries):
            if e['matched']: continue
            bparts = name_tokens(e['name'])
            hits = sum(1 for tp in tparts if any(tp in bp or bp in tp for bp in bparts))
            score = hits/len(tparts)
            if hits>=2 and score>=0.6 and score>best: best,bi=score,i
        if bi is not None:
            entries[bi]['matched']=True; return entries[bi]['amt']
    return 0

def prorated(rent, days_in_month, checkin_day):
    days = days_in_month - checkin_day + 1
    return round((rent/days_in_month)*days), days

# Build per-tenant data
results = []
for row in tenants:
    name, phone, prop, rent, deposit, checkin, checkout, room, status, tid = row
    rent = float(rent or 0); deposit = float(deposit or 0)
    if isinstance(checkin, datetime.datetime): checkin = checkin.date()
    if isinstance(checkout, datetime.datetime): checkout = checkout.date()
    building = 'HULK' if 'HULK' in prop else 'THOR'

    # April
    exited_pre_apr = checkout and checkout < datetime.date(2026,4,1)
    joined_apr = checkin and checkin.year==2026 and checkin.month==4
    if exited_pre_apr:
        apr_due=0; apr_basis='Exited'
    elif joined_apr:
        pr,d = prorated(rent,APR_DAYS,checkin.day)
        apr_due=deposit+pr; apr_basis=f'Dep+Prorated ({d}d)'
    else:
        apr_due=rent; apr_basis='Full Rent'
    ap = apr_pay.get(tid,{'upi':0,'cash':0})
    apr_upi, apr_cash = ap['upi'], ap['cash']
    apr_total = apr_upi+apr_cash; apr_bal = apr_due-apr_total

    # May
    exited_pre_may = checkout and checkout < datetime.date(2026,5,1)
    joined_may = checkin and checkin.year==2026 and checkin.month==5
    if exited_pre_may:
        may_due=0; may_basis='Exited'
    elif joined_may:
        pr,d = prorated(rent,MAY_DAYS,checkin.day)
        may_due=deposit+pr; may_basis=f'Dep+Prorated ({d}d)'
    else:
        may_due=rent; may_basis='Full Rent'
    may_upi = find_may_bank(name,phone,may_entries)
    may_cash_v = may_cash.get(tid,0)
    may_total = may_upi+may_cash_v; may_bal = may_due-may_total

    results.append(dict(
        building=building, name=name, room=room, rent=rent, checkin=checkin,
        apr_due=apr_due, apr_basis=apr_basis, apr_upi=apr_upi, apr_cash=apr_cash,
        apr_total=apr_total, apr_bal=apr_bal,
        may_due=may_due, may_basis=may_basis, may_upi=may_upi, may_cash=may_cash_v,
        may_total=may_total, may_bal=may_bal,
    ))

# Styles
HDR_FILL    = PatternFill('solid', fgColor='1F3864')
HDR_FONT    = Font(bold=True, color='FFFFFF', size=10)
PARTIAL_FILL= PatternFill('solid', fgColor='FFF2CC')
NOPAY_FILL  = PatternFill('solid', fgColor='FCE4D6')
TOTAL_FILL  = PatternFill('solid', fgColor='BDD7EE')
TOTAL_FONT  = Font(bold=True, size=10)
NORM_FONT   = Font(size=10)
thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

COL_HEADERS = ['#','Building','Name','Room','Check-in','Rent','Amount Due','Basis',
               'UPI Paid','Cash Paid','Total Paid','Balance','Status']
COL_WIDTHS  = [4, 8, 28, 6, 12, 10, 12, 20, 10, 10, 10, 10, 15]
NUM_COLS    = {6,7,9,10,11,12}   # 1-indexed columns with numbers

def write_month_sheet(ws, label, rows_out):
    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(COL_HEADERS))}1')
    ws['A1'] = label
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28
    # Headers
    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
    ws.row_dimensions[2].height = 22

    row_n = 3
    t_due=t_upi=t_cash=t_tot=t_bal = 0
    partial = [r for r in rows_out if r['total']>0]
    nopay   = [r for r in rows_out if r['total']==0]

    for section, rows_s, fill in [('— Partial Payment (sorted by balance) —', partial, PARTIAL_FILL),
                                    ('— No Payment —',                           nopay,   NOPAY_FILL)]:
        if not rows_s: continue
        ws.merge_cells(f'A{row_n}:{get_column_letter(len(COL_HEADERS))}{row_n}')
        ws.cell(row=row_n,column=1,value=section).font = Font(bold=True,italic=True,size=10,color='595959')
        ws.cell(row=row_n,column=1).alignment = LEFT
        row_n += 1
        for r in rows_s:
            vals = [r['no'], r['building'], r['name'], r['room'],
                    r['checkin'], r['rent'], r['due'], r['basis'],
                    r['upi'], r['cash'], r['total'], r['bal'], r['status']]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=ci, value=v)
                cell.fill=fill; cell.border=BORDER; cell.font=NORM_FONT
                if ci in NUM_COLS:
                    cell.alignment=RIGHT
                    cell.number_format='#,##0'
                elif ci==5:  # date
                    cell.alignment=CENTER
                    cell.number_format='DD-MMM-YYYY'
                else:
                    cell.alignment=LEFT
            row_n += 1
            t_due+=r['due']; t_upi+=r['upi']; t_cash+=r['cash']; t_tot+=r['total']; t_bal+=r['bal']

    # Total row
    totals = [None,'','TOTAL','',None,None,t_due,None,t_upi,t_cash,t_tot,t_bal,'']
    for ci, v in enumerate(totals, 1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        cell.fill=TOTAL_FILL; cell.font=TOTAL_FONT; cell.border=BORDER
        if ci in NUM_COLS and isinstance(v,(int,float)):
            cell.alignment=RIGHT; cell.number_format='#,##0'

    for ci,w in enumerate(COL_WIDTHS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'
    return t_bal


MONTHS = [
    ('April', 'apr_due','apr_basis','apr_upi','apr_cash','apr_total','apr_bal',
     'data/reports/April_Outstanding_2026.xlsx'),
    ('May',   'may_due','may_basis','may_upi','may_cash','may_total','may_bal',
     'data/reports/May_Outstanding_2026.xlsx'),
]

for month_name, due_k, basis_k, upi_k, cash_k, total_k, bal_k, filepath in MONTHS:
    wb_out = openpyxl.Workbook()
    if wb_out.active:
        wb_out.remove(wb_out.active)

    all_outstanding = []
    for bname in ['HULK', 'THOR']:
        brows = [r for r in results if r['building']==bname]
        sheet_rows = []
        for r in sorted(brows, key=lambda x: -x[bal_k]):
            if r[due_k] <= 0 or r[bal_k] <= 0: continue
            sheet_rows.append({
                'no': 0,
                'building': bname,
                'name': r['name'],
                'room': r['room'],
                'checkin': r['checkin'],
                'rent': r['rent'],
                'due': r[due_k],
                'basis': r[basis_k],
                'upi': r[upi_k],
                'cash': r[cash_k],
                'total': r[total_k],
                'bal': r[bal_k],
                'status': 'No Payment' if r[total_k]==0 else f'{int(r[bal_k]):,} outstanding',
            })
        for i, r in enumerate(sheet_rows, 1): r['no'] = i
        all_outstanding.extend(sheet_rows)

        ws = wb_out.create_sheet(f'{bname}')
        write_month_sheet(ws, f'{bname} — Outstanding Dues — {month_name} 2026  (as of 11 May 2026)', sheet_rows)

    # Combined sheet (both buildings sorted by balance)
    combined_sorted = sorted(all_outstanding, key=lambda x: -x['bal'])
    for i, r in enumerate(combined_sorted, 1): r['no'] = i
    ws_all = wb_out.create_sheet('ALL Buildings', 0)  # first sheet
    total_bal = write_month_sheet(ws_all, f'HULK + THOR — Outstanding Dues — {month_name} 2026  (as of 11 May 2026)', combined_sorted)

    wb_out.save(filepath)
    print(f'Saved {filepath}  ({len(combined_sorted)} outstanding tenants, total balance Rs.{total_bal:,.0f})')

print('\nDone.')
