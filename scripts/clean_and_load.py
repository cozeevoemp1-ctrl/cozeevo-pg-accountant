"""
Clean Excel → Google Sheet loader.
Uses EXACT same clean_num() and occupancy logic as full_report.py.

Rules (from SHEET_LOGIC.md):
- Cash/UPI: from columns only, never comments
- Balance: clean text (exit dates=0, returns=0, Chandra=parallel amount)
- Occupancy: Active + checkin <= month_end. No-show: ALL, no date filter.
- Status: from rent status column as-is
"""
import sys, os, re, time
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from collections import Counter
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

EXCEL_FILE = "Cozeevo Monthly stay (4).xlsx"
NEW_SHEET = "1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw"
CREDS = "credentials/gsheets_service_account.json"
TOTAL_BEDS = 291

# ── Shared functions (identical to full_report.py) ───────────────────────────

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ('%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: continue
    return None

def clean_status(v):
    s = str(v).strip().upper() if v else ''
    if s in ('CHECKIN', 'CHECK IN', ''): return 'Active'
    if 'EXIT' in s: return 'Exited'
    if 'NO SHOW' in s: return 'No-show'
    if 'CANCEL' in s: return 'Cancelled'
    return 'Active'

def rent_status(v):
    s = str(v).strip().upper() if v else ''
    if not s: return ''
    if 'PAID' in s and 'NOT' not in s and 'PARTIAL' not in s: return 'PAID'
    if 'PARTIAL' in s: return 'PARTIAL'
    if 'NOT PAID' in s or 'UNPAID' in s: return 'UNPAID'
    if 'EXIT' in s: return 'EXIT'
    if 'NO SHOW' in s: return 'NO SHOW'
    if 'ADVANCE' in s: return 'ADVANCE'
    if 'CANCEL' in s: return 'CANCELLED'
    return s

CHANDRA_RE = re.compile(r'received\s+by\s+chandra', re.I)
EXIT_PATS = [
    re.compile(r'(?:march|feb|jan|apr)\s*\d+\s*(?:st|nd|rd|th)?\s*exit', re.I),
    re.compile(r'exit\s+on\s+', re.I),
]

def clean_num(raw, is_balance=False, cash_num=0, upi_num=0, rent=0):
    """Clean a cell. Returns (number, note, chandra_amt, lakshmi_amt).
    IDENTICAL to full_report.py — do not modify one without the other."""
    if raw is None: return 0, '', 0, 0
    s = str(raw).strip()
    if not s: return 0, '', 0, 0
    try: return float(s), '', 0, 0
    except: pass

    # Chandra in balance col: return 0 for balance, track separately
    if is_balance and CHANDRA_RE.search(s):
        amt = cash_num if cash_num > 0 else (upi_num if upi_num > 0 else rent)
        return 0, s, amt, 0
    # Chandra in cash col: return 0 for cash, track separately
    if not is_balance and CHANDRA_RE.search(s):
        sp = re.match(r'(\d[\d,]*)\s*Received.*?/\s*(\d[\d,]*)\s*(.*)', s, re.I)
        if sp:
            a1 = float(sp.group(1).replace(',', ''))
            a2 = float(sp.group(2).replace(',', ''))
            return 0, s, a1, a2
        amt = upi_num if upi_num > 0 else rent
        return 0, s, amt, 0
    # Exit dates = 0
    for p in EXIT_PATS:
        if p.search(s): return 0, s, 0, 0
    # Return = refund, not balance
    if re.match(r'return\s+\d+', s, re.I): return 0, s, 0, 0
    # paid in X = 0
    if re.match(r'paid\s+in\s+', s, re.I): return 0, s, 0, 0
    # hitachi = 0
    if re.match(r'hitachi', s, re.I): return 0, s, 0, 0
    # dash = 0
    if s == '-': return 0, '-', 0, 0
    # Expression: 516*16=8256
    eq = re.search(r'=\s*([\d,.]+)\s*$', s)
    if eq: return float(eq.group(1).replace(',', '')), s, 0, 0
    # Number + text: "5500 on april 1st", "9000 deposit"
    m = re.search(r'([\d,]+)', s)
    if m: return float(m.group(1).replace(',', '')), s, 0, 0
    return 0, s, 0, 0

def safe_str(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'):
        try: return str(int(float(s)))
        except: pass
    return s

# ── Google Sheet helpers ─────────────────────────────────────────────────────

def get_ws(sp, name, rows=300, cols=20):
    try:
        ws = sp.worksheet(name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sp.add_worksheet(name, rows=rows, cols=cols)
    return ws

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Step 1: Reading Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['History']

    tenants = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name or not str(name).strip(): continue

        rm = sn(ws.cell(row, 10).value)
        rf = sn(ws.cell(row, 11).value)
        ry = sn(ws.cell(row, 12).value)
        rent = ry if ry > 0 else (rf if rf > 0 else rm)

        tenants.append({
            'name': str(name).strip(),
            'room': safe_str(ws.cell(row, 1).value),
            'phone': safe_str(ws.cell(row, 4).value),
            'gender': safe_str(ws.cell(row, 3).value),
            'block': safe_str(ws.cell(row, 18).value),
            'floor': safe_str(ws.cell(row, 19).value),
            'sharing': safe_str(ws.cell(row, 13).value),
            'checkin_raw': safe_str(ws.cell(row, 5).value),
            'checkin': parse_date(ws.cell(row, 5).value),
            'status': clean_status(ws.cell(row, 17).value),
            'rent_monthly': rm, 'current_rent': rent,
            'deposit': sn(ws.cell(row, 7).value),
            'booking': sn(ws.cell(row, 6).value),
            'maintenance': sn(ws.cell(row, 8).value),
            'staff': safe_str(ws.cell(row, 16).value),
            'comment': safe_str(ws.cell(row, 15).value),
            'refund_status': safe_str(ws.cell(row, 39).value),
            'refund_amount': sn(ws.cell(row, 40).value),
            # Raw payment values (cleaned later per-month)
            'dec_st': rent_status(ws.cell(row, 21).value),
            'jan_st': rent_status(ws.cell(row, 22).value),
            'jan_bal': ws.cell(row, 23).value, 'jan_cash': ws.cell(row, 24).value, 'jan_upi': ws.cell(row, 25).value,
            'feb_st': rent_status(ws.cell(row, 26).value),
            'feb_bal': ws.cell(row, 28).value, 'feb_cash': ws.cell(row, 29).value, 'feb_upi': ws.cell(row, 30).value,
            'mar_st': rent_status(ws.cell(row, 27).value),
            'mar_bal': ws.cell(row, 31).value, 'mar_cash': ws.cell(row, 32).value, 'mar_upi': ws.cell(row, 33).value,
        })

    print(f"  {len(tenants)} tenants parsed")

    # Global occupancy (same for all months)
    all_noshow = [t for t in tenants if t['status'] == 'No-show']
    ns_prem = sum(1 for t in all_noshow if t['sharing'].lower() == 'premium')
    ns_beds = (len(all_noshow) - ns_prem) + ns_prem * 2

    # ── Step 2: Write to Google Sheet ────────────────────────────────────
    print("\nStep 2: Writing to Google Sheet...")
    creds = Credentials.from_service_account_file(CREDS,
        scopes=['https://www.googleapis.com/auth/spreadsheets'])
    gc = gspread.authorize(creds)
    sp = gc.open_by_key(NEW_SHEET)

    # ── TENANTS tab ──────────────────────────────────────────────────────
    ws_t = get_ws(sp, "TENANTS", 300, 18)
    t_headers = ["Room", "Name", "Phone", "Gender", "Building", "Floor",
                 "Sharing", "Check-in", "Status", "Monthly Rent", "Current Rent",
                 "Deposit", "Booking", "Maintenance", "Staff", "Comments",
                 "Refund Status", "Refund Amount"]
    t_rows = [[
        t['room'], t['name'], t['phone'], t['gender'], t['block'], t['floor'],
        t['sharing'], t['checkin_raw'], t['status'], t['rent_monthly'], t['current_rent'],
        t['deposit'], t['booking'], t['maintenance'], t['staff'], t['comment'],
        t['refund_status'], t['refund_amount'],
    ] for t in tenants]

    ws_t.update(values=[t_headers] + t_rows, range_name="A1", value_input_option="USER_ENTERED")
    ws_t.format("A1:R1", {"textFormat": {"bold": True},
                           "backgroundColor": {"red": 0.85, "green": 0.95, "blue": 0.85}})
    ws_t.freeze(rows=1)
    try: ws_t.set_basic_filter(f"A1:R{1 + len(t_rows)}")
    except: pass
    print(f"  TENANTS: {len(t_rows)} rows")
    time.sleep(8)

    # ── Monthly tabs ─────────────────────────────────────────────────────
    months_cfg = [
        ("DECEMBER 2025", date(2025,12,1), date(2025,12,31), 'dec_st', None, None, None),
        ("JANUARY 2026",  date(2026,1,1),  date(2026,1,31),  'jan_st', 'jan_cash', 'jan_upi', 'jan_bal'),
        ("FEBRUARY 2026", date(2026,2,1),  date(2026,2,28),  'feb_st', 'feb_cash', 'feb_upi', 'feb_bal'),
        ("MARCH 2026",    date(2026,3,1),  date(2026,3,31),  'mar_st', 'mar_cash', 'mar_upi', 'mar_bal'),
    ]

    all_month_stats = {}

    for label, ms, me, stk, ck, uk, bk in months_cfg:
        ws_m = get_ws(sp, label, 300, 14)
        headers = ["Room", "Name", "Building", "Sharing", "Rent",
                   "Cash", "UPI", "Total Paid", "Balance",
                   "Status", "Check-in", "Event", "Notes"]

        # Clean payments for this month
        m_rows = []
        cash_total = 0; upi_total = 0; bal_total = 0
        chandra_total = 0; lakshmi_total = 0
        new_checkins = 0; exits = 0

        for t in tenants:
            st = t.get(stk, '')
            has_payment = False
            cash = 0; upi = 0; bal = 0; notes_parts = []

            if ck:
                cash_num = sn(t[ck]) if isinstance(t[ck], (int, float)) else 0
                upi_num = sn(t[uk]) if isinstance(t[uk], (int, float)) else 0

                cash, cn, cc, cl = clean_num(t[ck], False, 0, upi_num, t['current_rent'])
                upi, un, _, _ = clean_num(t[uk])
                bal, bn, bc, bl = clean_num(t[bk], True, cash_num, upi_num, t['current_rent'])

                if cn: notes_parts.append(cn)
                if un: notes_parts.append(un)
                if bn: notes_parts.append(bn)

                chandra_total += cc + bc
                lakshmi_total += cl + bl
                has_payment = cash > 0 or upi > 0

            # Who appears in this month?
            # Rules:
            # - NO SHOW in rent status = "not here yet", skip UNLESS actual no-show
            # - EXIT: only show in FIRST month it appears, not repeated
            # - PAID/PARTIAL/UNPAID/ADVANCE: always show
            actual_noshow = (st == 'NO SHOW' and t['status'] == 'No-show')
            is_meaningful = st in ('PAID', 'PARTIAL', 'UNPAID', 'ADVANCE') or actual_noshow

            # EXIT: only include if this is the first month with EXIT
            if st == 'EXIT':
                prev_months = {'dec_st': None, 'jan_st': 'dec_st', 'feb_st': 'jan_st', 'mar_st': 'feb_st'}
                prev_key = prev_months.get(stk)
                prev_st = t.get(prev_key, '') if prev_key else ''
                if prev_st == 'EXIT':
                    # Previous month already had EXIT, skip this month
                    pass
                else:
                    is_meaningful = True

            if not (has_payment or is_meaningful):
                continue

            total_paid = cash + upi
            notes = " | ".join(notes_parts)

            # Event
            event = ""
            if t['checkin'] and ms <= t['checkin'] <= me:
                if t['status'] == 'No-show' or st == 'NO SHOW':
                    event = "NO-SHOW"
                else:
                    event = "NEW CHECK-IN"
                    new_checkins += 1
            if st == 'EXIT':
                event = "EXITED" if not event else event + " + EXITED"
                exits += 1

            cash_total += cash
            upi_total += upi
            bal_total += bal

            m_rows.append([
                t['room'], t['name'], t['block'], t['sharing'],
                t['current_rent'], cash, upi, total_paid, bal,
                st, t['checkin_raw'], event, notes,
            ])

        # Occupancy for this month
        active = [t for t in tenants if t['status'] == 'Active' and t['checkin'] and t['checkin'] <= me]
        prem_a = sum(1 for t in active if t['sharing'].lower() == 'premium')
        reg_a = len(active) - prem_a
        a_beds = reg_a + prem_a * 2
        vacant = TOTAL_BEDS - a_beds - ns_beds

        # THOR/HULK
        thor = [t for t in active if t['block'] == 'THOR']
        hulk = [t for t in active if t['block'] == 'HULK']
        tp = sum(1 for t in thor if t['sharing'].lower() == 'premium')
        hp = sum(1 for t in hulk if t['sharing'].lower() == 'premium')
        tb = (len(thor) - tp) + tp * 2
        hb = (len(hulk) - hp) + hp * 2

        # Status counts
        stc = Counter(t.get(stk, '') for t in tenants if t.get(stk, ''))

        collected = cash_total + upi_total
        tenant_dues = bal_total  # Balance column = tenant dues (Chandra tracked separately)

        # Summary rows
        summary = [
            [label, "", "", "", "", "", "", "", "", "", "", "", ""],
            [
                "Checked-in", f"{a_beds} beds ({reg_a}+{prem_a}P)",
                f"No-show: {ns_beds} beds ({len(all_noshow)}t)",
                f"Vacant: {vacant}", f"Occ: {a_beds/TOTAL_BEDS*100:.1f}%",
                "Cash", cash_total, "UPI", upi_total, "Total", collected,
                f"Bal: {bal_total:,.0f}", "",
            ],
            [
                f"THOR: {tb}b ({len(thor)}t,{tp}P)", f"HULK: {hb}b ({len(hulk)}t,{hp}P)",
                f"New: {new_checkins}", f"Exit: {exits}",
                f"PAID:{stc.get('PAID',0)}", f"PARTIAL:{stc.get('PARTIAL',0)}",
                f"UNPAID:{stc.get('UNPAID',0)}", f"NS:{stc.get('NO SHOW',0)}",
                f"EXIT:{stc.get('EXIT',0)}",
                f"Chandra:{chandra_total:,.0f}", f"Lakshmi:{lakshmi_total:,.0f}",
                f"TenantDues:{tenant_dues:,.0f}", "",
            ],
            headers,
        ]

        ws_m.update(values=summary + m_rows, range_name="A1", value_input_option="USER_ENTERED")
        ws_m.format("A1:M1", {"textFormat": {"bold": True, "fontSize": 13}})
        ws_m.format("A2:M3", {"textFormat": {"bold": True, "fontSize": 9},
                               "numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})
        ws_m.format("A4:M4", {"textFormat": {"bold": True},
                               "backgroundColor": {"red": 0.85, "green": 0.9, "blue": 1.0}})
        ws_m.freeze(rows=4)
        try: ws_m.set_basic_filter(f"A4:M{4 + len(m_rows)}")
        except: pass

        all_month_stats[label] = {
            'tenants': len(m_rows), 'beds': a_beds, 'noshow': ns_beds, 'vacant': vacant,
            'reg': reg_a, 'prem': prem_a,
            'cash': cash_total, 'upi': upi_total, 'collected': collected,
            'bal': bal_total, 'chandra': chandra_total, 'lakshmi': lakshmi_total,
            'tenant_dues': tenant_dues,
            'paid': stc.get('PAID', 0), 'partial': stc.get('PARTIAL', 0),
            'unpaid': stc.get('UNPAID', 0), 'ns_count': stc.get('NO SHOW', 0),
            'exit_count': stc.get('EXIT', 0),
            'new': new_checkins, 'exits': exits,
            'thor_beds': tb, 'hulk_beds': hb, 'thor_t': len(thor), 'hulk_t': len(hulk),
        }

        print(f"  {label}: {len(m_rows)}t | Beds:{a_beds} NS:{ns_beds} V:{vacant} | "
              f"Cash:{cash_total:,.0f} UPI:{upi_total:,.0f} | "
              f"Chandra:{chandra_total:,.0f} Dues:{tenant_dues:,.0f}")
        time.sleep(8)

    # ── DASHBOARD ────────────────────────────────────────────────────────
    print("  Building DASHBOARD...")
    ws_d = get_ws(sp, "DASHBOARD", 50, 8)

    # Use latest month (March)
    latest_label = "MARCH 2026"
    d = all_month_stats[latest_label]
    occ_pct = f"{d['beds']/TOTAL_BEDS*100:.1f}%"
    coll_pct = f"{d['collected']/d['collected']*100:.0f}%" if d['collected'] else "0%"

    rows = []
    # R1: Title
    rows.append(["COZEEVO OPERATIONS DASHBOARD", "", "", "", "", "", "", ""])
    # R2: blank
    rows.append(["", "", "", "", "", "", "", ""])
    # R3: Current month
    rows.append([latest_label, "", "", "", "", "", "", ""])
    # R4: blank
    rows.append(["", "", "", "", "", "", "", ""])

    # R5-R11: Occupancy + Collections side by side
    rows.append(["OCCUPANCY", "", "", "", "COLLECTIONS", "", "", ""])
    rows.append(["Revenue Beds", TOTAL_BEDS, "", "", "Cash", d['cash'], "", ""])
    rows.append(["Checked-in", d['beds'], f"{d['reg']} reg + {d['prem']} premium", "", "UPI", d['upi'], "", ""])
    rows.append(["No-show", d['noshow'], f"{len(all_noshow)} tenants", "", "Total Collected", d['collected'], "", ""])
    rows.append(["Vacant", d['vacant'], "", "", "", "", "", ""])
    rows.append(["Occupancy", occ_pct, "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", ""])

    # R12-R16: Status + Movement
    rows.append(["PAYMENT STATUS", "", "", "", "MOVEMENT", "", "", ""])
    rows.append(["Paid", d['paid'], "", "", "New Check-ins", d['new'], "", ""])
    rows.append(["Partial", d['partial'], "", "", "Exits", d['exits'], "", ""])
    rows.append(["Unpaid", d['unpaid'], "", "", "No-show", len(all_noshow), "", ""])
    rows.append(["", "", "", "", "", "", "", ""])

    # R17-R20: Balance breakdown
    rows.append(["OUTSTANDING", "", "", "", "THOR vs HULK", "", "", ""])
    rows.append(["Tenant Dues", d['tenant_dues'], "", "", "THOR", f"{d['thor_beds']} beds", f"{d['thor_t']} tenants", ""])
    rows.append(["Chandra (collected)", d['chandra'], "", "", "HULK", f"{d['hulk_beds']} beds", f"{d['hulk_t']} tenants", ""])
    rows.append(["Lakshmi (collected)", d['lakshmi'], "", "", "", "", "", ""])
    rows.append(["Total Outstanding", d['bal'], "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", ""])

    # R23+: Month-on-Month
    rows.append(["MONTH-ON-MONTH", "", "", "", "", "", "", ""])
    rows.append(["Month", "Tenants", "Beds", "Cash", "UPI", "Collected", "Balance", ""])
    for ml in ["DECEMBER 2025", "JANUARY 2026", "FEBRUARY 2026", "MARCH 2026"]:
        s = all_month_stats[ml]
        rows.append([ml, s['tenants'], s['beds'], s['cash'], s['upi'], s['collected'], s['bal'], ""])

    rows.append(["", "", "", "", "", "", "", ""])
    rows.append(["Updated: " + datetime.now().strftime("%Y-%m-%d %H:%M"), "", "", "", "", "", "", ""])

    ws_d.update(values=rows, range_name="A1", value_input_option="USER_ENTERED")

    # ── Formatting ───────────────────────────────────────────────────────
    # Title
    ws_d.format("A1:H1", {"textFormat": {"bold": True, "fontSize": 18, "foregroundColorStyle": {"rgbColor": {"red": 0.08, "green": 0.4, "blue": 0.75}}},
                           "backgroundColor": {"red": 0.96, "green": 0.96, "blue": 0.96}})
    ws_d.merge_cells("A1:H1")
    # Current month badge
    ws_d.format("A3:H3", {"textFormat": {"bold": True, "fontSize": 14, "foregroundColorStyle": {"rgbColor": {"red": 1, "green": 1, "blue": 1}}},
                           "backgroundColor": {"red": 0.08, "green": 0.4, "blue": 0.75}})
    ws_d.merge_cells("A3:H3")

    # Occupancy header
    ws_d.format("A5:C5", {"textFormat": {"bold": True, "fontSize": 11},
                           "backgroundColor": {"red": 0.84, "green": 0.92, "blue": 1.0}})
    # Collections header
    ws_d.format("E5:F5", {"textFormat": {"bold": True, "fontSize": 11},
                           "backgroundColor": {"red": 0.84, "green": 1.0, "blue": 0.84}})
    # Number formatting
    ws_d.format("B6:B10", {"textFormat": {"bold": True}, "numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})
    ws_d.format("F6:F8", {"textFormat": {"bold": True}, "numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})

    # Payment status header
    ws_d.format("A12:C12", {"textFormat": {"bold": True, "fontSize": 11},
                             "backgroundColor": {"red": 1.0, "green": 0.94, "blue": 0.84}})
    ws_d.format("E12:F12", {"textFormat": {"bold": True, "fontSize": 11},
                             "backgroundColor": {"red": 0.91, "green": 0.84, "blue": 0.96}})
    # Paid/partial/unpaid colors
    ws_d.format("A13:C13", {"backgroundColor": {"red": 0.91, "green": 0.97, "blue": 0.91}})  # green
    ws_d.format("A14:C14", {"backgroundColor": {"red": 1.0, "green": 0.97, "blue": 0.88}})   # yellow
    ws_d.format("A15:C15", {"backgroundColor": {"red": 0.99, "green": 0.93, "blue": 0.93}})  # red
    ws_d.format("B13:B15", {"textFormat": {"bold": True}})
    ws_d.format("F13:F15", {"textFormat": {"bold": True}})

    # Outstanding header
    ws_d.format("A17:C17", {"textFormat": {"bold": True, "fontSize": 11},
                             "backgroundColor": {"red": 0.98, "green": 0.85, "blue": 0.85}})
    ws_d.format("E17:G17", {"textFormat": {"bold": True, "fontSize": 11},
                             "backgroundColor": {"red": 0.93, "green": 0.93, "blue": 1.0}})
    ws_d.format("B18:B21", {"textFormat": {"bold": True}, "numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})

    # Month-on-month header
    r_mom = len(rows) - 6  # row number for MoM header
    ws_d.format(f"A{r_mom}:H{r_mom}", {"textFormat": {"bold": True, "fontSize": 12},
                                         "backgroundColor": {"red": 0.83, "green": 0.87, "blue": 0.95}})
    ws_d.merge_cells(f"A{r_mom}:H{r_mom}")
    ws_d.format(f"A{r_mom+1}:H{r_mom+1}", {"textFormat": {"bold": True},
                                              "backgroundColor": {"red": 0.92, "green": 0.94, "blue": 0.98}})
    # MoM data formatting
    for i in range(4):
        r = r_mom + 2 + i
        bg = {"red": 1, "green": 1, "blue": 1} if i % 2 == 0 else {"red": 0.97, "green": 0.97, "blue": 0.97}
        ws_d.format(f"A{r}:H{r}", {"backgroundColor": bg})
    ws_d.format(f"B{r_mom+2}:G{r_mom+5}", {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})

    # Borders
    for rng in ["A5:C10", "E5:F8", "A12:C15", "E12:F15", "A17:C21", "E17:G19"]:
        ws_d.format(rng, {"borders": {"top": {"style": "SOLID", "colorStyle": {"rgbColor": {"red": 0.83, "green": 0.83, "blue": 0.83}}},
                                       "bottom": {"style": "SOLID", "colorStyle": {"rgbColor": {"red": 0.83, "green": 0.83, "blue": 0.83}}},
                                       "left": {"style": "SOLID", "colorStyle": {"rgbColor": {"red": 0.83, "green": 0.83, "blue": 0.83}}},
                                       "right": {"style": "SOLID", "colorStyle": {"rgbColor": {"red": 0.83, "green": 0.83, "blue": 0.83}}}}})

    # Column widths
    for i, w in enumerate([180, 110, 170, 20, 180, 120, 120, 100], 1):
        ws_d.update_dimension_properties(
            ws_d, 'COLUMNS', i - 1, i,
            gspread.utils.cast_to_a1_notation(1, i)
        ) if False else None  # gspread doesn't support setColumnWidth directly

    # Footer
    ws_d.format(f"A{len(rows)}", {"textFormat": {"foregroundColorStyle": {"rgbColor": {"red": 0.62, "green": 0.62, "blue": 0.62}}, "fontSize": 9}})

    print("  DASHBOARD: done")

    # ── Cleanup old tabs ─────────────────────────────────────────────────
    for name in ["Sheet1", "ROOMS", "PAYMENTS", "CHANGES LOG", "MONTHLY VIEW",
                 "MONTHLY JAN 2026", "MONTHLY FEB 2026", "_LOOKUP", "COLLECT RENT"]:
        try: sp.del_worksheet(sp.worksheet(name))
        except: pass

    print(f"\n{'='*60}")
    print("DONE! Tabs: DASHBOARD, TENANTS, DECEMBER-MARCH")
    print("Next: paste gsheet_apps_script.js for live updates")


if __name__ == "__main__":
    main()
