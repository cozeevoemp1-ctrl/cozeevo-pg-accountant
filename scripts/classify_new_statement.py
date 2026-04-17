"""
scripts/classify_new_statement.py
Classify YES Bank CSV statement (Jan-Mar 2026) using pnl_classify rules.
Outputs: data/reports/bank_classified_jan_mar_2026.xlsx
"""
import csv, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from src.rules.pnl_classify import classify_txn
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

STATEMENT = "Statement-124563400000961-04-15-2026-18-24-23.csv"

# ── Parse CSV ────────────────────────────────────────────────────────────────
rows = []
with open(STATEMENT, "r") as f:
    for line in f:
        if line.startswith("Transaction Date"):
            break
    reader = csv.reader(f)
    for r in reader:
        if len(r) < 7:
            continue
        dt_str, _, desc, ref, wd, dep, bal = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        try:
            dt = datetime.strptime(dt_str.strip(), "%Y-%m-%d")
        except ValueError:
            continue
        wd = float(wd.replace(",", "").strip()) if wd.strip() else 0
        dep = float(dep.replace(",", "").strip()) if dep.strip() else 0
        if wd > 0:
            cat, sub = classify_txn(desc, "expense")
            rows.append((dt, desc.strip(), "expense", wd, cat, sub, ref.strip()))
        if dep > 0:
            cat, sub = classify_txn(desc, "income")
            rows.append((dt, desc.strip(), "income", dep, cat, sub, ref.strip()))

# ── Summary ──────────────────────────────────────────────────────────────────
by_cat = defaultdict(float)
by_month_cat = defaultdict(lambda: defaultdict(float))
unclassified_count = 0
for dt, desc, typ, amt, cat, sub, ref in rows:
    key = f"{typ}:{cat}"
    by_cat[key] += amt
    by_month_cat[dt.strftime("%Y-%m")][key] += amt
    if cat in ("Other Expenses", "Other Income"):
        unclassified_count += 1

print(f"Total: {len(rows)} txns, {unclassified_count} unclassified")
print()
print("=== EXPENSES ===")
total_exp = 0
for k in sorted(by_cat):
    if k.startswith("expense:"):
        print(f"  {k[8:]:<30} Rs.{by_cat[k]:>12,.0f}")
        total_exp += by_cat[k]
print(f"  {'TOTAL EXPENSES':<30} Rs.{total_exp:>12,.0f}")

print()
print("=== INCOME ===")
total_inc = 0
for k in sorted(by_cat):
    if k.startswith("income:"):
        print(f"  {k[7:]:<30} Rs.{by_cat[k]:>12,.0f}")
        total_inc += by_cat[k]
print(f"  {'TOTAL INCOME':<30} Rs.{total_inc:>12,.0f}")
print(f"  {'NET':<30} Rs.{total_inc - total_exp:>12,.0f}")

# ── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1a1a2e")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="F8F9FA")
WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
YLW_FILL = PatternFill("solid", fgColor="FFFACD")
CTR = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()

# ── Sheet 1: All Transactions ────────────────────────────────────────────────
ws = wb.active
ws.title = "All Transactions"
hdrs = ["Date", "Type", "Description", "Amount", "Category", "Sub-Category", "Reference"]
for col, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = CTR

for ri, (dt, desc, typ, amt, cat, sub, ref) in enumerate(
    sorted(rows, key=lambda x: x[0], reverse=True), 2
):
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    ws.cell(row=ri, column=1, value=dt.strftime("%d %b %Y")).fill = fill
    ws.cell(row=ri, column=2, value=typ.upper()).fill = fill
    ws.cell(row=ri, column=3, value=desc).fill = fill
    c = ws.cell(row=ri, column=4, value=amt)
    c.fill = fill
    c.number_format = "#,##0"
    ws.cell(row=ri, column=5, value=cat).fill = fill
    ws.cell(row=ri, column=6, value=sub).fill = fill
    ws.cell(row=ri, column=7, value=ref).fill = fill

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 80
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 30
ws.column_dimensions["G"].width = 30
ws.freeze_panes = "A2"

# ── Sheet 2: Monthly Summary ────────────────────────────────────────────────
ws2 = wb.create_sheet("Monthly Summary")
months = sorted(by_month_cat.keys())
cats_exp = sorted(set(k for k in by_cat if k.startswith("expense:")))
cats_inc = sorted(set(k for k in by_cat if k.startswith("income:")))

hdrs2 = ["Category"] + months + ["Total"]
for col, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = CTR

row = 2
ws2.cell(row=row, column=1, value="--- EXPENSES ---").font = Font(bold=True)
row += 1
for cat_key in cats_exp:
    label = cat_key[8:]
    ws2.cell(row=row, column=1, value=label)
    total = 0
    for ci, m in enumerate(months, 2):
        val = by_month_cat[m].get(cat_key, 0)
        c = ws2.cell(row=row, column=ci, value=val if val else "")
        if val:
            c.number_format = "#,##0"
        total += val
    c = ws2.cell(row=row, column=len(months) + 2, value=total)
    c.number_format = "#,##0"
    c.font = Font(bold=True)
    row += 1

# Total expenses row
ws2.cell(row=row, column=1, value="TOTAL EXPENSES").font = Font(bold=True, color="FF0000")
for ci, m in enumerate(months, 2):
    val = sum(by_month_cat[m].get(k, 0) for k in cats_exp)
    c = ws2.cell(row=row, column=ci, value=val)
    c.number_format = "#,##0"
    c.font = Font(bold=True, color="FF0000")
c = ws2.cell(row=row, column=len(months) + 2, value=total_exp)
c.number_format = "#,##0"
c.font = Font(bold=True, color="FF0000")
row += 2

ws2.cell(row=row, column=1, value="--- INCOME ---").font = Font(bold=True)
row += 1
for cat_key in cats_inc:
    label = cat_key[7:]
    ws2.cell(row=row, column=1, value=label)
    total = 0
    for ci, m in enumerate(months, 2):
        val = by_month_cat[m].get(cat_key, 0)
        c = ws2.cell(row=row, column=ci, value=val if val else "")
        if val:
            c.number_format = "#,##0"
        total += val
    c = ws2.cell(row=row, column=len(months) + 2, value=total)
    c.number_format = "#,##0"
    c.font = Font(bold=True)
    row += 1

ws2.cell(row=row, column=1, value="TOTAL INCOME").font = Font(bold=True, color="008000")
for ci, m in enumerate(months, 2):
    val = sum(by_month_cat[m].get(k, 0) for k in cats_inc)
    c = ws2.cell(row=row, column=ci, value=val)
    c.number_format = "#,##0"
    c.font = Font(bold=True, color="008000")
c = ws2.cell(row=row, column=len(months) + 2, value=total_inc)
c.number_format = "#,##0"
c.font = Font(bold=True, color="008000")

ws2.column_dimensions["A"].width = 30
for ci in range(2, len(months) + 3):
    ws2.column_dimensions[get_column_letter(ci)].width = 16
ws2.freeze_panes = "B2"

# ── Sheet 3: Unclassified Review ────────────────────────────────────────────
ws3 = wb.create_sheet("Unclassified Review")
hdrs3 = ["Date", "Type", "Description", "Amount", "Current Category", "Your Correction"]
for col, h in enumerate(hdrs3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = CTR

ri = 2
for dt, desc, typ, amt, cat, sub, ref in sorted(rows, key=lambda x: -x[3]):
    if cat not in ("Other Expenses", "Other Income"):
        continue
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    ws3.cell(row=ri, column=1, value=dt.strftime("%d %b %Y")).fill = fill
    ws3.cell(row=ri, column=2, value=typ.upper()).fill = fill
    ws3.cell(row=ri, column=3, value=desc).fill = fill
    c = ws3.cell(row=ri, column=4, value=amt)
    c.fill = fill
    c.number_format = "#,##0"
    ws3.cell(row=ri, column=5, value=f"{cat} / {sub}").fill = fill
    ws3.cell(row=ri, column=6, value="").fill = YLW_FILL
    ri += 1

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 80
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 25
ws3.column_dimensions["F"].width = 30
ws3.freeze_panes = "A2"

os.makedirs("data/reports", exist_ok=True)
out = "data/reports/bank_classified_jan_mar_2026.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
