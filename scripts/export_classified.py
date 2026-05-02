"""
scripts/export_classified.py
Run: venv/Scripts/python scripts/export_classified.py

Outputs:
  data/reports/unclassified_review.xlsx   — one sheet per month, yellow 'Your Correction' column
  data/reports/expense_classified_full.xlsx — full transaction log + monthly summary
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from src.rules.pnl_classify import classify_txn
from src.utils.inr_format import INR_NUMBER_FORMAT
from collections import defaultdict
from datetime import datetime

# ── helpers ────────────────────────────────────────────────────────────────
def parse_date(v):
    if hasattr(v, 'strftime'): return v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try: return datetime.strptime(s[:10], fmt)
        except: pass
    return None

def parse_amt(v):
    if v is None or v == '': return 0.0
    try: return float(str(v).replace(',', '').strip())
    except: return 0.0

def read_yes_bank(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hrow = None
    for i, row in enumerate(rows):
        if row and any('transaction date' in str(c).lower() for c in row if c):
            hrow = i; break
    if hrow is None: return []
    headers = [str(c).lower().strip() if c else '' for c in rows[hrow]]
    out = []
    for row in rows[hrow+1:]:
        if not any(row): continue
        d = dict(zip(headers, row))
        dt = parse_date(d.get('transaction date'))
        if not dt: continue
        wd  = parse_amt(d.get('withdrawals', ''))
        dep = parse_amt(d.get('deposits', ''))
        desc = str(d.get('description') or '')
        if wd  > 0: out.append((dt, desc, 'expense', wd))
        if dep > 0: out.append((dt, desc, 'income',  dep))
    return out

# ── CSV reader (YES Bank CSV format) ──────────────────────────────────────
import csv as _csv

def read_yes_bank_csv(path):
    out = []
    with open(path, 'r') as f:
        for line in f:
            if line.startswith('Transaction Date'):
                break
        reader = _csv.reader(f)
        for row in reader:
            if len(row) < 7: continue
            dt_str, _, desc, ref, wd, dep, bal = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            dt = parse_date(dt_str)
            if not dt: continue
            wd_val = parse_amt(wd)
            dep_val = parse_amt(dep)
            if wd_val > 0: out.append((dt, desc.strip(), 'expense', wd_val))
            if dep_val > 0: out.append((dt, desc.strip(), 'income', dep_val))
    return out

# ── load data ──────────────────────────────────────────────────────────────
# Priority order (highest first): CSV is most recent download and overrides
# overlapping Excel downloads for the same period. Dedupe by (date, amt, desc).
txns = []
sources = []

import glob as _glob
# CSVs first — newest downloads (covers Statement-*.csv and "YYYY statment.csv")
csv_files = sorted(set(_glob.glob('Statement-*.csv') + _glob.glob('*statment*.csv') + _glob.glob('*statement*.csv') + _glob.glob('april month.csv') + _glob.glob('april*.csv')), reverse=True)
for f in csv_files:
    t = read_yes_bank_csv(f)
    print('Loaded %d from %s' % (len(t), f))
    txns += t
    sources.append(f)

# Then Excel statements (only add if not already seen)
for f in ['2025 statement.xlsx', '2026 statment.xlsx']:
    if not os.path.exists(f):
        continue
    t = read_yes_bank(f)
    print('Loaded %d from %s' % (len(t), f))
    txns += t
    sources.append(f)

# Dedupe by (date, amount, description) — same UTR / same cheque should
# only count once even if it's present in both an Excel and a CSV dump.
seen = set()
deduped = []
for dt, desc, typ, amt in txns:
    key = (dt.strftime('%Y-%m-%d'), round(float(amt), 2), (desc or '').strip().lower())
    if key in seen:
        continue
    seen.add(key)
    deduped.append((dt, desc, typ, amt))
print('After dedupe: %d txns (removed %d duplicates)' % (len(deduped), len(txns) - len(deduped)))
txns = deduped

classified = []
for dt, desc, typ, amt in txns:
    cat, sub = classify_txn(desc, typ)
    classified.append((dt, desc, typ, amt, cat, sub))

# ── style constants ────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1a1a2e')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL  = PatternFill('solid', fgColor='F8F9FA')
WHT_FILL  = PatternFill('solid', fgColor='FFFFFF')
YLW_FILL  = PatternFill('solid', fgColor='FFFACD')
RED_FILL  = PatternFill('solid', fgColor='FFF3CD')
CTR       = Alignment(horizontal='center', vertical='center', wrap_text=False)

CATS = [
    'Property Rent', 'Electricity', 'Water', 'IT & Software', 'Internet & WiFi',
    'Food & Groceries', 'Fuel & Diesel', 'Staff & Labour', 'Furniture & Fittings',
    'Maintenance & Repairs', 'Cleaning Supplies', 'Waste Disposal',
    'Shopping & Supplies', 'Operational Expenses', 'Marketing',
    'Govt & Regulatory', 'Tenant Deposit Refund', 'Bank Charges',
    'Capital Investment', 'Non-Operating',
    'Other Expenses',
]

# ════════════════════════════════════════════════════════════════════════════
# EXCEL 1: Unclassified review — one sheet per month
# ════════════════════════════════════════════════════════════════════════════
months_unc = defaultdict(list)
for dt, desc, typ, amt, cat, sub in classified:
    if typ == 'expense' and cat == 'Other Expenses':
        months_unc[dt.strftime('%Y-%m')].append((dt, desc, amt, cat, sub))

wb1 = openpyxl.Workbook()
wb1.remove(wb1.active)

for month in sorted(months_unc.keys()):
    items = sorted(months_unc[month], key=lambda x: -x[2])
    ws = wb1.create_sheet(title=month)

    hdrs = ['Date', 'Description', 'Amount (Rs)', 'Current Category', 'Sub-category', 'Your Correction (fill this)']
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    ws.row_dimensions[1].height = 22

    total = 0
    for ri, (dt, desc, amt, cat, sub) in enumerate(items, 2):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        ws.cell(row=ri, column=1, value=dt.strftime('%d %b %Y')).fill = fill
        ws.cell(row=ri, column=2, value=desc).fill = fill
        c = ws.cell(row=ri, column=3, value=amt)
        c.fill = fill; c.number_format = INR_NUMBER_FORMAT
        ws.cell(row=ri, column=4, value=cat).fill = fill
        ws.cell(row=ri, column=5, value=sub).fill = fill
        ws.cell(row=ri, column=6, value='').fill = YLW_FILL  # yellow = fill in
        total += amt

    tr = len(items) + 2
    c = ws.cell(row=tr, column=1, value='TOTAL')
    c.font = Font(bold=True)
    c = ws.cell(row=tr, column=3, value=total)
    c.font = Font(bold=True); c.number_format = INR_NUMBER_FORMAT

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 72
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 28
    ws.freeze_panes = 'A2'

out1 = os.path.join('data', 'reports', 'unclassified_review.xlsx')
wb1.save(out1)
print('Saved: %s' % out1)

# ════════════════════════════════════════════════════════════════════════════
# EXCEL 2: Full classified breakdown
# ════════════════════════════════════════════════════════════════════════════
months = sorted(set(dt.strftime('%Y-%m') for dt, *_ in classified))
monthly_exp = defaultdict(lambda: defaultdict(float))
monthly_inc = defaultdict(lambda: defaultdict(float))
monthly_sub = defaultdict(float)  # (month, cat, sub) -> amt

INCOME_CATS = ['Rent Income', 'Advance Deposit', 'Other Income']

for dt, desc, typ, amt, cat, sub in classified:
    m = dt.strftime('%Y-%m')
    if typ == 'expense':
        monthly_exp[m][cat] += amt
        monthly_sub[(m, cat, sub)] += amt
    else:
        monthly_inc[m][cat] += amt

wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

# ── Styles for P&L ───────────────────────────────────────────────────────
SEC_FILL  = PatternFill('solid', fgColor='2d2d44')
SEC_FONT  = Font(bold=True, color='FFFFFF', size=11)
TOT_FONT  = Font(bold=True, size=11)
TOT_FONT_R = Font(bold=True, size=11, color='FF0000')
TOT_FONT_G = Font(bold=True, size=11, color='008B00')
NUM_FMT   = INR_NUMBER_FORMAT

def _write_row(ws, ri, label, vals, font=None, fill=None, indent=False):
    """Write a label + monthly values row."""
    lbl = ('  ' + label) if indent else label
    c = ws.cell(row=ri, column=1, value=lbl)
    if font: c.font = font
    if fill: c.fill = fill
    for ci, v in enumerate(vals, 2):
        c = ws.cell(row=ri, column=ci, value=round(v) if v else '')
        c.number_format = NUM_FMT
        if font: c.font = font
        if fill: c.fill = fill

# ── Sheet 1: Monthly P&L ────────────────────────────────────────────────
ws_sum = wb2.create_sheet('Monthly P&L')
hdr_row = [''] + months + ['TOTAL']
for col, h in enumerate(hdr_row, 1):
    c = ws_sum.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
ws_sum.row_dimensions[1].height = 22

ri = 2

# ── INCOME section ───────────────────────────────────────────────────────
c = ws_sum.cell(row=ri, column=1, value='INCOME')
c.fill = SEC_FILL; c.font = SEC_FONT
for col in range(2, len(months)+3):
    ws_sum.cell(row=ri, column=col).fill = SEC_FILL
ri += 1

inc_month_totals = [0.0] * (len(months) + 1)
for cat in INCOME_CATS:
    vals = []
    for m in months:
        v = monthly_inc[m].get(cat, 0)
        vals.append(v)
    vals.append(sum(vals))
    if sum(vals) == 0: continue
    for i, v in enumerate(vals): inc_month_totals[i] += v
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    _write_row(ws_sum, ri, cat, vals, indent=True, fill=fill)
    ri += 1

_write_row(ws_sum, ri, 'TOTAL INCOME', inc_month_totals, font=TOT_FONT_G)
ri += 2

# ── OPERATING EXPENSES section ───────────────────────────────────────────
c = ws_sum.cell(row=ri, column=1, value='OPERATING EXPENSES')
c.fill = SEC_FILL; c.font = SEC_FONT
for col in range(2, len(months)+3):
    ws_sum.cell(row=ri, column=col).fill = SEC_FILL
ri += 1

op_cats = [c for c in CATS if c != 'Non-Operating']
exp_month_totals = [0.0] * (len(months) + 1)
for cat in op_cats:
    vals = []
    for m in months:
        v = monthly_exp[m].get(cat, 0)
        vals.append(v)
    vals.append(sum(vals))
    if sum(vals) == 0: continue
    for i, v in enumerate(vals): exp_month_totals[i] += v
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    _write_row(ws_sum, ri, cat, vals, indent=True, fill=fill)
    ri += 1

_write_row(ws_sum, ri, 'TOTAL OPERATING EXPENSES', exp_month_totals, font=TOT_FONT_R)
ri += 2

# ── OPERATING PROFIT ─────────────────────────────────────────────────────
op_profit = [inc_month_totals[i] - exp_month_totals[i] for i in range(len(inc_month_totals))]
_write_row(ws_sum, ri, 'OPERATING PROFIT (EBITDA)', op_profit, font=TOT_FONT)
ri += 2

# ── NON-OPERATING section ────────────────────────────────────────────────
nonop_vals = []
for m in months:
    v = monthly_exp[m].get('Non-Operating', 0)
    nonop_vals.append(v)
nonop_vals.append(sum(nonop_vals))

if sum(nonop_vals) > 0:
    c = ws_sum.cell(row=ri, column=1, value='NON-OPERATING EXPENSES')
    c.fill = SEC_FILL; c.font = SEC_FONT
    for col in range(2, len(months)+3):
        ws_sum.cell(row=ri, column=col).fill = SEC_FILL
    ri += 1
    _write_row(ws_sum, ri, 'Loan Repayment / Transfers', nonop_vals, indent=True, fill=ALT_FILL)
    ri += 2

# ── NET PROFIT ───────────────────────────────────────────────────────────
net = [op_profit[i] - nonop_vals[i] for i in range(len(op_profit))]
_write_row(ws_sum, ri, 'NET PROFIT / (LOSS)', net, font=Font(bold=True, size=12))
ri += 2

# ── MARGINS ──────────────────────────────────────────────────────────────
c = ws_sum.cell(row=ri, column=1, value='OPERATING MARGIN %')
c.font = Font(bold=True, italic=True)
for ci, m in enumerate(months, 2):
    inc = inc_month_totals[ci-2]
    margin = (op_profit[ci-2] / inc * 100) if inc else 0
    c = ws_sum.cell(row=ri, column=ci, value=f'{margin:.1f}%')
    c.font = Font(italic=True)
inc = inc_month_totals[-1]
margin = (op_profit[-1] / inc * 100) if inc else 0
c = ws_sum.cell(row=ri, column=len(months)+2, value=f'{margin:.1f}%')
c.font = Font(bold=True, italic=True)

ws_sum.column_dimensions['A'].width = 32
for i in range(2, len(months)+3):
    ws_sum.column_dimensions[get_column_letter(i)].width = 15
ws_sum.freeze_panes = 'B2'

# ── Keep old monthly_cat for sub-category sheet ──────────────────────────
monthly_cat = monthly_exp

# ── Sheet 2: Sub-category breakdown ───────────────────────────────────────
ws_sub = wb2.create_sheet('Sub-category Breakdown')
sub_hdrs = ['Category', 'Sub-category'] + months + ['TOTAL']
for col, h in enumerate(sub_hdrs, 1):
    c = ws_sub.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
ws_sub.row_dimensions[1].height = 22

ri = 2
for cat in CATS:
    # Collect all subs across months
    all_subs = sorted(set(
        s for (m, c, s), v in monthly_sub.items()
        if c == cat and v > 0
    ))
    if not all_subs: continue

    cat_total = sum(monthly_cat[m].get(cat, 0) for m in months)
    if cat_total == 0: continue

    # Category header row
    c = ws_sub.cell(row=ri, column=1, value=cat)
    c.fill = PatternFill('solid', fgColor='2d2d44')
    c.font = Font(bold=True, color='FFFFFF')
    for col in range(2, len(months)+3):
        ws_sub.cell(row=ri, column=col).fill = PatternFill('solid', fgColor='2d2d44')
    ri += 1

    for sub in all_subs:
        sub_total = sum(monthly_sub.get((m, cat, sub), 0) for m in months)
        if sub_total == 0: continue
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        ws_sub.cell(row=ri, column=1, value='').fill = fill
        ws_sub.cell(row=ri, column=2, value=sub).fill = fill
        for ci, m in enumerate(months, 3):
            v = monthly_sub.get((m, cat, sub), 0)
            c = ws_sub.cell(row=ri, column=ci, value=round(v) if v else 0)
            c.fill = fill; c.number_format = INR_NUMBER_FORMAT
        c = ws_sub.cell(row=ri, column=len(months)+3, value=round(sub_total))
        c.fill = fill; c.number_format = INR_NUMBER_FORMAT; c.font = Font(bold=True)
        ri += 1

ws_sub.column_dimensions['A'].width = 26
ws_sub.column_dimensions['B'].width = 38
for i in range(3, len(months)+4):
    ws_sub.column_dimensions[get_column_letter(i)].width = 14
ws_sub.freeze_panes = 'C2'

# ── Sheet 3: Full transaction log ─────────────────────────────────────────
ws_txn = wb2.create_sheet('All Transactions')
txn_hdrs = ['Date', 'Month', 'Type', 'Category', 'Sub-category', 'Amount (Rs)', 'Description']
for col, h in enumerate(txn_hdrs, 1):
    c = ws_txn.cell(row=1, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR

for ri, (dt, desc, typ, amt, cat, sub) in enumerate(sorted(classified, key=lambda x: x[0]), 2):
    fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
    ws_txn.cell(row=ri, column=1, value=dt.strftime('%d %b %Y')).fill = fill
    ws_txn.cell(row=ri, column=2, value=dt.strftime('%Y-%m')).fill = fill
    ws_txn.cell(row=ri, column=3, value=typ).fill = fill
    ws_txn.cell(row=ri, column=4, value=cat).fill = fill
    ws_txn.cell(row=ri, column=5, value=sub).fill = fill
    c = ws_txn.cell(row=ri, column=6, value=amt)
    c.fill = fill; c.number_format = INR_NUMBER_FORMAT
    ws_txn.cell(row=ri, column=7, value=desc).fill = fill

ws_txn.column_dimensions['A'].width = 14
ws_txn.column_dimensions['B'].width = 10
ws_txn.column_dimensions['C'].width = 10
ws_txn.column_dimensions['D'].width = 26
ws_txn.column_dimensions['E'].width = 38
ws_txn.column_dimensions['F'].width = 14
ws_txn.column_dimensions['G'].width = 80
ws_txn.auto_filter.ref = 'A1:G1'
ws_txn.freeze_panes = 'A2'

out2 = os.path.join('data', 'reports', 'expense_classified_full.xlsx')
wb2.save(out2)
print('Saved: %s' % out2)

# ── Print updated table ────────────────────────────────────────────────────
print()
print('UPDATED CLASSIFICATION — INR')
print('='*95)
hdr = '%-26s' % 'Category'
for m in months: hdr += '%13s' % m[-5:]
hdr += '%13s' % 'TOTAL'
print(hdr)
print('-'*95)
for cat in CATS:
    rt = sum(monthly_cat[m].get(cat, 0) for m in months)
    if rt == 0: continue
    row = '%-26s' % cat
    for m in months: row += '%13s' % ('%d' % int(monthly_cat[m].get(cat, 0)))
    row += '%13s' % ('%d' % int(rt))
    print(row)
print('-'*95)
row = '%-26s' % 'TOTAL'
for m in months: row += '%13s' % ('%d' % int(sum(monthly_cat[m].values())))
row += '%13s' % ('%d' % int(sum(sum(monthly_cat[m].values()) for m in months)))
print(row)
