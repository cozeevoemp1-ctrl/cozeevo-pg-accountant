"""
Day-wise / short-stay importer.
Reads 'Daily Basis' from main Excel + 'Day wise' from April Excel.
Deduplicates by SHA-256(name + phone + checkin_date + room).
Writes to daywise_stays table + Google Sheet 'DAY WISE' tab.

Usage:
  python scripts/import_daywise.py              # dry run
  python scripts/import_daywise.py --write      # commit to DB
  python scripts/import_daywise.py --sheet      # update Google Sheet only
  python scripts/import_daywise.py --write --sheet  # both
"""
import sys, os, re, asyncio, hashlib
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date, datetime, timedelta
from decimal import Decimal
from dotenv import load_dotenv
import openpyxl

from sqlalchemy import select  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

load_dotenv()

from src.database.models import DaywiseStay  # noqa: E402

DATABASE_URL = os.environ["DATABASE_URL"]

MAIN_EXCEL = "data/raw/Cozeevo Monthly stay (4).xlsx"
APRIL_EXCEL = "april month.xlsx"


def safe_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    try: return float(s)
    except: pass
    m = re.search(r'([\d,]+)', s)
    if m:
        try: return float(m.group(1).replace(',', ''))
        except: pass
    return 0


def safe_str(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'):
        try: return str(int(float(s)))
        except: pass
    return s


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ('%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: continue
    return None


def make_hash(name, phone, checkin, room):
    raw = f"{name}|{phone}|{checkin}|{room}".lower().strip()
    return hashlib.sha256(raw.encode()).hexdigest()


def read_daywise_sheet(wb, sheet_name, source_file, has_phone=True):
    """Parse a Day wise sheet. Returns list of dicts."""
    ws = wb[sheet_name]
    records = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name or not str(name).strip():
            continue

        room = safe_str(ws.cell(row, 1).value)
        phone = ''
        if has_phone:
            phone = safe_str(ws.cell(row, 3).value)
            phone = re.sub(r'[^\d]', '', phone)

        checkin = parse_date(ws.cell(row, 4).value)
        booking = safe_num(ws.cell(row, 5).value)
        stay_period = safe_str(ws.cell(row, 6).value)
        num_days = safe_num(ws.cell(row, 7).value)
        maintenance = safe_num(ws.cell(row, 8).value)
        daily_rate = safe_num(ws.cell(row, 9).value)
        sharing = safe_num(ws.cell(row, 10).value)
        occupancy = safe_num(ws.cell(row, 11).value)
        paid_date = parse_date(ws.cell(row, 12).value)
        comments = safe_str(ws.cell(row, 13).value)
        staff = safe_str(ws.cell(row, 14).value)
        status = safe_str(ws.cell(row, 15).value) or 'EXIT'

        total = booking if booking > 0 else (num_days * daily_rate if num_days and daily_rate else 0)

        checkout = None
        if checkin and num_days and int(num_days) > 0:
            checkout = checkin + timedelta(days=int(num_days))

        h = make_hash(str(name).strip(), phone, checkin, room)

        records.append({
            'room_number': room,
            'guest_name': str(name).strip(),
            'phone': phone,
            'checkin_date': checkin,
            'checkout_date': checkout,
            'num_days': int(num_days) if num_days else None,
            'stay_period': stay_period,
            'sharing': int(sharing) if sharing else None,
            'occupancy': int(occupancy) if occupancy else None,
            'booking_amount': booking,
            'daily_rate': daily_rate,
            'total_amount': total,
            'maintenance': maintenance,
            'payment_date': paid_date,
            'assigned_staff': staff,
            'status': status.upper() if status else 'EXIT',
            'comments': comments,
            'source_file': source_file,
            'unique_hash': h,
        })

    return records


def read_all_daywise():
    """Read from both Excel files, deduplicate."""
    all_records = []
    seen_hashes = set()

    # 1. Main Excel — "Daily Basis" (no phone column)
    try:
        wb_main = openpyxl.load_workbook(MAIN_EXCEL, data_only=True)
        main_recs = read_daywise_sheet(wb_main, "Daily Basis",
                                        source_file="Cozeevo Monthly stay (4).xlsx",
                                        has_phone=False)
        for r in main_recs:
            if r['unique_hash'] not in seen_hashes:
                seen_hashes.add(r['unique_hash'])
                all_records.append(r)
        print(f"  Main Excel 'Daily Basis': {len(main_recs)} rows ({len(all_records)} after dedup)")
    except Exception as e:
        print(f"  WARNING: Could not read main Excel: {e}")

    # 2. April Excel — "Day wise" (has phone)
    try:
        wb_apr = openpyxl.load_workbook(APRIL_EXCEL, data_only=True)
        apr_recs = read_daywise_sheet(wb_apr, "Day wise",
                                      source_file="april month.xlsx",
                                      has_phone=True)
        added = 0
        for r in apr_recs:
            if r['unique_hash'] not in seen_hashes:
                seen_hashes.add(r['unique_hash'])
                all_records.append(r)
                added += 1
        print(f"  April Excel 'Day wise': {len(apr_recs)} rows ({added} new after dedup)")
    except Exception as e:
        print(f"  WARNING: Could not read April Excel: {e}")

    all_records.sort(key=lambda r: r['checkin_date'] or date.min)
    return all_records


async def import_to_db(records, write=False):
    """Insert into daywise_stays table, skip duplicates."""
    url = DATABASE_URL
    if url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+asyncpg://", 1)
    engine = create_async_engine(url, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    stats = {'inserted': 0, 'skipped_dup': 0, 'skipped_nodate': 0}

    async with Session() as session:
        for rec in records:
            if not rec['checkin_date']:
                stats['skipped_nodate'] += 1
                continue

            existing = await session.scalar(
                select(DaywiseStay).where(DaywiseStay.unique_hash == rec['unique_hash'])
            )
            if existing:
                stats['skipped_dup'] += 1
                continue

            session.add(DaywiseStay(
                room_number=rec['room_number'],
                guest_name=rec['guest_name'],
                phone=rec['phone'] or None,
                checkin_date=rec['checkin_date'],
                checkout_date=rec['checkout_date'],
                num_days=rec['num_days'],
                stay_period=rec['stay_period'],
                sharing=rec['sharing'],
                occupancy=rec['occupancy'],
                booking_amount=Decimal(str(rec['booking_amount'])),
                daily_rate=Decimal(str(rec['daily_rate'])),
                total_amount=Decimal(str(rec['total_amount'])),
                maintenance=Decimal(str(rec['maintenance'])),
                payment_date=rec['payment_date'],
                assigned_staff=rec['assigned_staff'] or None,
                status=rec['status'],
                comments=rec['comments'] or None,
                source_file=rec['source_file'],
                unique_hash=rec['unique_hash'],
            ))
            stats['inserted'] += 1

        if write:
            await session.commit()
            print("  ** COMMITTED to DB **")
        else:
            print("  ** DRY RUN — no changes saved **")

    await engine.dispose()
    return stats


def update_sheet(records):
    """Write DAY WISE tab to Google Sheet."""
    import gspread
    from google.oauth2.service_account import Credentials

    SHEET_ID = "1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw"
    CREDS_FILE = "credentials/gsheets_service_account.json"

    creds = Credentials.from_service_account_file(CREDS_FILE,
        scopes=['https://www.googleapis.com/auth/spreadsheets'])
    gc = gspread.authorize(creds)
    sp = gc.open_by_key(SHEET_ID)

    try:
        ws = sp.worksheet("DAY WISE")
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sp.add_worksheet("DAY WISE", rows=200, cols=16)

    headers = ["Room", "Guest Name", "Phone", "Check-in", "Stay Period",
               "Days", "Daily Rate", "Booking Amt", "Total", "Maintenance",
               "Sharing", "Staff", "Status", "Comments", "Source"]

    rows = []
    total_revenue = 0
    for rec in records:
        checkin_str = rec['checkin_date'].strftime('%Y-%m-%d') if rec['checkin_date'] else ''
        rows.append([
            rec['room_number'], rec['guest_name'], rec['phone'] or '',
            checkin_str, rec['stay_period'],
            rec['num_days'] or '', rec['daily_rate'] or '',
            rec['booking_amount'] or '', rec['total_amount'] or '',
            rec['maintenance'] or '',
            rec['sharing'] or '', rec['assigned_staff'] or '',
            rec['status'], rec['comments'] or '', rec['source_file'],
        ])
        total_revenue += rec['total_amount']

    n = len(rows)
    summary = [
        "DAY WISE STAYS", f"Total: {n} guests",
        f"Revenue: {total_revenue:,.0f}", "", "", "", "", "", "", "", "", "", "", "", "",
    ]

    ws.update(values=[summary, headers] + rows, range_name="A1", value_input_option="USER_ENTERED")
    ws.format("A1:O1", {"textFormat": {"bold": True, "fontSize": 13}})
    ws.format("A2:O2", {"textFormat": {"bold": True},
                         "backgroundColor": {"red": 1.0, "green": 0.95, "blue": 0.8}})
    ws.freeze(rows=2)
    try: ws.set_basic_filter(f"A2:O{2 + n}")
    except: pass
    print(f"  DAY WISE: {n} rows, revenue {total_revenue:,.0f}")

    return n


async def main():
    import argparse
    parser = argparse.ArgumentParser(description="Import day-wise short stays")
    parser.add_argument("--write", action="store_true", help="Commit to DB")
    parser.add_argument("--sheet", action="store_true", help="Update Google Sheet")
    args = parser.parse_args()

    print("=" * 60)
    print("DAY-WISE STAY IMPORT")
    print("=" * 60)

    print("\nStep 1: Reading Excel files...")
    records = read_all_daywise()
    print(f"  Total: {len(records)} unique records")

    print("\nStep 2: Importing to DB...")
    stats = await import_to_db(records, write=args.write)
    print(f"\n  Inserted:       {stats['inserted']}")
    print(f"  Skipped (dup):  {stats['skipped_dup']}")
    print(f"  Skipped (date): {stats['skipped_nodate']}")

    if args.sheet:
        print("\nStep 3: Writing Google Sheet...")
        update_sheet(records)

    print("\nDone!")


if __name__ == "__main__":
    asyncio.run(main())
