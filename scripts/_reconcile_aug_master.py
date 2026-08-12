"""
August-2026 MASTER reconciliation — one row per customer, every source side by side.

Sources
  RENT DUE   rent_schedule for 2026-08-01 (includes proration + deposit on the
             check-in month, per src/services/rent_schedule.py)
  CASH       receipt book  : 'Cash -Kiran.xlsx' -> CASH COLLECTION August-26
             app           : payments, mode=cash
             excel/sheet   : 'google sheet local excel.xlsx' -> 'Aug Cash'
  UPI        QR statements : 'Thor UPI-11-08-2026.csv' + 'Hulk UPI till 1th 08 2026.csv'
             app           : payments, mode=upi
             excel/sheet   : 'Aug UPI'

Window: 1-11 Aug 2026. Every source stops on the 11th (the month is not over),
so no source is being compared against a longer period than another.

Output: data/reports/Aug2026_MASTER_Reconciliation.xlsx

Matching is done ONCE here across all sources so cash and UPI land on the same
customer row. Helpers are imported from the two single-purpose scripts — the
name matcher is never re-implemented.
"""
from __future__ import annotations

import asyncio
import os
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
load_dotenv()

from _reconcile_aug_cash import (load_receipt, loose_similar, money, norm_name,  # noqa: E402
                                 norm_room, room_key, similar)
from _reconcile_aug_upi import ALIASES, NAME_FLOOR, NAME_MARGIN, load_gateway, name_score, phone10  # noqa: E402
from src.database.db_manager import get_session, init_engine  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
GSHEET_XLSX = ROOT / "data/uploads/csv/august 11.xlsx"
OUT = ROOT / "data/reports/Aug2026_MASTER_Reconciliation.xlsx"

WIN_FROM, WIN_TO = date(2026, 8, 1), date(2026, 8, 31)


class Cust:
    __slots__ = ("room", "names", "phones", "status", "stay_type", "due",
                 "cash_book", "cash_app", "cash_sheet",
                 "upi_qr", "upi_app", "upi_sheet", "notes", "pre_aug")

    def __init__(self, room, name, phone=""):
        self.room = room
        self.names = [name] if name else []
        self.phones = {phone} if phone else set()
        self.status = ""
        self.stay_type = ""
        self.due = 0.0
        self.cash_book = self.cash_app = self.cash_sheet = 0.0
        self.upi_qr = self.upi_app = self.upi_sheet = 0.0
        self.pre_aug = 0.0        # Aug rent the app booked on 21-31 July
        self.notes: list[str] = []

    @property
    def display(self):
        return max(self.names, key=len) if self.names else "(unknown)"

    @property
    def paid_app(self):
        return self.cash_app + self.upi_app

    @property
    def paid_external(self):
        """What the two independent outside records say arrived."""
        return self.cash_book + self.upi_qr

    def note(self, m):
        if m not in self.notes:
            self.notes.append(m)


def make_register():
    customers: list[Cust] = []
    by_phone: dict[str, Cust] = {}

    def get(room, name, phone="") -> Cust:
        room = room or ""
        if phone and phone in by_phone:
            c = by_phone[phone]
            if name and name not in c.names:
                c.names.append(name)
            if room and not c.room:
                c.room = room
            return c
        for c in customers:
            if c.room == room and name and any(similar(n, name) for n in c.names):
                if phone:
                    c.phones.add(phone)
                    by_phone[phone] = c
                return c
        c = Cust(room, name, phone)
        customers.append(c)
        if phone:
            by_phone[phone] = c
        return c

    return customers, by_phone, get


async def load_db():
    init_engine(os.environ["DATABASE_URL"])
    async with get_session() as s:
        reg = (await s.execute(text("""
            SELECT tc.id, r.room_number, t.name, t.phone, tc.status::text AS status,
                   tc.stay_type::text AS stay_type,
                   COALESCE(rs.rent_due,0) + COALESCE(rs.maintenance_due,0)
                     + COALESCE(rs.adjustment,0) AS due
            FROM tenancies tc
            JOIN tenants t ON t.id = tc.tenant_id
            LEFT JOIN rooms r ON r.id = tc.room_id
            LEFT JOIN rent_schedule rs
                   ON rs.tenancy_id = tc.id AND rs.period_month = DATE '2026-08-01'
            WHERE tc.checkin_date < DATE '2026-09-01'
              AND (tc.checkout_date IS NULL OR tc.checkout_date >= DATE '2026-08-01')
        """))).mappings().all()

        pays = (await s.execute(text("""
            SELECT r.room_number, t.name, t.phone, p.payment_date, p.amount,
                   p.payment_mode::text AS mode, p.for_type::text AS for_type
            FROM payments p
            JOIN tenancies tc ON tc.id = p.tenancy_id
            JOIN tenants   t  ON t.id  = tc.tenant_id
            LEFT JOIN rooms r ON r.id  = tc.room_id
            WHERE p.is_void = false
              AND p.payment_date >= :d1 AND p.payment_date <= :d2
        """), {"d1": WIN_FROM, "d2": WIN_TO})).mappings().all()

        day = (await s.execute(text("""
            SELECT room_number, guest_name, phone, total_amount, checkin_date
            FROM daywise_stays
            WHERE checkin_date < DATE '2026-09-01'
              AND (checkout_date IS NULL OR checkout_date >= DATE '2026-08-01')
        """))).mappings().all()

        # August rent collected in the last days of July. The receipt book's
        # first six rows are exactly this, so without it they look "missing".
        late_july = (await s.execute(text("""
            SELECT r.room_number, t.name, t.phone, p.amount, p.payment_date,
                   p.payment_mode::text AS mode
            FROM payments p
            JOIN tenancies tc ON tc.id = p.tenancy_id
            JOIN tenants   t  ON t.id  = tc.tenant_id
            LEFT JOIN rooms r ON r.id  = tc.room_id
            WHERE p.is_void = false
              AND p.payment_date >= DATE '2026-07-21' AND p.payment_date < DATE '2026-08-01'
              AND p.period_month = DATE '2026-08-01'
        """))).mappings().all()
    return reg, pays, day, late_july


def load_sheet():
    ws = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)["Sheet1"]
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            h.setdefault(str(v).strip(), c)
    out = []
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, h["Name"]).value
        if not nm:
            continue
        out.append({
            "room": norm_room(ws.cell(r, h["Room No"]).value),
            "name": norm_name(nm),
            "phone": phone10(ws.cell(r, h["Mobile Number"]).value),
            "cash": money(ws.cell(r, h["Aug Cash"]).value),
            "upi": money(ws.cell(r, h["Aug UPI"]).value),
            "status": ws.cell(r, h["Checkin/out"]).value or "",
        })
    return out


async def build():
    reg, pays, day, late_july = await load_db()
    sheet = load_sheet()
    receipt = load_receipt()
    gateway = [g for g in load_gateway() if WIN_FROM <= g["date"] <= WIN_TO]

    customers, by_phone, get = make_register()

    # 1. DB register + August rent due.
    #    A person can hold several tenancy rows (re-booking, cancelled
    #    pre-registration into placeholder room 000, no-show). They collapse
    #    onto one customer via phone, so the room and status shown must come
    #    from the LIVE tenancy — otherwise a cancelled room-000 row hides the
    #    real room and the receipt book can never be matched to them.
    rank = {"active": 0, "exited": 1, "no_show": 2, "cancelled": 3}

    def better(new_status, new_room, cur_status, cur_room) -> bool:
        if cur_room in ("", "000") and new_room not in ("", "000"):
            return True
        if new_room in ("", "000") and cur_room not in ("", "000"):
            return False
        return rank.get(new_status, 9) < rank.get(cur_status, 9)

    for m in reg:
        room = norm_room(m["room_number"])
        c = get(room, norm_name(m["name"]), phone10(m["phone"]))
        if not c.status or better(m["status"], room, c.status, c.room):
            c.status = m["status"]
            c.room = room
        c.stay_type = m["stay_type"] or "monthly"
        # only the live tenancy's rent schedule counts as this month's due
        if m["status"] in ("active", "exited"):
            c.due += float(m["due"] or 0)

    # 2. day-stay guests (separate table, no tenancy row)
    for m in day:
        c = get(norm_room(m["room_number"]), norm_name(m["guest_name"]), phone10(m["phone"]))
        c.stay_type = "daily"
        c.due += float(m["total_amount"] or 0)
        c.note("day-stay guest")

    # 3. app payments
    for m in pays:
        c = get(norm_room(m["room_number"]), norm_name(m["name"]), phone10(m["phone"]))
        if m["mode"] == "cash":
            c.cash_app += float(m["amount"])
        elif m["mode"] == "upi":
            c.upi_app += float(m["amount"])
        else:
            c.note(f"app payment with no mode: Rs {float(m['amount']):,.0f}")

    # 3b. August rent the app recorded in late July
    for m in late_july:
        c = get(norm_room(m["room_number"]), norm_name(m["name"]), phone10(m["phone"]))
        c.pre_aug += float(m["amount"])

    # 4. excel / google sheet
    for x in sheet:
        c = get(x["room"], x["name"], x["phone"])
        c.cash_sheet += x["cash"]
        c.upi_sheet += x["upi"]
        if not c.status and x["status"]:
            c.status = str(x["status"])

    # 5. receipt book (cash) — room number is handwritten, so fall back to name
    for rec in receipt:
        parts = rec["parts"] if len(rec["parts"]) > 1 else [rec["name"]]
        matched = []
        for p in parts:
            hit = [c for c in customers if c.room == rec["room"] and any(similar(n, p) for n in c.names)]
            if not hit:
                hit = [c for c in customers if any(similar(n, p) for n in c.names)]
                if len(hit) > 1:
                    live = [c for c in hit if c.cash_app or c.cash_sheet]
                    hit = live if len(live) == 1 else []
            if len(hit) == 1:
                matched.append(hit[0])
        if not matched:
            c = get(rec["room"] or "", rec["name"])
            c.cash_book += rec["amount"]
            c.note("receipt book only - no matching customer")
            continue
        weights = [max(c.cash_app, c.cash_sheet, 1.0) for c in matched]
        tw = sum(weights)
        for c, w in zip(matched, weights):
            c.cash_book += round(rec["amount"] * w / tw, 2)
            if len(matched) > 1:
                c.note(f"joint receipt '{rec['raw_name']}' Rs {int(rec['amount']):,} split")

    # 6. UPI QR statements
    for x in gateway:
        c = None
        room = ALIASES.get(x["name"])
        if room:
            cand = [z for z in customers if z.room == room and (z.upi_app or z.upi_sheet)] \
                or [z for z in customers if z.room == room]
            if cand:
                c = max(cand, key=lambda z: z.upi_app + z.upi_sheet)
        if c is None and x["phone"] and x["phone"] in by_phone:
            c = by_phone[x["phone"]]
        if c is None and x["name"]:
            scored = sorted(((max((name_score(n, x["name"]) for n in z.names), default=0.0), z)
                             for z in customers if z.names), key=lambda z: -z[0])
            if scored and scored[0][0] >= NAME_FLOOR:
                runner = scored[1][0] if len(scored) > 1 else 0.0
                if scored[0][0] - runner >= NAME_MARGIN:
                    c = scored[0][1]
                else:
                    tie = [z for sc, z in scored if sc >= scored[0][0] - NAME_MARGIN]
                    if len({z.room for z in tie}) == 1:
                        c = max(tie, key=lambda z: z.upi_app + z.upi_sheet)
                    else:
                        live = [z for z in tie if z.upi_app or z.upi_sheet]
                        c = live[0] if len(live) == 1 else None
        if c is None:
            c = get("", x["name"])
            c.note(f"UPI from '{x['payer']}' - no matching customer")
        c.upi_qr += x["amount"]

    # 7. consolidation — rows created purely from the receipt book or a QR
    #    statement (no tenancy, no rent due) are usually the same person under a
    #    different spelling or a mis-written room. Merge them into the real
    #    customer where the link is unambiguous; leave genuine third-party
    #    payers standing alone.
    real = [c for c in customers if c.status or c.due]
    orphans = [c for c in customers if not (c.status or c.due)]
    for o in orphans:
        cands = [c for c in real if loose_similar(o.display, c.display)]
        if len(cands) > 1 and o.room:
            same = [c for c in cands if c.room == o.room]
            if same:
                cands = same
        if len(cands) > 1:
            amt = o.cash_book or o.upi_qr
            exact = [c for c in cands
                     if round(c.cash_app) == round(amt) or round(c.cash_sheet) == round(amt)
                     or round(c.upi_app) == round(amt) or round(c.upi_sheet) == round(amt)]
            if len(exact) == 1:
                cands = exact
            else:
                # several people share the amount — let the NAME decide.
                # 'Darshit' -> 'Darshit Khakhkhar' beats every other 13,500.
                pool = exact or cands
                ranked = sorted(((name_score(o.display, c.display), c) for c in pool),
                                key=lambda z: -z[0])
                if ranked and ranked[0][0] >= NAME_FLOOR and (
                        len(ranked) == 1 or ranked[0][0] - ranked[1][0] >= NAME_MARGIN):
                    cands = [ranked[0][1]]
                else:
                    cands = []
        if len(cands) != 1:
            continue
        t = cands[0]
        t.cash_book += o.cash_book
        t.upi_qr += o.upi_qr
        t.cash_app += o.cash_app
        t.upi_app += o.upi_app
        t.cash_sheet += o.cash_sheet
        t.upi_sheet += o.upi_sheet
        t.pre_aug += o.pre_aug
        t.note(f"'{o.display.title()}'"
               + (f" room {o.room}" if o.room else " (no room)")
               + " merged - same person, different spelling")
        customers.remove(o)

    return customers


# ------------------------------------------------------------------ output

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = PatternFill("solid", fgColor="1F3864")
CASHF = PatternFill("solid", fgColor="2E7D32")
UPIF = PatternFill("solid", fgColor="0277BD")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BAND = PatternFill("solid", fgColor="DDEBF7")
INR = '#,##0;[Red]-#,##0;"-"'

COLS = ["Room", "Customer", "Type", "Status", "Rent Due Aug",
        "Cash: Receipt Book", "Cash: App", "Cash: Excel",
        "UPI: QR Stmt", "UPI: App", "UPI: Excel",
        "Total in App", "Outside Records", "App - Outside",
        "Balance vs Due", "Issue", "Notes"]


def issues(c: Cust) -> str:
    out = []
    if round(c.cash_book) != round(c.cash_app):
        out.append("cash")
    if round(c.upi_qr) != round(c.upi_app):
        out.append("upi")
    if not out:
        return ""
    gap = c.paid_app - c.paid_external
    if gap < -1:
        return "MONEY MISSING IN APP"
    if gap > 1:
        return "APP HAS EXTRA"
    return "MODE/SPLIT DIFF"


def write(customers):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Aug 2026"

    def has_money(c):
        return (c.cash_book or c.cash_app or c.cash_sheet
                or c.upi_qr or c.upi_app or c.upi_sheet)

    def is_live(c):
        """Active monthly tenants and day-stay guests only.

        59 tenancies are marked 'exited' with a NULL checkout_date, so a
        date-based filter lets them through and bills them a full month
        (Suraj Prasana, room 106, Rs 29,000). Status is the reliable test.
        """
        st = str(c.status).strip().lower().replace(" ", "_")
        return st == "active" or c.stay_type == "daily" or st == "checkin"

    rows = [c for c in customers if is_live(c) and (c.due or has_money(c))]
    rows.sort(key=lambda c: (room_key(c.room), c.display))

    # everyone else who nonetheless has money against them — kept visible
    others = [c for c in customers if not is_live(c) and has_money(c)]
    others.sort(key=lambda c: -(c.cash_book + c.upi_qr + c.cash_app + c.upi_app))

    ws.append(["AUGUST 2026 — MASTER RECONCILIATION (1–11 Aug, every source)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Outside Records = Receipt Book cash + QR statement UPI, i.e. the two records that do not come from the app"])
    ws["A2"].font = Font(italic=True, color="808080")
    ws.append(COLS)
    for i in range(1, len(COLS) + 1):
        cell = ws.cell(3, i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = CASHF if 6 <= i <= 8 else UPIF if 9 <= i <= 11 else NAVY
    ws.row_dimensions[3].height = 30

    for c in rows:
        iss = issues(c)
        ws.append([c.room, c.display.title(), c.stay_type or "monthly", c.status,
                   c.due or None,
                   c.cash_book or None, c.cash_app or None, c.cash_sheet or None,
                   c.upi_qr or None, c.upi_app or None, c.upi_sheet or None,
                   c.paid_app or None, c.paid_external or None,
                   c.paid_app - c.paid_external,
                   c.paid_app - c.due, iss, "; ".join(c.notes)])
        r = ws.max_row
        for i in range(1, len(COLS) + 1):
            ws.cell(r, i).border = BORDER
        for i in list(range(5, 16)):
            ws.cell(r, i).number_format = INR
        if iss == "MONEY MISSING IN APP":
            for i in range(1, len(COLS) + 1):
                ws.cell(r, i).fill = RED
        elif iss:
            ws.cell(r, 16).fill = AMBER

    last = ws.max_row
    ws.append([])
    ws.append(["", "TOTAL", "", ""] + [f"=SUM({get_column_letter(i)}4:{get_column_letter(i)}{last})"
                                       for i in range(5, 16)] + ["", ""])
    tr = ws.max_row
    for i in range(1, len(COLS) + 1):
        ws.cell(tr, i).font = Font(bold=True)
        ws.cell(tr, i).fill = BAND
        ws.cell(tr, i).border = BORDER
    for i in range(5, 16):
        ws.cell(tr, i).number_format = INR

    widths = [8, 28, 9, 10, 12, 13, 11, 11, 12, 11, 11, 12, 13, 13, 13, 21, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:Q{last}"

    # ---------------------------------------------------- Tab 2: action list
    a = wb.create_sheet("Action List")
    a.append(["MONEY THAT ARRIVED BUT IS NOT IN THE APP — work this list"])
    a["A1"].font = Font(bold=True, size=14)
    a.append(["Outside records = receipt book (cash) + QR statements (UPI). Both are independent of the app."])
    a["A2"].font = Font(italic=True, color="808080")
    acols = ["Room", "Customer", "Status", "Type", "Outside Says", "App Has",
             "Gap", "Excel Says", "Verdict", "What to do"]
    a.append(acols)
    for i in range(1, len(acols) + 1):
        cell = a.cell(3, i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    items = []
    for c in rows + others:
        for kind, outside, inapp, excel in (
            ("CASH", c.cash_book, c.cash_app, c.cash_sheet),
            ("UPI", c.upi_qr, c.upi_app, c.upi_sheet),
        ):
            gap = outside - inapp
            if gap <= 1:
                continue
            # order matters: a no-show who paid is a different problem from a
            # mis-written amount, and they can look alike (small booking
            # advance in the app vs a full rent outside).
            if c.pre_aug and abs(c.pre_aug - gap) < 1:
                verdict = "OK - booked in late July"
                todo = "No action. The app dated it 21-31 Jul."
            elif c.status in ("no_show", "cancelled") or str(c.status).upper() in ("NO SHOW", "CANCELLED"):
                verdict = "PAID BUT MARKED NO-SHOW"
                todo = "They did move in. Convert the tenancy to active, then log the payment."
            elif not c.room:
                verdict = "UNKNOWN PAYER"
                todo = "Identify who paid, then log it against their room."
            elif kind == "CASH" and inapp > 0 and outside >= inapp * 5:
                verdict = "CHECK RECEIPT BOOK"
                todo = "Written amount looks wrong (extra digit). Verify against the book."
            else:
                verdict = "MISSING - LOG IT"
                todo = "Money arrived and is not in the app. Enter the payment."
            items.append((c, kind, outside, inapp, gap, excel, verdict, todo))

    items.sort(key=lambda z: (z[6].startswith("OK"), -z[4]))
    for c, kind, outside, inapp, gap, excel, verdict, todo in items:
        a.append([c.room, c.display.title(), c.status, kind, outside, inapp or None,
                  gap, excel or None, verdict, todo])
        r = a.max_row
        for i in range(1, len(acols) + 1):
            a.cell(r, i).border = BORDER
        for i in (5, 6, 7, 8):
            a.cell(r, i).number_format = INR
        if verdict.startswith("OK"):
            a.cell(r, 9).fill = PatternFill("solid", fgColor="C6EFCE")
        elif verdict.startswith("CHECK") or verdict.startswith("UNKNOWN"):
            a.cell(r, 9).fill = AMBER
        else:
            a.cell(r, 9).fill = RED
            a.cell(r, 9).font = Font(bold=True)

    alast = a.max_row
    a.append([])
    a.append(["", "TOTAL", "", "", f"=SUM(E4:E{alast})", f"=SUM(F4:F{alast})",
              f"=SUM(G4:G{alast})", "", "", ""])
    for i in range(1, len(acols) + 1):
        a.cell(a.max_row, i).font = Font(bold=True)
        a.cell(a.max_row, i).fill = BAND
    for i in (5, 6, 7):
        a.cell(a.max_row, i).number_format = INR
    for i, w in enumerate([8, 28, 11, 7, 13, 12, 12, 12, 26, 46], start=1):
        a.column_dimensions[get_column_letter(i)].width = w
    a.freeze_panes = "A4"
    a.auto_filter.ref = f"A3:J{alast}"

    # ------------------------------------------- Tab 3: not active, but paid
    o = wb.create_sheet("Not Active But Paid")
    o.append(["EXITED / NO-SHOW / CANCELLED CUSTOMERS WITH MONEY AGAINST THEM"])
    o["A1"].font = Font(bold=True, size=14)
    o.append(["Excluded from the main table on purpose. Money here still needs explaining."])
    o["A2"].font = Font(italic=True, color="808080")
    ocols = ["Room", "Customer", "Status", "Cash: Receipt", "Cash: App", "Cash: Excel",
             "UPI: QR Stmt", "UPI: App", "UPI: Excel", "Total in App", "Outside Records", "Notes"]
    o.append(ocols)
    for i in range(1, len(ocols) + 1):
        cell = o.cell(3, i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in others:
        o.append([c.room, c.display.title(), c.status,
                  c.cash_book or None, c.cash_app or None, c.cash_sheet or None,
                  c.upi_qr or None, c.upi_app or None, c.upi_sheet or None,
                  c.paid_app or None, c.paid_external or None, "; ".join(c.notes)])
        r = o.max_row
        for i in range(1, len(ocols) + 1):
            o.cell(r, i).border = BORDER
        for i in range(4, 12):
            o.cell(r, i).number_format = INR
        if c.paid_external > c.paid_app + 1:
            for i in range(1, 4):
                o.cell(r, i).fill = RED
    for i, w in enumerate([8, 28, 11, 13, 11, 12, 12, 11, 12, 12, 14, 40], start=1):
        o.column_dimensions[get_column_letter(i)].width = w
    o.freeze_panes = "C4"
    if o.max_row > 3:
        o.auto_filter.ref = f"A3:L{o.max_row}"

    wb.save(OUT)
    return rows, others


async def main():
    customers = await build()
    rows, others = write(customers)

    def tot(f):
        return sum(f(c) for c in rows)

    print(f"active + day-stay customers: {len(rows)}")
    print(f"excluded (exited/no-show/cancelled) but holding money: {len(others)}")
    print(f"{'rent due Aug':<22} {tot(lambda c: c.due):>12,.0f}")
    print(f"{'cash  receipt book':<22} {tot(lambda c: c.cash_book):>12,.0f}")
    print(f"{'cash  app':<22} {tot(lambda c: c.cash_app):>12,.0f}")
    print(f"{'cash  excel':<22} {tot(lambda c: c.cash_sheet):>12,.0f}")
    print(f"{'upi   qr statement':<22} {tot(lambda c: c.upi_qr):>12,.0f}")
    print(f"{'upi   app':<22} {tot(lambda c: c.upi_app):>12,.0f}")
    print(f"{'upi   excel':<22} {tot(lambda c: c.upi_sheet):>12,.0f}")
    print()
    counts = defaultdict(lambda: [0, 0.0])
    for c in rows:
        i = issues(c)
        if i:
            counts[i][0] += 1
            counts[i][1] += c.paid_app - c.paid_external
    for k, (n, v) in sorted(counts.items(), key=lambda z: -z[1][0]):
        print(f"  {k:<24} {n:>4} customers   net Rs {v:>12,.0f}")
    print(f"\nwrote {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
