"""
Dues Detail Export — all tenants with pending dues (from DB)
Columns: Building, Name, Room, Check-in Date, May Check-in?, Stay Type,
         Agreed Rent, Security Deposit, Booking Amount Paid,
         Total Charged, Total Paid, Outstanding Dues, Months Pending

Output: data/reports/Dues_Detail_2026_05.xlsx
"""
import asyncio, os, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from datetime import date
from dotenv import load_dotenv
load_dotenv()

import asyncpg, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── styles ────────────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill('solid', fgColor='1F3864')
HDR_FONT    = Font(bold=True, color='FFFFFF', size=10)
MAY_FILL    = PatternFill('solid', fgColor='E2EFDA')   # green = May check-in
OLD_FILL    = PatternFill('solid', fgColor='FCE4D6')   # red = no payment at all
PART_FILL   = PatternFill('solid', fgColor='FFF2CC')   # yellow = partial
TOTAL_FILL  = PatternFill('solid', fgColor='BDD7EE')
TOTAL_FONT  = Font(bold=True, size=10)
NORM_FONT   = Font(size=10)
thin        = Side(style='thin', color='BBBBBB')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT       = Alignment(horizontal='right', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
NUM_FMT     = '#,##0'
DATE_FMT    = 'DD-MMM-YYYY'

COL_HEADERS = [
    '#', 'Building', 'Name', 'Room', 'Phone',
    'Check-in Date', 'May Check-in?', 'Stay Type',
    'Agreed Rent', 'Security Deposit', 'Booking Paid',
    'May Rent Due', 'Rent Paid (May)', 'Dep Still Owed', 'Outstanding', 'Pending Months',
]
COL_WIDTHS = [4, 9, 28, 6, 14, 14, 13, 10, 12, 16, 13, 13, 14, 14, 12, 32]
NUM_COLS   = {9, 10, 11, 12, 13, 14, 15}  # 1-indexed money cols


async def main():
    db_url = os.getenv('DATABASE_URL').replace('postgresql+asyncpg', 'postgresql')
    conn = await asyncpg.connect(db_url)

    # Match KPI logic exactly: current month rent dues + unpaid deposit
    # period = May 2026
    rows = await conn.fetch("""
        WITH
        period AS (SELECT DATE '2026-05-01' AS dt),
        next_p AS (SELECT DATE '2026-06-01' AS dt),

        -- May rent paid per tenancy
        rent_paid AS (
            SELECT p.tenancy_id, SUM(p.amount) AS paid
            FROM payments p, period
            WHERE p.is_void = FALSE
              AND p.for_type = 'rent'
              AND p.period_month = period.dt
            GROUP BY p.tenancy_id
        ),
        -- Deposit paid THIS calendar month (offsets rent_due on new check-ins)
        dep_paid_period AS (
            SELECT p.tenancy_id, SUM(p.amount) AS paid
            FROM payments p, period, next_p
            WHERE p.is_void = FALSE
              AND p.for_type = 'deposit'
              AND p.payment_date >= period.dt
              AND p.payment_date < next_p.dt
            GROUP BY p.tenancy_id
        ),
        -- Total deposit ever paid per tenancy
        dep_paid_total AS (
            SELECT p.tenancy_id, SUM(p.amount) AS paid
            FROM payments p
            WHERE p.is_void = FALSE
              AND p.for_type = 'deposit'
            GROUP BY p.tenancy_id
        ),
        -- Also grab booking payments ever made
        booking_paid_total AS (
            SELECT p.tenancy_id, SUM(p.amount) AS paid
            FROM payments p
            WHERE p.is_void = FALSE
              AND p.for_type = 'booking'
            GROUP BY p.tenancy_id
        )

        SELECT
            t.name,
            t.phone,
            r.room_number,
            pr.name                          AS building,
            tn.id                            AS tenancy_id,
            tn.checkin_date,
            tn.stay_type,
            tn.agreed_rent,
            tn.security_deposit,
            tn.booking_amount                AS booking_amount_agreed,
            COALESCE(bk.paid, 0)             AS booking_paid,

            -- May rent schedule
            rs.rent_due,
            COALESCE(rs.adjustment, 0)       AS adjustment,
            rs.rent_due + COALESCE(rs.adjustment, 0) AS effective_due,

            -- Payments
            COALESCE(rp.paid, 0)             AS rent_paid,
            COALESCE(dp.paid, 0)             AS dep_paid_period,
            COALESCE(dt.paid, 0)             AS dep_paid_total,

            -- Rent dues for May
            GREATEST(0,
                rs.rent_due + COALESCE(rs.adjustment, 0)
                - COALESCE(rp.paid, 0)
                - COALESCE(dp.paid, 0)
            )                                AS rent_dues,

            -- Deposit still owed (deposit agreed - all deposit paid - booking_amount)
            GREATEST(0,
                tn.security_deposit
                - COALESCE(dt.paid, 0)
                - tn.booking_amount
            )                                AS dep_due,

            -- Total outstanding = rent_dues + dep_due
            GREATEST(0,
                rs.rent_due + COALESCE(rs.adjustment, 0)
                - COALESCE(rp.paid, 0)
                - COALESCE(dp.paid, 0)
            )
            +
            GREATEST(0,
                tn.security_deposit
                - COALESCE(dt.paid, 0)
                - tn.booking_amount
            )                                AS outstanding,

            -- Pending months (all time, for context)
            COALESCE((
                SELECT STRING_AGG(TO_CHAR(rs2.period_month, 'Mon YYYY'), ', '
                                  ORDER BY rs2.period_month)
                FROM rent_schedule rs2
                WHERE rs2.tenancy_id = tn.id
                  AND rs2.status IN ('pending', 'partial')
            ), '')                           AS months_pending

        FROM tenancies tn
        JOIN tenants    t  ON t.id  = tn.tenant_id
        JOIN rooms      r  ON r.id  = tn.room_id
        JOIN properties pr ON pr.id = r.property_id
        JOIN rent_schedule rs ON rs.tenancy_id = tn.id
        CROSS JOIN period
        LEFT JOIN rent_paid          rp ON rp.tenancy_id = tn.id
        LEFT JOIN dep_paid_period    dp ON dp.tenancy_id = tn.id
        LEFT JOIN dep_paid_total     dt ON dt.tenancy_id = tn.id
        LEFT JOIN booking_paid_total bk ON bk.tenancy_id = tn.id
        WHERE tn.status = 'active'
          AND r.is_staff_room = FALSE
          AND tn.stay_type != 'daily'
          AND rs.period_month = period.dt
        ORDER BY pr.name, outstanding DESC, t.name
    """)

    await conn.close()

    # ── build output rows ─────────────────────────────────────────────────────
    data = []
    for row in rows:
        outstanding = float(row['outstanding'])
        if outstanding <= 0:
            continue
        may_checkin = (row['checkin_date'] and
                       row['checkin_date'].month == 5 and
                       row['checkin_date'].year == 2026)
        building = (row['building'] or '').upper()
        if 'HULK' in building:
            bname = 'HULK'
        elif 'THOR' in building:
            bname = 'THOR'
        else:
            bname = building or '?'

        rent_paid   = float(row['rent_paid'] or 0)
        dep_period  = float(row['dep_paid_period'] or 0)
        total_paid  = rent_paid + dep_period

        data.append({
            'building':     bname,
            'name':         row['name'],
            'room':         row['room_number'],
            'phone':        row['phone'] or '',
            'checkin':      row['checkin_date'],
            'may_checkin':  'YES' if may_checkin else '',
            'stay_type':    row['stay_type'] or 'monthly',
            'agreed_rent':  float(row['agreed_rent'] or 0),
            'security_dep': float(row['security_deposit'] or 0),
            'booking_paid': float(row['booking_paid'] or 0),
            'effective_due': float(row['effective_due'] or 0),
            'rent_dues':    float(row['rent_dues'] or 0),
            'dep_due':      float(row['dep_due'] or 0),
            'total_paid':   total_paid,
            'outstanding':  outstanding,
            'months':       row['months_pending'] or '',
            'may_flag':     may_checkin,
        })

    print(f"Total tenants with outstanding dues: {len(data)}")
    print(f"  May check-ins in list: {sum(1 for d in data if d['may_flag'])}")
    total_due = sum(d['outstanding'] for d in data)
    print(f"  Total outstanding: Rs.{total_due:,.0f}")

    # ── write Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for bfilter in ['ALL', 'HULK', 'THOR']:
        rows_out = data if bfilter == 'ALL' else [d for d in data if d['building'] == bfilter]
        if not rows_out:
            continue

        ws = wb.create_sheet(bfilter)

        # Title row
        ws.merge_cells(f'A1:{get_column_letter(len(COL_HEADERS))}1')
        title = f'{bfilter} — Dues Detail — May 2026  (generated {date.today().strftime("%d %b %Y")})'
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=13, color='1F3864')
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Header row
        for ci, h in enumerate(COL_HEADERS, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[2].height = 22

        # Data rows
        tot_due = tot_paid = tot_dep_due = tot_out = 0
        for i, d in enumerate(rows_out, 1):
            row_n = i + 2
            if d['may_flag']:
                fill = MAY_FILL
            elif d['total_paid'] == 0:
                fill = OLD_FILL
            else:
                fill = PART_FILL

            vals = [
                i,
                d['building'],
                d['name'],
                d['room'],
                d['phone'],
                d['checkin'],
                d['may_checkin'],
                d['stay_type'],
                d['agreed_rent'],
                d['security_dep'],
                d['booking_paid'],
                d['effective_due'],
                d['total_paid'],
                d['dep_due'],
                d['outstanding'],
                d['months'],
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=ci, value=v)
                cell.fill = fill; cell.border = BORDER; cell.font = NORM_FONT
                if ci in NUM_COLS:
                    cell.alignment = RIGHT; cell.number_format = NUM_FMT
                elif ci == 6:   # check-in date
                    cell.alignment = CENTER; cell.number_format = DATE_FMT
                elif ci == 7:   # May check-in flag
                    cell.alignment = CENTER
                    if d['may_flag']:
                        cell.font = Font(bold=True, color='375623', size=10)
                else:
                    cell.alignment = LEFT

            tot_due     += d['effective_due']
            tot_paid    += d['total_paid']
            tot_dep_due += d['dep_due']
            tot_out     += d['outstanding']

        # Totals row
        tot_row = len(rows_out) + 3
        totals = [None, '', 'TOTAL', '', '', None, '', '', None, None, None,
                  tot_due, tot_paid, tot_dep_due, tot_out, '']
        for ci, v in enumerate(totals, 1):
            cell = ws.cell(row=tot_row, column=ci, value=v)
            cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER
            if ci in NUM_COLS and isinstance(v, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

        # Column widths
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A3'

    os.makedirs('data/reports', exist_ok=True)
    out_path = 'data/reports/Dues_Detail_2026_05.xlsx'
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # ── console summary ───────────────────────────────────────────────────────
    print("\nTop 10 by outstanding:")
    for d in sorted(data, key=lambda x: -x['outstanding'])[:10]:
        flag = ' [MAY]' if d['may_flag'] else ''
        print(f"  {d['name']:<30}  Room {d['room']:>5}  Rs.{d['outstanding']:>8,.0f}{flag}")


asyncio.run(main())
