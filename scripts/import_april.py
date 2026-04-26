"""
April payment drop-and-reload importer.
Reads 'april month.xlsx' — the TRUTH for April 2026.

What it does:
  1. DROPS all existing April rent_schedule + payments (period_month = 2026-04-01)
  2. Reloads from Excel: rent_schedule + payments from cols 22 (cash), 23 (UPI), 24 (balance)
  3. Only loads pure numeric values as payments; text extracted as notes
  4. Creates missing tenants/tenancies if not in DB
  5. Updates tenancy.notes with permanent comments, rent_schedule.notes with monthly
  6. Updates Google Sheet: APRIL 2026 tab

Usage:
  python scripts/import_april.py              # dry run
  python scripts/import_april.py --write      # commit to DB
  python scripts/import_april.py --sheet      # update Google Sheet only
  python scripts/import_april.py --write --sheet  # both
"""
import sys, os, re, asyncio
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date, datetime
from decimal import Decimal
from dotenv import load_dotenv
import openpyxl

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

load_dotenv()

from src.database.models import (
    Tenant, Tenancy, RentSchedule, Payment, Room, Property,
    RentStatus, PaymentMode, PaymentFor, TenancyStatus, StayType,
)

DATABASE_URL = os.environ["DATABASE_URL"]

APRIL_FILE = "April Month Collection.xlsx"
APRIL_PERIOD = date(2026, 4, 1)

# ── Column mapping ─────────────────────────────────────────────────────
COL_ROOM = 1
COL_NAME = 2
COL_GENDER = 3
COL_PHONE = 4
COL_CHECKIN = 5
COL_BOOKING = 6
COL_DEPOSIT = 7
COL_MAINTENANCE = 8
COL_RENT_MONTHLY = 10
COL_RENT_FEB = 11
COL_RENT_MAY = 12
COL_SHARING = 13
COL_COMMENT = 15
COL_STAFF = 16
COL_INOUT = 17
COL_BLOCK = 18
COL_FLOOR = 19
COL_APR_STATUS = 20
COL_MAR_BALANCE = 21
COL_APR_CASH = 22
COL_APR_UPI = 23
COL_APR_BALANCE = 24
COL_FOOD = 25
COL_COMPLAINTS = 26
COL_VACATION = 27


# ── Comment classification ─────────────────────────────────────────────
_PERMANENT_SIGNALS = re.compile(
    r"(?:always|agreed|checkout|planned|lease|contract|first.*months|from.*month|"
    r"company|student|parent|no due|refund|exit|lock.?in|deposit|registration|"
    r"switch|moved|bed|room|luggage|vacation|key|GST|installment)",
    re.I,
)
_MONTHLY_SIGNALS = re.compile(
    r"(?:will pay|by \d+|next week|balance|partial|collected|pending|"
    r"april|apr|sent to|received by|chandra|no response|march|half)",
    re.I,
)


def classify_comment(comment):
    if not comment or not comment.strip():
        return "permanent", "", ""
    c = comment.strip()
    is_perm = bool(_PERMANENT_SIGNALS.search(c))
    is_month = bool(_MONTHLY_SIGNALS.search(c))
    if is_perm and is_month:
        return "both", c, c
    if is_month:
        return "monthly", "", c
    return "permanent", c, ""


def _rent_status_to_enum(st):
    s = str(st).strip().upper() if st else ''
    if not s:
        return None
    if s == 'PAID':
        return RentStatus.paid
    if 'PARTIAL' in s:
        return RentStatus.partial
    if 'EXIT' in s or 'EXITED' in s:
        return RentStatus.exit
    if 'NO SHOW' in s or 'CANCEL' in s:
        return RentStatus.na
    return RentStatus.pending


def extract_numeric_and_text(v):
    """Extract pure numeric value and any text from a cell.
    Returns (number, text_note).
    Only returns number if the cell is purely numeric (int/float).
    Text cells return 0 + the text as a note."""
    if v is None:
        return 0, ''
    if isinstance(v, (int, float)):
        return float(v), ''
    s = str(v).strip()
    try:
        return float(s), ''
    except ValueError:
        pass
    # It's text — return 0 and the text as a note
    return 0, s


def safe_str(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'):
        try: return str(int(float(s)))
        except: pass
    return s


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ('%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: continue
    return None


def _norm_phone(raw):
    if not raw: return ""
    digits = re.sub(r'\D', '', raw)
    if digits.startswith('91') and len(digits) == 12:
        digits = digits[2:]
    if len(digits) == 10 and digits[0] in '6789':
        return f"+91{digits}"
    return ""


def read_april(excel_file=APRIL_FILE):
    """Parse 'April month' sheet. Returns list of dicts for ALL rows."""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['Long term']
    records = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, COL_NAME).value
        if not name or not str(name).strip():
            continue

        inout = safe_str(ws.cell(row, COL_INOUT).value).upper()
        phone_raw = safe_str(ws.cell(row, COL_PHONE).value)
        phone_db = _norm_phone(phone_raw)  # +91XXXXXXXXXX format

        apr_status_raw = safe_str(ws.cell(row, COL_APR_STATUS).value)
        apr_status = _rent_status_to_enum(apr_status_raw)

        rm = float(ws.cell(row, COL_RENT_MONTHLY).value or 0) if isinstance(ws.cell(row, COL_RENT_MONTHLY).value, (int, float)) else 0
        rf = float(ws.cell(row, COL_RENT_FEB).value or 0) if isinstance(ws.cell(row, COL_RENT_FEB).value, (int, float)) else 0
        ry = float(ws.cell(row, COL_RENT_MAY).value or 0) if isinstance(ws.cell(row, COL_RENT_MAY).value, (int, float)) else 0
        april_rent = ry if ry > 0 else (rf if rf > 0 else rm)

        # Extract numeric + text from payment columns
        apr_cash, cash_text = extract_numeric_and_text(ws.cell(row, COL_APR_CASH).value)
        apr_upi, upi_text = extract_numeric_and_text(ws.cell(row, COL_APR_UPI).value)
        apr_bal_num, bal_text = extract_numeric_and_text(ws.cell(row, COL_APR_BALANCE).value)

        comment = safe_str(ws.cell(row, COL_COMMENT).value)
        mar_balance = safe_str(ws.cell(row, COL_MAR_BALANCE).value)

        ctype, perm_note, month_note = classify_comment(comment)

        # Build monthly notes from all text sources
        monthly_parts = []
        if month_note:
            monthly_parts.append(month_note)
        if cash_text:
            monthly_parts.append(f"Cash: {cash_text}")
        if upi_text:
            monthly_parts.append(f"UPI: {upi_text}")
        if bal_text:
            monthly_parts.append(f"Bal: {bal_text}")
        if mar_balance:
            monthly_parts.append(f"Mar bal: {mar_balance}")

        records.append({
            'name': str(name).strip(),
            'phone_raw': phone_raw,
            'phone_db': phone_db,
            'gender': safe_str(ws.cell(row, COL_GENDER).value),
            'room': safe_str(ws.cell(row, COL_ROOM).value),
            'block': safe_str(ws.cell(row, COL_BLOCK).value),
            'floor': safe_str(ws.cell(row, COL_FLOOR).value),
            'sharing': safe_str(ws.cell(row, COL_SHARING).value),
            'checkin_raw': safe_str(ws.cell(row, COL_CHECKIN).value),
            'checkin': parse_date(ws.cell(row, COL_CHECKIN).value),
            'deposit': float(ws.cell(row, COL_DEPOSIT).value or 0) if isinstance(ws.cell(row, COL_DEPOSIT).value, (int, float)) else 0,
            'booking': float(ws.cell(row, COL_BOOKING).value or 0) if isinstance(ws.cell(row, COL_BOOKING).value, (int, float)) else 0,
            'maintenance': float(ws.cell(row, COL_MAINTENANCE).value or 0) if isinstance(ws.cell(row, COL_MAINTENANCE).value, (int, float)) else 0,
            'food': safe_str(ws.cell(row, COL_FOOD).value),
            'complaints': safe_str(ws.cell(row, COL_COMPLAINTS).value),
            'vacation': safe_str(ws.cell(row, COL_VACATION).value),
            'staff': safe_str(ws.cell(row, COL_STAFF).value),
            'inout': inout,
            'apr_status': apr_status,
            'apr_status_raw': apr_status_raw,
            'april_rent': april_rent,
            'current_rent': april_rent,
            'apr_cash': apr_cash,
            'apr_upi': apr_upi,
            'apr_bal_num': apr_bal_num,  # numeric from balance col — treat as UPI payment
            'comment': comment,
            'permanent_note': perm_note if ctype in ('permanent', 'both') else comment,
            'monthly_note': " | ".join(monthly_parts) if monthly_parts else '',
            'mar_balance': mar_balance,
        })

    return records


def _clean_status(inout):
    s = inout.upper() if inout else ''
    if s in ('CHECKIN', 'CHECK IN', ''): return TenancyStatus.active
    if 'EXIT' in s: return TenancyStatus.exited
    if 'NO SHOW' in s: return TenancyStatus.no_show
    if 'CANCEL' in s: return TenancyStatus.cancelled
    return TenancyStatus.active


async def import_to_db(records, write=False):
    """Drop April data, reload from Excel. Create missing tenants."""
    url = DATABASE_URL
    if url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+asyncpg://", 1)
    engine = create_async_engine(url, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    stats = {'matched': 0, 'created_tenant': 0, 'created_tenancy': 0,
             'skipped': 0, 'rent_schedule': 0, 'payments': 0,
             'dropped_rs': 0, 'dropped_pay': 0, 'updated_notes': 0,
             'no_match': []}

    async with Session() as session:
        session.autoflush = False

        # ── Step 1: DROP existing April rent_schedule + payments ─────
        from sqlalchemy import delete as sa_delete, func
        rs_count = await session.scalar(
            select(func.count()).select_from(RentSchedule).where(RentSchedule.period_month == APRIL_PERIOD)
        )
        pay_count = await session.scalar(
            select(func.count()).select_from(Payment).where(
                Payment.period_month == APRIL_PERIOD,
                Payment.for_type == PaymentFor.rent,
            )
        )
        stats['dropped_rs'] = rs_count or 0
        stats['dropped_pay'] = pay_count or 0

        if not write:
            print(f"  Would drop: {stats['dropped_rs']} rent_schedule + {stats['dropped_pay']} payments for April")
        else:
            await session.execute(
                sa_delete(Payment).where(
                    Payment.period_month == APRIL_PERIOD,
                    Payment.for_type == PaymentFor.rent,
                )
            )
            await session.execute(
                sa_delete(RentSchedule).where(RentSchedule.period_month == APRIL_PERIOD)
            )
            await session.flush()
            print(f"  Dropped: {stats['dropped_rs']} rent_schedule + {stats['dropped_pay']} payments for April")

        # Pre-load room lookup for creating tenancies
        rooms_result = await session.execute(select(Room))
        room_map = {}
        for r in rooms_result.scalars().all():
            room_map[r.room_number] = r

        # ── Step 2: Process EVERY row ────────────────────────────────
        seen_tenancy_ids = set()

        for rec in records:
            is_exit = 'EXIT' in rec['inout']
            is_cancelled = 'CANCEL' in rec['inout']
            has_payment = rec['apr_cash'] > 0 or rec['apr_upi'] > 0
            gets_rent_schedule = not is_exit and not is_cancelled

            # EXIT/CANCELLED with no payment: skip
            if (is_exit or is_cancelled) and not has_payment:
                continue

            phone_db = rec['phone_db']
            if not phone_db:
                stats['skipped'] += 1
                stats['no_match'].append(f"{rec['name']} — no valid phone")
                continue

            # ── Find or create tenant (match by phone+name to handle shared phones) ─
            norm_name = rec['name'].strip().title()
            tenant = await session.scalar(
                select(Tenant).where(Tenant.phone == phone_db, Tenant.name == norm_name)
            )
            if not tenant:
                # Fallback: phone-only match, but ONLY if name is similar (not a different person)
                phone_match = await session.scalar(
                    select(Tenant).where(Tenant.phone == phone_db)
                )
                if phone_match and phone_match.name.lower().split()[0] == norm_name.lower().split()[0]:
                    tenant = phone_match  # minor name variation (e.g. casing)
                # else: different person sharing same phone — will create new tenant below

            if not tenant:
                # Try matching by placeholder phone (from prior imports)
                safe = re.sub(r'[^a-zA-Z0-9]', '', rec['name'])[:12]
                tenant = await session.scalar(
                    select(Tenant).where(Tenant.phone.like(f"NOPHONE_{rec['room']}_%"),
                                         Tenant.name == norm_name)
                )
                if not tenant:
                    # Also try name-only match in same room
                    from sqlalchemy import and_
                    tenant = await session.scalar(
                        select(Tenant).where(Tenant.name == norm_name)
                        .join(Tenancy, Tenancy.tenant_id == Tenant.id)
                        .join(Room, Room.id == Tenancy.room_id)
                        .where(Room.room_number == rec['room'])
                    )

            if not tenant and write:
                # Create new tenant — use placeholder phone if phone already taken
                existing_phone = await session.scalar(
                    select(Tenant).where(Tenant.phone == phone_db)
                )
                actual_phone = phone_db
                if existing_phone:
                    # Shared phone — use placeholder
                    safe = re.sub(r'[^a-zA-Z0-9]', '', rec['name'])[:12]
                    actual_phone = f"NOPHONE_{rec['room']}_{safe}"
                    # Check placeholder also not taken
                    existing_placeholder = await session.scalar(
                        select(Tenant).where(Tenant.phone == actual_phone)
                    )
                    if existing_placeholder:
                        actual_phone = f"NOPHONE_{rec['room']}_{safe}_{int(datetime.now().timestamp()) % 10000}"
                tenant = Tenant(
                    name=norm_name,
                    phone=actual_phone,
                    gender='female' if str(rec['gender']).strip().lower() == 'female' else 'male',
                    notes=rec['permanent_note'] or None,
                )
                session.add(tenant)
                await session.flush()
                stats['created_tenant'] += 1
                print(f"    Created tenant: {rec['name']} ({actual_phone})")
            elif not tenant:
                stats['no_match'].append(f"{rec['name']} ({phone_db}) — not in DB, will create with --write")
                # Still count payments for dry-run totals
                if rec['apr_cash'] > 0: stats['payments'] += 1
                if rec['apr_upi'] > 0: stats['payments'] += 1
                # apr_bal_num is outstanding dues, not payment
                if gets_rent_schedule and rec['apr_status']:
                    stats['rent_schedule'] += 1
                continue

            # ── Find or create tenancy ───────────────────────────────
            tenancy = await session.scalar(
                select(Tenancy).where(
                    Tenancy.tenant_id == tenant.id,
                    Tenancy.status == TenancyStatus.active,
                )
            )
            if not tenancy:
                # Try most recent tenancy (for EXIT/NO SHOW)
                tenancy = await session.scalar(
                    select(Tenancy).where(
                        Tenancy.tenant_id == tenant.id,
                    ).order_by(Tenancy.checkin_date.desc())
                )

            if not tenancy and write:
                # Create tenancy — find room
                room_str = re.sub(r'\.0$', '', str(rec['room']))
                room_obj = room_map.get(room_str)
                if not room_obj:
                    # Try without leading zeros, with G prefix, etc.
                    for k, v in room_map.items():
                        if k.lstrip('0') == room_str.lstrip('0'):
                            room_obj = v
                            break

                if not room_obj:
                    stats['no_match'].append(f"{rec['name']} — room {rec['room']} not in DB")
                    continue

                tenancy = Tenancy(
                    tenant_id=tenant.id,
                    room_id=room_obj.id,
                    stay_type=StayType.monthly,
                    status=_clean_status(rec['inout']),
                    checkin_date=rec['checkin'] or APRIL_PERIOD,
                    agreed_rent=Decimal(str(rec['april_rent'])),
                    security_deposit=Decimal(str(rec['deposit'])),
                    booking_amount=Decimal(str(rec['booking'])),
                    maintenance_fee=Decimal(str(rec['maintenance'])),
                    notes=rec['permanent_note'] or None,
                    entered_by="excel_import",
                )
                session.add(tenancy)
                await session.flush()
                stats['created_tenancy'] += 1
                print(f"    Created tenancy: {rec['name']} room {rec['room']}")
            elif not tenancy:
                stats['no_match'].append(f"{rec['name']} ({phone_db}) — no tenancy, will create with --write")
                if rec['apr_cash'] > 0: stats['payments'] += 1
                if rec['apr_upi'] > 0: stats['payments'] += 1
                # apr_bal_num is outstanding dues, not payment
                if gets_rent_schedule and rec['apr_status']:
                    stats['rent_schedule'] += 1
                continue

            stats['matched'] += 1

            # ── Update tenancy status + sharing_type from Excel ─────
            new_status = _clean_status(rec['inout'])
            if tenancy.status != new_status:
                tenancy.status = new_status
            # Update sharing_type
            sharing_raw = str(rec.get('sharing', '') or '').lower()
            from src.database.models import SharingType as _ST
            new_sharing = None
            if 'prem' in sharing_raw: new_sharing = _ST.premium
            elif 'single' in sharing_raw: new_sharing = _ST.single
            elif 'double' in sharing_raw: new_sharing = _ST.double
            elif 'triple' in sharing_raw: new_sharing = _ST.triple
            if new_sharing and tenancy.sharing_type != new_sharing:
                tenancy.sharing_type = new_sharing

            # ── Dry run: just count ──────────────────────────────────
            if not write:
                if gets_rent_schedule and rec['apr_status']:
                    stats['rent_schedule'] += 1
                if rec['apr_cash'] > 0: stats['payments'] += 1
                if rec['apr_upi'] > 0: stats['payments'] += 1
                # apr_bal_num is outstanding dues, not payment
                if gets_rent_schedule and rec['permanent_note'] and rec['permanent_note'] != (tenancy.notes or ''):
                    stats['updated_notes'] += 1
                continue

            # ── Rent schedule (CHECKIN, NO SHOW, blank — not EXIT/CANCELLED) ─
            if gets_rent_schedule and rec['apr_status'] and tenancy.id not in seen_tenancy_ids:
                seen_tenancy_ids.add(tenancy.id)
                session.add(RentSchedule(
                    tenancy_id=tenancy.id,
                    period_month=APRIL_PERIOD,
                    rent_due=Decimal(str(rec['april_rent'])),
                    maintenance_due=Decimal("0"),
                    status=rec['apr_status'],
                    due_date=APRIL_PERIOD,
                    notes=rec['monthly_note'] or None,
                ))
                stats['rent_schedule'] += 1

            # ── Cash payment ─────────────────────────────────────────
            if rec['apr_cash'] > 0:
                session.add(Payment(
                    tenancy_id=tenancy.id,
                    amount=Decimal(str(rec['apr_cash'])),
                    payment_date=APRIL_PERIOD,
                    payment_mode=PaymentMode.cash,
                    for_type=PaymentFor.rent,
                    period_month=APRIL_PERIOD,
                    notes=f"April Excel [{rec['inout'] or 'CHECKIN'}]",
                ))
                stats['payments'] += 1

            # ── UPI payment ──────────────────────────────────────────
            if rec['apr_upi'] > 0:
                session.add(Payment(
                    tenancy_id=tenancy.id,
                    amount=Decimal(str(rec['apr_upi'])),
                    payment_date=APRIL_PERIOD,
                    payment_mode=PaymentMode.upi,
                    for_type=PaymentFor.rent,
                    period_month=APRIL_PERIOD,
                    notes=f"April Excel [{rec['inout'] or 'CHECKIN'}]",
                ))
                stats['payments'] += 1

            # ── Balance column numeric → NOT a payment, it's outstanding dues
            # Stored as adjustment on rent_schedule (already in monthly_note)

            # ── Update tenancy notes ─────────────────────────────────
            if gets_rent_schedule:
                new_notes = rec['permanent_note']
                if new_notes and new_notes != (tenancy.notes or ''):
                    tenancy.notes = new_notes
                    stats['updated_notes'] += 1

        if write:
            await session.commit()
            print("  ** COMMITTED to DB **")
        else:
            print("  ** DRY RUN — no changes saved **")

    await engine.dispose()
    return stats


def sync_tenants_master(records, write=False):
    """Add new tenants from import to TENANTS master tab.
    Compares Excel names against existing TENANTS rows.
    Only adds CHECKIN/NO SHOW — not EXIT/CANCELLED."""
    import gspread
    from google.oauth2.service_account import Credentials

    SHEET_ID = "1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw"
    CREDS_FILE = "credentials/gsheets_service_account.json"

    creds = Credentials.from_service_account_file(CREDS_FILE,
        scopes=['https://www.googleapis.com/auth/spreadsheets'])
    gc = gspread.authorize(creds)
    sp = gc.open_by_key(SHEET_ID)
    ws = sp.worksheet("TENANTS")

    data = ws.get_all_values()
    existing = set()
    for row in data[1:]:
        if row[1]:
            existing.add(str(row[1]).strip().lower())

    # TENANTS columns: Room,Name,Phone,Gender,Building,Floor,Sharing,Check-in,Status,Agreed Rent,Deposit,Booking,Maintenance,Notice Date,Expected Exit
    new_rows = []
    for rec in records:
        name = rec['name'].strip()
        if name.lower() in existing:
            continue
        if 'EXIT' in rec['inout'] or 'CANCEL' in rec['inout']:
            continue

        status = 'Active'
        if 'NO SHOW' in rec['inout']:
            status = 'No-show'

        new_rows.append([
            rec['room'],
            name,
            rec['phone_raw'],
            rec['gender'],
            rec['block'],
            rec['floor'],
            rec['sharing'],
            rec['checkin_raw'],
            status,
            rec['current_rent'],
            rec['deposit'],
            rec['booking'],
            rec['maintenance'],
            '',  # Notice Date
            '',  # Expected Exit
        ])
        existing.add(name.lower())  # prevent duplicates within same import

    if new_rows and write:
        ws.append_rows(new_rows, value_input_option='USER_ENTERED')

    if new_rows:
        for r in new_rows:
            print(f"    + {r[0]:5} {r[1]:25} {r[6]:10} {r[8]}")

    return len(new_rows)


def update_sheet(records):
    """Update APRIL 2026 tab on Google Sheet (new 17-col format matching gsheets.py).

    Columns: Room, Name, Phone, Building, Sharing, Rent Due,
             Cash, UPI, Total Paid, Balance, Status,
             Check-in, Notice Date, Event, Notes, Prev Due, Entered By
    """
    import gspread
    from google.oauth2.service_account import Credentials

    SHEET_ID = "1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw"
    CREDS_FILE = "credentials/gsheets_service_account.json"

    creds = Credentials.from_service_account_file(CREDS_FILE,
        scopes=['https://www.googleapis.com/auth/spreadsheets'])
    gc = gspread.authorize(creds)
    sp = gc.open_by_key(SHEET_ID)

    try:
        ws = sp.worksheet("APRIL 2026")
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sp.add_worksheet("APRIL 2026", rows=300, cols=18)

    headers = ["Room", "Name", "Phone", "Building", "Sharing", "Rent Due",
               "Cash", "UPI", "Total Paid", "Balance", "Status",
               "Check-in", "Notice Date", "Event", "Notes", "Prev Due",
               "Entered By"]

    # Include ALL rows from Excel — exits, cancelled, no-shows, everything
    rows = []
    for rec in records:
        is_exit = 'EXIT' in rec['inout']
        is_cancelled = 'CANCEL' in rec['inout']

        cash = rec['apr_cash']
        upi = rec['apr_upi']
        tp = cash + upi
        rent = rec['current_rent']
        prev_due = 0
        try:
            prev_due = float(rec['mar_balance']) if rec['mar_balance'] else 0
        except (ValueError, TypeError):
            pass
        bal = rent + prev_due - tp

        # Use Excel status directly
        status_raw = rec['apr_status_raw'].upper().strip() if rec['apr_status_raw'] else ''
        if is_exit:
            status = "EXIT"
        elif is_cancelled:
            status = "CANCELLED"
        elif 'NO SHOW' in rec['inout']:
            status = "NO SHOW"
        elif 'PAID' in status_raw and 'UNPAID' not in status_raw and 'NOT' not in status_raw:
            status = "PAID"
        elif 'PARTIAL' in status_raw:
            status = "PARTIAL"
        else:
            status = "UNPAID"

        rows.append([
            rec['room'], rec['name'], rec['phone_raw'],
            rec['block'], rec['sharing'], rent,
            cash, upi, tp, bal, status,
            rec['checkin_raw'], "",  # Notice Date — empty for import
            rec['inout'], rec['comment'],
            prev_due, "Excel Import",
        ])

    n = len(rows)

    # Build summary from the rows we just prepared (using Excel statuses, not recalculated)
    TOTAL_BEDS = 291
    cash_total = sum(r[6] for r in rows)
    upi_total = sum(r[7] for r in rows)
    collected = cash_total + upi_total
    balance_total = sum(r[9] for r in rows if r[10] not in ("EXIT", "NO SHOW"))

    beds = 0
    regular = 0
    premium = 0
    noshow_count = 0
    paid = partial = unpaid = new_checkins = exits = 0
    thor_beds = thor_t = hulk_beds = hulk_t = 0

    for r in rows:
        # r: [room, name, phone, building, sharing, rent, cash, upi, tp, bal, status, ...]
        status = r[10]
        building = str(r[3]).upper()
        sharing = str(r[4]).lower()
        event = str(r[13]).upper() if len(r) > 13 else ""

        if status == "PAID": paid += 1
        elif status == "PARTIAL": partial += 1
        elif status == "UNPAID": unpaid += 1

        if status == "EXIT":
            exits += 1
            continue
        if status == "NO SHOW" or "NO SHOW" in event:
            noshow_count += 1
            continue

        bed_count = 2 if sharing == "premium" else 1
        beds += bed_count
        if sharing == "premium":
            premium += 1
        else:
            regular += 1
        if building == "THOR":
            thor_beds += bed_count
            thor_t += 1
        else:
            hulk_beds += bed_count
            hulk_t += 1

    vacant = TOTAL_BEDS - beds - noshow_count
    occ_pct = f"{beds / TOTAL_BEDS * 100:.1f}" if TOTAL_BEDS > 0 else "0"

    row2 = [
        "Checked-in", f"{beds} beds ({regular}+{premium}P)",
        f"No-show: {noshow_count}", f"Vacant: {vacant}", f"Occ: {occ_pct}%",
        "Cash", cash_total, "UPI", upi_total, "Total", collected,
        f"Bal: {int(balance_total)}", "", "", "", "", "",
    ]
    row3 = [
        f"THOR: {thor_beds}b ({thor_t}t)", f"HULK: {hulk_beds}b ({hulk_t}t)",
        f"New: {new_checkins}", f"Exit: {exits}", "",
        f"PAID:{paid}", f"PARTIAL:{partial}", f"UNPAID:{unpaid}",
        "", "", "", "", "", "", "", "", "",
    ]

    summary = [
        ["APRIL 2026"] + [""] * 16,
        row2,
        row3,
        headers,
    ]

    ws.update(values=summary + rows, range_name="A1", value_input_option="USER_ENTERED")
    ws.format("A1:Q1", {"textFormat": {"bold": True, "fontSize": 13}})
    ws.format("A2:Q3", {"textFormat": {"bold": True, "fontSize": 9},
                         "numberFormat": {"type": "NUMBER", "pattern": "#,##0"}})
    ws.format("A4:Q4", {"textFormat": {"bold": True},
                         "backgroundColor": {"red": 0.85, "green": 0.9, "blue": 1.0}})
    ws.freeze(rows=4)
    try: ws.set_basic_filter(f"A4:Q{4 + n}")
    except: pass

    print(f"  APRIL 2026: {n} rows written to Sheet (17-col format)")
    return n


async def main():
    import argparse
    parser = argparse.ArgumentParser(description="Import April payment data (drop + reload)")
    parser.add_argument("--write", action="store_true", help="Commit to DB (default: dry run)")
    parser.add_argument("--sheet", action="store_true", help="Update Google Sheet")
    parser.add_argument("--file", default=APRIL_FILE, help="Excel file path")
    args = parser.parse_args()

    print("=" * 60)
    print("APRIL PAYMENT IMPORT (DROP + RELOAD)")
    print("=" * 60)

    print("\nStep 1: Reading April Excel...")
    records = read_april(args.file)
    checkin = sum(1 for r in records if r['inout'] in ('CHECKIN', 'CHECK IN'))
    exits = sum(1 for r in records if 'EXIT' in r['inout'])
    noshow = sum(1 for r in records if 'NO SHOW' in r['inout'])
    print(f"  {len(records)} rows: {checkin} CHECKIN, {exits} EXIT, {noshow} NO SHOW")

    # Show target totals
    t_cash = sum(r['apr_cash'] for r in records)
    t_upi = sum(r['apr_upi'] for r in records)
    t_bal = sum(r['apr_bal_num'] for r in records)
    print(f"  Target: Cash={t_cash:,.0f} + UPI={t_upi:,.0f} + Bal={t_bal:,.0f} = {t_cash+t_upi+t_bal:,.0f}")

    print("\nStep 2: Importing to DB...")
    stats = await import_to_db(records, write=args.write)

    print(f"\n  Dropped RS:        {stats['dropped_rs']}")
    print(f"  Dropped payments:  {stats['dropped_pay']}")
    print(f"  Matched:           {stats['matched']}")
    print(f"  Created tenants:   {stats['created_tenant']}")
    print(f"  Created tenancies: {stats['created_tenancy']}")
    print(f"  Rent schedules:    {stats['rent_schedule']}")
    print(f"  Payments added:    {stats['payments']}")
    print(f"  Notes updated:     {stats['updated_notes']}")
    print(f"  Skipped:           {stats['skipped']}")

    if stats['no_match']:
        print(f"\n  ISSUES ({len(stats['no_match'])}):")
        for nm in stats['no_match']:
            print(f"    - {nm}")

    if args.sheet:
        print("\nStep 3: Updating Google Sheet (APRIL 2026)...")
        update_sheet(records)

        print("\nStep 4: Syncing new tenants to TENANTS master tab...")
        new_count = sync_tenants_master(records, write=args.write)
        print(f"  {new_count} new tenants added to TENANTS tab.")

    print("\nDone!")


if __name__ == "__main__":
    asyncio.run(main())
