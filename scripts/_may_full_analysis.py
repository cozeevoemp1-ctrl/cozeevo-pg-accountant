"""
May 2026 — Full outstanding dues analysis.

Sources:
  THOR UPI (April):  thor bank statement april til now.xlsx  (April 1-30)
  THOR UPI (May):    thor bank statement april til now.xlsx  (May 1-9)
  HULK UPI (May):    Hulk may 11th bank statement.xlsx
  Cash (both):       MAY 2026 Google Sheet  (receptionist record)
  April carry-over:  MAY 2026 Google Sheet  Prev Due column  (receptionist)
  Tenant details:    MAY 2026 + APRIL 2026 Google Sheet

Logic per tenant:
  May dues  = rent (or prorated+deposit if joined May)
  April carry-over = Prev Due from May Sheet
  Total due = May dues + April carry-over
  UPI paid  = bank match (THOR Apr file / HULK May file), fallback to Sheet UPI
  Cash paid = Sheet Cash column
  Balance   = Total due - UPI paid - Cash paid

Output:
  data/reports/May_Full_Analysis_2026.xlsx
    Sheet "ALL"   — all outstanding (balance > 0), sorted by balance desc
    Sheet "HULK"  — HULK only
    Sheet "THOR"  — THOR only
    Sheet "UNMATCHED" — tenants with balance > 0 and ZERO bank match found
                         (neither phone nor name found in bank statement)
"""
import os, csv, re, openpyxl
from datetime import datetime, date
import gspread
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ── helpers ───────────────────────────────────────────────────────────────────
def extract_phone_from_vpa(vpa):
    m = re.match(r'^(\d{10})(?:-\d+)?@', str(vpa))
    return m.group(1) if m else None

def normalize_phone(ph):
    if not ph: return None
    ph = re.sub(r'\D', '', str(ph))
    if ph.startswith('91') and len(ph) == 12: ph = ph[2:]
    return ph if len(ph) == 10 else None

def name_tokens(s):
    return [p for p in re.sub(r'[^A-Z0-9 ]', ' ', s.upper()).split() if len(p) > 1]

def safe_float(v):
    try: return float(str(v).replace(',', '').replace('Rs.', '').strip() or 0)
    except: return 0

# ── load bank transactions ────────────────────────────────────────────────────
def load_thor_bank():
    """Load THOR bank statement (April 1 – May 9). Returns (april_txns, may_txns)."""
    wb = openpyxl.load_workbook('thor bank statement april til now.xlsx')
    ws = wb.worksheets[0]
    apr, may = [], []
    seen_rrn = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[11] != 'SUCCESS' or not row[3]: continue
        rrn = row[0]
        if rrn in seen_rrn: continue
        seen_rrn.add(rrn)
        txn_date = row[1].date() if isinstance(row[1], datetime) else row[1]
        vpa  = str(row[7]) if row[7] else ''
        name = str(row[8]).strip().upper() if row[8] else ''
        entry = {'amt': float(row[3]), 'phone': extract_phone_from_vpa(vpa),
                 'vpa': vpa, 'name': name, 'date': txn_date, 'used': False}
        if txn_date and txn_date <= date(2026, 4, 30):
            apr.append(entry)
        else:
            may.append(entry)
    return apr, may

def load_hulk_may_bank():
    wb = openpyxl.load_workbook('Hulk may 11th bank statement.xlsx')
    seen_rrn = set()
    entries = []
    for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
        if row[11] != 'SUCCESS' or not row[3]: continue
        rrn = row[0]
        if rrn in seen_rrn: continue
        seen_rrn.add(rrn)
        vpa  = str(row[7]) if row[7] else ''
        name = str(row[8]).strip().upper() if row[8] else ''
        entries.append({'amt': float(row[3]), 'phone': extract_phone_from_vpa(vpa),
                        'vpa': vpa, 'name': name, 'used': False})
    return entries

print('Loading bank statements...')
thor_apr_txns, thor_may_txns = load_thor_bank()
hulk_may_txns = load_hulk_may_bank()
print(f'  THOR April txns: {len(thor_apr_txns)}, total Rs.{sum(t["amt"] for t in thor_apr_txns):,.0f}')
print(f'  THOR May txns:   {len(thor_may_txns)}, total Rs.{sum(t["amt"] for t in thor_may_txns):,.0f}')
print(f'  HULK May txns:   {len(hulk_may_txns)}, total Rs.{sum(t["amt"] for t in hulk_may_txns):,.0f}')

# ── match function ────────────────────────────────────────────────────────────
def find_total_in_bank(tenant_name, tenant_phone, txns):
    """Sum ALL bank transactions matching this tenant (phone or name). Marks used."""
    tphone = normalize_phone(tenant_phone)
    name_up = tenant_name.strip().upper()
    tparts  = name_tokens(name_up)
    total   = 0
    matched_names = []

    # 1. Phone match — take ALL with this phone
    if tphone:
        for t in txns:
            if not t['used'] and normalize_phone(t['phone']) == tphone:
                total += t['amt']; t['used'] = True
                matched_names.append(t['name'])

    if total > 0:
        return total, ', '.join(set(matched_names))

    # 2. Exact name match
    for t in txns:
        if not t['used'] and t['name'] == name_up:
            t['used'] = True
            return t['amt'], t['name']

    # 3. Fuzzy name (≥2 word hits, score ≥0.6)
    if len(tparts) >= 2:
        best, bi = 0, None
        for i, t in enumerate(txns):
            if t['used']: continue
            bparts = name_tokens(t['name'])
            hits   = sum(1 for tp in tparts if any(tp in bp or bp in tp for bp in bparts))
            score  = hits / len(tparts)
            if hits >= 2 and score >= 0.6 and score > best:
                best, bi = score, i
        if bi is not None:
            txns[bi]['used'] = True
            return txns[bi]['amt'], txns[bi]['name']

    return 0, ''

# ── load Google Sheet ─────────────────────────────────────────────────────────
print('Loading Google Sheet...')
creds = Credentials.from_service_account_file(
    'credentials/gsheets_service_account.json',
    scopes=['https://www.googleapis.com/auth/spreadsheets']
)
gc = gspread.authorize(creds)
sh = gc.open_by_key('1Hp5dTM7TcDEq75jgHEjvwtBjOolruGfQ7CVMzVqjdGw')

def read_sheet(ws_name):
    ws = sh.worksheet(ws_name)
    rows = ws.get_all_values()
    H = {h: i for i, h in enumerate(rows[7])}
    out = []
    for row in rows[8:]:
        if len(row) <= H.get('UPI', 9): continue
        room = row[H['Room']].strip(); name = row[H['Name']].strip()
        if not room or not name: continue
        rent_due = safe_float(row[H['Rent Due']])
        if rent_due <= 0: continue
        building_raw = row[H['Building']].strip() if 'Building' in H else ''
        bname = 'HULK' if 'HULK' in building_raw.upper() else 'THOR'
        prev_due = safe_float(row[H['Prev Due']]) if 'Prev Due' in H and len(row) > H['Prev Due'] else 0
        rent = safe_float(row[H['Rent']]) if 'Rent' in H and len(row) > H['Rent'] else 0
        checkin_raw = row[H['Check-in']].strip() if 'Check-in' in H and len(row) > H['Check-in'] else ''
        checkin = None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try: checkin = datetime.strptime(checkin_raw, fmt).date(); break
            except: pass
        deposit_raw = row[H['Deposit']].strip() if 'Deposit' in H and len(row) > H['Deposit'] else '0'
        out.append({
            'room': room, 'name': name, 'phone': row[H['Phone']].strip(),
            'building': bname, 'rent': rent, 'deposit': safe_float(deposit_raw),
            'rent_due': rent_due, 'prev_due': prev_due,
            'sheet_cash': safe_float(row[H['Cash']]),
            'sheet_upi':  safe_float(row[H['UPI']]),
            'sheet_bal':  safe_float(row[H['Balance']]),
            'status':     row[H['Status']].strip(),
            'checkin':    checkin,
        })
    return out

may_tenants = read_sheet('MAY 2026')
print(f'  May Sheet tenants with dues: {len(may_tenants)}')

# ── match each tenant against bank ───────────────────────────────────────────
results = []
for t in may_tenants:
    is_thor = t['building'] == 'THOR'

    # May UPI: THOR → thor_may_txns, HULK → hulk_may_txns
    may_bank_pool = thor_may_txns if is_thor else hulk_may_txns
    may_bank_upi, may_bank_name = find_total_in_bank(t['name'], t['phone'], may_bank_pool)

    # April UPI (for THOR only — we have the bank data)
    apr_bank_upi = 0
    apr_bank_name = ''
    if is_thor:
        apr_bank_upi, apr_bank_name = find_total_in_bank(t['name'], t['phone'], thor_apr_txns)

    # Effective UPI = max(sheet, bank)
    eff_upi = max(t['sheet_upi'], may_bank_upi)
    upi_corrected = may_bank_upi > t['sheet_upi'] + 100

    # Total paid = effective UPI + cash
    total_paid = eff_upi + t['sheet_cash']

    # Balance = rent_due + prev_due - total_paid
    balance = t['rent_due'] + t['prev_due'] - total_paid

    # Was April carry-over verified by bank? (THOR only)
    apr_note = ''
    if is_thor and apr_bank_upi > 0:
        apr_note = f'Apr bank: {apr_bank_upi:,.0f}'

    results.append({
        'room': t['room'], 'name': t['name'], 'phone': t['phone'],
        'building': t['building'], 'rent': t['rent'], 'deposit': t['deposit'],
        'checkin': t['checkin'],
        'rent_due': t['rent_due'], 'prev_due': t['prev_due'],
        'cash': t['sheet_cash'], 'upi': eff_upi,
        'total_paid': total_paid, 'balance': balance,
        'sheet_upi': t['sheet_upi'], 'may_bank_upi': may_bank_upi,
        'apr_bank_upi': apr_bank_upi,
        'upi_corrected': upi_corrected,
        'bank_matched_name': may_bank_name,
        'apr_bank_name': apr_bank_name,
        'apr_note': apr_note,
        'no_bank_match': may_bank_upi == 0,
    })

outstanding = [r for r in results if r['balance'] > 0]
outstanding.sort(key=lambda x: -x['balance'])

print(f'\nOutstanding: {len(outstanding)} tenants, total Rs.{sum(r["balance"] for r in outstanding):,.0f}')
upi_fixed = [r for r in outstanding if r['upi_corrected']]
no_match  = [r for r in outstanding if r['no_bank_match'] and r['cash'] == 0]
print(f'  UPI corrected from bank (not in Sheet): {len(upi_fixed)}')
print(f'  No bank match + no cash at all: {len(no_match)}')

# ── Excel styles ──────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F3864')
HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
BAL_FILL   = PatternFill('solid', fgColor='FFF2CC')   # yellow — has some payment
NOPAY_FILL = PatternFill('solid', fgColor='FCE4D6')   # red — zero payment
CORR_FILL  = PatternFill('solid', fgColor='E2EFDA')   # green — UPI corrected from bank
TOTAL_FILL = PatternFill('solid', fgColor='BDD7EE')
TOTAL_FONT = Font(bold=True, size=10)
NORM_FONT  = Font(size=10)
thin       = Side(style='thin', color='BBBBBB')
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT      = Alignment(horizontal='right',  vertical='center')
LEFT       = Alignment(horizontal='left',   vertical='center')

COLS = ['#', 'Building', 'Name', 'Room', 'Check-in', 'Rent', 'Deposit',
        'May Due', 'Apr Carry', 'Total Due', 'Cash', 'UPI', 'Total Paid', 'Balance', 'Status']
WIDTHS = [4, 8, 28, 6, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 22]
NUM_COLS = {6, 7, 8, 9, 10, 11, 12, 13, 14}  # 1-indexed

def write_tab(ws, label, rows_in):
    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
    ws['A1'] = label
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(COLS, 1):
        c = ws.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22

    row_n = 3
    totals = {'may_due':0, 'prev_due':0, 'cash':0, 'upi':0, 'total_paid':0, 'balance':0}

    partial = [r for r in rows_in if r['total_paid'] > 0]
    nopay   = [r for r in rows_in if r['total_paid'] == 0]

    for section, rows_s, base_fill in [
        ('— Partial / Carry-over —', partial, BAL_FILL),
        ('— No Payment —',           nopay,   NOPAY_FILL),
    ]:
        if not rows_s: continue
        ws.merge_cells(f'A{row_n}:{get_column_letter(len(COLS))}{row_n}')
        ws.cell(row_n, 1, section).font = Font(bold=True, italic=True, size=10, color='595959')
        ws.cell(row_n, 1).alignment = LEFT
        row_n += 1
        for r in rows_s:
            fill = CORR_FILL if r['upi_corrected'] else base_fill
            total_due = r['rent_due'] + r['prev_due']
            status_txt = ('No Payment' if r['total_paid'] == 0
                          else f'Bal {int(r["balance"]):,}')
            if r['upi_corrected']:
                status_txt += ' *bank'
            vals = [r['no'], r['building'], r['name'], r['room'], r['checkin'],
                    r['rent'], r['deposit'],
                    r['rent_due'], r['prev_due'], total_due,
                    r['cash'], r['upi'], r['total_paid'], r['balance'], status_txt]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row_n, ci, v)
                cell.fill = fill; cell.border = BORDER; cell.font = NORM_FONT
                if ci in NUM_COLS:
                    cell.alignment = RIGHT; cell.number_format = '#,##0'
                elif ci == 5:
                    cell.alignment = CENTER; cell.number_format = 'DD-MMM-YYYY'
                else:
                    cell.alignment = LEFT
            row_n += 1
            totals['may_due']   += r['rent_due']
            totals['prev_due']  += r['prev_due']
            totals['cash']      += r['cash']
            totals['upi']       += r['upi']
            totals['total_paid']+= r['total_paid']
            totals['balance']   += r['balance']

    total_due_sum = totals['may_due'] + totals['prev_due']
    total_vals = [None, '', 'TOTAL', '', None, None, None,
                  totals['may_due'], totals['prev_due'], total_due_sum,
                  totals['cash'], totals['upi'], totals['total_paid'], totals['balance'], '']
    for ci, v in enumerate(total_vals, 1):
        cell = ws.cell(row_n, ci, v)
        cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER
        if ci in NUM_COLS and isinstance(v, (int, float)):
            cell.alignment = RIGHT; cell.number_format = '#,##0'

    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'
    return totals['balance']

def write_unmatched_tab(ws, rows_in):
    """Separate tab: tenants with balance but ZERO bank UPI match and ZERO cash."""
    UCOLS  = ['#', 'Building', 'Name', 'Room', 'Phone', 'Check-in', 'Agreed Rent',
              'Deposit', 'May Due', 'Apr Carry', 'Total Due', 'Cash', 'Note']
    UWIDTH = [4, 8, 28, 6, 14, 12, 11, 10, 10, 10, 10, 10, 35]
    UNUM   = {7, 8, 9, 10, 11, 12}

    ws.merge_cells(f'A1:{get_column_letter(len(UCOLS))}1')
    ws['A1'] = 'NO PAYMENT FOUND — Bank + Cash both zero (May 2026)'
    ws['A1'].font = Font(bold=True, size=13, color='C00000')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(UCOLS, 1):
        c = ws.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER

    row_n = 3
    for r in rows_in:
        total_due = r['rent_due'] + r['prev_due']
        note = 'No UPI in bank, no cash in sheet'
        if r['apr_bank_upi'] > 0:
            note += f'; Apr bank found {r["apr_bank_upi"]:,.0f}'
        vals = [r['no'], r['building'], r['name'], r['room'], r['phone'],
                r['checkin'], r['rent'], r['deposit'],
                r['rent_due'], r['prev_due'], total_due, r['cash'], note]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row_n, ci, v)
            cell.fill = NOPAY_FILL; cell.border = BORDER; cell.font = NORM_FONT
            if ci in UNUM:
                cell.alignment = RIGHT; cell.number_format = '#,##0'
            elif ci == 6:
                cell.alignment = CENTER; cell.number_format = 'DD-MMM-YYYY'
            else:
                cell.alignment = LEFT
        row_n += 1

    for ci, w in enumerate(UWIDTH, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'

# ── Build Excel ───────────────────────────────────────────────────────────────
for i, r in enumerate(outstanding, 1): r['no'] = i

wb_out = openpyxl.Workbook()
if wb_out.active: wb_out.remove(wb_out.active)

# ALL sheet
ws_all = wb_out.create_sheet('ALL')
tot_all = write_tab(ws_all, 'HULK + THOR — May 2026 Outstanding (as of 11 May 2026)', outstanding)

# Per-building
for bname in ['HULK', 'THOR']:
    brows = [r for r in outstanding if r['building'] == bname]
    for i, r in enumerate(brows, 1): r['no'] = i
    ws = wb_out.create_sheet(bname)
    write_tab(ws, f'{bname} — May 2026 Outstanding (as of 11 May 2026)', brows)

# Unmatched tab
no_match_list = [r for r in outstanding if r['no_bank_match'] and r['cash'] == 0]
for i, r in enumerate(no_match_list, 1): r['no'] = i
ws_unm = wb_out.create_sheet('NO PAYMENT FOUND')
write_unmatched_tab(ws_unm, no_match_list)

os.makedirs('data/reports', exist_ok=True)
filepath = 'data/reports/May_Full_Analysis_2026.xlsx'
wb_out.save(filepath)
print(f'\nSaved: {filepath}')
print(f'  ALL tab:            {len(outstanding)} tenants, Rs.{tot_all:,.0f} total balance')
hulk_bal = sum(r["balance"] for r in outstanding if r["building"]=="HULK")
thor_bal = sum(r["balance"] for r in outstanding if r["building"]=="THOR")
print(f'  HULK:               {sum(1 for r in outstanding if r["building"]=="HULK")} tenants, Rs.{hulk_bal:,.0f}')
print(f'  THOR:               {sum(1 for r in outstanding if r["building"]=="THOR")} tenants, Rs.{thor_bal:,.0f}')
print(f'  NO PAYMENT FOUND:   {len(no_match_list)} tenants (zero bank + zero cash)')
print(f'\nLegend: Yellow=partial, Red=no payment, Green=UPI updated from bank statement')

# ── Also print the NO PAYMENT list to console ─────────────────────────────────
print(f'\n{"="*100}')
print(f'  TENANTS WITH ZERO PAYMENT FOUND (no UPI in bank, no cash in sheet) — {len(no_match_list)}')
print(f'{"="*100}')
print(f'  {"#":>3}  {"Bldg":5}  {"Name":<28}  {"Room":>5}  {"Phone":<14}  {"Checkin":10}  {"Rent":>7}  {"Dep":>7}  {"May Due":>8}  {"Apr Carry":>9}  {"Total":>8}')
print('  ' + '-'*120)
for r in no_match_list:
    cin = str(r['checkin']) if r['checkin'] else 'N/A'
    total = r['rent_due'] + r['prev_due']
    print(f'  {r["no"]:>3}  {r["building"]:5}  {r["name"]:<28}  {r["room"]:>5}  {r["phone"]:<14}  {cin:10}  {r["rent"]:>7,.0f}  {r["deposit"]:>7,.0f}  {r["rent_due"]:>8,.0f}  {r["prev_due"]:>9,.0f}  {total:>8,.0f}')
print()
