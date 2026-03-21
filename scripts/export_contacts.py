"""Export all Cozeevo contacts from WhatsApp chats to Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Contacts"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
cat_font = Font(bold=True, size=11, color="2F5496")
cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = ["Name", "Phone", "Category", "Sub-Category", "Supply / Role", "Payment Status / Notes"]
widths = [35, 22, 20, 22, 50, 45]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
ws.freeze_panes = "A2"

# Section header marker
SECTION = "__SECTION__"

contacts = [
    # ── PARTNERS ──
    [SECTION, "PARTNERS / INVESTORS"],
    ["Kiran Kumar Pemmasani (Pk)", "+91 7845952289", "Partner", "Co-founder", "Admin, payments, marketing, vendor negotiations", ""],
    ["Jitendra (Paisanurture.in, CFP)", "8978888551", "Partner", "Senior Advisor", "Financial planning, BNI Arka member, vendor sourcing", ""],
    ["Lakshmi Gorjala (Gundoos)", "7358341775", "Partner", "On-site Manager", "Day-to-day ops, rent collection, check-ins", ""],
    ["Prabhakaran Pemmasani", "", "Partner", "Investor", "Capital provider, Pk father", ""],
    ["Ashokan Perumal", "", "Partner", "Finance", "Added Mar 2026, spending reviews", ""],
    ["Chandrasekhar (Chandu)", "+91 85488 84455", "Partner", "Ops Manager", "On-site ops lead, staff/kitchen/procurement", ""],

    # ── COZEEVO NUMBERS ──
    [SECTION, "COZEEVO BUSINESS NUMBERS"],
    ["Cozeevo Reception", "8548884455", "Business", "Reception", "Main reception line", ""],
    ["Cozeevo Second Number", "8548884466", "Business", "Backup", "Second number", ""],

    # ── BUILDING OWNER ──
    [SECTION, "BUILDING OWNER SIDE"],
    ["Raghu", "", "Building Owner", "Landlord", "Rent 19.85L/month (30% bank + 70% cash), electricity disputes", ""],
    ["Suma", "", "Building Owner", "Owner wife", "Receives rent payments, bank beneficiary", ""],
    ["Manju (Raghu brother)", "", "Building Owner", "Family", "Cash rent collection", ""],

    # ── FAMILY / ADVISORS ──
    [SECTION, "FAMILY / ADVISORS"],
    ["Narendra", "", "Advisor", "Family / Owner rep", "Mediates with landlord Raghu, lease, police, brand", ""],
    ["Dhana", "", "Advisor", "Family", "Building issues, brand discussion", ""],
    ["Chandra", "", "Staff", "Family ops", "Gas contacts, kitchen mgmt, liaison with Raghu", ""],
    ["Kiran Prabhu (Bellandur)", "", "Advisor", "External", "Contacts for housekeeping, inverter/battery (Babai Store HSR)", ""],

    # ── STAFF ──
    [SECTION, "STAFF (CURRENT & FORMER)"],
    ["Lokesh / Loki", "", "Staff", "Receptionist / Supervisor", "On-ground, vendor coordination, active", ""],
    ["Akhil Reddy", "", "Staff", "Site coordinator", "Kitchen, construction, payments, active", ""],
    ["Anurudh V", "+91 90927 80783", "Staff", "Gym manager", "Gym equipment, CCTV, water leakage reports", ""],
    ["Chef Som / Vijaykumar Kumar", "+91 99723 34783", "Staff", "Head Cook (2nd)", "Salary 25k->26k, active", ""],
    ["Kitchen helper (Aunty)", "", "Staff", "Kitchen", "Salary 13-14k", ""],
    ["Gourav Aditya", "", "Staff", "Receptionist", "Hired ~12 Mar 2026, 25k + 10% after probation", ""],
    ["Naresh", "", "Staff (former)", "Receptionist", "Left / unavailable 2 months", ""],
    ["BK / Bikey", "", "Staff (former)", "Receptionist", "Let go after 7 days, 21k salary", ""],
    ["Imrana (~immusoni)", "", "Staff (former)", "Receptionist", "Removed ~Dec 1, keys collected", ""],
    ["Feroz", "", "Staff (former)", "Cleaner", "Left without notice, salary withheld", ""],
    ["Nandhini", "734-8950447", "Staff candidate", "Receptionist", "Age 29, good English poor Hindi, 25k->30k", ""],

    # ── FURNITURE & FITTINGS ──
    [SECTION, "VENDORS - FURNITURE & FITTINGS"],
    ["Ibrahim", "+91 9845227633", "Vendor", "Chairs", "150 chairs THOR + 83 HULK. Wholesale: +91 99011 60099", "THOR paid. HULK paid, ~23 pending delivery"],
    ["Jubair (Grace Traders)", "", "Vendor", "Bed frames / Cots", "170 cots (3x6@4600 + 2.5x6@4200) = 7,74,000. Bank: Union Bank 395301010031026", "Settled"],
    ["Royal Furniture", "9900322679", "Vendor", "Wardrobes", "Iron wardrobes 2,700-3,000/unit", ""],
    ["Kondala Rao", "+91 96326 94840", "Vendor", "Wardrobe carpenter", "Wardrobe repairs", ""],
    ["Mubeen", "", "Vendor", "Shoe racks", "1,675/rack, 88 total (~1.47L)", "Fully paid"],
    ["Somnath", "96327 84999", "Vendor", "Carpenter / Plumber", "Foldable study tables @1,650, plumbing, 1,500/day", ""],
    ["Lavanya", "+91 97433 17333", "Vendor", "Study tables / Chairs", "Imported tables @1,650 each", "Sample paid"],
    ["Abhishek", "+91 99245 40656", "Vendor", "Second-hand furniture", "Side tables, key locks (Chikpet)", ""],
    ["Wakefit", "", "Vendor", "Mattresses & Pillows", "100 mattresses + pillows = 3,39,693+", "Paid"],

    # ── ELECTRONICS ──
    [SECTION, "VENDORS - ELECTRONICS & APPLIANCES"],
    ["Karan (Usha Trading Co)", "+91 98445 32900", "Vendor", "Electronics", "TVs BPL 9800/unit, washing machines, microwaves, fridges, dispensers", "Credit note 45,891 pending"],
    ["John (Unisol)", "+91 98454 19873", "Vendor", "CCTV", "CC camera system installation", "1L paid + 1.02L pending"],
    ["Veeramani (Unisol)", "+91 98860 04169", "Vendor", "CCTV", "CCTV setup partner", ""],
    ["Praveen (RO & Sensors)", "+91 99865 93093", "Vendor", "Automation / RO", "Water tank sensors 6,900 each, RO repairs", ""],
    ["Ramesh (Hot Water)", "+49 9902 618311", "Vendor", "Hot water tech", "Heat pump / hot water system. Often unreachable", ""],
    ["Shiva (Motor Automation)", "+91 98804 79792", "Vendor", "Automation", "Motor automation", ""],

    # ── CONSTRUCTION ──
    [SECTION, "VENDORS - CONSTRUCTION & FABRICATION"],
    ["Shyam Reddy", "+91 97312 60340", "Vendor", "Fabricator", "Gas shelter 20k, study tables 1,650/unit, blue drums", "Multiple payments"],
    ["Govinda Nayaku", "+91 93422 05440", "Vendor", "Gas piping", "Copper piping setup", "45k advance paid"],
    ["Imran Saki", "9901650624", "Vendor", "Gas cage fabricator", "Storage cage. Contact from Govinda", ""],
    ["Dinesh (Scaffoldings)", "+91 99002 32936", "Vendor", "Scaffolding rental", "Profile light install", "8,000 paid"],
    ["PRITHVI Steel Kraft (RAJ)", "+91 72597 52457", "Vendor", "Kitchen steel items", "Steel utensils", "1,00,000 paid"],
    ["Dattatreya", "+91 93796 87293", "Vendor", "Gym flooring", "648 sft rubber @135/sft = 87,480", "60k advance"],
    ["Manju (Cozeevo Engineer)", "+91 97383 61211", "Vendor", "Builder engineer", "Building construction, plumbing, painting coord", ""],

    # ── INTERIOR ──
    [SECTION, "VENDORS - INTERIOR & DESIGN"],
    ["Sachin Jain (BNI Pragati)", "+91 70229 70608", "Vendor", "Interior / Cushion", "Cushion, vertical garden, bar stools", "Multiple installments"],
    ["Shalini", "+91 74060 36220", "Vendor", "Wallpaper / Stickers", "Installation + printing", "52,540 total"],
    ["Balu.S (BNI Pragathi)", "+91 98459 34818", "Vendor", "Painter", "Reception, gym, building painting", "+5000 extra for redo"],
    ["Naveen (Architect)", "+91 89040 86669", "Vendor", "Architect", "Design, plant placement, wallpaper", ""],
    ["Prakash Reddy", "+91 89044 02406", "Vendor", "Photo frames", "Custom 20x30 inch frames", ""],
    ["Suresh (Signages)", "+91 77609 60636", "Vendor", "Signage", "Sign boards, room/floor numbers", "43,426 + 9,000 GST"],
    ["Biju", "+91 97390 01672", "Vendor", "Plants", "Natural plants for reception", ""],
    ["Krishnaveni (Vcare Plants)", "+91 96637 53024", "Vendor", "Plants", "Alternative supplier", ""],

    # ── ELECTRICAL & PLUMBING ──
    [SECTION, "VENDORS - ELECTRICAL & PLUMBING"],
    ["Shiva (Architect electrician)", "+91 90353 38896", "Vendor", "Electrician", "Light installation. 20K agreed, 10K paid, 15K pending", ""],
    ["Pavan (Electrician)", "+91 83103 18123", "Vendor", "Electrician", "Room light install/repairs", ""],
    ["Keshav (Electrician WF)", "+91 98804 01360", "Vendor", "Electrician", "Room repairs, wiring fixes", ""],
    ["Md Akbar", "76448 67491", "Vendor", "Electrician", "Local electrician", ""],
    ["Mahesh", "+91 76248 07174", "Vendor", "Electrician", "Tamil, apartment background", ""],
    ["Gopal Ele", "+91 99019 50658", "Vendor", "Electrician", "From VCF", ""],
    ["Alam", "+91 87955 14149", "Vendor", "Electrician", "Local/emergency", ""],
    ["Dilip (Plumber)", "+91 78996 01416", "Vendor", "Plumber", "Emergency plumbing, hot water. Brother: Jayant", ""],
    ["Srinivas / Shriniwash", "+91 97402 13289", "Vendor", "Plumber", "Blockage removal 400-500", ""],
    ["Chandan (Plumber WF)", "+91 73488 77664", "Vendor", "Plumber", "Whitefield area", ""],
    ["Jaffar Mithri", "+91 83101 06797", "Vendor", "Plumber", "From VCF", ""],
    ["Nayak Plumber BMR", "+91 78994 21056", "Vendor", "Plumber", "From VCF", ""],
    ["Sarav Sree Ram", "+91 93379 29447", "Vendor", "Plumber", "AECS Layout plumbing", ""],
    ["Jayanth (Plumber WF)", "+91 99165 15779", "Vendor", "Plumber", "Dilip brother", ""],
    ["Rambabu (Lift Mechanic)", "+91 90031 57415", "Vendor", "Lift mechanic", "Frequent breakdowns", ""],

    # ── WIFI ──
    [SECTION, "VENDORS - INTERNET & WIFI"],
    ["Sachin (WiFi old)", "+91 99765 35858", "Vendor", "WiFi provider", "1,000/floor/month. Slow. Replaced for Hulk", ""],
    ["Thiyagu (Airwire)", "+91 77950 76250", "Vendor", "WiFi new (Hulk)", "Frequently unreliable", ""],
    ["Sreeram (Airwire owner)", "+91 99453 44115", "Vendor", "WiFi escalation", "Airwire owner, escalation contact", ""],

    # ── POWER ──
    [SECTION, "VENDORS - POWER / DIESEL / GENERATOR"],
    ["Suhail", "+91 98861 48862", "Vendor", "Diesel supplier", "Jitendra ref. Multiple vehicles. Mar 2026 crisis", ""],
    ["Vikram", "+91 74115 35239", "Vendor", "Diesel pump", "Generator diesel", ""],
    ["Jayalakshmi", "+91 98450 66039", "Vendor", "DG / Power", "Commercial generator rental", ""],
    ["Sudheesh (OJUS DJI)", "+91 96866 99609", "Vendor", "DG service", "Main service for OJUS generator", ""],
    ["Jai Bescom", "+91 94498 74375", "Vendor", "BESCOM", "Electricity contact", ""],

    # ── FOOD ──
    [SECTION, "VENDORS - FOOD & KITCHEN"],
    ["Omkar", "9663049651", "Vendor", "Vegetables", "KR Market delivery to PGs", ""],
    ["Ram Kumar (HP Gas)", "8619377620", "Vendor", "Gas delivery", "Sakthi Gas HP distributor", ""],
    ["HP Gas booking", "8310745974", "Vendor", "Gas booking", "Commercial cylinders", ""],

    # ── MANPOWER ──
    [SECTION, "VENDORS - MANPOWER & HOUSEKEEPING"],
    ["Sakthi Vel", "+91 97400 74470", "Vendor", "Cleaner supplier", "15k salary + food/accom, 10% commission", ""],
    ["Rohit / Poojayya (Rock Power)", "+91 97407 93471", "Vendor", "Manpower agency", "Security 21k, housekeeping 20k/person. BNI ARKA", ""],
    ["Vergiese", "+91 9900631199", "Vendor", "Manpower agency", "Staff recruitment", ""],
    ["Lokesh (Blue Collar)", "9901007990", "Vendor", "Manpower agency", "Cleaning staff recruitment", ""],
    ["Nirdesh", "6363293493", "Vendor", "Housekeeping", "Via Kiran Prabhu (Bellandur)", ""],
    ["Ganesh (family)", "+91 6364544461", "Vendor", "Cleaner couple", "12k/person, needed separate room", ""],
    ["Abhishek Kumar (cleaner)", "+91 6362 712 216", "Vendor", "Cleaner couple", "Joined Mar 2026", ""],

    # ── MARKETING ──
    [SECTION, "VENDORS - MARKETING & MEDIA"],
    ["Viplab", "89184 31221 (GPay)", "Vendor", "Marketing / Video", "25k initial + 10k asked. Find my PG collab", ""],
    ["Pg Marketing guy", "+91 90196 53917", "Vendor", "Marketing", "From VCF", ""],
    ["Nithin", "9164794919", "Vendor", "Printing", "Standees, good quality", ""],
    ["Prateek", "90364 09106", "Vendor", "T-shirts / Uniforms", "Staff uniforms", ""],
    ["Sandeep Gowda", "+91 96321 39796", "Vendor", "Projector / Sound", "T20 World Cup event rental", ""],

    # ── SPORTS ──
    [SECTION, "VENDORS - SPORTS & RECREATION"],
    ["9balls India", "", "Vendor", "Sports equipment", "Pool 72K, Foosball 33K, TT 24K, Subsoccer 42K, Carrom 7.5K, Chess 6K + GST = 1,98,075", ""],

    # ── FIRE SAFETY ──
    [SECTION, "VENDORS - FIRE & SAFETY"],
    ["Pramod (Pragathi Fire)", "", "Vendor", "Fire extinguishers", "14 ABC 6kg + 1 CO2 + 1 K-type = ~32,700", "Paid"],

    # ── WASTE & WATER ──
    [SECTION, "VENDORS - WASTE & WATER"],
    ["Pavan (Garbage)", "+91 6366 411 789", "Vendor", "Garbage collector", "Waste collection", ""],
    ["Venkatesh (Water)", "7349663198", "Vendor", "Water tanker", "HWS supply", ""],
    ["Vinay Kumar (Tanker)", "+91 97396 47672", "Vendor", "Water tanker", "BM Shoma area", ""],
    ["Muneshwara (Tanker)", "+91 85531 67678", "Vendor", "Water tanker", "Tanker supply", ""],

    # ── KEY & LOCKS ──
    [SECTION, "VENDORS - KEY & LOCKS"],
    ["Key Shop BMR", "+91 91488 09732 / +91 90154 04182", "Vendor", "Key shop", "Key cutting, locks", ""],

    # ── PROFESSIONAL ──
    [SECTION, "PROFESSIONAL SERVICES"],
    ["Ashok (Auditor)", "+91 98440 36556", "Professional", "Auditor", "For Dhana/Narendra", ""],
    ["CA Manesh (BNI ARKA)", "", "Professional", "Chartered Accountant", "", ""],
    ["Roopesh", "", "Professional", "GST / Tax consultant", "Jitendra knows", ""],

    # ── POLICE / GOVT ──
    [SECTION, "POLICE / GOVERNMENT"],
    ["Pradeep (HAL Police)", "", "Government", "Police constable", "HAL station. Wanted 6k, negotiated to 2k", ""],
    ["Subhan (Police WF)", "+91 90604 77309", "Government", "Police", "Whitefield police contact", ""],
    ["Manju (BBMP)", "8747884323", "Government", "BBMP contact", "Trade license", ""],

    # ── CORPORATE ──
    [SECTION, "CORPORATE / BUSINESS CONTACTS"],
    ["Satish Kumar J (Celestial Systems)", "skjayakumar@celestialsys.com", "Corporate", "Booking", "Celestial Systems / Hitachi. Signed", ""],
    ["Chidambaram", "", "Corporate", "Hitachi admin", "In progress", ""],
    ["Venkata Rao", "+91 9502506702", "Industry", "Other PG operator", "PG industry contact", ""],
    ["Anitha (BNI Arka)", "+91 94838 59240", "BNI", "Reference", "Contact shared 15 Feb", ""],

    # ── TENANTS / LEADS ──
    [SECTION, "TENANTS / LEADS (mentioned in chat)"],
    ["Ajay Gupta", "8305901822", "Lead", "Double sharing", "", ""],
    ["Sandeep", "9561093025", "Lead", "Single sharing", "2 rooms from Feb 1", ""],
    ["Shyam", "7093624566", "Lead", "Double sharing", "First week Jan", ""],
    ["Santosh", "6301307100", "Lead", "3 sharing", "2 beds", ""],
    ["Shreyansh", "6393770307", "Lead", "3 sharing", "Jan 1, 3 beds", ""],
    ["Sanskar", "+91 70000 64421", "Lead", "3 people room", "Inquiry", ""],
]

row = 2
for c in contacts:
    if c[0] == SECTION:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=c[1])
        cell.font = cat_font
        cell.fill = cat_fill
        cell.border = border
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = cat_fill
            ws.cell(row=row, column=col).border = border
        row += 1
        continue
    for col, val in enumerate(c, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

# ── BANK ACCOUNTS sheet ──
ws2 = wb.create_sheet("Bank Accounts")
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 25
ws2.column_dimensions["D"].width = 15
ws2.column_dimensions["E"].width = 35

for col, h in enumerate(["Account Name", "Bank", "Account Number", "IFSC", "Purpose"], 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

banks = [
    ["Grace Traders (Usha Trading Co)", "Union Bank", "395301010031026", "", "Cot payments"],
    ["Sri Lakshmi Chandrasekhar Svc Stn", "SBI", "34891736112", "SBIN0006997", "Cash-to-bank conversion (6L)"],
    ["Kumar U.C", "SBI", "41357570548", "", "Vendor payment, Ph: 8105427895"],
]
for r, bank in enumerate(banks, 2):
    for col, val in enumerate(bank, 1):
        cell = ws2.cell(row=r, column=col, value=val)
        cell.border = border

wb.save("data/whatsapp_chats/cozeevo_all_contacts.xlsx")
print(f"Saved: data/whatsapp_chats/cozeevo_all_contacts.xlsx")
print(f"Total rows: {row - 2}")
