"""Show the 29 tenants source sheet flags with April Balance > 0.

For each: source balance, our DB balance, gap, reason.
Output: data/reports/april_balance_29_tenants.xlsx
"""
import asyncio
import os
import sys
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv()

import gspread
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from google.oauth2.service_account import Credentials
from sqlalchemy import select
from src.database.db_manager import init_engine, get_session
from src.database.models import Payment, PaymentFor, RentSchedule, Tenancy, Tenant, Room
from src.utils.money import inr as _inr


def pn(v):
    if not v:
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


async def main():
    # ── Read source sheet ──
    creds = Credentials.from_service_account_file(
        "credentials/gsheets_service_account.json",
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key("1Vr_fSIOuuKBK4MWF-FVqgAIUPbqun3POszaXYfj-Ea0")
    rows = sh.worksheet("Long term").get_all_values()[1:]

    # Extract ONLY rows with April Balance > 0
    src_rows = []
    for r in rows:
        if len(r) < 24:
            continue
        room = (r[0] or "").strip()
        name = (r[1] or "").strip()
        if not name:
            continue
        bal = pn(r[23])
        if bal <= 0:
            continue
        src_rows.append({
            "room": room,
            "name": name,
            "comment": (r[14] or "").strip(),
            "inout": (r[16] or "").strip(),
            "march_bal": pn(r[20]),
            "april_status": (r[19] or "").strip(),
            "cash": pn(r[21]),
            "upi": pn(r[22]),
            "src_balance": bal,
        })

    print(f"Source rows with Balance > 0: {len(src_rows)}")
    print(f"Source total: Rs.{_inr(sum(r['src_balance'] for r in src_rows))}")

    # ── DB lookup ──
    init_engine(os.environ["DATABASE_URL"])
    apr = date(2026, 4, 1)
    may = date(2026, 5, 1)

    async with get_session() as s:
        # Load maps
        rs_all = (await s.execute(
            select(RentSchedule).where(RentSchedule.period_month == apr)
        )).scalars().all()
        rs_map = {rs.tenancy_id: rs for rs in rs_all}

        pay_all = (await s.execute(
            select(Payment).where(
                Payment.period_month == apr,
                Payment.is_void == False,
                Payment.for_type == PaymentFor.rent,
            )
        )).scalars().all()
        cash_m, upi_m, prepaid_m = {}, {}, {}
        for p in pay_all:
            in_apr = p.payment_date and apr <= p.payment_date < may
            mode = p.payment_mode.value if hasattr(p.payment_mode, "value") else str(p.payment_mode or "cash")
            if in_apr:
                if mode.lower() == "cash":
                    cash_m[p.tenancy_id] = cash_m.get(p.tenancy_id, 0) + float(p.amount)
                else:
                    upi_m[p.tenancy_id] = upi_m.get(p.tenancy_id, 0) + float(p.amount)
            else:
                prepaid_m[p.tenancy_id] = prepaid_m.get(p.tenancy_id, 0) + float(p.amount)

        bk_m, dep_m = {}, {}
        for p in (await s.execute(
            select(Payment).where(Payment.is_void == False, Payment.for_type == PaymentFor.booking)
        )).scalars():
            bk_m[p.tenancy_id] = bk_m.get(p.tenancy_id, 0) + float(p.amount)
        for p in (await s.execute(
            select(Payment).where(Payment.is_void == False, Payment.for_type == PaymentFor.deposit)
        )).scalars():
            dep_m[p.tenancy_id] = dep_m.get(p.tenancy_id, 0) + float(p.amount)

        # For each source row, find tenancy
        data = []
        for s_row in src_rows:
            # Lookup tenant by name + room
            room_num = s_row["room"]
            name_lc = s_row["name"].lower()

            rm = (await s.execute(
                select(Room).where(Room.room_number == room_num)
            )).scalar_one_or_none()

            our_due = 0
            our_cash = our_upi = our_prepaid = our_booking = our_deposit = 0
            our_balance = 0
            status = ""
            checkin = ""
            found = False

            if rm:
                tenancies = (await s.execute(
                    select(Tenancy).where(Tenancy.room_id == rm.id)
                )).scalars().all()
                for t in tenancies:
                    tn = await s.get(Tenant, t.tenant_id)
                    if tn and tn.name.lower() == name_lc:
                        found = True
                        status = t.status.value if hasattr(t.status, "value") else str(t.status)
                        checkin = t.checkin_date.strftime("%Y-%m-%d") if t.checkin_date else ""
                        is_fm = bool(t.checkin_date and t.checkin_date.replace(day=1) == apr)
                        rs = rs_map.get(t.id)
                        our_due = float(rs.rent_due) if rs else 0
                        our_cash = cash_m.get(t.id, 0)
                        our_upi = upi_m.get(t.id, 0)
                        our_prepaid = prepaid_m.get(t.id, 0)
                        our_booking = bk_m.get(t.id, 0) if is_fm else 0
                        our_deposit = dep_m.get(t.id, 0) if is_fm else 0
                        our_paid = our_cash + our_upi + our_prepaid + our_booking + our_deposit
                        our_balance = max(0, our_due - our_paid)
                        break

            gap = our_balance - s_row["src_balance"]
            reason = ""
            if not found:
                reason = "NOT IN DB for that room"
            elif our_due == 0:
                reason = "No RentSchedule (future checkin or not billed)"
            elif gap < -100:
                reason = "We show less pending (deposit/booking covers more)"
            elif gap > 100:
                reason = "We show more pending"

            data.append({
                **s_row,
                "status": status,
                "checkin": checkin,
                "our_due": int(our_due),
                "our_cash": int(our_cash),
                "our_upi": int(our_upi),
                "our_prepaid": int(our_prepaid),
                "our_booking": int(our_booking),
                "our_deposit": int(our_deposit),
                "our_balance": int(our_balance),
                "gap": int(gap),
                "reason": reason,
            })

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "29 Source Balance Tenants"

    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F8F9FA")
    red_fill = PatternFill("solid", fgColor="FFE5E5")
    green_fill = PatternFill("solid", fgColor="E5FFE5")
    gold_fill = PatternFill("solid", fgColor="FFD700")

    headers = [
        "Room", "Name", "In/Out", "Status", "Check-in", "Comment",
        "Src Cash", "Src UPI", "Src Balance",
        "Our Due", "Our Cash", "Our UPI", "Our Prepaid", "Our Booking", "Our Deposit",
        "Our Balance", "Gap (Ours − Src)", "Reason",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    data.sort(key=lambda d: -d["src_balance"])
    for ri, d in enumerate(data, 2):
        vals = [
            d["room"], d["name"], d["inout"], d["status"], d["checkin"], d["comment"],
            int(d["cash"]), int(d["upi"]), int(d["src_balance"]),
            d["our_due"], d["our_cash"], d["our_upi"], d["our_prepaid"], d["our_booking"], d["our_deposit"],
            d["our_balance"], d["gap"], d["reason"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if 7 <= ci <= 17 and isinstance(v, int):
                c.number_format = '#,##0'
            if ri % 2 == 0:
                c.fill = alt_fill
        gc = ws.cell(row=ri, column=17)
        if d["gap"] > 100:
            gc.fill = red_fill
        elif d["gap"] < -100:
            gc.fill = green_fill

    tr = len(data) + 2
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col, key in [(7, "cash"), (8, "upi"), (9, "src_balance"),
                     (10, "our_due"), (11, "our_cash"), (12, "our_upi"),
                     (13, "our_prepaid"), (14, "our_booking"), (15, "our_deposit"),
                     (16, "our_balance"), (17, "gap")]:
        tot = sum(d[key] for d in data)
        c = ws.cell(row=tr, column=col, value=int(tot))
        c.font = Font(bold=True)
        c.number_format = '#,##0'
        c.fill = gold_fill

    widths = [8, 28, 10, 10, 12, 40,
              10, 10, 11, 10, 11, 11, 11, 11, 11,
              12, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "C2"

    out = os.path.join("data", "reports", "april_balance_29_tenants.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"\nSrc Balance total:  Rs.{_inr(sum(d['src_balance'] for d in data))}")
    print(f"Our Balance total:  Rs.{_inr(sum(d['our_balance'] for d in data))}")
    print(f"Net gap:            Rs.{_inr(sum(d['gap'] for d in data))}")

    # Classification
    not_in_db = [d for d in data if "NOT IN DB" in d["reason"]]
    no_rs = [d for d in data if "No RentSchedule" in d["reason"]]
    less_pending = [d for d in data if "less pending" in d["reason"]]
    more_pending = [d for d in data if "more pending" in d["reason"]]
    match = [d for d in data if not d["reason"]]

    print(f"\nBreakdown:")
    print(f"  NOT IN DB:                  {len(not_in_db)}")
    print(f"  No RentSchedule:            {len(no_rs)} (future checkins etc)")
    print(f"  We show less pending:       {len(less_pending)}")
    print(f"  We show more pending:       {len(more_pending)}")
    print(f"  Exact match:                {len(match)}")


asyncio.run(main())
