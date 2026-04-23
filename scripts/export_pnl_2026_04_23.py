"""Export the corrected accrual-basis P&L and opex breakdown for Oct'25-Apr'26.

Produces: data/reports/PnL_Accrual_2026_04_23.xlsx

Rules applied (from memory/sop_pnl.md Step 4 — updated 2026-04-23):
- Property Rent: accrual Jan+Feb+Mar+Apr = Rs.21.32L/mo each. Nov-Dec zero.
- Water (Manoj B): accrual Mar+Apr = Rs.42,500/mo. No water before March.
- Internet: keep pre-Feb bank as-is; amortise Rs.15,514/mo Feb onwards.
- Police: Rs.3K/mo accrual Jan onwards.
- Waste: Rs.3.5K/mo (classifier rule handles).
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from utils.inr_format import INR_NUMBER_FORMAT

OUT = Path(__file__).parent.parent / "data" / "reports" / "PnL_Accrual_2026_04_23.xlsx"

MONTHS = ["Oct'25", "Nov'25", "Dec'25", "Jan'26", "Feb'26", "Mar'26"]

# ---- Income (from DB, per sop_pnl.md Step 3) ----
income = {
    "Rent Cash":    [0,       0,        0,        300572,   653300,   1094220],
    "Rent UPI":     [0,       722908,   1350548,  1335224,  2324048,  2889193],
    "Other Income": [0,       100,      0,        14000,    28048,    28014],
}

# Owner capital contribution — NOT revenue, shown separately below P&L
capital_contributions = {
    "Owner startup advance (Kiran, from pocket)": [500000, 0, 0, 0, 0, 0],
}

# ---- Opex (accrual, per sop_pnl.md Step 4) ----
opex = {
    "Property Rent (accrual Rs.21.32L/mo, Jan onwards)":  [0, 0, 0,       2132000, 2132000, 2132000],
    "Electricity":                                         [0, 0, 0,       131554,  134538,  96617],
    "Water (Manoj Mar bill Rs.42.5K + bank tanker Rs.8K)": [0, 0, 0,       0,       0,       50500],
    "IT & Software":                                       [0, 0, 3480,    10620,   0,       0],
    "Internet & WiFi (bank pre-Feb + Rs.15.5K Feb+)":      [0, 0, 40946,   70952,   15514,   15514],
    "Food & Groceries":                                    [0, 1086, 34435, 201558, 94931,   237747],
    "Fuel & Diesel":                                       [0, 0, 0,       9099,    104366,  346308],
    "Staff & Labour":                                      [0, 1000, 125935, 112063, 219715, 155481],
    "Maintenance & Repairs":                               [0, 0, 0,       0,       500,     18470],
    "Cleaning Supplies":                                   [0, 0, 4414,    1400,    700,     4566],
    "Waste Disposal (Pavan Rs.3.5K/mo)":                   [0, 0, 0,       3000,    3500,    3500],
    "Shopping & Supplies":                                 [0, 2730, 10530, 12036,  6127,    6184],
    "Operational Expenses":                                [0, 0, 47769,   10315,   2174,    3237],
    "Marketing":                                           [0, 0, 39500,   17895,   3620,    27700],
    "Govt & Regulatory (incl Police Rs.3K accrual Jan+Feb)":[0, 0, 75716,   88073,   3000,    3000],
    "Bank Charges":                                        [0, 0, 0,       149,     0,       0],
    "Other Expenses":                                      [0, 10000, 83556, 6564,  23308,   98500],
}

# ---- Excluded from opex (informational) ----
excluded = {
    "Furniture & Fittings (CAPEX)":      [0, 50000, 110191, 203815, 1185397, 331],
    "CCTV Installation (CAPEX)":         [0, 82000, 0, 0, 0, 0],
    "Tenant Deposit Refund (liability)": [0, 10000, 21500, 53944, 74532, 118671],
    "Loan Repayment / Transfers (non-op)":[0, 450000, 0, 0, 600000, 2090000],
}

# ---- April 2026 partial (fixed lines only) ----
april_fixed = {
    "Property Rent": 2132000,
    "Internet & WiFi (amortised)": 15514,
    "Waste Disposal": 3500,
    "Police": 3000,
    "Water (Manoj accrual)": 42500,
}
april_revenue = 4239579
april_other_opex_est = 1000000


def sum_row(row):
    return [sum(row[:i+1]) if False else row[i] for i in range(len(row))] + [sum(row)]


def rupees(n):
    if n == 0 or n is None:
        return ""
    return f"{n:,.0f}"


def main():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: P&L Summary ----
    ws = wb.active
    ws.title = "P&L Summary"
    bold = Font(bold=True)
    hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    header = [""] + MONTHS + ["TOTAL"]
    ws.append(header)
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    ws.append(["INCOME"])
    ws[ws.max_row][0].font = bold
    for label, row in income.items():
        ws.append([label] + row + [sum(row)])
    rev_row = [sum(col) for col in zip(*income.values())]
    ws.append(["Revenue"] + rev_row + [sum(rev_row)])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = total_fill

    ws.append([])
    ws.append(["OPERATING EXPENSES (accrual)"])
    ws[ws.max_row][0].font = bold
    for label, row in opex.items():
        ws.append([label] + row + [sum(row)])
    opex_row = [sum(col) for col in zip(*opex.values())]
    ws.append(["Total Opex"] + opex_row + [sum(opex_row)])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = total_fill

    ws.append([])
    profit_row = [r - o for r, o in zip(rev_row, opex_row)]
    ws.append(["NET PROFIT (accrual)"] + profit_row + [sum(profit_row)])
    for c in ws[ws.max_row]:
        c.font = bold

    margin_row = [f"{(p/r*100):.1f}%" if r else "-" for p, r in zip(profit_row, rev_row)]
    ws.append(["Margin %"] + margin_row + [f"{(sum(profit_row)/sum(rev_row)*100):.1f}%"])

    ws.append([])
    ws.append(["CAPITAL CONTRIBUTIONS (not P&L — balance-sheet / owner's equity)"])
    ws[ws.max_row][0].font = bold
    for label, row in capital_contributions.items():
        ws.append([label] + row + [sum(row)])

    ws.append([])
    ws.append(["EXCLUDED (not in opex)"])
    ws[ws.max_row][0].font = bold
    for label, row in excluded.items():
        ws.append([label] + row + [sum(row)])

    # Format numbers as rupees
    for row in ws.iter_rows(min_row=2, max_col=len(header)):
        for i, cell in enumerate(row):
            if i == 0 or cell.value is None or isinstance(cell.value, str):
                continue
            cell.number_format = INR_NUMBER_FORMAT

    ws.column_dimensions["A"].width = 55
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 14

    # ---- Sheet 2: April 2026 (partial) ----
    ws2 = wb.create_sheet("April 2026 (partial)")
    ws2.append(["April 2026 — revenue complete, opex partial"])
    ws2[1][0].font = bold
    ws2.append([])
    ws2.append(["Revenue", april_revenue])
    ws2.append([])
    ws2.append(["Fixed-line opex (accrual)"])
    ws2[ws2.max_row][0].font = bold
    fixed_sum = 0
    for k, v in april_fixed.items():
        ws2.append([k, v])
        fixed_sum += v
    ws2.append(["Fixed-line total", fixed_sum])
    ws2[ws2.max_row][0].font = bold
    ws2.append([])
    ws2.append(["Other opex (estimated at Mar run-rate)", april_other_opex_est])
    ws2.append(["Projected total opex", fixed_sum + april_other_opex_est])
    ws2.append([])
    ws2.append(["Apr estimated profit", april_revenue - fixed_sum - april_other_opex_est])
    ws2[ws2.max_row][0].font = bold
    for row in ws2.iter_rows():
        for i, cell in enumerate(row):
            if i >= 1 and isinstance(cell.value, (int, float)):
                cell.number_format = INR_NUMBER_FORMAT
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 16

    # ---- Sheet 3: Rules & Overrides ----
    ws3 = wb.create_sheet("Rules Applied")
    rules = [
        ("Basis", "Accrual — expense in month of service, not payment"),
        ("Source of income", "DB payments table (SUM by period_month, payment_mode, where for_type=rent, is_void=false)"),
        ("Property Rent", "Rs.21,32,000/mo accrual. Jan+Feb+Mar+Apr. Nov-Dec = zero (notice period). Bank UPI replaced by accrual (do not add on top)."),
        ("Payment cycle — rent", "Each month's rent paid on 10th of FOLLOWING month. Jan rent -> 10 Feb, Feb rent -> 10 Mar, etc."),
        ("Water — Manoj B (9535665407)", "Rs.42,500/mo accrual starting MARCH only. Paid one month behind. Mar usage -> April payment."),
        ("Water — bank tanker Rs.8K Mar", "Kept SEPARATE, additive to Manoj's accrual (not replacement)."),
        ("Internet pre-Feb", "Keep bank classifier numbers (Sachin 9976535858 + Q786467820 + others). Do NOT erase."),
        ("Internet Feb onwards", "Airwire Rs.1.13L + WiFi Vendor Rs.1.04L = 14-month prepayment covering BOTH buildings. Amortise Rs.15,514/mo Feb 2026 - Mar 2027. Replace Feb bank spike."),
        ("Police", "Rs.3,000/mo cash accrual Jan onwards. Injected where not in bank."),
        ("Waste Disposal (Pavan 6366411789)", "Rs.3,500/mo. Classifier catches it; no override."),
        ("Excluded from opex", "Furniture & Fittings (capex), CCTV (capex), Tenant Deposit Refunds (liability), Loan Repayments (non-operating)."),
    ]
    ws3.append(["Rule", "Value"])
    for c in ws3[1]:
        c.fill = hdr_fill
        c.font = hdr_font
    for k, v in rules:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 100
    for row in ws3.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
