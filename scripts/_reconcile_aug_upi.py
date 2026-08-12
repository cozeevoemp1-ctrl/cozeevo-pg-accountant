"""
August-2026 UPI reconciliation: gateway settlement reports vs App vs Google Sheet.

  GATEWAY  data/uploads/csv/Thor UPI.csv     (real CSV)
           data/uploads/csv/Hulk  upi.csv    (actually an XLSX despite the name)
           Columns: RRN, Date, Time, TXN_AMOUNT, Payer_VPA, Payer_Name, Settlement_Status
           NOTE: neither file carries a room number — the room is derived by
           matching the payer's phone (from the VPA) or name to a tenant.

  APP      DB payments, payment_mode='upi', is_void=false, Aug-2026 window
  GSHEET   'google sheet local excel.xlsx' -> col 'Aug UPI'

Output: data/reports/Aug2026_UPI_3Way_Reconciliation.xlsx
  Tab 1 'Transactions'   — every gateway txn, 1..n, with matched room + tenant
  Tab 2 'By Tenant'      — per tenant: Gateway | App | Google Sheet, side by side
  Tab 3 'Summary'        — totals, variances, duplicate-file warning

Matching order: phone from VPA -> tenant phone (strongest), then fuzzy name.
Reuses the name matcher from _reconcile_aug_cash.py — do not duplicate it.
"""
from __future__ import annotations

import asyncio
import csv
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
load_dotenv()

from difflib import SequenceMatcher  # noqa: E402
from _reconcile_aug_cash import money, norm_name, norm_room, room_key, similar  # noqa: E402
from src.database.db_manager import get_session, init_engine  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
THOR_CSV = ROOT / "data/uploads/csv/Thor UPI-11-08-2026.csv"   # supersedes 'Thor UPI.csv'
HULK_XLSX = ROOT / "data/uploads/csv/Hulk UPI till 1th 08 2026.csv"   # misnamed: real format is xlsx
GSHEET_XLSX = ROOT / "data/uploads/csv/august 11.xlsx"
OUT = ROOT / "data/reports/Aug2026_UPI_3Way_Reconciliation.xlsx"

APP_FROM, APP_TO = date(2026, 7, 21), date(2026, 8, 31)


def phone10(s) -> str:
    """Last 10 digits of anything that looks like an Indian mobile number."""
    if not s:
        return ""
    digits = re.sub(r"\D", "", str(s))
    if len(digits) >= 10:
        tail = digits[-10:]
        if tail[0] in "6789":
            return tail
    return ""


def vpa_phone(vpa: str) -> str:
    """Many VPAs are literally the mobile number: 9885732833@ibl, 6303164102-3@ybl."""
    if not vpa:
        return ""
    local = str(vpa).split("@")[0]
    local = re.sub(r"-\d$", "", local)          # strip the -1 / -2 / -3 suffix
    return phone10(local) if re.fullmatch(r"\d{10,}", local) else ""


# ------------------------------------------------------------------ loaders


def load_gateway() -> list[dict]:
    rows: list[dict] = []

    with open(THOR_CSV, newline="", encoding="utf-8-sig") as f:
        for d in csv.DictReader(f):
            rows.append({
                "building": "THOR",
                "rrn": str(d["RRN"]).strip(),
                "date": datetime.strptime(d["Date"].strip(), "%d %b %Y").date(),
                "time": d["Time"].strip(),
                "amount": money(d["TXN_AMOUNT"]),
                "vpa": (d["Payer_VPA"] or "").strip(),
                "payer": (d["Payer_Name"] or "").strip(),
                "status": (d["Settlement_Status"] or "").strip(),
            })

    # openpyxl refuses the .csv extension, so read a temp copy named .xlsx
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "hulk.xlsx"
        shutil.copyfile(HULK_XLSX, tmp)
        ws = openpyxl.load_workbook(tmp, data_only=True).active
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None:
                continue
            d = ws.cell(r, 2).value
            t = ws.cell(r, 3).value
            rows.append({
                "building": "HULK",
                "rrn": str(ws.cell(r, 1).value).strip(),
                "date": d.date() if isinstance(d, datetime) else d,
                "time": t.strftime("%H:%M") if hasattr(t, "strftime") else str(t)[:5],
                "amount": money(ws.cell(r, 4).value),
                "vpa": str(ws.cell(r, 8).value or "").strip(),
                "payer": str(ws.cell(r, 9).value or "").strip(),
                "status": str(ws.cell(r, 12).value or "").strip(),
            })

    for x in rows:
        x["name"] = norm_name(x["payer"])
        x["phone"] = vpa_phone(x["vpa"])

    # Some payers appear in BOTH files on the same day for the same amount.
    # RRN is unique per UPI transaction and none are shared between the files,
    # so these are two genuine transactions — a tenant splitting the rent across
    # the Thor and Hulk QRs (confirmed by Kiran 2026-08-11). Count both; the
    # flag is informational only so the pairs stay easy to eyeball.
    seen: dict[tuple, list[dict]] = defaultdict(list)
    for x in rows:
        seen[(x["date"], x["amount"], x["name"])].append(x)
    for group in seen.values():
        both = len({g["building"] for g in group}) > 1
        for g in group:
            g["dup"] = "split across both QRs" if both else (
                "same amount x2, same file" if len(group) > 1 else "")
            g["dup_keep"] = True     # never drop a distinct RRN
    return rows


def load_gsheet() -> list[dict]:
    ws = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)["Sheet1"]
    hdr = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            hdr.setdefault(str(h).strip(), c)
    out = []
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, hdr["Name"]).value
        if not nm:
            continue
        out.append({
            "room": norm_room(ws.cell(r, hdr["Room No"]).value),
            "raw_name": str(nm).strip(),
            "name": norm_name(nm),
            "phone": phone10(ws.cell(r, hdr["Mobile Number"]).value),
            "amount": money(ws.cell(r, hdr["Aug UPI"]).value),
        })
    return out


async def load_app(win_from: date, win_to: date) -> tuple[list[dict], list[dict]]:
    """-> (upi payments in window, every tenancy as a name/phone/room register)

    win_to is the LAST date the gateway files actually cover. The statements are
    a part-month export (Thor to 09-Aug, Hulk to 11-Aug), so comparing them
    against a full month of app payments would invent a fake shortfall.
    """
    init_engine(os.environ["DATABASE_URL"])
    async with get_session() as s:
        pays = (await s.execute(text(
            """
            SELECT r.room_number, t.name, t.phone, p.payment_date, p.amount,
                   p.for_type::text AS for_type, p.upi_reference
            FROM payments p
            JOIN tenancies tc ON tc.id = p.tenancy_id
            JOIN tenants   t  ON t.id  = tc.tenant_id
            LEFT JOIN rooms r ON r.id  = tc.room_id
            WHERE p.is_void = false AND p.payment_mode = 'upi'
              AND p.payment_date >= :d1 AND p.payment_date <= :d2
            """), {"d1": win_from, "d2": win_to})).mappings().all()

        reg = (await s.execute(text(
            """
            SELECT DISTINCT r.room_number, t.name, t.phone
            FROM tenancies tc
            JOIN tenants t ON t.id = tc.tenant_id
            LEFT JOIN rooms r ON r.id = tc.room_id
            WHERE tc.status IN ('active','exited')
            """))).mappings().all()

    payments = [{
        "room": norm_room(m["room_number"]), "raw_name": (m["name"] or "").strip(),
        "name": norm_name(m["name"]), "phone": phone10(m["phone"]),
        "date": m["payment_date"], "amount": float(m["amount"]),
        "for_type": m["for_type"],
    } for m in pays]

    register = [{
        "room": norm_room(m["room_number"]), "raw_name": (m["name"] or "").strip(),
        "name": norm_name(m["name"]), "phone": phone10(m["phone"]),
    } for m in reg]
    return payments, register


# ------------------------------------------------------------------ model


def name_score(a: str, b: str) -> float:
    """0..1 similarity between two normalised full names.

    The gateway's Payer_Name is the bank account holder, which is often a
    reordered or fuller version of the tenant's name
    ('VIJAYA MOHAN REDDY NAGIREDDY' vs 'Vijay Nagi Reddy'), so this scores the
    WHOLE name rather than accepting any single shared token — otherwise
    'ABHISHEK JAIN' matches every Abhishek and every Jain in the register.
    """
    if not a or not b:
        return 0.0
    fa, fb = a.replace(" ", ""), b.replace(" ", "")
    flat_r = SequenceMatcher(None, fa, fb).ratio()

    ta, tb = set(a.split()), set(b.split())
    if not ta or not tb:
        return flat_r
    # token overlap, allowing near-miss tokens (NAGIREDDY ~ NAGI + REDDY)
    hits = 0
    for x in ta:
        if any(x == y or (len(x) >= 4 and len(y) >= 4
                          and SequenceMatcher(None, x, y).ratio() >= 0.85)
               for y in tb):
            hits += 1
    tok_r = hits / min(len(ta), len(tb))
    return max(flat_r, tok_r * 0.95)


NAME_FLOOR = 0.70     # below this, never call it a match
NAME_MARGIN = 0.06    # best must beat runner-up by this much

# Payers whose bank-account name is far enough from the tenant name that no
# scorer will ever link them, but where the person is unambiguous from the
# amount + room + date. Keyed by normalised Payer_Name -> room number.
# Deliberately conservative: third parties whose tenant is genuinely unknown
# (ANATOMIZ3D MEDTECH, SULEKHA DEVI, M NAGESWAR) are left unmatched on purpose.
ALIASES = {
    "SAURABH KUMAR MAHRA": "406",        # 406 Saurabh Kumar Mahra
    "K SINDHU": "410",                   # 410 Sindhu Krishnamoorthy
    "DINTAKURTHI YASWANTH": "302",       # 302 D Yaswanth
    "YASWANTH DINTAKURTHI": "302",
    "BELLIAPPA I P": "413",              # 413 Shannon Muthanna I B
    "VEENA VIRUPAXAPPA TUPPAD": "623",   # 623 Veena T
    "JHANVI JOSHI": "214",               # 214 Jahnavi
    "KODIMELA ANURAAG": "215",           # 215 Kodimela Anuraag
    "SOUMYA AGRAWAL": "206",             # 206 Soumya Agarwal
}


class Tenant:
    __slots__ = ("room", "names", "phones", "gateway", "app", "gsheet", "notes")

    def __init__(self, room, name, phone=""):
        self.room = room
        self.names = [name] if name else []
        self.phones = {phone} if phone else set()
        self.gateway = self.app = self.gsheet = 0.0
        self.notes: list[str] = []

    @property
    def display(self):
        return max(self.names, key=len) if self.names else "(unknown)"

    def note(self, m):
        if m not in self.notes:
            self.notes.append(m)


def build(gateway, app_pays, register, gsheet):
    tenants: list[Tenant] = []
    by_phone: dict[str, Tenant] = {}

    def get(room, name, phone="") -> Tenant:
        if phone and phone in by_phone:
            t = by_phone[phone]
            if name and name not in t.names:
                t.names.append(name)
            if room and not t.room:
                t.room = room
            return t
        for t in tenants:
            if t.room == room and name and any(similar(n, name) for n in t.names):
                if phone:
                    t.phones.add(phone)
                    by_phone[phone] = t
                return t
        t = Tenant(room, name, phone)
        tenants.append(t)
        if phone:
            by_phone[phone] = t
        return t

    # register first so rooms/phones exist before money is attached
    for r in register:
        get(r["room"], r["name"], r["phone"])
    for g in gsheet:
        t = get(g["room"], g["name"], g["phone"])
        t.gsheet += g["amount"]
    for p in app_pays:
        t = get(p["room"], p["name"], p["phone"])
        t.app += p["amount"]
        if p["for_type"] != "rent":
            t.note(f"app row is {p['for_type']} Rs {int(p['amount']):,}")

    # an app UPI payment of the same amount within a few days is strong
    # corroboration when the payer's name is not the tenant's name
    pay_index: dict[float, list[dict]] = defaultdict(list)
    for p in app_pays:
        pay_index[round(p["amount"], 2)].append(p)

    def by_amount_date(x) -> Tenant | None:
        cands = []
        for p in pay_index.get(round(x["amount"], 2), []):
            if p["date"] and abs((p["date"] - x["date"]).days) <= 3:
                owner = get(p["room"], p["name"], p["phone"])
                if owner not in cands:
                    cands.append(owner)
        return cands[0] if len(cands) == 1 else None

    # gateway txns -> tenant: phone, then whole-name score, then amount+date
    for x in gateway:
        t = None
        how = ""
        alias_room = ALIASES.get(x["name"])
        if alias_room:
            cands = [c for c in tenants if c.room == alias_room and (c.app or c.gsheet)]
            if not cands:
                cands = [c for c in tenants if c.room == alias_room]
            if cands:
                t, how = max(cands, key=lambda c: c.app + c.gsheet), "manual alias"

        if t is None and x["phone"] and x["phone"] in by_phone:
            t, how = by_phone[x["phone"]], "phone"

        if t is None and x["name"]:
            scored = sorted(
                ((max((name_score(n, x["name"]) for n in c.names), default=0.0), c)
                 for c in tenants if c.names),
                key=lambda z: -z[0],
            )
            if scored and scored[0][0] >= NAME_FLOOR:
                runner = scored[1][0] if len(scored) > 1 else 0.0
                if scored[0][0] - runner >= NAME_MARGIN:
                    t, how = scored[0][1], f"name {scored[0][0]:.2f}"
                else:
                    tie = [c for sc, c in scored if sc >= scored[0][0] - NAME_MARGIN]
                    # the register holds one row per tenancy, so the same person
                    # can appear twice (re-booking, phone changed). Same room +
                    # same name is one human, not an ambiguity.
                    rooms = {c.room for c in tie}
                    if len(rooms) == 1:
                        t = max(tie, key=lambda c: (c.app + c.gsheet, len(c.display)))
                        how = f"name {scored[0][0]:.2f} (dup register rows)"
                    else:
                        exact = [c for c in tie if round(c.app) == round(x["amount"])
                                 or round(c.gsheet) == round(x["amount"])]
                        if len(exact) == 1:
                            t, how = exact[0], "name+amount"
                        else:
                            # only the tenants who actually have UPI money this
                            # window are plausible owners of a UPI credit
                            live = [c for c in tie if c.app or c.gsheet]
                            if len(live) == 1:
                                t, how = live[0], "name+has-upi"
                            else:
                                x["match_how"] = f"ambiguous ({len(tie)})"

        if t is None:
            t = by_amount_date(x)
            if t is not None:
                how = "amount+date"

        if t is None:
            x["room"], x["tenant"], x["match_how"] = "", "", x.get("match_how", "no match")
            continue
        x["room"], x["tenant"], x["match_how"] = t.room, t.display.title(), how
        t.gateway += x["amount"]
        if x["dup"] == "split across both QRs":
            t.note(f"paid Rs {int(x['amount']):,} to each QR on {x['date']:%d-%b}")
    return tenants


# ------------------------------------------------------------------ writer

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = PatternFill("solid", fgColor="1F3864")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BAND = PatternFill("solid", fgColor="DDEBF7")
INR = '#,##0;[Red]-#,##0;"-"'


def hdr(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write(gateway, tenants):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ---- Tab 1: every transaction, numbered 1..n
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["AUGUST 2026 UPI — ALL GATEWAY TRANSACTIONS (THOR + HULK)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    cols = ["#", "Building", "Date", "Time", "Room", "Tenant", "Payer Name (gateway)",
            "Payer VPA", "Amount", "Status", "Matched By", "Duplicate?"]
    ws.append(cols)
    hdr(ws, 3, len(cols))
    for i, x in enumerate(sorted(gateway, key=lambda z: (z["date"], z["time"])), start=1):
        ws.append([i, x["building"], x["date"], x["time"], x.get("room", ""),
                   x.get("tenant", ""), x["payer"], x["vpa"], x["amount"],
                   x["status"], x.get("match_how", ""), x.get("dup", "")])
        r = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(r, c).border = BORDER
        ws.cell(r, 3).number_format = "dd-mmm"
        ws.cell(r, 9).number_format = INR
        if not x.get("tenant"):
            ws.cell(r, 6).fill = RED
            ws.cell(r, 11).fill = RED
        if x.get("dup"):
            ws.cell(r, 12).fill = AMBER
        if x["status"] != "SUCCESS":
            ws.cell(r, 10).fill = AMBER
    last = ws.max_row
    ws.append([])
    ws.append(["", "", "", "", "", "", "TOTAL", "", f"=SUM(I4:I{last})", "", "", ""])
    for c in range(1, len(cols) + 1):
        ws.cell(ws.max_row, c).font = Font(bold=True)
        ws.cell(ws.max_row, c).fill = BAND
    ws.cell(ws.max_row, 9).number_format = INR
    for i, w in enumerate([5, 10, 10, 7, 8, 26, 30, 30, 12, 10, 14, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{last}"

    # ---- Tab 2: per tenant, 3 columns side by side
    t2 = wb.create_sheet("By Tenant")
    t2.append(["AUGUST 2026 UPI — GATEWAY vs APP vs GOOGLE SHEET"])
    t2["A1"].font = Font(bold=True, size=14)
    t2.append([])
    c2 = ["Room", "Tenant", "UPI Statement", "App", "Google Sheet",
          "Stmt - App", "Stmt - Sheet", "Match?", "Notes"]
    t2.append(c2)
    hdr(t2, 3, len(c2))
    rows = sorted([t for t in tenants if (t.gateway or t.app or t.gsheet)],
                  key=lambda t: (room_key(t.room), t.display))
    for t in rows:
        ok = round(t.gateway) == round(t.app) == round(t.gsheet)
        t2.append([t.room, t.display.title(), t.gateway or None, t.app or None,
                   t.gsheet or None, t.gateway - t.app, t.gateway - t.gsheet,
                   "" if ok else "CHECK", "; ".join(t.notes)])
        r = t2.max_row
        for c in range(1, len(c2) + 1):
            t2.cell(r, c).border = BORDER
        for c in range(3, 8):
            t2.cell(r, c).number_format = INR
        if not ok:
            for c in range(1, len(c2) + 1):
                t2.cell(r, c).fill = RED
    lastb = t2.max_row
    t2.append([])
    t2.append(["", "TOTAL"] + [f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{lastb})"
                               for c in range(3, 8)] + ["", ""])
    for c in range(1, len(c2) + 1):
        t2.cell(t2.max_row, c).font = Font(bold=True)
        t2.cell(t2.max_row, c).fill = BAND
    for c in range(3, 8):
        t2.cell(t2.max_row, c).number_format = INR
    for i, w in enumerate([9, 30, 15, 13, 14, 13, 14, 10, 42], start=1):
        t2.column_dimensions[get_column_letter(i)].width = w
    t2.freeze_panes = "C4"
    t2.auto_filter.ref = f"A3:I{lastb}"

    # ---- Tab 3: summary
    s = wb.create_sheet("Summary")
    s.append(["AUGUST 2026 UPI — SUMMARY"])
    s["A1"].font = Font(bold=True, size=14)
    s.append([])
    kept = gateway
    dups = [x for x in gateway if x["dup"] == "split across both QRs"]
    thor = [x for x in gateway if x["building"] == "THOR"]
    hulk = [x for x in gateway if x["building"] == "HULK"]
    s.append(["Source", "Txns", "Total (Rs)"])
    hdr(s, 3, 3)
    for label, grp in [("Thor UPI.csv", thor), ("Hulk upi.csv", hulk),
                       ("Both files combined", gateway),
                       ("  of which: split across both QRs", dups)]:
        s.append([label, len(grp), sum(x["amount"] for x in grp)])
        s.cell(s.max_row, 3).number_format = INR
    s.append([])
    tot_g = sum(x["amount"] for x in kept)
    tot_a = sum(t.app for t in tenants)
    tot_s = sum(t.gsheet for t in tenants)
    s.append(["Comparison", "", "Total (Rs)"])
    hdr(s, s.max_row, 3)
    for label, v in [("UPI statement (net)", tot_g), ("App", tot_a), ("Google Sheet", tot_s)]:
        s.append([label, "", v])
        s.cell(s.max_row, 3).number_format = INR
    s.append([])
    for label, v in [("Statement - App", tot_g - tot_a),
                     ("Statement - Sheet", tot_g - tot_s),
                     ("App - Sheet", tot_a - tot_s)]:
        s.append([label, "", v])
        s.cell(s.max_row, 3).number_format = INR
        s.cell(s.max_row, 3).font = Font(bold=True)
    s.append([])
    nomatch = [x for x in gateway if not x.get("tenant")]
    s.append(["Unmatched gateway txns (no tenant found)", len(nomatch),
              sum(x["amount"] for x in nomatch)])
    s.cell(s.max_row, 3).number_format = INR
    s.cell(s.max_row, 1).fill = RED
    for i, w in enumerate([40, 12, 18], start=1):
        s.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    return rows, kept, dups, nomatch


async def main():
    gateway = load_gateway()
    gsheet = load_gsheet()
    win_from = min(x["date"] for x in gateway)
    win_to = max(x["date"] for x in gateway)
    app_pays, register = await load_app(win_from, win_to)
    tenants = build(gateway, app_pays, register, gsheet)
    rows, kept, dups, nomatch = write(gateway, tenants)

    tot_g = sum(x["amount"] for x in kept)
    tot_a = sum(t.app for t in tenants)
    tot_s = sum(t.gsheet for t in tenants)
    print(f"window        : {win_from} -> {win_to}  (gateway coverage, part month)")
    print(f"gateway raw   : {len(gateway):>4} txns  Rs {sum(x['amount'] for x in gateway):>12,.0f}")
    print(f"  duplicates  : {len(dups):>4} txns  Rs {sum(x['amount'] for x in dups):>12,.0f}")
    print(f"gateway net   : {len(kept):>4} txns  Rs {tot_g:>12,.0f}")
    print(f"app (upi)     : {len(app_pays):>4} rows  Rs {tot_a:>12,.0f}")
    print(f"google sheet  :        Rs {tot_s:>12,.0f}")
    print()
    print(f"statement - app   : Rs {tot_g - tot_a:>12,.0f}")
    print(f"statement - sheet : Rs {tot_g - tot_s:>12,.0f}")
    print(f"app - sheet       : Rs {tot_a - tot_s:>12,.0f}")
    print(f"\nunmatched gateway txns: {len(nomatch)}  Rs {sum(x['amount'] for x in nomatch):,.0f}")
    print(f"tenants with a CHECK  : {sum(1 for t in rows if not (round(t.gateway)==round(t.app)==round(t.gsheet)))} of {len(rows)}")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
