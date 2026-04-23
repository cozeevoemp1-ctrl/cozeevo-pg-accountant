"""Gap report to Excel: per-tenant comparison of bot's Pending vs source's April Balance.

Output: data/reports/april_pending_gaps.xlsx
Sheets:
  1. Summary — totals
  2. Gap rows — every tenant where OurBalance != SourceBalance
  3. Match rows — where both agree (audit trail)
"""
import asyncio
import os
import sys
from datetime import date
from decimal import Decimal

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv()

import gspread
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from google.oauth2.service_account import Credentials
from sqlalchemy import select
from src.database.db_manager import init_engine, get_session
from src.database.models import (
    Payment, PaymentFor, RentSchedule, Tenancy, Tenant, Room,
)


def pn(v):
    if not v:
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


async def main():
    # ── Read source sheet ──
    creds = Credentials.from_service_account_file(
        "credentials/gsheets_service_account.json",
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key("1Vr_fSIOuuKBK4MWF-FVqgAIUPbqun3POszaXYfj-Ea0")
    rows = sh.worksheet("Long term").get_all_values()[1:]
    src = {}
    for r in rows:
        if len(r) < 24:
            continue
        room, name = (r[0] or "").strip(), (r[1] or "").strip()
        if not name:
            continue
        src[(room, name.lower())] = {
            "cash": pn(r[21]),
            "upi": pn(r[22]),
            "balance": pn(r[23]),
            "comment": (r[14] or "").strip(),
            "inout": (r[16] or "").strip(),
        }

    # ── Read DB ──
    init_engine(os.environ["DATABASE_URL"])
    apr = date(2026, 4, 1)
    may = date(2026, 5, 1)

    async with get_session() as s:
        rs_rows = (await s.execute(
            select(RentSchedule).where(RentSchedule.period_month == apr)
        )).scalars().all()

        # Per-tenancy Apr cash/UPI (only payments dated in April — display)
        pay_rows = (await s.execute(
            select(Payment).where(
                Payment.period_month == apr,
                Payment.is_void == False,
                Payment.for_type == PaymentFor.rent,
            )
        )).scalars().all()
        pay_map = {}
        prepaid_map = {}
        for p in pay_rows:
            in_apr = p.payment_date and apr <= p.payment_date < may
            mode = p.payment_mode.value if hasattr(p.payment_mode, "value") else str(p.payment_mode or "cash")
            if in_apr:
                if mode.lower() == "cash":
                    pay_map.setdefault(p.tenancy_id, {"cash": 0, "upi": 0})["cash"] += float(p.amount)
                else:
                    pay_map.setdefault(p.tenancy_id, {"cash": 0, "upi": 0})["upi"] += float(p.amount)
            else:
                prepaid_map[p.tenancy_id] = prepaid_map.get(p.tenancy_id, 0) + float(p.amount)

        # Booking + deposit maps
        book_rows = (await s.execute(
            select(Payment).where(
                Payment.is_void == False, Payment.for_type == PaymentFor.booking
            )
        )).scalars().all()
        bk_map = {}
        for p in book_rows:
            bk_map[p.tenancy_id] = bk_map.get(p.tenancy_id, 0) + float(p.amount)

        dep_rows = (await s.execute(
            select(Payment).where(
                Payment.is_void == False, Payment.for_type == PaymentFor.deposit
            )
        )).scalars().all()
        dep_map = {}
        for p in dep_rows:
            dep_map[p.tenancy_id] = dep_map.get(p.tenancy_id, 0) + float(p.amount)

        data = []
        for rs in rs_rows:
            t = await s.get(Tenancy, rs.tenancy_id)
            if not t:
                continue
            tn = await s.get(Tenant, t.tenant_id)
            rm = await s.get(Room, t.room_id)
            is_fm = bool(t.checkin_date and t.checkin_date.replace(day=1) == apr)

            our_due = float(rs.rent_due or 0)
            cash = pay_map.get(t.id, {}).get("cash", 0)
            upi = pay_map.get(t.id, {}).get("upi", 0)
            prepaid = prepaid_map.get(t.id, 0)
            booking = bk_map.get(t.id, 0) if is_fm else 0
            deposit_cr = dep_map.get(t.id, 0) if is_fm else 0
            our_paid = cash + upi + prepaid + booking + deposit_cr
            our_bal = max(0, our_due - our_paid)

            key = (rm.room_number, tn.name.lower())
            s_row = src.get(key, {})
            s_cash = s_row.get("cash", 0)
            s_upi = s_row.get("upi", 0)
            s_bal = s_row.get("balance", 0)
            s_due = s_cash + s_upi + s_bal
            comment = s_row.get("comment", "")
            inout = s_row.get("inout", "")

            gap = our_bal - s_bal
            data.append({
                "room": rm.room_number,
                "name": tn.name,
                "inout": inout,
                "fm": "FM" if is_fm else "",
                "our_due": int(our_due),
                "cash_db": int(cash),
                "upi_db": int(upi),
                "prepaid_db": int(prepaid),
                "booking_db": int(booking),
                "deposit_db": int(deposit_cr),
                "our_paid": int(our_paid),
                "our_balance": int(our_bal),
                "src_cash": int(s_cash),
                "src_upi": int(s_upi),
                "src_balance": int(s_bal),
                "src_due_implied": int(s_due),
                "gap_our_minus_src": int(gap),
                "comment": comment,
            })

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F8F9FA")
    red_fill = PatternFill("solid", fgColor="FFE5E5")
    green_fill = PatternFill("solid", fgColor="E5FFE5")

    headers = [
        "Room", "Name", "In/Out", "FM", "Our Due",
        "DB Cash (Apr)", "DB UPI (Apr)", "DB Prepaid", "DB Booking", "DB Deposit",
        "Our Paid", "Our Balance",
        "Src Cash", "Src UPI", "Src Balance", "Src Implied Due",
        "Gap (Ours - Src)", "Comment",
    ]

    def write_sheet(ws, rows):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for ri, d in enumerate(rows, 2):
            vals = [
                d["room"], d["name"], d["inout"], d["fm"], d["our_due"],
                d["cash_db"], d["upi_db"], d["prepaid_db"], d["booking_db"], d["deposit_db"],
                d["our_paid"], d["our_balance"],
                d["src_cash"], d["src_upi"], d["src_balance"], d["src_due_implied"],
                d["gap_our_minus_src"], d["comment"],
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                if ci >= 5 and ci <= 17 and isinstance(v, int):
                    c.number_format = '#,##,##,##0;(#,##,##,##0);"-"'
                if ri % 2 == 0:
                    c.fill = alt_fill
            # highlight gap column
            gap_col = 17
            gc = ws.cell(row=ri, column=gap_col)
            if d["gap_our_minus_src"] > 0:
                gc.fill = red_fill
            elif d["gap_our_minus_src"] < 0:
                gc.fill = green_fill
        # column widths
        widths = [8, 28, 12, 5, 11, 12, 12, 11, 11, 11, 11, 12, 11, 11, 12, 14, 14, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "C2"

    # Sheet 1: Summary
    ws1 = wb.create_sheet("Summary")
    our_bal_total = sum(d["our_balance"] for d in data)
    src_bal_total = sum(d["src_balance"] for d in data)
    our_due_total = sum(d["our_due"] for d in data)
    src_due_total = sum(d["src_due_implied"] for d in data)
    with_gap = [d for d in data if abs(d["gap_our_minus_src"]) >= 1]

    rows_summary = [
        ["Metric", "Value"],
        ["Total tenancies", len(data)],
        ["", ""],
        ["OUR Rent Due (April)", our_due_total],
        ["OUR Balance (what we say pending)", our_bal_total],
        ["", ""],
        ["SOURCE Implied Due (cash+upi+balance)", src_due_total],
        ["SOURCE Balance (what source sheet says pending)", src_bal_total],
        ["", ""],
        ["Our Balance - Src Balance (net)", our_bal_total - src_bal_total],
        ["Tenants with any gap", len(with_gap)],
    ]
    for ri, (k, v) in enumerate(rows_summary, 1):
        c1 = ws1.cell(row=ri, column=1, value=k)
        c2 = ws1.cell(row=ri, column=2, value=v)
        if ri == 1:
            c1.fill = hdr_fill; c1.font = hdr_font
            c2.fill = hdr_fill; c2.font = hdr_font
        if isinstance(v, int):
            c2.number_format = '#,##,##,##0;(#,##,##,##0);"-"'
    ws1.column_dimensions['A'].width = 50
    ws1.column_dimensions['B'].width = 20

    # Sheet 2: Gap rows
    ws2 = wb.create_sheet("Gap rows")
    write_sheet(ws2, sorted(with_gap, key=lambda d: -abs(d["gap_our_minus_src"])))

    # Sheet 3: Match rows (audit)
    ws3 = wb.create_sheet("Match rows")
    matches = [d for d in data if abs(d["gap_our_minus_src"]) < 1]
    write_sheet(ws3, matches)

    out = os.path.join("data", "reports", "april_pending_gaps.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Total tenancies: {len(data)}")
    print(f"With gaps:       {len(with_gap)}")
    print(f"Our balance sum: Rs.{our_bal_total:,}")
    print(f"Src balance sum: Rs.{src_bal_total:,}")
    print(f"Net gap:         Rs.{our_bal_total - src_bal_total:+,}")


asyncio.run(main())
