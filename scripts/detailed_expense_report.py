"""
Deep expense classifier + full report for the YES Bank statement.
Reads the already-extracted Excel and produces a categorised Excel + console summary.
"""
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "Statement-124563400000961-03-10-2026-20-15-08 (1)_extracted.xlsx"
OUT = "expense_report_detailed.xlsx"

# ── Detailed category rules (checked in order — first match wins) ─────────────
# Each rule: (category, sub_category, list_of_keywords_in_description)
RULES = [
    # ── INCOME / INWARD (deposits — label for reference) ──
    ("INCOME", "Rent Collection (UPI Settlement)",  ["upi collection settlement", "115063600001082"]),
    ("INCOME", "NEFT / Bank Transfer In",           ["neft", "rtgs", "imps"]),

    # ── PROPERTY RENT (biggest outflow) ──
    ("Property Rent", "Monthly Rent - Vakkal Sravani",      ["vakkal", "sravani"]),
    ("Property Rent", "Monthly Rent - R Suma",               ["r suma", "rsuma"]),
    ("Property Rent", "Monthly Rent - Raghu Nandha",         ["raghu nandha", "raghunandha"]),
    ("Property Rent", "Monthly Rent - Chandrasekhar",        ["chandrasekhar"]),
    ("Property Rent", "Monthly Rent - Bharathi",             ["bharathi"]),
    ("Property Rent", "DG Rent",                             ["dg rent"]),
    ("Property Rent", "Other Property Rent",                 ["jan rent", "feb rent", "mar rent", "monthly rent", "property rent"]),

    # ── ELECTRICITY ──
    ("Electricity", "BESCOM Bill",   ["bescom", "besco"]),

    # ── INTERNET / WIFI ──
    ("Internet & WiFi", "Airwire Broadband",  ["airwire", "airwirebroadband"]),
    ("Internet & WiFi", "WiFi Vendor",        ["wifi", "wi-fi", "broadband", "internet"]),

    # ── FURNITURE & FITTINGS ──
    ("Furniture & Fittings", "Wakefit - Mattresses",   ["wakefit"]),
    ("Furniture & Fittings", "Bedsheets / Linen",      ["bedsheet", "bed sheet", "linen"]),
    ("Furniture & Fittings", "Shoe Racks / Storage",   ["shoe rack", "rack"]),
    ("Furniture & Fittings", "Curtains",               ["curtain"]),
    ("Furniture & Fittings", "Bedframes",              ["bedframe", "bed frame", "grace trader"]),
    ("Furniture & Fittings", "Other Furniture",        ["furniture", "sofa", "chair", "table", "almirah", "refurbish"]),

    # ── FOOD & GROCERIES ──
    ("Food & Groceries", "Grocery Vendor (Virani Trading)", ["virani"]),
    ("Food & Groceries", "Food Supplies (Vyapar)",          ["vyapar"]),
    ("Food & Groceries", "Milk / Daily Supplies",           ["milk", "dairy"]),
    ("Food & Groceries", "Grocery / Kirana",                ["grocer", "kirana", "grocery", "grocery"]),
    ("Food & Groceries", "Restaurant / Canteen",            ["restaurant", "canteen", "food"]),
    ("Food & Groceries", "Gas Cylinders",                   ["cylinder", "lpg", "gas cylinder"]),

    # ── FUEL ──
    ("Fuel & Diesel", "Diesel (9888751222 vendor)", ["9888751222", "diesel"]),
    ("Fuel & Diesel", "Petrol / Fuel",              ["petrol", "fuel", "hp petrol", "indian oil"]),

    # ── STAFF SALARY ──
    ("Staff & Labour", "Salary - Arjun",            ["arjun"]),
    ("Staff & Labour", "Salary - Phiros / Phirose", ["phiros", "phirose"]),
    ("Staff & Labour", "Salary - Lokesh",           ["lokesh"]),
    ("Staff & Labour", "Salary - Housekeeping",     ["housekeep", "housekeeper", "cleaning staff"]),
    ("Staff & Labour", "Salary - Other",            ["salary", "neft-yesob", "net-neft-yes", "joshi arjun"]),
    ("Staff & Labour", "Labour / Helper Payment",   ["helper", "labour", "labor", "worker", "kshitij"]),
    ("Staff & Labour", "Cleaning Service",          ["cleaning", "clean"]),

    # ── MAINTENANCE & REPAIRS ──
    ("Maintenance & Repairs", "Plumber / Electrician",  ["electrician", "plumber", "plumbing"]),
    ("Maintenance & Repairs", "Pest Control",            ["pest"]),
    ("Maintenance & Repairs", "Repairs / Handyman",      ["repair", "handyman", "fix", "broken"]),
    ("Maintenance & Repairs", "Painting",                ["paint"]),
    ("Maintenance & Repairs", "General Maintenance",     ["maintenance", "maintain"]),

    # ── GOVT / REGULATORY ──
    ("Govt & Regulatory", "BBMP Tax / Charges",    ["bbmp"]),
    ("Govt & Regulatory", "Govt Registration Fee", ["edcs", "directorate", "registration"]),
    ("Govt & Regulatory", "GST / Tax",             ["gst", "tax"]),

    # ── DEPOSIT REFUND TO TENANT ──
    ("Tenant Deposit Refund", "Security Deposit Refund", ["deposit refund", "refund deposit", "security deposit", "sd refund", "deposit return"]),

    # ── MARKETING ──
    ("Marketing", "Marketing / Ads", ["marketing", "advertisement", "ads", "promotion"]),

    # ── AMAZON / SHOPPING ──
    ("Shopping & Supplies", "Amazon Purchases",   ["amazon"]),
    ("Shopping & Supplies", "Flipkart / Meesho",  ["flipkart", "meesho", "myntra"]),
    ("Shopping & Supplies", "3D Boards / Signage",["3d bo", "signage", "sign board"]),
    ("Shopping & Supplies", "Other Supplies",     ["supply", "supplies"]),

    # ── TENANT / PERSON TRANSFERS ──
    ("Person Transfers", "Sameer / Staff Transfer",    ["sameer"]),
    ("Person Transfers", "Kshitij Departure",          ["kshitij"]),
    ("Person Transfers", "Biplab Payment",             ["biplab"]),
    ("Person Transfers", "Sanket Payment",             ["sanket"]),
    ("Person Transfers", "Manisha Payment",            ["manisha"]),
    ("Person Transfers", "Person Transfer (Other)",    ["payment from phone", "payment for", "payment to", "transfer"]),

    # ── BANK / FINANCIAL ──
    ("Bank & Financial", "RTGS Transfer Out",    ["rtgs"]),
    ("Bank & Financial", "NEFT Transfer Out",    ["neft"]),
    ("Bank & Financial", "IMPS Transfer Out",    ["imps"]),
    ("Bank & Financial", "Usha Trading (TV/Equipment)", ["usha trading", "usha t rading"]),
]


def classify(desc: str) -> tuple[str, str]:
    if not isinstance(desc, str):
        return ("Uncategorised", "Unknown")
    d = desc.lower()
    for cat, sub, keywords in RULES:
        if cat in ("INCOME",):
            continue
        for kw in keywords:
            if kw in d:
                return (cat, sub)
    return ("Uncategorised", "Review Needed")


def classify_deposit(desc: str) -> tuple[str, str]:
    if not isinstance(desc, str):
        return ("Other Income", "Unknown")
    d = desc.lower()
    for cat, sub, keywords in RULES:
        if cat != "INCOME":
            continue
        for kw in keywords:
            if kw in d:
                return (cat, sub)
    return ("Other Income", "UPI / Direct Transfer")


# ── colour map ────────────────────────────────────────────────────────────────
CAT_COLORS = {
    "Property Rent":       "FFD7D7",
    "Electricity":         "FFE4B5",
    "Internet & WiFi":     "B0E0E6",
    "Furniture & Fittings":"E0D0F0",
    "Food & Groceries":    "C8EFC8",
    "Fuel & Diesel":       "F5CBA7",
    "Staff & Labour":      "D6EAF8",
    "Maintenance & Repairs":"FDEBD0",
    "Govt & Regulatory":   "D5DBDB",
    "Tenant Deposit Refund":"FADBD8",
    "Marketing":           "FDFFD6",
    "Shopping & Supplies": "E8DAEF",
    "Person Transfers":    "D0ECE7",
    "Bank & Financial":    "EAECEE",
    "Uncategorised":       "FFFFFF",
}
THIN   = Side(style="thin",   color="CCCCCC")
MED    = Side(style="medium", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_BG = "1F497D"


def hcell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c


def dcell(ws, row, col, val, bg, bold=False, color="000000", align="left", num_fmt=None):
    c = ws.cell(row, col, val)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.border    = BORDER
    c.font      = Font(name="Calibri", size=9, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "left"))
    if num_fmt:
        c.number_format = num_fmt
    return c


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── main ─────────────────────────────────────────────────────────────────────
def main():
    df = pd.read_excel(SRC)
    df['Withdrawals'] = df['Withdrawals'].apply(lambda x: x if isinstance(x, float) else 0)
    df['Deposits']    = df['Deposits'].apply(lambda x: x if isinstance(x, float) else 0)
    df['_date']       = pd.to_datetime(df['Transaction Date'], format='%Y-%m-%d', errors='coerce')
    df['Month']       = df['_date'].dt.strftime('%b %Y')

    # Classify expenses
    df['Expense_Cat'] = ""
    df['Expense_Sub'] = ""
    mask_wd = df['Withdrawals'] > 0
    df.loc[mask_wd, ['Expense_Cat', 'Expense_Sub']] = [
        list(classify(d)) for d in df.loc[mask_wd, 'Description']
    ]

    # Classify deposits
    mask_dep = df['Deposits'] > 0
    df.loc[mask_dep, ['Expense_Cat', 'Expense_Sub']] = [
        list(classify_deposit(d)) for d in df.loc[mask_dep, 'Description']
    ]

    wb = Workbook()
    wb.remove(wb.active)

    # ── SHEET 1: Full Expense Detail (withdrawals only) ───────────────────────
    ws1 = wb.create_sheet("All Expenses Detail")
    exp = df[df['Withdrawals'] > 0].copy().sort_values('_date').reset_index(drop=True)

    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value     = "Full Expense Detail — YES Bank (Dec 2025 – Mar 2026)"
    t.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 22

    for c, h in enumerate(["#", "Date", "Amount (Rs)", "Category", "Sub-Category", "Description", "Ref / UTR", "Month"], 1):
        hcell(ws1, 2, c, h)

    for i, (_, row) in enumerate(exp.iterrows(), 1):
        r    = i + 2
        cat  = row['Expense_Cat']
        sub  = row['Expense_Sub']
        bg   = CAT_COLORS.get(cat, "FFFFFF")
        dcell(ws1, r, 1, i,                              bg, align="center")
        dcell(ws1, r, 2, row['Transaction Date'],        bg)
        dcell(ws1, r, 3, row['Withdrawals'],             bg, align="right", num_fmt="#,##0.00", color="CC0000")
        dcell(ws1, r, 4, cat,                            bg, bold=True)
        dcell(ws1, r, 5, sub,                            bg)
        dcell(ws1, r, 6, str(row['Description'])[:120], bg)
        dcell(ws1, r, 7, str(row.get('utr',''))[:30],   bg)
        dcell(ws1, r, 8, row['Month'],                   bg, align="center")

    # Total row
    tr = len(exp) + 3
    ws1.cell(tr, 4, "TOTAL").font = Font(bold=True, name="Calibri")
    c = ws1.cell(tr, 3, exp['Withdrawals'].sum())
    c.number_format = "#,##0.00"; c.font = Font(bold=True, name="Calibri", color="CC0000")
    c.alignment = Alignment(horizontal="right")
    for col in range(1, 9):
        ws1.cell(tr, col).fill   = PatternFill("solid", fgColor="D9D9D9")
        ws1.cell(tr, col).border = BORDER

    set_widths(ws1, [4, 12, 14, 22, 30, 55, 22, 10])
    ws1.freeze_panes = "A3"

    # ── SHEET 2: Category Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Category Summary")

    cat_sum = exp.groupby(['Expense_Cat', 'Expense_Sub']).agg(
        Count=('Withdrawals', 'count'),
        Total=('Withdrawals', 'sum'),
    ).reset_index().sort_values(['Expense_Cat', 'Total'], ascending=[True, False])

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "Expense Category Summary"
    t2.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t2.fill      = PatternFill("solid", fgColor=HDR_BG)
    t2.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 22

    for c, h in enumerate(["Category", "Sub-Category", "Txns", "Total (Rs)", "% of Spend"], 1):
        hcell(ws2, 2, c, h)

    total_spend = exp['Withdrawals'].sum()
    prev_cat = None
    ri = 3
    cat_totals = exp.groupby('Expense_Cat')['Withdrawals'].sum().sort_values(ascending=False)

    for cat, grp in cat_sum.groupby('Expense_Cat', sort=False):
        bg = CAT_COLORS.get(cat, "FFFFFF")
        cat_total = grp['Total'].sum()
        pct_cat   = cat_total / total_spend * 100

        # Category header row
        for col in range(1, 6):
            ws2.cell(ri, col).fill   = PatternFill("solid", fgColor=bg)
            ws2.cell(ri, col).border = BORDER
        ws2.cell(ri, 1, cat).font = Font(name="Calibri", bold=True, size=10)
        ws2.cell(ri, 1).fill = PatternFill("solid", fgColor=bg)
        ws2.cell(ri, 1).border = BORDER
        ws2.cell(ri, 4, cat_total).number_format = "#,##0.00"
        ws2.cell(ri, 4).font = Font(name="Calibri", bold=True)
        ws2.cell(ri, 4).alignment = Alignment(horizontal="right")
        ws2.cell(ri, 4).fill = PatternFill("solid", fgColor=bg); ws2.cell(ri, 4).border = BORDER
        ws2.cell(ri, 5, f"{pct_cat:.1f}%").font = Font(name="Calibri", bold=True)
        ws2.cell(ri, 5).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 5).fill = PatternFill("solid", fgColor=bg); ws2.cell(ri, 5).border = BORDER
        ws2.cell(ri, 3, int(grp['Count'].sum())).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 3).fill = PatternFill("solid", fgColor=bg); ws2.cell(ri, 3).border = BORDER
        ri += 1

        # Sub-category rows
        for _, row in grp.iterrows():
            pct = row['Total'] / total_spend * 100
            dcell(ws2, ri, 1, "",                     "FAFAFA")
            dcell(ws2, ri, 2, row['Expense_Sub'],     "FAFAFA")
            dcell(ws2, ri, 3, int(row['Count']),      "FAFAFA", align="center")
            dcell(ws2, ri, 4, row['Total'],            "FAFAFA", align="right", num_fmt="#,##0.00")
            dcell(ws2, ri, 5, f"{pct:.1f}%",          "FAFAFA", align="center")
            ri += 1

    # Grand total
    dcell(ws2, ri, 1, "GRAND TOTAL", "D9D9D9", bold=True)
    dcell(ws2, ri, 2, "",             "D9D9D9")
    dcell(ws2, ri, 3, len(exp),       "D9D9D9", bold=True, align="center")
    dcell(ws2, ri, 4, total_spend,    "D9D9D9", bold=True, align="right", num_fmt="#,##0.00", color="CC0000")
    dcell(ws2, ri, 5, "100.0%",       "D9D9D9", bold=True, align="center")

    set_widths(ws2, [24, 36, 8, 16, 10])
    ws2.freeze_panes = "A3"

    # ── SHEET 3: Monthly x Category Pivot ────────────────────────────────────
    ws3 = wb.create_sheet("Monthly x Category")

    months = ['Dec 2025', 'Jan 2026', 'Feb 2026', 'Mar 2026']
    exp['Month'] = pd.Categorical(exp['Month'], categories=months, ordered=True)
    pivot = exp.groupby(['Expense_Cat', 'Month'], observed=True)['Withdrawals'].sum().unstack(fill_value=0).reset_index()
    pivot['Total'] = pivot[months].sum(axis=1)
    pivot = pivot.sort_values('Total', ascending=False)

    ws3.merge_cells(f"A1:{get_column_letter(len(pivot.columns))}1")
    t3 = ws3["A1"]
    t3.value     = "Monthly Spend by Category"
    t3.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t3.fill      = PatternFill("solid", fgColor=HDR_BG)
    t3.alignment = Alignment(horizontal="center")

    for c, h in enumerate(["Category"] + months + ["Total"], 1):
        hcell(ws3, 2, c, h)

    for ri2, (_, row) in enumerate(pivot.iterrows(), 3):
        cat = row['Expense_Cat']
        bg  = CAT_COLORS.get(cat, "FFFFFF")
        dcell(ws3, ri2, 1, cat, bg, bold=True)
        for ci2, m in enumerate(months, 2):
            val = row.get(m, 0)
            dcell(ws3, ri2, ci2, val if val > 0 else None, bg, align="right", num_fmt="#,##0")
        dcell(ws3, ri2, len(months)+2, row['Total'], bg, bold=True, align="right", num_fmt="#,##0")

    # Grand total row
    tr3 = len(pivot) + 3
    dcell(ws3, tr3, 1, "TOTAL", "D9D9D9", bold=True)
    for ci2, m in enumerate(months, 2):
        val = pivot[m].sum()
        dcell(ws3, tr3, ci2, val, "D9D9D9", bold=True, align="right", num_fmt="#,##0")
    dcell(ws3, tr3, len(months)+2, pivot['Total'].sum(), "D9D9D9", bold=True, align="right", num_fmt="#,##0", color="CC0000")

    set_widths(ws3, [24] + [14]*len(months) + [14])
    ws3.freeze_panes = "A3"

    # ── SHEET 4: Expenses > 10K flagged for review ───────────────────────────
    ws4 = wb.create_sheet("Big Expenses (>10K)")
    big = exp[exp['Withdrawals'] >= 10000].sort_values('Withdrawals', ascending=False).reset_index(drop=True)

    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value     = f"All Expenses Above Rs 10,000  ({len(big)} transactions — Rs {big['Withdrawals'].sum():,.0f} total)"
    t4.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    t4.fill      = PatternFill("solid", fgColor=HDR_BG)
    t4.alignment = Alignment(horizontal="center")

    for c, h in enumerate(["#", "Date", "Amount (Rs)", "Category", "Sub-Category", "Description", "Month"], 1):
        hcell(ws4, 2, c, h)

    for i, (_, row) in enumerate(big.iterrows(), 1):
        r   = i + 2
        cat = row['Expense_Cat']
        bg  = CAT_COLORS.get(cat, "FFFFFF")
        dcell(ws4, r, 1, i,                              bg, align="center")
        dcell(ws4, r, 2, row['Transaction Date'],        bg)
        dcell(ws4, r, 3, row['Withdrawals'],             bg, bold=True, align="right", num_fmt="#,##0.00", color="CC0000")
        dcell(ws4, r, 4, cat,                            bg, bold=True)
        dcell(ws4, r, 5, row['Expense_Sub'],             bg)
        dcell(ws4, r, 6, str(row['Description'])[:100], bg)
        dcell(ws4, r, 7, row['Month'],                   bg, align="center")

    set_widths(ws4, [4, 12, 14, 22, 30, 60, 10])
    ws4.freeze_panes = "A3"

    wb.save(OUT)

    # Console summary
    print("=" * 60)
    print("  EXPENSE REPORT GENERATED")
    print("=" * 60)
    print(f"  File  : {OUT}")
    print(f"  Rows  : {len(exp)} expense transactions")
    print(f"  Total : Rs {exp['Withdrawals'].sum():,.0f}")
    print()
    print("  TOP CATEGORIES:")
    cat_s = exp.groupby('Expense_Cat')['Withdrawals'].sum().sort_values(ascending=False)
    for cat, amt in cat_s.items():
        pct = amt / exp['Withdrawals'].sum() * 100
        print(f"    {str(cat):<26} Rs {amt:>10,.0f}  ({pct:4.1f}%)")
    print()
    print("  Sheets: All Expenses Detail | Category Summary | Monthly x Category | Big Expenses >10K")
    print("=" * 60)


if __name__ == "__main__":
    main()
