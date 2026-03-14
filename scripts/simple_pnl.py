"""Simple plain P&L summary — no colours, no borders, just data."""
import sys
sys.path.insert(0, '.')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ── reuse rules from pnl_report ───────────────────────────────────────────────
from scripts.pnl_report import EXPENSE_RULES, INCOME_RULES, SRC, MONTHS

def classify(desc, rules):
    d = (desc or "").lower()
    for cat, sub, keywords in rules:
        if not keywords:
            continue
        for kw in keywords:
            if kw in d:
                return cat, sub
    return rules[-1][0], rules[-1][1]

# ── load & classify ───────────────────────────────────────────────────────────
df = pd.read_excel(SRC)
df['Withdrawals'] = df['Withdrawals'].apply(lambda x: x if isinstance(x, float) else 0)
df['Deposits']    = df['Deposits'].apply(lambda x: x if isinstance(x, float) else 0)
df['_date']       = pd.to_datetime(df['Transaction Date'], format='%Y-%m-%d', errors='coerce')
df['Month']       = pd.Categorical(df['_date'].dt.strftime('%b %Y'), categories=MONTHS, ordered=True)

exp_df = df[df['Withdrawals'] > 0].copy()
exp_df[['Cat','Sub']] = pd.DataFrame(
    [classify(d, EXPENSE_RULES) for d in exp_df['Description']], index=exp_df.index)

inc_df = df[df['Deposits'] > 0].copy()
inc_df[['Cat','Sub']] = pd.DataFrame(
    [classify(d, INCOME_RULES) for d in inc_df['Description']], index=inc_df.index)

# ── build pivots ──────────────────────────────────────────────────────────────
inc_pivot = inc_df.pivot_table(index='Cat', columns='Month', values='Deposits',
                                aggfunc='sum', fill_value=0, observed=True)
inc_pivot.columns = [str(c) for c in inc_pivot.columns]
inc_pivot['TOTAL'] = inc_pivot.sum(axis=1)

exp_pivot = exp_df.pivot_table(index='Cat', columns='Month', values='Withdrawals',
                                aggfunc='sum', fill_value=0, observed=True)
exp_pivot.columns = [str(c) for c in exp_pivot.columns]
exp_pivot['TOTAL'] = exp_pivot.sum(axis=1)

# ── write Excel ───────────────────────────────────────────────────────────────
OUT = "Simple_PnL.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "P&L Summary"

cols = MONTHS + ['TOTAL']

def bold(cell):
    cell.font = Font(bold=True)

# ── INCOME ────────────────────────────────────────────────────────────────────
r = 1
ws.cell(r, 1, "INCOME").font = Font(bold=True, size=12)
r += 1

# header
ws.cell(r, 1, "Category")
for ci, m in enumerate(cols, 2):
    ws.cell(r, ci, m)
bold(ws.cell(r, 1))
for ci in range(2, len(cols)+2):
    bold(ws.cell(r, ci))
r += 1

inc_totals = {m: 0.0 for m in cols}
for cat in inc_pivot.sort_values('TOTAL', ascending=False).index:
    ws.cell(r, 1, str(cat))
    for ci, m in enumerate(cols, 2):
        v = inc_pivot.loc[cat, m] if m in inc_pivot.columns else 0
        ws.cell(r, ci, round(v))
        inc_totals[m] += v
    r += 1

# total row
ws.cell(r, 1, "TOTAL INCOME")
bold(ws.cell(r, 1))
for ci, m in enumerate(cols, 2):
    ws.cell(r, ci, round(inc_totals[m]))
    bold(ws.cell(r, ci))
r += 2

# ── EXPENSES ──────────────────────────────────────────────────────────────────
ws.cell(r, 1, "EXPENSES").font = Font(bold=True, size=12)
r += 1

ws.cell(r, 1, "Category")
for ci, m in enumerate(cols, 2):
    ws.cell(r, ci, m)
bold(ws.cell(r, 1))
for ci in range(2, len(cols)+2):
    bold(ws.cell(r, ci))
r += 1

exp_totals = {m: 0.0 for m in cols}
for cat in exp_pivot.sort_values('TOTAL', ascending=False).index:
    ws.cell(r, 1, str(cat))
    for ci, m in enumerate(cols, 2):
        v = exp_pivot.loc[cat, m] if m in exp_pivot.columns else 0
        ws.cell(r, ci, round(v))
        exp_totals[m] += v
    r += 1

# total row
ws.cell(r, 1, "TOTAL EXPENSES")
bold(ws.cell(r, 1))
for ci, m in enumerate(cols, 2):
    ws.cell(r, ci, round(exp_totals[m]))
    bold(ws.cell(r, ci))
r += 2

# ── NET ───────────────────────────────────────────────────────────────────────
ws.cell(r, 1, "NET PROFIT / (LOSS)")
bold(ws.cell(r, 1))
for ci, m in enumerate(cols, 2):
    net = round(inc_totals[m] - exp_totals[m])
    ws.cell(r, ci, net)
    bold(ws.cell(r, ci))

# column widths
ws.column_dimensions['A'].width = 26
for ci in range(2, len(cols)+2):
    ws.column_dimensions[get_column_letter(ci)].width = 14

# ── SHEET 2: Tenant Deposit Refunds detail ────────────────────────────────────
ws2 = wb.create_sheet("Deposit Refunds")
ws2.cell(1, 1, "Tenant Deposit Refunds — All Transactions").font = Font(bold=True, size=12)
headers = ["Date", "Amount (Rs)", "Sub-Category", "Description"]
for ci, h in enumerate(headers, 1):
    ws2.cell(2, ci, h).font = Font(bold=True)

refund_df = exp_df[exp_df['Cat'] == 'Tenant Deposit Refund'].sort_values('_date').reset_index(drop=True)
for i, (_, row_data) in enumerate(refund_df.iterrows(), 3):
    ws2.cell(i, 1, str(row_data['Transaction Date']))
    ws2.cell(i, 2, round(row_data['Withdrawals']))
    ws2.cell(i, 3, row_data['Sub'])
    ws2.cell(i, 4, str(row_data['Description'])[:100])

tr = len(refund_df) + 3
ws2.cell(tr, 1, "TOTAL").font = Font(bold=True)
ws2.cell(tr, 2, round(refund_df['Withdrawals'].sum())).font = Font(bold=True)

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 60

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Tenant Deposit Refunds: Rs {refund_df['Withdrawals'].sum():,.0f} ({len(refund_df)} rows)")
