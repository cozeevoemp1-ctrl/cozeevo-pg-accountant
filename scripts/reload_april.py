"""
scripts/reload_april.py
========================
Full reload of APRIL 2026 Google Sheet tab from:
  - Excel (Cozeevo Monthly stay (4).xlsx) — permanent comments, staff
  - March 2026 Sheet — prev balance, prev notes
  - Current April Sheet — preserve existing payments
  - DB — tenancy notes

Output: 17 columns (A-Q) new format:
  Room, Name, Phone, Building, Sharing, Rent Due,
  Cash, UPI, Total Paid, Balance, Status,
  Check-in, Notice Date, Event, Notes, Prev Due, Entered By

Usage:
    python scripts/reload_april.py              # dry run
    python scripts/reload_april.py --write      # actually write to Sheet
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))
load_dotenv()


def pn(val) -> float:
    """Parse numeric, return 0 for non-numeric."""
    if not val:
        return 0.0
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_date(val) -> str:
    """Format a date value to DD/MM/YYYY."""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    # Handle "2025-12-31 00:00:00" format
    if " 00:00:00" in s:
        s = s.replace(" 00:00:00", "")
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        return s


def read_excel(path: str) -> dict:
    """
    Read Excel History sheet. Returns dict keyed by (room, name_lower).
    Each value: {room, name, phone, gender, building, floor, sharing,
                 checkin, rent, booking, deposit, maintenance, comments,
                 staff, status, march_bal, march_cash, march_upi}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["History"]

    data = {}
    for r in range(2, ws.max_row + 1):
        inout = str(ws.cell(r, 17).value or "").strip().upper()
        # Include CHECKIN + empty IN/OUT with a checkin date (new entries not yet tagged)
        if inout == "EXIT" or inout == "NO SHOW":
            continue
        if inout != "CHECKIN" and inout != "":
            continue
        # Skip rows with no checkin date and no name
        if not ws.cell(r, 5).value and inout == "":
            continue

        room_raw = str(ws.cell(r, 1).value or "").replace(".0", "").strip()
        name = str(ws.cell(r, 2).value or "").strip()
        if not room_raw or not name:
            continue

        phone = str(ws.cell(r, 4).value or "").replace(".0", "").strip()
        gender = str(ws.cell(r, 3).value or "").strip()
        checkin = ws.cell(r, 5).value
        building = str(ws.cell(r, 18).value or "").strip().upper()
        floor = str(ws.cell(r, 19).value or "").replace(".0", "").strip()
        sharing = str(ws.cell(r, 13).value or "").strip()
        rent = pn(ws.cell(r, 12).value) or pn(ws.cell(r, 11).value) or pn(ws.cell(r, 10).value)
        booking = pn(ws.cell(r, 6).value)
        deposit = pn(ws.cell(r, 7).value)
        maintenance = pn(ws.cell(r, 8).value)
        comments = str(ws.cell(r, 15).value or "").strip()
        staff = str(ws.cell(r, 16).value or "").strip()

        # March data
        march_bal = pn(ws.cell(r, 32).value)
        march_cash = pn(ws.cell(r, 33).value)
        march_upi = pn(ws.cell(r, 34).value)

        # April rent column (col 28) — may have "PAID"/"NOT PAID" or a number
        april_col = str(ws.cell(r, 28).value or "").strip()

        key = (room_raw, name.lower())
        data[key] = {
            "room": room_raw,
            "name": name,
            "phone": phone,
            "gender": gender,
            "building": building,
            "floor": floor,
            "sharing": sharing,
            "checkin": fmt_date(checkin),
            "rent": rent,
            "booking": booking,
            "deposit": deposit,
            "maintenance": maintenance,
            "comments": comments,
            "staff": staff,
            "march_bal": march_bal,
            "april_status": april_col,
        }

    wb.close()
    return data


def read_march_sheet() -> dict:
    """Read March 2026 Google Sheet for balance + notes."""
    from src.integrations.gsheets import _get_worksheet_sync, _safe_parse_numeric

    try:
        ws = _get_worksheet_sync("MARCH 2026")
    except Exception:
        print("WARNING: MARCH 2026 tab not found, skipping prev dues")
        return {}

    vals = ws.get_all_values()
    hdr = vals[3] if len(vals) > 3 else []
    is_new = "phone" in str(hdr[2] if len(hdr) > 2 else "").lower()
    bal_col = 9 if is_new else 8
    notes_col = 14 if is_new else 12

    data = {}
    for i in range(4, len(vals)):
        row = vals[i]
        if not row[0] or not row[1]:
            continue
        room = str(row[0]).strip()
        name = str(row[1]).strip().lower()
        bal = _safe_parse_numeric(str(row[bal_col]) if bal_col < len(row) else "0")
        notes = str(row[notes_col]).strip() if notes_col < len(row) else ""
        data[(room, name)] = {"balance": bal, "notes": notes}

    return data


def read_current_april() -> dict:
    """Read current April 2026 Sheet for existing payment data."""
    from src.integrations.gsheets import _get_worksheet_sync, _safe_parse_numeric

    try:
        ws = _get_worksheet_sync("APRIL 2026")
    except Exception:
        print("WARNING: APRIL 2026 tab not found")
        return {}

    vals = ws.get_all_values()
    hdr = vals[3] if len(vals) > 3 else []
    is_new = "phone" in str(hdr[2] if len(hdr) > 2 else "").lower()

    data = {}
    for i in range(4, len(vals)):
        row = vals[i]
        if not row[0] or not row[1]:
            continue
        room = str(row[0]).strip()
        name = str(row[1]).strip().lower()

        if is_new:
            cash = _safe_parse_numeric(str(row[6]) if len(row) > 6 else "0")
            upi = _safe_parse_numeric(str(row[7]) if len(row) > 7 else "0")
            notes = str(row[14]).strip() if len(row) > 14 else ""
            event = str(row[13]).strip() if len(row) > 13 else ""
            notice = str(row[12]).strip() if len(row) > 12 else ""
            entered_by = str(row[16]).strip() if len(row) > 16 else ""
        else:
            cash = _safe_parse_numeric(str(row[5]) if len(row) > 5 else "0")
            upi = _safe_parse_numeric(str(row[6]) if len(row) > 6 else "0")
            notes = str(row[12]).strip() if len(row) > 12 else ""
            event = str(row[11]).strip() if len(row) > 11 else ""
            notice = ""
            entered_by = ""

        data[(room, name)] = {
            "cash": cash,
            "upi": upi,
            "notes": notes,
            "event": event,
            "notice": notice,
            "entered_by": entered_by,
        }

    return data


async def read_db_notes() -> dict:
    """Read tenancy notes from DB. Returns dict keyed by (room_number, name_lower)."""
    from src.database.db_manager import get_session, init_db
    from src.database.models import Room, Tenancy, Tenant
    from sqlalchemy import select
    from sqlalchemy.orm import joinedload

    await init_db(os.getenv("SUPABASE_DB_URL") or os.getenv("DATABASE_URL"))
    async with get_session() as s:
        rooms = (await s.execute(select(Room))).scalars().all()
        room_map = {r.id: r.room_number for r in rooms}

        result = await s.execute(
            select(Tenancy).options(joinedload(Tenancy.tenant)).where(
                Tenancy.status == "active"
            )
        )
        tenancies = result.unique().scalars().all()

        data = {}
        for t in tenancies:
            rn = room_map.get(t.room_id, str(t.room_id))
            name = t.tenant.name.lower() if t.tenant else ""
            notes = str(t.notes or "").strip()
            if notes.lower() in ("no due", "no dues", "none", ""):
                notes = ""
            data[(rn, name)] = notes

    return data


def build_april_rows(
    excel_data: dict,
    march_data: dict,
    april_data: dict,
    db_notes: dict,
) -> list[list]:
    """Build 17-column rows for APRIL 2026."""
    from src.integrations.gsheets import _safe_parse_numeric

    rows = []
    for key, ex in sorted(excel_data.items(), key=lambda x: (x[0][0], x[0][1])):
        room, name_lower = key
        name = ex["name"]

        # Get data from each source
        march = march_data.get(key, {"balance": 0, "notes": ""})
        april = april_data.get(key, {"cash": 0, "upi": 0, "notes": "", "event": "", "notice": "", "entered_by": ""})
        db_note = db_notes.get(key, "")

        # Rent
        rent = ex["rent"]

        # Prev due from March balance
        prev_due = march["balance"] if march["balance"] > 0 else 0

        # Existing payments from current April sheet
        cash = april["cash"]
        upi = april["upi"]
        tp = cash + upi

        # Balance
        total_due = rent + prev_due
        balance = total_due - tp

        # Status
        event = april["event"]
        if event.upper() in ("EXIT", "NO-SHOW") or ex.get("april_status", "").upper() == "EXIT":
            status = event.upper() if event else "EXIT"
            balance = 0
        elif tp <= 0:
            status = "UNPAID"
        elif balance <= 0:
            status = "PAID"
        else:
            status = "PARTIAL"

        # Notes: combine DB permanent notes + March carried notes + existing April notes
        notes_parts = []
        if db_note:
            notes_parts.append(db_note)
        if prev_due > 0 and march["notes"]:
            notes_parts.append(f"[Mar] {march['notes'][:80]}")
        # Add existing April notes (payment logs) if they exist
        existing_april_notes = april["notes"]
        if existing_april_notes:
            # Don't duplicate DB notes or March notes already in there
            for part in existing_april_notes.split(" | "):
                part = part.strip()
                if part and part not in notes_parts and not any(p in part for p in notes_parts if p):
                    notes_parts.append(part)

        notes = " | ".join(notes_parts) if notes_parts else ""

        # Notice date
        notice = april["notice"]

        # Entered by from existing or staff
        entered_by = april["entered_by"] or ex["staff"]

        rows.append([
            room,                   # A: Room
            name,                   # B: Name
            ex["phone"],            # C: Phone
            ex["building"],         # D: Building
            ex["sharing"],          # E: Sharing
            rent,                   # F: Rent Due
            cash,                   # G: Cash
            upi,                    # H: UPI
            tp,                     # I: Total Paid
            balance,                # J: Balance
            status,                 # K: Status
            ex["checkin"],          # L: Check-in
            notice,                 # M: Notice Date
            event,                  # N: Event
            notes,                  # O: Notes
            prev_due,               # P: Prev Due
            entered_by,             # Q: Entered By
        ])

    return rows


def write_to_sheet(rows: list[list]) -> None:
    """Write rows to APRIL 2026 Sheet."""
    from src.integrations.gsheets import _get_worksheet_sync, _refresh_summary_sync

    ws = _get_worksheet_sync("APRIL 2026")

    headers = [
        "Room", "Name", "Phone", "Building", "Sharing", "Rent Due",
        "Cash", "UPI", "Total Paid", "Balance", "Status",
        "Check-in", "Notice Date", "Event", "Notes", "Prev Due", "Entered By",
    ]

    # Clear everything below row 1
    max_rows = max(ws.row_count, len(rows) + 10)
    ws.batch_clear([f"A2:Q{max_rows}"])

    # Write headers (row 4) — leave rows 2-3 for summary
    ws.update(values=[[""] * 17], range_name="A2:Q2", value_input_option="USER_ENTERED")
    ws.update(values=[[""] * 17], range_name="A3:Q3", value_input_option="USER_ENTERED")
    ws.update(values=[headers], range_name="A4:Q4", value_input_option="USER_ENTERED")

    # Write data (row 5+)
    if rows:
        ws.update(
            values=rows,
            range_name=f"A5:Q{4 + len(rows)}",
            value_input_option="USER_ENTERED",
        )

    # Refresh summary
    _refresh_summary_sync("APRIL 2026")
    print(f"Written {len(rows)} rows to APRIL 2026 + summary refreshed")


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true", help="Actually write to Sheet")
    args = parser.parse_args()

    excel_path = str(Path(__file__).parent.parent / "Cozeevo Monthly stay (4).xlsx")
    print(f"Reading Excel: {excel_path}")
    excel_data = read_excel(excel_path)
    print(f"  Active tenants: {len(excel_data)}")

    print("Reading March 2026 Sheet...")
    march_data = read_march_sheet()
    with_march_dues = sum(1 for v in march_data.values() if v["balance"] > 0)
    print(f"  March dues: {with_march_dues} tenants")

    print("Reading current April 2026 Sheet...")
    april_data = read_current_april()
    with_payments = sum(1 for v in april_data.values() if v["cash"] + v["upi"] > 0)
    print(f"  Existing payments: {with_payments} tenants")

    print("Reading DB notes...")
    db_notes = await read_db_notes()
    with_notes = sum(1 for v in db_notes.values() if v)
    print(f"  DB notes: {with_notes} tenants")

    print("\nBuilding rows...")
    rows = build_april_rows(excel_data, march_data, april_data, db_notes)
    print(f"  Total rows: {len(rows)}")

    # Stats
    paid = sum(1 for r in rows if r[10] == "PAID")
    partial = sum(1 for r in rows if r[10] == "PARTIAL")
    unpaid = sum(1 for r in rows if r[10] == "UNPAID")
    with_prev = sum(1 for r in rows if r[15] and r[15] > 0)
    with_notes_count = sum(1 for r in rows if r[14])
    with_entered = sum(1 for r in rows if r[16])
    print(f"  PAID: {paid}, PARTIAL: {partial}, UNPAID: {unpaid}")
    print(f"  With prev dues: {with_prev}")
    print(f"  With notes: {with_notes_count}")
    print(f"  With entered_by: {with_entered}")

    # Show sample rows
    print("\n--- Sample rows ---")
    for r in rows[:5]:
        print(f"  Room {r[0]} {r[1]}: Rent={r[5]} Cash={r[6]} UPI={r[7]} Bal={r[9]} "
              f"St={r[10]} PrevDue={r[15]} Notes={str(r[14])[:50]} By={r[16]}")

    # Show rows with prev dues
    print("\n--- Rows with prev dues ---")
    for r in rows:
        if r[15] and r[15] > 0:
            print(f"  Room {r[0]} {r[1]}: Rent={r[5]} PrevDue={r[15]} Bal={r[9]} "
                  f"Notes={str(r[14])[:60]}")

    if args.write:
        print("\nWriting to Sheet...")
        write_to_sheet(rows)
        print("DONE!")
    else:
        print("\n[DRY RUN] Add --write to actually update the Sheet")


if __name__ == "__main__":
    asyncio.run(main())
