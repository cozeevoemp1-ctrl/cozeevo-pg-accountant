"""
Short version of tenant agreed terms — only rows with REAL agreed terms,
with rent history columns to cross-check.

Run: PYTHONPATH=. PYTHONUTF8=1 venv/Scripts/python scripts/tenant_agreed_terms_short.py
Output: tenant_terms_short_<date>.xlsx
"""
import sys, openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXCEL_PATH = "Cozeevo Monthly stay (3).xlsx"
TODAY = date.today()
ACTIVE_STATUSES = {"CHECKIN", "NO SHOW"}
SKIP_COMMENTS = {"no due", "no dues", "-", "", "none", "exit", "n/a"}


def main():
    wb_src = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb_src["History"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    # Output columns: key info + rent history + agreed terms
    OUT = [
        ("Room No",        "Room",       9),
        ("Name",           "Name",       24),
        ("IN/OUT",         "Status",     10),
        ("Checkin date",   "Checkin",    12),
        ("Monthly Rent",   "Rent",       10),
        ("DEC RENT",       "DEC",        10),
        ("JAN RENT",       "JAN",        10),
        ("FEB RENT",       "FEB",        10),
        ("MARCH RENT",     "MAR",        10),
        ("March Balance",  "Balance",    10),
        ("Assigned Staff", "Staff",      14),
        ("Comments",       "Agreed Terms", 55),
    ]

    # Filter: CHECKIN or NO SHOW AND meaningful comment
    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[idx["IN/OUT"]]
        if status not in ACTIVE_STATUSES:
            continue
        comment = row[idx["Comments"]]
        comment_str = str(comment).strip() if comment else ""
        if comment_str.lower() in SKIP_COMMENTS:
            continue
        if not comment_str:
            continue
        rows_out.append(row)

    # Build output workbook
    wb = openpyxl.Workbook()
    ws_out = wb.active
    ws_out.title = "Agreed Terms"

    # Header
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (_, label, width) in enumerate(OUT, 1):
        cell = ws_out.cell(row=1, column=ci, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws_out.column_dimensions[get_column_letter(ci)].width = width
    ws_out.row_dimensions[1].height = 28
    ws_out.freeze_panes = "A2"

    # Data
    for ri, row in enumerate(rows_out, 2):
        status = row[idx["IN/OUT"]]
        for ci, (src_col, _, _) in enumerate(OUT, 1):
            val = row[idx[src_col]] if src_col in idx else ""
            # Format date
            if hasattr(val, "strftime"):
                val = val.strftime("%d %b %Y")
            cell = ws_out.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == len(OUT)))

            # Row shading: NO SHOW = light orange
            if status == "NO SHOW":
                cell.fill = PatternFill("solid", fgColor="FCE4D6")

        # Highlight Agreed Terms cell yellow
        terms_cell = ws_out.cell(row=ri, column=len(OUT))
        terms_cell.fill = PatternFill("solid", fgColor="FFF2CC")

    # Summary row count
    print(f"  Rows with real agreed terms: {len(rows_out)}")
    for row in rows_out:
        name = row[idx["Name"]]
        room = row[idx["Room No"]]
        comment = str(row[idx["Comments"]]).strip()[:70]
        print(f"    Room {room} | {name}: {comment}")

    out_path = f"tenant_terms_short_{TODAY}.xlsx"
    wb.save(out_path)
    print(f"\n  Saved: {out_path}")


if __name__ == "__main__":
    main()
