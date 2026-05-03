"""
scripts/export_other_expenses_detail.py
---------------------------------------
Drill-down for every "Other Expenses" bank transaction in the P&L.

For each transaction that the classifier couldn't recognise it parses the
raw description string into split columns:
  Txn Type | UTR / Ref | Payee Name | Payee UPI | Note

Yes Bank UPI-out format:
  UPI/{UTR}/{PAYEE_NAME}/{PAYEE_VPA}/{NOTE}

NEFT-out format (Yes Bank):
  NEFT Cr-{BANK}-{SENDER}-{RECIPIENT}-{REF}  or
  YIB-NEFT-{REF}-{NAME}

Output: data/reports/other_expenses_detail.xlsx
  Sheet 1  — All months combined, sorted by date
  Sheet 2+ — One sheet per calendar month

Run:
  venv/Scripts/python scripts/export_other_expenses_detail.py
"""

import sys, os, glob as _glob, re
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

from src.parsers.yes_bank import parse_date, parse_amt, read_yes_bank_csv
from src.rules.pnl_classify import classify_txn
from src.utils.inr_format import INR_NUMBER_FORMAT

# ── Month label mapping (matches P&L) ─────────────────────────────────────────
def _month_label(dt) -> str:
    return dt.strftime("%b'%y")   # e.g. "Nov'25"

# ── Yes Bank description parser ────────────────────────────────────────────────
UTR_RE  = re.compile(r"\b(\d{10,16})\b")
UPI_RE  = re.compile(r"[a-zA-Z0-9.\-_+]+@[a-zA-Z]+")

def parse_yesbank_desc(desc: str) -> dict:
    """
    Split a Yes Bank description string into structured columns.

    UPI-out:   UPI/{UTR}/{PAYEE_NAME}/{PAYEE_VPA}/{NOTE}
    NEFT-out:  NEFT Cr-{BANK}-{SENDER}-{RECIPIENT}-{REF}
               YIB-NEFT-{REF}-{NAME}
    IMPS-out:  IMPS/{REF}/{NAME}/{ACCT}
    """
    result = {
        "txn_type": "",
        "utr_ref":  "",
        "payee_name": "",
        "payee_upi":  "",
        "note":       "",
    }
    if not desc or not isinstance(desc, str):
        return result

    d = desc.strip()
    dl = d.lower()

    # ── transaction type ──────────────────────────────────────────────────────
    if "upi" in dl:
        result["txn_type"] = "UPI"
    elif "neft" in dl:
        result["txn_type"] = "NEFT"
    elif "imps" in dl:
        result["txn_type"] = "IMPS"
    elif "rtgs" in dl:
        result["txn_type"] = "RTGS"
    elif "atm" in dl or "cash" in dl:
        result["txn_type"] = "CASH"

    # ── UPI format: UPI/UTR/NAME/VPA/NOTE ────────────────────────────────────
    # Yes Bank uses "/" as separator for UPI transactions
    if result["txn_type"] == "UPI" and "/" in d:
        parts = [p.strip() for p in d.split("/")]
        # parts[0] usually = "UPI" or starts with UPI
        # parts[1] = UTR (10-16 digits)
        # parts[2] = payee name
        # parts[3] = payee UPI (VPA)
        # parts[4+] = note
        if len(parts) >= 2:
            utr_candidate = parts[1].strip() if len(parts) > 1 else ""
            if re.match(r"^\d{10,16}$", utr_candidate):
                result["utr_ref"] = utr_candidate
                if len(parts) > 2:
                    result["payee_name"] = parts[2].strip()
                if len(parts) > 3:
                    vpa = parts[3].strip()
                    if "@" in vpa:
                        result["payee_upi"] = vpa
                    else:
                        # sometimes name continues here
                        if result["payee_name"]:
                            result["payee_name"] += " " + vpa
                        else:
                            result["payee_name"] = vpa
                if len(parts) > 4:
                    result["note"] = "/".join(parts[4:]).strip()

    # ── NEFT format: NEFT Cr-BANK-SENDER-RECIPIENT-REF ───────────────────────
    elif result["txn_type"] == "NEFT":
        # Try "YIB-NEFT-REF-NAME" format first
        yib_m = re.match(r"(?:YIB-NEFT|NET-NEFT|YESOB)-(\w+)[-/](.+)$", d, re.IGNORECASE)
        if yib_m:
            result["utr_ref"]    = yib_m.group(1)
            result["payee_name"] = yib_m.group(2).strip()
        else:
            # Standard "NEFT Cr-BANKNAME-SENDER-RECIPIENT-REF" (inward) or
            # outward: "NEFT/REF/NAME" — split on dash/slash
            sep = "-" if "-" in d else "/"
            parts = [p.strip() for p in d.split(sep)]
            # find the UTR (long alpha-num ref near end)
            for p in reversed(parts):
                if re.match(r"^[A-Z0-9]{14,22}$", p):
                    result["utr_ref"] = p
                    break
            # payee name is the last readable segment before the ref
            clean = [p for p in parts if p and not re.match(r"^[A-Z0-9]{14,22}$", p)
                     and p.upper() not in ("NEFT", "RTGS", "CR", "DR")]
            if clean:
                result["payee_name"] = clean[-1]

    # ── IMPS format: IMPS/REF/NAME/ACCT ──────────────────────────────────────
    elif result["txn_type"] == "IMPS" and "/" in d:
        parts = [p.strip() for p in d.split("/")]
        if len(parts) >= 2:
            result["utr_ref"] = parts[1]
        if len(parts) >= 3:
            result["payee_name"] = parts[2]

    # ── fallback: extract UTR and UPI from anywhere in description ────────────
    if not result["utr_ref"]:
        m = UTR_RE.search(d)
        if m:
            result["utr_ref"] = m.group(1)

    if not result["payee_upi"]:
        upis = UPI_RE.findall(d)
        if upis:
            result["payee_upi"] = upis[-1]  # last one is usually payee

    return result


# ── Load bank statement files (same logic as export_classified.py) ─────────────
txns = []

csv_files = sorted(
    set(
        _glob.glob("Statement-*.csv")
        + _glob.glob("*statment*.csv")
        + _glob.glob("*statement*.csv")
        + _glob.glob("april month.csv")
        + _glob.glob("april*.csv")
    ),
    reverse=True,
)
for f in csv_files:
    t = read_yes_bank_csv(f)
    print(f"Loaded {len(t)} from {f}")
    txns += t

for f in ["2025 statement.xlsx", "2026 statment.xlsx"]:
    if not os.path.exists(f):
        continue
    import openpyxl as _opx
    from src.parsers.yes_bank import parse_date as _pd, parse_amt as _pa
    wb_s = _opx.load_workbook(f, data_only=True)
    ws_s = wb_s.active
    rows_s = list(ws_s.iter_rows(values_only=True))
    hrow = None
    for i, row in enumerate(rows_s):
        if row and any("transaction date" in str(c).lower() for c in row if c):
            hrow = i
            break
    if hrow is not None:
        headers = [str(c).lower().strip() if c else "" for c in rows_s[hrow]]
        for row in rows_s[hrow + 1:]:
            if not any(row):
                continue
            d2 = dict(zip(headers, row))
            dt = _pd(d2.get("transaction date"))
            if not dt:
                continue
            wd  = _pa(d2.get("withdrawals", ""))
            dep = _pa(d2.get("deposits", ""))
            desc2 = str(d2.get("description") or "")
            if wd  > 0:
                txns.append((dt, desc2, "expense", wd))
            if dep > 0:
                txns.append((dt, desc2, "income", dep))
    print(f"Loaded from {f}")

# Dedupe
seen = set()
deduped = []
for dt, desc, typ, amt in txns:
    key = (dt.strftime("%Y-%m-%d"), round(float(amt), 2), (desc or "").strip().lower())
    if key in seen:
        continue
    seen.add(key)
    deduped.append((dt, desc, typ, amt))
print(f"After dedupe: {len(deduped)} txns (removed {len(txns) - len(deduped)} dups)")
txns = deduped

# ── Filter to Other Expenses only ─────────────────────────────────────────────
other = []
for dt, desc, typ, amt in txns:
    if typ != "expense":
        continue
    cat, sub = classify_txn(desc, typ)
    if cat == "Other Expenses":
        parsed = parse_yesbank_desc(desc)
        other.append({
            "date":       dt,
            "month":      _month_label(dt),
            "amount":     amt,
            "txn_type":   parsed["txn_type"],
            "utr_ref":    parsed["utr_ref"],
            "payee_name": parsed["payee_name"],
            "payee_upi":  parsed["payee_upi"],
            "note":       parsed["note"],
            "description": desc,
        })

other.sort(key=lambda x: x["date"])

# Group by month
by_month = defaultdict(list)
for row in other:
    by_month[row["date"].strftime("%Y-%m")].append(row)

print(f"\nOther Expenses transactions: {len(other)}")
for m in sorted(by_month):
    total = sum(r["amount"] for r in by_month[m])
    print(f"  {m}: {len(by_month[m])} txns  Rs {total:,.0f}")

# ── Excel output ───────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F7FF")
WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="D9D9D9")
CTR      = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT    = Alignment(horizontal="right",  vertical="center")
NUM_FMT  = INR_NUMBER_FORMAT

COLS = [
    ("#",           4,  "center"),
    ("Date",        12, "center"),
    ("Month",       9,  "center"),
    ("Amount (Rs)", 14, "right"),
    ("Txn Type",    9,  "center"),
    ("UTR / Ref",   18, "left"),
    ("Payee Name",  28, "left"),
    ("Payee UPI",   30, "left"),
    ("Note",        28, "left"),
    ("Full Description", 60, "left"),
]


def _write_sheet(ws, rows, title):
    # title row
    n_cols = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = title
    t.font  = Font(bold=True, size=12, color="FFFFFF")
    t.fill  = HDR_FILL
    t.alignment = CTR
    ws.row_dimensions[1].height = 20

    # header row
    for ci, (h, w, _) in enumerate(COLS, 1):
        c = ws.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # data rows
    total = 0.0
    for ri, row in enumerate(rows, 1):
        r    = ri + 2
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        vals = [
            ri,
            row["date"].strftime("%d %b %Y"),
            row["month"],
            row["amount"],
            row["txn_type"],
            row["utr_ref"],
            row["payee_name"],
            row["payee_upi"],
            row["note"],
            row["description"],
        ]
        for ci, (val, (_, _, align)) in enumerate(zip(vals, COLS), 1):
            c = ws.cell(r, ci, val)
            c.fill = fill
            if ci == 4:  # amount
                c.number_format = NUM_FMT
                c.font = Font(size=9, color="CC0000")
                c.alignment = RIGHT
            else:
                c.font = Font(size=9)
                c.alignment = Alignment(
                    horizontal=align,
                    vertical="center",
                    wrap_text=(align == "left"),
                )
        total += row["amount"]

    # total row
    tr = len(rows) + 3
    for ci in range(1, n_cols + 1):
        ws.cell(tr, ci).fill = TOT_FILL
    ws.cell(tr, 1, "TOTAL").font = Font(bold=True, size=10)
    ws.cell(tr, 1).fill = TOT_FILL
    c = ws.cell(tr, 4, total)
    c.number_format = NUM_FMT
    c.font = Font(bold=True, size=10, color="CC0000")
    c.alignment = RIGHT
    c.fill = TOT_FILL

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"
    return total


wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: all months combined
ws_all = wb.create_sheet("All Other Expenses")
grand_total = _write_sheet(
    ws_all, other,
    f"Other Expenses — All Months (Oct'25–Apr'26) — {len(other)} transactions"
)

# Per-month sheets
for m_key in sorted(by_month.keys()):
    rows_m = by_month[m_key]
    dt0 = rows_m[0]["date"]
    label = dt0.strftime("%b '%y")  # "Nov '25"
    ws_m = wb.create_sheet(label)
    month_total = _write_sheet(
        ws_m, rows_m,
        f"Other Expenses — {label}  ({len(rows_m)} transactions)"
    )

OUT = os.path.join("data", "reports", "other_expenses_detail.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)

print(f"\nSaved: {OUT}")
print(f"Grand total Other Expenses: Rs {grand_total:,.0f}")
print(f"Sheets: All Other Expenses + {len(by_month)} monthly sheets")
