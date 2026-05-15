"""
scripts/ebitda_matrix_jun2026.py
---------------------------------
EBITDA sensitivity matrix for Jun 2026 rent increase planning.

Layout (matches reference format):
  Columns: Metric | Rent/bed | 80%(238b) | 85%(252b) | 90%(267b) | 95%(282b) | 100%(297b)
  Rows:    Section label (merged) + one row per price point
  Sections: EBITDA | After GST 12% | Net-Net (GST+IT 8%)
  Two scenario tables (current vs future property rent)

OPEX methodology:
  Fixed (property rent): Scenario 1 = Rs.21,32,000 | Scenario 2 = Rs.22,14,000
  Variable: Apr'26 actual non-rent OPEX = Rs.9,93,067 at 270 beds (91% occ)
            Scaled: variable_opex = 9,93,067 x (occupied_beds / 270)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Parameters ────────────────────────────────────────────────────────────────
TOTAL_BEDS       = 297
APR_NONRENT_OPEX = 993_067   # April 2026 actual non-rent OPEX
APR_REF_BEDS     = 270       # April reference occupied beds (91% occ)

PROPERTY_RENT_S1 = 2_132_000  # Current
PROPERTY_RENT_S2 = 2_214_000  # Jun 2026 increase (+82K)

PRICES     = [14_000, 13_500, 13_000]
OCC_LEVELS = [0.80, 0.85, 0.90, 0.95, 1.00]
GST_RATE   = 0.12
IT_RATE    = 0.08

# ── Helpers ───────────────────────────────────────────────────────────────────
def beds(occ: float) -> int:
    return round(occ * TOTAL_BEDS)

def total_opex(occ: float, prop_rent: int) -> float:
    var = APR_NONRENT_OPEX * beds(occ) / APR_REF_BEDS
    return prop_rent + var

def revenue(occ: float, price: int) -> float:
    return beds(occ) * price

def ebitda(occ: float, price: int, prop_rent: int) -> float:
    return revenue(occ, price) - total_opex(occ, prop_rent)

def after_gst(val: float) -> float:
    return val * (1 - GST_RATE)

def net_net(val: float) -> float:
    return after_gst(val) * (1 - IT_RATE)

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL      = PatternFill("solid", fgColor="1F4E78")
HDR_FONT      = Font(bold=True, color="FFFFFF", size=10)
S1_TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
S2_TITLE_FILL = PatternFill("solid", fgColor="375623")
EBITDA_FILL   = PatternFill("solid", fgColor="FFFFFF")
GST_FILL      = PatternFill("solid", fgColor="E2EFDA")
NET_FILL      = PatternFill("solid", fgColor="FCE4D6")
LABEL_FILLS   = [EBITDA_FILL, GST_FILL, NET_FILL]
LABEL_COLORS  = ["1F4E78", "375623", "833C00"]
BLANK_FILL    = PatternFill("solid", fgColor="F2F2F2")
thin_side     = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
NUM_FMT       = u'[>=1000000]##\\,##\\,##0;[>=100000]##\\,##0;##,##0'
CENTER        = Alignment(horizontal="center", vertical="center")
RIGHT         = Alignment(horizontal="right", vertical="center")
LEFT_WRAP     = Alignment(horizontal="left", vertical="center", wrap_text=True)

SECTIONS = [
    ("EBITDA",               EBITDA_FILL, "1F4E78", lambda v: v),
    ("After GST 12%",        GST_FILL,    "375623", after_gst),
    ("Net-Net\n(GST+IT 8%)", NET_FILL,    "833C00", net_net),
]


def set_cell(ws, row, col, value=None, fill=None, font=None, alignment=None, border=None, number_format=None):
    c = ws.cell(row, col, value)
    if fill:       c.fill       = fill
    if font:       c.font       = font
    if alignment:  c.alignment  = alignment
    if border:     c.border     = border
    if number_format: c.number_format = number_format
    return c


def write_scenario(ws, start_row: int, title: str, prop_rent: int, title_fill) -> int:
    r = start_row
    total_cols = 2 + len(OCC_LEVELS)  # col A=Metric, col B=Rent/bed, cols C-G=occupancies

    # ── Scenario title ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    set_cell(ws, r, 1, title,
             fill=title_fill, font=Font(bold=True, color="FFFFFF", size=11),
             alignment=CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Assumption note ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    var_rate = round(APR_NONRENT_OPEX / APR_REF_BEDS)
    note = (f"Property Rent Rs.{prop_rent:,}  |  Variable OPEX: Apr'26 actual Rs.{APR_NONRENT_OPEX:,} "
            f"at {APR_REF_BEDS} beds (Rs.{var_rate:,}/bed), scaled proportionally by occupied beds")
    set_cell(ws, r, 1, note,
             fill=PatternFill("solid", fgColor="EBF3FB"),
             font=Font(italic=True, size=8),
             alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[r].height = 14
    r += 1

    # ── Column header ─────────────────────────────────────────────────────────
    set_cell(ws, r, 1, "Metric",     fill=HDR_FILL, font=HDR_FONT, alignment=CENTER, border=BORDER)
    set_cell(ws, r, 2, "Rent/bed",   fill=HDR_FILL, font=HDR_FONT, alignment=CENTER, border=BORDER)
    for ci, occ in enumerate(OCC_LEVELS, 3):
        b = beds(occ)
        set_cell(ws, r, ci, f"{int(occ*100)}%\n({b} beds)",
                 fill=HDR_FILL, font=HDR_FONT,
                 alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER)
    ws.row_dimensions[r].height = 28
    r += 1

    # ── Data sections ─────────────────────────────────────────────────────────
    for s_label, s_fill, s_color, s_fn in SECTIONS:
        n_prices = len(PRICES)

        # Merge section label vertically across all price rows
        ws.merge_cells(start_row=r, start_column=1, end_row=r + n_prices - 1, end_column=1)
        set_cell(ws, r, 1, s_label,
                 fill=s_fill,
                 font=Font(bold=True, color=s_color, size=10),
                 alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER)

        for pi, price in enumerate(PRICES):
            data_row = r + pi
            # Price label
            set_cell(ws, data_row, 2, f"Rs.{price:,}",
                     fill=s_fill,
                     font=Font(size=9),
                     alignment=Alignment(horizontal="right", vertical="center"),
                     border=BORDER)
            # Values
            for ci, occ in enumerate(OCC_LEVELS, 3):
                base_val = ebitda(occ, price, prop_rent)
                val = round(s_fn(base_val))
                color = "9C0006" if val < 0 else "000000"
                set_cell(ws, data_row, ci, val,
                             fill=s_fill,
                             font=Font(bold=(pi == 0), color=color, size=9),
                             alignment=RIGHT,
                             border=BORDER,
                             number_format=NUM_FMT)
            ws.row_dimensions[data_row].height = 16

        r += n_prices

        # Blank separator between sections
        for col in range(1, total_cols + 1):
            set_cell(ws, r, col, fill=BLANK_FILL)
        ws.row_dimensions[r].height = 6
        r += 1

    r += 1  # extra gap between scenarios
    return r


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "EBITDA Matrix Jun 2026"

    # Title
    total_cols = 2 + len(OCC_LEVELS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t1 = ws.cell(1, 1, "EBITDA Sensitivity Matrix — Jun 2026 Rent Planning")
    t1.font = Font(bold=True, size=13, color="FFFFFF")
    t1.fill = PatternFill("solid", fgColor="0D2B45")
    t1.alignment = CENTER
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    t2 = ws.cell(2, 1,
                 "297 total beds  |  Variable OPEX: Apr'26 actual Rs.9,93,067 at 270 beds (Rs.3,678/bed), scaled by occupancy"
                 "  |  GST 12%  |  IT 8%  |  Two scenarios: current vs Jun +Rs.82K property rent")
    t2.font = Font(italic=True, size=8)
    t2.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    for col_idx in range(3, 3 + len(OCC_LEVELS)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    row = 4
    row = write_scenario(ws, row, "SCENARIO 1 — Current Property Rent  Rs.21,32,000 / month", PROPERTY_RENT_S1, S1_TITLE_FILL)
    row = write_scenario(ws, row, "SCENARIO 2 — Post-increase Property Rent  Rs.22,14,000 / month  (+Rs.82,000)", PROPERTY_RENT_S2, S2_TITLE_FILL)

    out = "scripts/ebitda_matrix_jun2026.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    build()
