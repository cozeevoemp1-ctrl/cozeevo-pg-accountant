"""
Tenant Agreed Terms export.
Reads the Excel, filters CHECKIN + NO SHOW tenants, copies Comments as "Agreed Terms",
outputs a clean Excel for record-keeping.

Run: PYTHONPATH=. PYTHONUTF8=1 venv/Scripts/python scripts/tenant_agreed_terms.py
Output: tenant_agreed_terms_<date>.xlsx
"""
import sys, openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXCEL_PATH = "Cozeevo Monthly stay (3).xlsx"
TODAY = date.today()

# Columns to include in output
OUT_COLS = [
    "Room No", "Name", "Gender", "Mobile Number",
    "Checkin date", "Monthly Rent", "Security Deposit", "Maintenance",
    "Assigned Staff", "IN/OUT", "Comments",
]

ACTIVE_STATUSES = {"CHECKIN", "NO SHOW"}


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_agreed(cell):
    """Highlight the Agreed Terms column."""
    cell.fill = PatternFill("solid", fgColor="FFF2CC")
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb_src = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb_src["History"]

    headers = [c.value for c in ws[1]]

    # Build index map
    idx = {h: i for i, h in enumerate(headers) if h}
    in_out_i   = idx["IN/OUT"]
    comments_i = idx["Comments"]
    name_i     = idx["Name"]
    room_i     = idx["Room No"]

    # Collect CHECKIN + NO SHOW rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[in_out_i] in ACTIVE_STATUSES:
            rows.append(row)

    print(f"  Active + No-show tenants found: {len(rows)}")

    # Build output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Agreed Terms"

    # Output headers: original cols + Agreed Terms
    out_header = OUT_COLS + ["Agreed Terms"]
    for col_i, h in enumerate(out_header, 1):
        cell = ws_out.cell(row=1, column=col_i, value=h)
        style_header(cell)

    ws_out.row_dimensions[1].height = 30

    # Write data
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, row in enumerate(rows, 2):
        status = row[in_out_i]
        comment = row[comments_i] or ""
        comment_str = str(comment).strip() if comment else ""

        for col_i, col_name in enumerate(out_header, 1):
            if col_name == "Agreed Terms":
                value = comment_str
            elif col_name == "Maintenance":
                value = row[idx.get("Maintence", -1)] if "Maintence" in idx else ""
            elif col_name in idx:
                value = row[idx[col_name]]
            else:
                value = ""

            cell = ws_out.cell(row=row_i, column=col_i, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "Agreed Terms"))

            # Colour row by status
            if status == "NO SHOW":
                cell.fill = PatternFill("solid", fgColor="FCE4D6")  # light orange
            if col_name == "Agreed Terms" and comment_str and comment_str.lower() not in (
                "no due", "no dues", "-", "", "none"
            ):
                style_agreed(cell)

    # Column widths
    widths = {
        "Room No": 9, "Name": 24, "Gender": 8, "Mobile Number": 14,
        "Checkin date": 13, "Monthly Rent": 13, "Security Deposit": 15,
        "Maintenance": 12, "Assigned Staff": 14, "IN/OUT": 10,
        "Comments": 35, "Agreed Terms": 55,
    }
    for col_i, col_name in enumerate(out_header, 1):
        ws_out.column_dimensions[get_column_letter(col_i)].width = widths.get(col_name, 15)

    # Freeze header row
    ws_out.freeze_panes = "A2"

    # Add legend sheet
    ws_legend = wb_out.create_sheet("Legend")
    ws_legend["A1"] = "Colour Legend"
    ws_legend["A1"].font = Font(bold=True)
    ws_legend["A2"] = "Orange row"
    ws_legend["A2"].fill = PatternFill("solid", fgColor="FCE4D6")
    ws_legend["B2"] = "NO SHOW — booked but never arrived"
    ws_legend["A3"] = "Yellow cell (Agreed Terms)"
    ws_legend["A3"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws_legend["B3"] = "Has meaningful agreed terms (not just 'No Due')"
    ws_legend.column_dimensions["A"].width = 28
    ws_legend.column_dimensions["B"].width = 50

    out_path = f"tenant_agreed_terms_{TODAY}.xlsx"
    wb_out.save(out_path)
    print(f"  Saved: {out_path}  ({len(rows)} rows)")

    # Summary
    with_terms = [r for r in rows if r[comments_i] and
                  str(r[comments_i]).strip().lower() not in ("no due", "no dues", "-", "", "none")]
    print(f"  Rows with meaningful agreed terms: {len(with_terms)}")
    print()
    print("  Sample agreed terms:")
    for r in with_terms[:10]:
        name = r[name_i]
        room = r[room_i]
        comm = str(r[comments_i]).strip()[:80]
        print(f"    Room {room} | {name}: {comm}")


if __name__ == "__main__":
    main()
