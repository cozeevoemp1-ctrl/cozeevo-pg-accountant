"""
scripts/_import_personal_sbi.py
Import SBI personal account (XX0167) expenses from two sources:
  1. PhonePe_Transaction_Statement.pdf  (Jan–Apr 2026, 59 debits)
  2. Paytm_UPI_Statement_08_Jan'26_-_30_Apr'26.xlsx  (27 debits)

These were paid by partner from personal SBI account for PG business expenses
and need to be reimbursed from the company account.

Account-name in DB: PERSONAL_SBI_0167
The 0961 (Yes Bank company) rows in both files are SKIPPED — already in THOR CSV.

Run:
  venv/Scripts/python scripts/_import_personal_sbi.py          # dry-run
  venv/Scripts/python scripts/_import_personal_sbi.py --write  # insert + save Excel
"""
from __future__ import annotations

import asyncio
import hashlib
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, ".")
from src.rules.pnl_classify import classify_txn

DB = "postgresql+asyncpg://postgres:Anchorstrong123!@db.oxiqomoilqwfxjauxhzp.supabase.co:5432/postgres"
WRITE = "--write" in sys.argv

PHONEPE_PDF = "PhonePe_Transaction_Statement.pdf"
PHONEPE_PWD = "7358341775"
PAYTM_XLSX = "Paytm_UPI_Statement_08_Jan'26_-_30_Apr'26.xlsx"
ACCOUNT_NAME = "PERSONAL_SBI_0167"

# SBI account suffix — both files identify this account differently
SBI_SUFFIXES = {"0167", "67"}  # PhonePe uses "0167", Paytm uses "- 67"

MONTHS = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
          "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

# Confirmed personal expenses — exclude from P&L (not PG business)
PERSONAL_EXCLUDE = [
    "motherhood hospital", "sps motherhood", "born babies",
    "apollo pharmacy", "apollopharmacy",
    "zippycubs",
    "california burrito", "kfc", "gopizza", "hms host",
    "cafe coffee day", "american sweet corn",
    "hennes n mauritz",  # H&M
    "myntra",
    "trends bangalore",
    "sharif foot wear",
    "velidi venkata chaithanya",  # parlour
    "paid to kiran",              # personal transfer to Kiran
]

# Confirmed reclassifications (description keyword → override category/sub)
RECLASSIFY: dict[str, tuple[str, str]] = {
    "house hunt":                        ("Marketing",           "Influencer / Channel Marketing Fee"),
    "anumola yoga anil kumar":           ("Tenant Deposit Refund", "Deposit Refund — Anumola Yoga Anil Kumar"),
    "aahil rafiq":                       ("Tenant Deposit Refund", "Deposit Refund — Aahil Rafiq"),
    "p deepa":                           ("Tenant Deposit Refund", "Deposit Refund — P Deepa"),
    "notion online solutions":           ("Food & Groceries",    "Notion Online Solutions (PG Food)"),
    "shubh chikan":                      ("Food & Groceries",     "Shubh Chikan (PG Food)"),
    "spar hypermarkets":                 ("Food & Groceries",     "Spar Hypermarkets (PG Groceries)"),
    "subramani":                         ("Staff & Labour",       "Subramani (Worker)"),
    "tuniki lavanya":                    ("Other Expenses",       "Worker Payment"),
}


def make_hash(date_val, amount: float, utr: str, desc: str) -> str:
    key = f"PERSONAL_SBI_0167|{date_val}|{amount}|{utr or desc}"
    return hashlib.sha256(key.encode()).hexdigest()


def parse_phonepe() -> list[dict]:
    """Extract SBI 0167 debit transactions from PhonePe PDF."""
    txns = []
    with pdfplumber.open(PHONEPE_PDF, password=PHONEPE_PWD) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or "").split("\n")
            for idx, line in enumerate(lines):
                line = line.strip()
                # Pattern: "Jan 04, 2026 Description Debit INR [amount]"
                date_m = re.match(
                    r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{2}),\s+(\d{4})\s+(.+?)\s+Debit\s+INR\s*([\d,\.]*)",
                    line,
                )
                if not date_m:
                    continue
                # Look ahead up to 6 lines for account, UTR, possibly amount
                block = "\n".join(lines[idx: idx + 7])
                acct_m = re.search(r"Debited from XX(\d+)", block)
                if not acct_m or acct_m.group(1) != "0167":
                    continue

                month_str, day_str, year_str = date_m.group(1), date_m.group(2), date_m.group(3)
                desc = date_m.group(4).strip()
                amt_str = date_m.group(5).strip()

                # Amount sometimes overflows to transaction-id line
                if not amt_str:
                    amt_m2 = re.search(r"Transaction ID\s*:\s*\S+\s+([\d,\.]+)", block)
                    amt_str = amt_m2.group(1) if amt_m2 else "0"

                amount = float(amt_str.replace(",", "")) if amt_str else 0.0
                if amount <= 0:
                    continue

                utr_m = re.search(r"UTR No\s*:\s*(\S+)", block)
                utr = utr_m.group(1) if utr_m else ""

                txn_date = datetime(int(year_str), MONTHS[month_str], int(day_str)).date()
                txns.append({
                    "txn_date": txn_date,
                    "description": desc,
                    "amount": amount,
                    "utr": utr,
                    "source_file": "phonepe",
                })
    return txns


def parse_paytm() -> list[dict]:
    """Extract SBI 0167 debit transactions from Paytm XLSX."""
    txns = []
    wb = openpyxl.load_workbook(PAYTM_XLSX, data_only=True)
    ws = wb["Passbook Payment History"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        date_str = str(row[0] or "").strip()     # "27/04/2026"
        desc = str(row[2] or "").strip()          # "Money sent to X"
        account = str(row[4] or "").strip()       # "State Bank Of India - 67"
        amount_str = str(row[5] or "").strip()    # "-89.00"
        utr = str(row[6] or "").strip()           # UPI ref no

        if "67" not in account:
            continue
        if not amount_str.startswith("-"):
            continue  # skip credits

        amount = float(amount_str.replace(",", "").replace("-", ""))
        if amount <= 0:
            continue

        try:
            txn_date = datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError:
            continue

        # Use remarks if description is generic
        remarks = str(row[8] or "").strip()
        if remarks and len(remarks) > 3:
            desc = f"{desc} ({remarks})"

        txns.append({
            "txn_date": txn_date,
            "description": desc,
            "amount": amount,
            "utr": utr,
            "source_file": "paytm",
        })
    return txns


def is_personal(desc: str) -> bool:
    d = desc.lower()
    return any(kw in d for kw in PERSONAL_EXCLUDE)


def apply_reclassify(desc: str, cat: str, sub: str) -> tuple[str, str]:
    d = desc.lower()
    for kw, (new_cat, new_sub) in RECLASSIFY.items():
        if kw in d:
            return new_cat, new_sub
    return cat, sub


def build_records(raw: list[dict]) -> tuple[list[dict], list[dict]]:
    """Dedup by UTR, exclude personal, classify+reclassify, return (kept, excluded)."""
    seen_utrs: set[str] = set()
    records, excluded = [], []
    for t in raw:
        utr = t["utr"]
        if utr and utr in seen_utrs:
            continue
        if utr:
            seen_utrs.add(utr)

        if is_personal(t["description"]):
            excluded.append(t)
            continue

        cat, sub = classify_txn(t["description"], "expense")
        cat, sub = apply_reclassify(t["description"], cat, sub)
        h = make_hash(t["txn_date"], t["amount"], utr, t["description"])
        records.append({
            "txn_date": t["txn_date"],
            "description": t["description"],
            "amount": t["amount"],
            "txn_type": "expense",
            "category": cat,
            "sub_category": sub,
            "upi_reference": utr or None,
            "account_name": ACCOUNT_NAME,
            "source_file": t["source_file"],
            "unique_hash": h,
        })
    return records, excluded


def print_summary(records: list[dict]) -> None:
    by_month: dict = defaultdict(lambda: defaultdict(float))
    for r in records:
        m = r["txn_date"].strftime("%Y-%m")
        by_month[m][r["category"]] += r["amount"]

    for m in sorted(by_month):
        total = sum(by_month[m].values())
        print(f"\n  {m}  total={total:>10,.0f}")
        for cat, amt in sorted(by_month[m].items(), key=lambda x: -x[1]):
            print(f"    {cat:<45} {amt:>10,.0f}")

    grand = sum(r["amount"] for r in records)
    print(f"\n  GRAND TOTAL PERSONAL SBI EXPENSES: Rs.{grand:,.0f}")

    print(f"\n  ALL TRANSACTIONS ({len(records)}):")
    for r in sorted(records, key=lambda x: x["txn_date"]):
        print(f"  {r['txn_date']}  {r['amount']:>10,.0f}  [{r['category'][:20]:<20}]  {r['description'][:55]}  [{r['source_file']}]")


def save_reimbursement_excel(records: list[dict]) -> None:
    """Save reimbursement Excel for review / company payment."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SBI 0167 Reimbursement"

        headers = ["Date", "Description", "Category", "Sub-Category", "Amount (Rs.)", "Source", "UTR / Ref"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")

        total = 0
        for r in sorted(records, key=lambda x: x["txn_date"]):
            ws.append([
                r["txn_date"].strftime("%d %b %Y"),
                r["description"],
                r["category"],
                r["sub_category"],
                r["amount"],
                r["source_file"],
                r["upi_reference"] or "",
            ])
            total += r["amount"]

        ws.append(["", "TOTAL", "", "", total, "", ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 5).font = Font(bold=True)

        # Column widths
        widths = [12, 55, 30, 30, 14, 10, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        out = "data/reports/SBI_0167_Reimbursement.xlsx"
        Path("data/reports").mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"\n  Reimbursement Excel saved: {out}")
    except Exception as e:
        print(f"  Warning: Could not save Excel: {e}")


async def run() -> None:
    print("Parsing PhonePe PDF...")
    phonepe_txns = parse_phonepe()
    print(f"  Found {len(phonepe_txns)} SBI 0167 debits")

    print("Parsing Paytm XLSX...")
    paytm_txns = parse_paytm()
    print(f"  Found {len(paytm_txns)} SBI 0167 debits")

    all_raw = phonepe_txns + paytm_txns
    records, excluded = build_records(all_raw)

    print(f"\n{'DRY RUN' if not WRITE else 'WRITING'} — {len(records)} business transactions ({len(excluded)} personal excluded)\n")
    print_summary(records)

    if not WRITE:
        save_reimbursement_excel(records)
        print("\n  Run with --write to insert into DB.")
        return

    engine = create_async_engine(DB)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with Session() as session:
        result = await session.execute(
            sa.text("""
                INSERT INTO bank_uploads (phone, file_path, row_count, new_count, from_date, to_date, status, uploaded_at)
                VALUES (:phone, :file_path, :row_count, :new_count, :from_date, :to_date, 'processed', NOW())
                RETURNING id
            """),
            {
                "phone": "system",
                "file_path": "PhonePe_Statement + Paytm_XLSX (SBI 0167 personal)",
                "row_count": len(records),
                "new_count": len(records),
                "from_date": min(r["txn_date"] for r in records),
                "to_date": max(r["txn_date"] for r in records),
            },
        )
        upload_id = result.scalar()
        print(f"\n  Created bank_uploads id={upload_id}")

        inserted = skipped = 0
        for r in records:
            existing = await session.scalar(
                sa.text("SELECT id FROM bank_transactions WHERE unique_hash = :h"),
                {"h": r["unique_hash"]},
            )
            if existing:
                skipped += 1
                continue
            await session.execute(
                sa.text("""
                    INSERT INTO bank_transactions
                      (upload_id, txn_date, description, amount, txn_type, category, sub_category,
                       upi_reference, source, unique_hash, account_name)
                    VALUES
                      (:upload_id, :txn_date, :description, :amount, :txn_type, :category, :sub_category,
                       :upi_reference, :source, :unique_hash, :account_name)
                """),
                {**{k: v for k, v in r.items() if k != "source_file"}, "upload_id": upload_id,
                 "source": f"personal_sbi_{r['source_file']}"},
            )
            inserted += 1

        await session.commit()
        print(f"  Inserted: {inserted}, Skipped (dupes): {skipped}")

    await engine.dispose()
    save_reimbursement_excel(records)  # type: ignore[arg-type]


asyncio.run(run())
