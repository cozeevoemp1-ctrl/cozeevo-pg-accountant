"""
P&L Report — Income vs Expenses, classified and categorised by month.
Reads all bank statement sources (Excel + CSV) and uses shared pnl_classify rules.

Run:  venv/Scripts/python scripts/pnl_report.py
Output:  PnL_Report.xlsx  (3 sheets: P&L Summary, Income Detail, Expense Detail)
"""
import csv as _csv
import glob
import re
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.rules.pnl_classify import classify_txn

OUT = "PnL_Report.xlsx"

# ── Helpers ──────────────────────────────────────────────────────────────────
def parse_date(v):
    if hasattr(v, 'strftime'): return v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            from datetime import datetime
            return datetime.strptime(s[:10], fmt)
        except: pass
    return None

def parse_amt(v):
    if v is None or v == '': return 0.0
    try: return float(str(v).replace(',', '').strip())
    except: return 0.0

# ── Read YES Bank Excel ─────────────────────────────────────────────────────
def read_yes_bank_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hrow = None
    for i, row in enumerate(rows):
        if row and any('transaction date' in str(c).lower() for c in row if c):
            hrow = i; break
    if hrow is None: return []
    headers = [str(c).lower().strip() if c else '' for c in rows[hrow]]
    out = []
    for row in rows[hrow+1:]:
        if not any(row): continue
        d = dict(zip(headers, row))
        dt = parse_date(d.get('transaction date'))
        if not dt: continue
        wd = parse_amt(d.get('withdrawals', ''))
        dep = parse_amt(d.get('deposits', ''))
        desc = str(d.get('description') or '')
        if wd > 0: out.append({'date': dt, 'desc': desc, 'amount': wd, 'type': 'expense'})
        if dep > 0: out.append({'date': dt, 'desc': desc, 'amount': dep, 'type': 'income'})
    return out

# ── Read YES Bank CSV ────────────────────────────────────────────────────────
def read_yes_bank_csv(path):
    out = []
    with open(path, 'r') as f:
        for line in f:
            if line.startswith('Transaction Date'): break
        reader = _csv.reader(f)
        for row in reader:
            if len(row) < 7: continue
            dt = parse_date(row[0])
            if not dt: continue
            wd = parse_amt(row[4]); dep = parse_amt(row[5])
            desc = row[2].strip()
            if wd > 0: out.append({'date': dt, 'desc': desc, 'amount': wd, 'type': 'expense'})
            if dep > 0: out.append({'date': dt, 'desc': desc, 'amount': dep, 'type': 'income'})
    return out

# ── Load all sources ─────────────────────────────────────────────────────────
all_txns = []
# 2025 statement = Oct-Dec 2025, 2026 statment = Jan 2026 only
# CSV covers Jan-Mar 2026, so skip 2026 Excel to avoid duplicates
for f in ['2025 statement.xlsx']:
    if os.path.exists(f):
        t = read_yes_bank_xlsx(f)
        print(f'Loaded {len(t)} from {f}')
        all_txns += t

for f in glob.glob('Statement-*.csv'):
    t = read_yes_bank_csv(f)
    print(f'Loaded {len(t)} from {f}')
    all_txns += t

print(f'Total: {len(all_txns)} txns')

# ── Classify ─────────────────────────────────────────────────────────────────
for txn in all_txns:
    cat, sub = classify_txn(txn['desc'], txn['type'])
    txn['cat'] = cat
    txn['sub'] = sub

df = pd.DataFrame(all_txns)
df['_date'] = pd.to_datetime(df['date'])
df['Month'] = df['_date'].dt.strftime('%b %Y')
# Sort months chronologically
month_order = sorted(df['Month'].unique(), key=lambda x: pd.to_datetime(x, format='%b %Y'))
MONTHS = month_order

exp_df = df[df['type'] == 'expense'].copy()
inc_df = df[df['type'] == 'income'].copy()

# ── Styling ──────────────────────────────────────────────────────────────────
HDR  = "1F497D"
GRN  = "1A5276"
RED  = "922B21"
THIN = Side(style="thin", color="CCCCCC")
MED  = Side(style="medium", color="666666")
B    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BM   = Border(left=THIN, right=THIN, top=THIN, bottom=MED)

INC_BG  = "D5F5E3"
EXP_BG  = "FDEDEC"
TTL_BG  = "D6EAF8"
NET_BG_P= "A9DFBF"
NET_BG_N= "F1948A"

def fw(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(hex):
    return PatternFill("solid", fgColor=hex)

def hcell(ws, r, c, v, bg=HDR, color="FFFFFF", bold=True, size=10, align="center"):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name="Calibri", bold=bold, color=color, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = B
    return cell

def dcell(ws, r, c, v, bg, bold=False, color="000000", align="left", fmt=None):
    cell = ws.cell(r, c, v if v != 0 else None)
    cell.font      = Font(name="Calibri", bold=bold, color=color, size=9)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = B
    if fmt:
        cell.number_format = fmt
    return cell

def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


# ── Build Excel ──────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

ncols = 2 + len(MONTHS) + 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — P&L SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("P&L Summary")

# Title
ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
c = ws["A1"]
period_str = f"{MONTHS[0]} – {MONTHS[-1]}" if len(MONTHS) > 1 else MONTHS[0]
c.value     = f"Profit & Loss Statement — LAKSHMI GORJALA / YES Bank ({period_str})"
c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
c.fill      = PatternFill("solid", fgColor=HDR)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Column headers
hcell(ws, 2, 1, "Section",   HDR)
hcell(ws, 2, 2, "Category",  HDR)
for i, m in enumerate(MONTHS, 3):
    hcell(ws, 2, i, m, HDR)
hcell(ws, 2, ncols, "TOTAL",  HDR)
ws.row_dimensions[2].height = 22

row = 3

def section_header(label, bg):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c2 = ws.cell(row, 1, label)
    c2.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c2.fill      = PatternFill("solid", fgColor=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border    = B
    ws.row_dimensions[row].height = 18

# ── INCOME ────────────────────────────────────────────────────────────────
section_header("  INCOME", GRN)
row += 1

inc_pivot = inc_df.pivot_table(index='cat', columns='Month', values='amount',
                                aggfunc='sum', fill_value=0, observed=True)
inc_pivot['TOTAL'] = inc_pivot.sum(axis=1)
inc_pivot = inc_pivot.sort_values('TOTAL', ascending=False)

inc_section_total = {m: 0.0 for m in MONTHS}; inc_section_total['TOTAL'] = 0.0

for cat, pivot_row in inc_pivot.iterrows():
    bg = INC_BG
    ws.cell(row, 1, "").fill = PatternFill("solid", fgColor=bg); ws.cell(row,1).border=B
    dcell(ws, row, 2, str(cat), bg, bold=True, color="1A5276")
    for i, m in enumerate(MONTHS, 3):
        v = pivot_row.get(m, 0)
        dcell(ws, row, i, v if v else None, bg, align="right", fmt="#,##0", color="1A5276")
        inc_section_total[m] += v
    dcell(ws, row, ncols, pivot_row['TOTAL'], bg, bold=True, align="right", fmt="#,##0", color="1A5276")
    inc_section_total['TOTAL'] += pivot_row['TOTAL']
    row += 1

# Income total
dcell(ws, row, 1, "TOTAL INCOME", TTL_BG, bold=True, color="1A5276")
ws.cell(row, 1).border = BM
ws.merge_cells(f"A{row}:B{row}")
for i, m in enumerate(MONTHS, 3):
    c2 = dcell(ws, row, i, inc_section_total[m], TTL_BG, bold=True, align="right", fmt="#,##0", color="1A5276")
    c2.border = BM
c2 = dcell(ws, row, ncols, inc_section_total['TOTAL'], TTL_BG, bold=True, align="right", fmt="#,##0", color="1A5276")
c2.border = BM
ws.row_dimensions[row].height = 16
row += 2

# ── EXPENSES ──────────────────────────────────────────────────────────────
section_header("  EXPENSES", RED)
row += 1

exp_pivot = exp_df.pivot_table(index='cat', columns='Month', values='amount',
                                aggfunc='sum', fill_value=0, observed=True)
exp_pivot['TOTAL'] = exp_pivot.sum(axis=1)
exp_pivot = exp_pivot.sort_values('TOTAL', ascending=False)

exp_section_total = {m: 0.0 for m in MONTHS}; exp_section_total['TOTAL'] = 0.0

for cat, pivot_row in exp_pivot.iterrows():
    bg = EXP_BG
    ws.cell(row, 1, "").fill = PatternFill("solid", fgColor=bg); ws.cell(row,1).border=B
    dcell(ws, row, 2, str(cat), bg, bold=True, color="922B21")
    for i, m in enumerate(MONTHS, 3):
        v = pivot_row.get(m, 0)
        dcell(ws, row, i, v if v else None, bg, align="right", fmt="#,##0", color="922B21")
        exp_section_total[m] += v
    dcell(ws, row, ncols, pivot_row['TOTAL'], bg, bold=True, align="right", fmt="#,##0", color="922B21")
    exp_section_total['TOTAL'] += pivot_row['TOTAL']
    row += 1

# Expense total
dcell(ws, row, 1, "TOTAL EXPENSES", TTL_BG, bold=True, color="922B21")
ws.cell(row, 1).border = BM
ws.merge_cells(f"A{row}:B{row}")
for i, m in enumerate(MONTHS, 3):
    c2 = dcell(ws, row, i, exp_section_total[m], TTL_BG, bold=True, align="right", fmt="#,##0", color="922B21")
    c2.border = BM
c2 = dcell(ws, row, ncols, exp_section_total['TOTAL'], TTL_BG, bold=True, align="right", fmt="#,##0", color="922B21")
c2.border = BM
ws.row_dimensions[row].height = 16
row += 2

# ── NET PROFIT / LOSS ─────────────────────────────────────────────────────
ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
c3 = ws.cell(row, 1, "NET PROFIT / LOSS")
c3.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
c3.fill      = PatternFill("solid", fgColor="1A5276")
c3.alignment = Alignment(horizontal="left", vertical="center")
c3.border    = B
ws.row_dimensions[row].height = 18
row += 1

dcell(ws, row, 1, "", "FFFFFF"); ws.cell(row,1).border = B
dcell(ws, row, 2, "", "FFFFFF"); ws.cell(row,2).border = B
for i, m in enumerate(MONTHS, 3):
    net = inc_section_total[m] - exp_section_total[m]
    bg  = NET_BG_P if net >= 0 else NET_BG_N
    col_str = "1A5276" if net >= 0 else "922B21"
    dcell(ws, row, i, net, bg, bold=True, align="right", fmt="#,##0", color=col_str)

total_net = inc_section_total['TOTAL'] - exp_section_total['TOTAL']
bg_net = NET_BG_P if total_net >= 0 else NET_BG_N
dcell(ws, row, ncols, total_net, bg_net, bold=True, align="right", fmt="#,##0",
      color="1A5276" if total_net >= 0 else "922B21")
ws.row_dimensions[row].height = 18

widths(ws, [20, 28] + [14]*len(MONTHS) + [14])
ws.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INCOME DETAIL
# ════════════════════════════════════════════════════════════════════════════
ws_inc = wb.create_sheet("Income Detail")
ws_inc.merge_cells("A1:G1")
h = ws_inc["A1"]
h.value = "Income Transactions Detail"
h.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
h.fill = PatternFill("solid", fgColor=GRN)
h.alignment = Alignment(horizontal="center")

for c2, hd in enumerate(["#","Date","Amount (Rs)","Category","Sub-Category","Description","Month"],1):
    hcell(ws_inc, 2, c2, hd, GRN)

inc_sorted = inc_df.sort_values('_date').reset_index(drop=True)
for i, (_, r2) in enumerate(inc_sorted.iterrows(), 1):
    ro = i + 2
    dcell(ws_inc, ro, 1, i,                      INC_BG, align="center")
    dcell(ws_inc, ro, 2, r2['_date'].strftime('%Y-%m-%d') if pd.notna(r2['_date']) else '', INC_BG)
    dcell(ws_inc, ro, 3, r2['amount'],            INC_BG, bold=True, align="right", fmt="#,##0", color="1A5276")
    dcell(ws_inc, ro, 4, r2['cat'],              INC_BG, bold=True)
    dcell(ws_inc, ro, 5, r2['sub'],              INC_BG)
    dcell(ws_inc, ro, 6, str(r2['desc'])[:100],  INC_BG)
    dcell(ws_inc, ro, 7, str(r2.get('Month','')), INC_BG, align="center")

tr = len(inc_sorted) + 3
ws_inc.cell(tr, 4, "TOTAL INCOME").font = Font(bold=True, name="Calibri")
c5 = ws_inc.cell(tr, 3, inc_sorted['amount'].sum())
c5.number_format = "#,##0"; c5.font = Font(bold=True, name="Calibri", color="1A5276")
c5.alignment = Alignment(horizontal="right")
for col in range(1, 8):
    ws_inc.cell(tr, col).fill   = PatternFill("solid", fgColor="D5F5E3")
    ws_inc.cell(tr, col).border = B

widths(ws_inc, [4, 12, 14, 22, 30, 60, 10])
ws_inc.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — EXPENSE DETAIL
# ════════════════════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("Expense Detail")
ws_exp.merge_cells("A1:G1")
h2 = ws_exp["A1"]
h2.value = "Expense Transactions Detail"
h2.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
h2.fill = PatternFill("solid", fgColor=RED)
h2.alignment = Alignment(horizontal="center")

for c2, hd in enumerate(["#","Date","Amount (Rs)","Category","Sub-Category","Description","Month"],1):
    hcell(ws_exp, 2, c2, hd, RED)

exp_sorted = exp_df.sort_values('_date').reset_index(drop=True)
for i, (_, r2) in enumerate(exp_sorted.iterrows(), 1):
    ro = i + 2
    dcell(ws_exp, ro, 1, i,                      EXP_BG, align="center")
    dcell(ws_exp, ro, 2, r2['_date'].strftime('%Y-%m-%d') if pd.notna(r2['_date']) else '', EXP_BG)
    dcell(ws_exp, ro, 3, r2['amount'],            EXP_BG, bold=True, align="right", fmt="#,##0", color="922B21")
    dcell(ws_exp, ro, 4, r2['cat'],              EXP_BG, bold=True)
    dcell(ws_exp, ro, 5, r2['sub'],              EXP_BG)
    dcell(ws_exp, ro, 6, str(r2['desc'])[:100],  EXP_BG)
    dcell(ws_exp, ro, 7, str(r2.get('Month','')), EXP_BG, align="center")

tr2 = len(exp_sorted) + 3
ws_exp.cell(tr2, 4, "TOTAL EXPENSES").font = Font(bold=True, name="Calibri")
c6 = ws_exp.cell(tr2, 3, exp_sorted['amount'].sum())
c6.number_format = "#,##0"; c6.font = Font(bold=True, name="Calibri", color="922B21")
c6.alignment = Alignment(horizontal="right")
for col in range(1, 8):
    ws_exp.cell(tr2, col).fill   = PatternFill("solid", fgColor="FADBD8")
    ws_exp.cell(tr2, col).border = B

widths(ws_exp, [4, 12, 14, 22, 30, 60, 10])
ws_exp.freeze_panes = "A3"

wb.save(OUT)

# ── Console P&L ──────────────────────────────────────────────────────────────
tot_inc = inc_sorted['amount'].sum()
tot_exp = exp_sorted['amount'].sum()

W = 14

def mrow(label, vals):
    line = "  " + label[:34].ljust(34)
    for m in MONTHS:
        v = vals.get(m, 0)
        cell = f"{v:,.0f}" if v else ""
        line += cell.rjust(W+2)
    tot = vals.get("TOTAL", 0)
    line += f"{tot:,.0f}".rjust(W+2)
    print(line)

SEP = "=" * (36 + (W+2)*(len(MONTHS)+1))
print()
print(SEP)
print("  COZEEVO P&L  —  YES Bank / LAKSHMI GORJALA")
print(SEP)

hdr = f"  {'Category':<34}"
for m in MONTHS:
    hdr += m.rjust(W+2)
hdr += "TOTAL".rjust(W+2)
print(hdr)
print("-" * len(hdr))

# INCOME
print("  INCOME")
inc_pivot2 = inc_df.pivot_table(index='cat', columns='Month', values='amount',
                                 aggfunc='sum', fill_value=0, observed=True)
inc_pivot2['TOTAL'] = inc_pivot2.sum(axis=1)
inc_m_totals = {m: inc_df[inc_df['Month']==m]['amount'].sum() for m in MONTHS}
inc_m_totals['TOTAL'] = tot_inc
for cat in inc_pivot2.sort_values('TOTAL', ascending=False).index:
    row_vals = {m: inc_pivot2.loc[cat, m] if m in inc_pivot2.columns else 0 for m in MONTHS}
    row_vals['TOTAL'] = inc_pivot2.loc[cat, 'TOTAL']
    mrow(f"  {cat}", row_vals)
print("-" * len(hdr))
mrow("  TOTAL INCOME", inc_m_totals)

print()
print("  EXPENSES")
exp_pivot2 = exp_df.pivot_table(index='cat', columns='Month', values='amount',
                                 aggfunc='sum', fill_value=0, observed=True)
exp_pivot2['TOTAL'] = exp_pivot2.sum(axis=1)
exp_pivot2 = exp_pivot2.sort_values('TOTAL', ascending=False)
exp_m_totals = {m: exp_df[exp_df['Month']==m]['amount'].sum() for m in MONTHS}
exp_m_totals['TOTAL'] = tot_exp
for cat in exp_pivot2.index:
    row_vals = {m: exp_pivot2.loc[cat, m] if m in exp_pivot2.columns else 0 for m in MONTHS}
    row_vals['TOTAL'] = exp_pivot2.loc[cat, 'TOTAL']
    mrow(f"  {cat}", row_vals)
print("-" * len(hdr))
mrow("  TOTAL EXPENSES", exp_m_totals)

print()
net_vals = {m: inc_m_totals[m] - exp_m_totals[m] for m in MONTHS}
net_vals['TOTAL'] = tot_inc - tot_exp
mrow("  NET PROFIT / (LOSS)", net_vals)
print(SEP)
print(f"  Saved: {OUT}")
