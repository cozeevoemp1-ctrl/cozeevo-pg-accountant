"""Per-tenant breakdown of April 2026 Pending.

Shows every tenant with positive balance and the formula components:
  Balance = Rent Due + Prev Due − (Cash + UPI + Prepaid + Booking + Deposit)

Saves data/reports/april_pending_breakdown.xlsx
"""
import asyncio
import os
import sys
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from sqlalchemy import select
from src.database.db_manager import init_engine, get_session
from src.database.models import Payment, PaymentFor, RentSchedule, Tenancy, Tenant, Room


async def main():
    init_engine(os.environ["DATABASE_URL"])
    apr = date(2026, 4, 1)
    may = date(2026, 5, 1)
    mar = date(2026, 3, 1)

    async with get_session() as s:
        rs_rows = (await s.execute(
            select(RentSchedule).where(RentSchedule.period_month == apr)
        )).scalars().all()

        # Rent payments for April
        pay_rows = (await s.execute(
            select(Payment).where(
                Payment.period_month == apr,
                Payment.is_void == False,
                Payment.for_type == PaymentFor.rent,
            )
        )).scalars().all()
        cash_map, upi_map, prepaid_map = {}, {}, {}
        for p in pay_rows:
            in_apr = p.payment_date and apr <= p.payment_date < may
            mode = p.payment_mode.value if hasattr(p.payment_mode, "value") else str(p.payment_mode or "cash")
            if in_apr:
                if mode.lower() == "cash":
                    cash_map[p.tenancy_id] = cash_map.get(p.tenancy_id, 0) + float(p.amount)
                else:
                    upi_map[p.tenancy_id] = upi_map.get(p.tenancy_id, 0) + float(p.amount)
            else:
                prepaid_map[p.tenancy_id] = prepaid_map.get(p.tenancy_id, 0) + float(p.amount)

        # Booking + deposit (first-month only when applied)
        bk_map = {}
        for p in (await s.execute(
            select(Payment).where(
                Payment.is_void == False, Payment.for_type == PaymentFor.booking
            )
        )).scalars():
            bk_map[p.tenancy_id] = bk_map.get(p.tenancy_id, 0) + float(p.amount)

        dep_map = {}
        for p in (await s.execute(
            select(Payment).where(
                Payment.is_void == False, Payment.for_type == PaymentFor.deposit
            )
        )).scalars():
            dep_map[p.tenancy_id] = dep_map.get(p.tenancy_id, 0) + float(p.amount)

        # Prev Due (March shortfall)
        mar_rs = (await s.execute(
            select(RentSchedule).where(RentSchedule.period_month == mar)
        )).scalars().all()
        mar_pays = (await s.execute(
            select(Payment).where(
                Payment.period_month == mar,
                Payment.is_void == False,
                Payment.for_type == PaymentFor.rent,
            )
        )).scalars().all()
        mar_paid = {}
        for p in mar_pays:
            mar_paid[p.tenancy_id] = mar_paid.get(p.tenancy_id, 0) + float(p.amount)
        prev_due_map = {}
        for rs in mar_rs:
            bal = float(rs.rent_due or 0) - mar_paid.get(rs.tenancy_id, 0)
            if bal > 0:
                prev_due_map[rs.tenancy_id] = bal

        data = []
        for rs in rs_rows:
            t = await s.get(Tenancy, rs.tenancy_id)
            if not t:
                continue
            tn = await s.get(Tenant, t.tenant_id)
            rm = await s.get(Room, t.room_id)
            is_fm = bool(t.checkin_date and t.checkin_date.replace(day=1) == apr)

            rent_due = float(rs.rent_due or 0)
            prev_due = prev_due_map.get(t.id, 0)
            cash = cash_map.get(t.id, 0)
            upi = upi_map.get(t.id, 0)
            prepaid = prepaid_map.get(t.id, 0)
            booking = bk_map.get(t.id, 0) if is_fm else 0
            deposit = dep_map.get(t.id, 0) if is_fm else 0
            eff_paid = cash + upi + prepaid + booking + deposit
            balance = rent_due + prev_due - eff_paid
            if balance <= 0:
                continue
            data.append({
                "room": rm.room_number,
                "name": tn.name,
                "fm": "FM" if is_fm else "",
                "rent_due": int(rent_due),
                "prev_due": int(prev_due),
                "cash": int(cash),
                "upi": int(upi),
                "prepaid": int(prepaid),
                "booking": int(booking),
                "deposit": int(deposit),
                "eff_paid": int(eff_paid),
                "balance": int(balance),
                "status": t.status.value if hasattr(t.status, "value") else str(t.status),
                "checkin": t.checkin_date.strftime("%Y-%m-%d") if t.checkin_date else "",
            })

    # Sort by balance desc
    data.sort(key=lambda d: -d["balance"])

    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "April Pending Breakdown"

    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F8F9FA")
    red_fill = PatternFill("solid", fgColor="FFE5E5")
    fm_fill = PatternFill("solid", fgColor="FFF4C7")

    headers = [
        "Room", "Name", "FM", "Status", "Check-in",
        "Rent Due", "Prev Due",
        "Cash (Apr)", "UPI (Apr)", "Prepaid", "Booking", "Deposit",
        "Eff Paid", "Balance (pending)",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, d in enumerate(data, 2):
        vals = [
            d["room"], d["name"], d["fm"], d["status"], d["checkin"],
            d["rent_due"], d["prev_due"],
            d["cash"], d["upi"], d["prepaid"], d["booking"], d["deposit"],
            d["eff_paid"], d["balance"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if ci >= 6 and isinstance(v, int):
                c.number_format = '#,##,##,##0;(#,##,##,##0);"-"'
            if ri % 2 == 0:
                c.fill = alt_fill
            if ci == 3 and d["fm"]:
                c.fill = fm_fill
            if ci == 14:
                c.fill = red_fill
                c.font = Font(bold=True)

    # Totals
    tr = len(data) + 2
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col, key in enumerate(["rent_due", "prev_due", "cash", "upi", "prepaid", "booking", "deposit", "eff_paid", "balance"], 6):
        tot = sum(d[key] for d in data)
        c = ws.cell(row=tr, column=col, value=tot)
        c.font = Font(bold=True)
        c.number_format = '#,##,##,##0;(#,##,##,##0);"-"'
        c.fill = PatternFill("solid", fgColor="FFD700")

    # Widths
    widths = [8, 30, 5, 10, 12, 11, 10, 11, 11, 10, 10, 10, 11, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "C2"

    out = os.path.join("data", "reports", "april_pending_breakdown.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Tenants with Pending > 0: {len(data)}")
    print(f"Total Pending: Rs.{sum(d['balance'] for d in data):,}")


asyncio.run(main())
