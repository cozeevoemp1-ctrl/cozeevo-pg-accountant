"""
Styled Excel exporter using openpyxl.
Generates a multi-sheet workbook with charts.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from loguru import logger


class ExcelExporter:
    def __init__(self, exports_dir: Path):
        self.exports_dir = exports_dir

    async def export(self, data: dict, period: str = "monthly") -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference, PieChart
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        filename = f"report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = self.exports_dir / filename

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, data, period)

        # ── Sheet 2: Category Breakdown ───────────────────────────────────
        ws_cat = wb.create_sheet("Categories")
        self._write_categories(ws_cat, data)

        # ── Sheet 3: Rent ─────────────────────────────────────────────────
        rent = data.get("rent_summary")
        if rent:
            ws_rent = wb.create_sheet("Rent")
            self._write_rent(ws_rent, rent)

        # ── Sheet 4: Salary ───────────────────────────────────────────────
        sal = data.get("salary_summary")
        if sal:
            ws_sal = wb.create_sheet("Salaries")
            self._write_salary(ws_sal, sal)

        wb.save(out_path)
        logger.info(f"Excel exported: {out_path}")
        return str(out_path)

    # ── Sheet writers ─────────────────────────────────────────────────────

    def _write_summary(self, ws, data: dict, period: str):
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        title = data.get("month_name") or f"{period.title()} Report"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:B1")

        header_fill = PatternFill("solid", fgColor="1F4E79")
        rows = [
            ("Metric", "Value"),
            ("Total Income (₹)", data.get("total_income", 0)),
            ("Total Expense (₹)", data.get("total_expense", 0)),
            ("Net Income (₹)", data.get("net_income", 0)),
            ("Transactions", data.get("txn_count", 0)),
            ("Period", f"{data.get('start_date', '')} → {data.get('end_date', '')}"),
        ]
        for i, (label, val) in enumerate(rows, start=3):
            ws.cell(i, 1, label)
            ws.cell(i, 2, val)
            if i == 3:
                ws.cell(i, 1).fill = header_fill
                ws.cell(i, 2).fill = header_fill
                ws.cell(i, 1).font = Font(bold=True, color="FFFFFF")
                ws.cell(i, 2).font = Font(bold=True, color="FFFFFF")

    def _write_categories(self, ws, data: dict):
        from openpyxl.styles import Font, PatternFill
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        headers = ["Category", "Income (₹)", "Expense (₹)", "Net (₹)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")

        by_cat = data.get("by_category", {})
        for row_i, (cat, vals) in enumerate(
            sorted(by_cat.items(), key=lambda x: x[1].get("expense", 0), reverse=True),
            start=2
        ):
            inc = vals.get("income", 0)
            exp = vals.get("expense", 0)
            ws.cell(row_i, 1, cat)
            ws.cell(row_i, 2, inc)
            ws.cell(row_i, 3, exp)
            ws.cell(row_i, 4, inc - exp)

    def _write_rent(self, ws, rent: dict):
        from openpyxl.styles import Font, PatternFill
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10

        headers = ["Tenant", "Room", "Rent (₹)", "Paid?"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="375623")

        for i, d in enumerate(rent.get("details", []), start=2):
            ws.cell(i, 1, d["customer"])
            ws.cell(i, 2, d.get("room", ""))
            ws.cell(i, 3, d["expected"])
            ws.cell(i, 4, "✓" if d["paid"] else "✗")

        # Totals
        last = len(rent.get("details", [])) + 2
        ws.cell(last, 1, "TOTAL").font = Font(bold=True)
        ws.cell(last, 3, rent.get("expected", 0)).font = Font(bold=True)

    def _write_salary(self, ws, sal: dict):
        from openpyxl.styles import Font, PatternFill
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10

        headers = ["Employee", "Role", "Salary (₹)", "Paid?"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7030A0")

        for i, d in enumerate(sal.get("details", []), start=2):
            ws.cell(i, 1, d["employee"])
            ws.cell(i, 2, d.get("role", ""))
            ws.cell(i, 3, d["expected"])
            ws.cell(i, 4, "✓" if d["paid"] else "✗")
