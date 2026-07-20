"""
src/reports/pnl_builder.py
---------------------------
Canonical P&L builder — single source of truth shared by:
  - scripts/export_pnl_2026_05_02.py  (local regeneration)
  - src/api/v2/finance.py  GET /finance/pnl/excel  (PWA download)

Both produce identical output. Verified figures as of 2026-05-17.
See memory/sop_pnl.md for full methodology.

Workbook tabs:
  1. P&L — Full (incl Cash)   — all items, canonical report
  2. Bank — Digital Only       — same layout; cash income + cash rent excluded
  3. Rules Applied             — methodology notes
"""
from __future__ import annotations

import io
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.inr_format import INR_NUMBER_FORMAT
from src.reports.pnl_verified_data import (
    MONTHS,
    INCOME,
    CAPITAL_CONTRIBUTIONS,
    OPEX,
    EXCLUDED,
    DEPOSIT_RECEIVED,
    DEPOSIT_REFUNDED,
    DEPOSITS,
    BANK_BALANCE_THOR,
    BANK_BALANCE_HULK,
    BANK_CLOSING_BALANCE_THOR,
    BANK_CLOSING_BALANCE_HULK,
    CASH_IN_HAND,
    THOR_CAP_IN,
    THOR_CAP_OUT,
    RECON_NOTES,
    KIRAN_REVIEW_FLAGS,
    RULES_APPLIED,
    KEY_OPEX_WATER,
    KEY_OPEX_WASTE,
)

# All verified monthly income/opex/deposit/bank-balance figures — and the real
# names/phone numbers/amounts in the reconciliation + rules tabs — live in
# src/reports/pnl_verified_data.py. A demo deployment ships
# src/reports/pnl_verified_data_stub.py (same names, empty values) in its place.

# ── Cash items to exclude for the Bank-Only tab ───────────────────────────────
# Income key containing this fragment is skipped
_CASH_INCOME_KEY  = "Cash (physical"
# OPEX key containing this fragment is skipped
_CASH_RENT_KEY    = "Property Rent — Cash paid"


# ── Dynamic-month translation (DB → SOP-format line keys) ─────────────────────
# When new months are appended from the DB, their values attach to these exact
# verified line keys so they land in the same rows as the hardcoded history.
_KEY_INCOME_THOR = "THOR — Bank Income (UPI + NEFT)"
_KEY_INCOME_HULK = "HULK — Bank Income (UPI + cheque)"
_KEY_INCOME_CASH = "Cash (physical — both buildings combined)"
_KEY_RENT_CASH   = "Property Rent — Cash paid (Jan rent in Feb, Feb in Mar, Mar in Apr)"
_KEY_CASH_EXP    = "Cash Expenses (paid in cash — manual entry)"
_KEY_EXCL_REFUND = "Tenant Deposit Refund (balance sheet)"
_KEY_EXCL_NONOP  = "Cash Exchange Repayments via Bank (non-op)"
_KEY_DEP_SEC     = "Security Deposits — refundable (must return to active tenants)"
_KEY_DEP_MAINT   = "  Maintenance Fee retained (non-refundable, by exit month)"

# bank_transactions.category → verified OPEX line key
_DB_CAT_TO_OPEX_KEY = {
    "Property Rent":         "Property Rent — Bank UPI/RTGS paid",
    "Electricity":           "Electricity",
    "Water":                 KEY_OPEX_WATER,
    "IT & Software":         "IT & Software",
    "Internet & WiFi":       "Internet & WiFi (cash — Jan Airwire UPI, Feb 8x Razorpay, Mar-Dec Rs.0)",
    "Food & Groceries":      "Food & Groceries",
    "Fuel & Diesel":         "Fuel & Diesel",
    "Staff & Labour":        "Staff & Labour",
    "Maintenance & Repairs": "Maintenance & Repairs",
    "Cleaning Supplies":     "Cleaning Supplies",
    "Waste Disposal":        KEY_OPEX_WASTE,
    "Shopping & Supplies":   "Shopping & Supplies",
    "Furniture & Fittings":  "Furniture & Supplies",
    "Capital Investment":    "Furniture & Supplies",
    "Marketing":             "Marketing",
    "Govt & Regulatory":     "Govt & Regulatory (incl Police Rs.3K accrual Jan+)",
    "Bank Charges":          "Bank Charges",
    "Other Expenses":        "Other Expenses",
}


def _dynamic_line_values(d: dict):
    """Translate one DB-month record into (income, opex, excluded) contributions
    keyed by the verified SOP line names."""
    income = {
        _KEY_INCOME_THOR: d.get("income_thor", 0),
        _KEY_INCOME_HULK: d.get("income_hulk", 0),
        _KEY_INCOME_CASH: d.get("cash", 0),
    }
    opex: Dict[str, float] = {}
    for cat, amt in (d.get("opex_by_cat") or {}).items():
        key = _DB_CAT_TO_OPEX_KEY.get(cat, "Other Expenses")
        opex[key] = opex.get(key, 0) + amt
    if d.get("rent_paid_cash"):
        opex[_KEY_RENT_CASH] = opex.get(_KEY_RENT_CASH, 0) + d["rent_paid_cash"]
    if d.get("cash_expense"):
        opex[_KEY_CASH_EXP] = opex.get(_KEY_CASH_EXP, 0) + d["cash_expense"]
    excluded = {
        _KEY_EXCL_REFUND: d.get("dep_refunded", 0),
    }
    # Show non-operating money by name (buyout, hand loans, capital returns) instead
    # of one blind lump. Falls back to the lump line if no breakdown was provided.
    detail = d.get("non_op_detail") or {}
    if detail:
        for sub, amt in sorted(detail.items(), key=lambda kv: -kv[1]):
            excluded[f"└ {sub} (non-op)"] = amt
    else:
        excluded[_KEY_EXCL_NONOP] = d.get("non_op", 0)
    return income, opex, excluded


def _extend_dict(base: Dict[str, List[int]], per_month: List[dict]) -> Dict[str, List[int]]:
    """Append one column per dynamic month to every verified line; add any
    dynamic-only keys as new rows (verified months = 0)."""
    n_verified = len(MONTHS)
    out: Dict[str, List[int]] = {k: list(v) + [pm.get(k, 0) for pm in per_month] for k, v in base.items()}
    extra: List[str] = []
    for pm in per_month:
        for k in pm:
            if k not in base and k not in extra:
                extra.append(k)
    for k in extra:
        out[k] = [0] * n_verified + [pm.get(k, 0) for pm in per_month]
    return out


# ── Shared P&L sheet writer ───────────────────────────────────────────────────

def _write_pnl_tab(
    ws: Worksheet,
    income_dict: Dict[str, List[int]],
    opex_dict: Dict[str, List[int]],
    tab_note: str = "",
    *,
    months: List[str] | None = None,
    deposits: Dict[str, List[int]] | None = None,
    excluded: Dict[str, List[int]] | None = None,
    capital: Dict[str, List[int]] | None = None,
    bank_thor_close: int | None = None,
    bank_hulk_close: int | None = None,
    cash_in_hand: Dict[str, int] | None = None,
    sec_owed: int = 2315175,
    snapshot_label: str = "Apr 30",
) -> None:
    """
    Render a complete P&L into *ws*.

    income_dict / opex_dict are the (possibly filtered) subsets of INCOME / OPEX.
    tab_note is shown in the top-left header cell when provided.

    All month-varying inputs default to the verified module globals, so calling
    with no keyword args reproduces the canonical Oct'25–May'26 report byte-for-byte.
    Pass extended dicts + a longer `months` list to append dynamic (DB) months.
    """
    months          = months          if months          is not None else MONTHS
    deposits        = deposits         if deposits        is not None else DEPOSITS
    excluded        = excluded         if excluded        is not None else EXCLUDED
    capital         = capital          if capital         is not None else CAPITAL_CONTRIBUTIONS
    bank_thor_close = bank_thor_close  if bank_thor_close is not None else BANK_CLOSING_BALANCE_THOR
    bank_hulk_close = bank_hulk_close  if bank_hulk_close is not None else BANK_CLOSING_BALANCE_HULK
    cash_in_hand    = cash_in_hand     if cash_in_hand    is not None else CASH_IN_HAND
    bold       = Font(bold=True)
    hdr_fill   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font   = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    flag_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ctr        = Alignment(horizontal="center")

    header = [tab_note or "", "Op"] + months + ["TOTAL"]
    ws.append(header)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = ctr

    # ── 1. INCOME ──────────────────────────────────────────────────────────────
    acct_fill  = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    close_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    acct_font  = Font(bold=True, color="1F4E78")

    def _acct_row(label, values, fill):
        ws.append([label, ""] + values + [None])
        for c in ws[ws.max_row]:
            c.fill = fill
            c.font = acct_font

    def _get(key):
        return income_dict.get(key, [0] * len(months))

    ws.append(["INCOME", ""])
    ws[ws.max_row][0].font = bold

    # ── THOR group (bank balances live only on the reconciliation tab) ────────
    row = _get("THOR — Bank Income (UPI + NEFT)")
    if row:
        ws.append(["THOR — Bank Income (UPI + NEFT)", "+"] + row + [sum(row)])
    row = _get("THOR — transferred to HULK acct (reclassification)")
    if any(row):
        ws.append(["  THOR — transferred to HULK acct (reclassification)", "−"] + row + [sum(row)])

    ws.append([])

    # ── Cash ──────────────────────────────────────────────────────────────────
    cash_key = "Cash (physical — both buildings combined)"
    if cash_key in income_dict:
        row = income_dict[cash_key]
        ws.append([cash_key, "+"] + row + [sum(row)])

    ws.append([])

    # ── HULK group ────────────────────────────────────────────────────────────
    row = _get("HULK — Bank Income (UPI + cheque)")
    if row:
        ws.append(["HULK — Bank Income (UPI + cheque)", "+"] + row + [sum(row)])
    row = _get("HULK — received from THOR acct (reclassification)")
    if any(row):
        ws.append(["HULK — received from THOR acct (reclassification)", "+"] + row + [sum(row)])

    ws.append([])

    rev_row = [sum(col) for col in zip(*income_dict.values())]
    ws.append(["Total Gross Inflows", "="] + rev_row + [sum(rev_row)])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = total_fill

    sec_dep_collected   = deposits["Security Deposits — refundable (must return to active tenants)"]
    maint_fee_collected = deposits["  Maintenance Fee retained (non-refundable, by exit month)"]
    sec_dep_neg         = [-v for v in sec_dep_collected]
    # Each monthly column = deposits COLLECTED that month (a flow, not a stock).
    # TOTAL col = SUM of the monthly flows = total deposits collected/held over the period.
    dep_refunded     = excluded["Tenant Deposit Refund (balance sheet)"]
    dep_refunded_neg = [-v for v in dep_refunded]
    closing_sec_dep  = sum(sec_dep_neg)        # total refundable deposits collected
    closing_maint    = sum(maint_fee_collected)

    ws.append(["  Less: Security Deposits held (active tenants — must return at exit)", "−"]
              + sec_dep_neg + [closing_sec_dep])
    ws[ws.max_row][0].font = Font(italic=True)

    ws.append(["     └ Maintenance Fee retained from tenants EXITED this month (non-refundable — yours to keep)", "(kept)"]
              + list(maint_fee_collected) + [closing_maint])
    ws[ws.max_row][0].font = Font(italic=True, color="375623")
    for c in ws[ws.max_row][1:]:
        if isinstance(c.value, (int, float)):
            c.font = Font(italic=True, color="375623")

    ws.append(["  Less: Deposits Refunded to Exited Tenants (already paid back)", "−"]
              + dep_refunded_neg + [sum(dep_refunded_neg)])
    ws[ws.max_row][0].font = Font(italic=True)

    true_rev_row = [r + s + d for r, s, d in zip(rev_row, sec_dep_neg, dep_refunded_neg)]
    # TOTAL col = sum of the monthly True Revenue figures
    true_rev_total = sum(true_rev_row)
    ws.append(["True Rent Revenue (excl. all deposit pass-throughs)", "="] + true_rev_row + [true_rev_total])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="375623")
    ws.append([])

    # ── 2. BORROWED MONEY / OWNER ADVANCES ────────────────────────────────────
    ws.append(["BORROWED MONEY — Owner loans & advances (to be repaid, not P&L)", ""])
    ws[ws.max_row][0].font = bold
    for label, row in capital.items():
        ws.append([label, "↑"] + row + [sum(row)])
    borrowed_row = [sum(col) for col in zip(*capital.values())]
    ws.append(["Total Borrowed Money (owners must be repaid this)", "="] + borrowed_row + [sum(borrowed_row)])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="9C0006")
        if isinstance(c.value, (int, float)):
            c.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.append([])

    # ── 3. OPERATING EXPENSES ──────────────────────────────────────────────────
    ws.append(["OPERATING EXPENSES (accrual)", ""])
    ws[ws.max_row][0].font = bold
    for label, row in opex_dict.items():
        ws.append([label, "−"] + row + [sum(row)])
        if "TBD" in label or "⚠" in label:
            for c in ws[ws.max_row]:
                c.fill = flag_fill

    ws.append(["EXCLUDED FROM OPEX (balance sheet items — not costs)", ""])
    ws[ws.max_row][0].font = Font(italic=True)
    for label, row in excluded.items():
        ws.append(["  " + label, "(B/S)"] + row + [sum(row)])

    opex_row = [sum(col) for col in zip(*opex_dict.values())]
    ws.append(["Total Opex", "="] + opex_row + [sum(opex_row)])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = total_fill
    ws.append([])

    # ── 4. EBITDA ──────────────────────────────────────────────────────────────
    op_profit_row = [r - o for r, o in zip(true_rev_row, opex_row)]
    op_profit_total = true_rev_total - sum(opex_row)
    ws.append(["NET OPERATING PROFIT (True Revenue − All Opex incl. Furniture & Supplies)", "="] + op_profit_row + [op_profit_total])
    for c in ws[ws.max_row]:
        c.font = bold

    op_margin_row = [f"{(p/r*100):.1f}%" if r else "-" for p, r in zip(op_profit_row, true_rev_row)]
    ws.append(["Operating Margin %", ""] + op_margin_row
              + [f"{(op_profit_total/true_rev_total*100):.1f}%" if true_rev_total else "-"])
    ws.append([])

    # ── 5. ADJUSTED PROFIT (after deducting borrowed money) ───────────────────
    borrowed_neg = [-v for v in borrowed_row]
    ws.append(["  Less: Borrowed Money to repay (owner loans — must be paid back)", "−"]
              + borrowed_neg + [sum(borrowed_neg)])
    ws[ws.max_row][0].font = Font(italic=True, color="9C0006")
    for c in ws[ws.max_row][2:]:
        if isinstance(c.value, (int, float)):
            c.font = Font(italic=True, color="9C0006")

    adjusted_row = [p - b for p, b in zip(op_profit_row, borrowed_row)]
    adjusted_total = op_profit_total - sum(borrowed_row)
    ws.append(["ADJUSTED NET PROFIT (after repaying all owner loans)", "="]
              + adjusted_row + [adjusted_total])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="375623" if adjusted_total >= 0 else "9C0006")
        if isinstance(c.value, (int, float)):
            c.fill = PatternFill(start_color="E2EFDA" if adjusted_total >= 0 else "FCE4D6",
                                 end_color="E2EFDA" if adjusted_total >= 0 else "FCE4D6",
                                 fill_type="solid")
    ws.append([])

    # ── 7. BALANCE SHEET ITEMS ────────────────────────────────────────────────
    # Pure refundable security deposits owed to active tenants
    # Combined total ₹33,83,875 included maintenance ₹10,68,700 — maintenance is non-refundable
    # True refundable liability = ₹33,83,875 − ₹10,68,700 = ₹23,15,175
    _sec_collected = sec_owed
    _bank_total    = bank_thor_close + bank_hulk_close
    _cash_total    = sum(cash_in_hand.values())

    # value lands one column before TOTAL (matches verified layout for 8 months)
    def _bs_row(label, value):
        ws.append([label] + [""] * (len(months) - 1) + [value])

    ws.append([f"BALANCE SHEET ITEMS ({snapshot_label})", ""])
    ws[ws.max_row][0].font = bold
    _bs_row(f"Bank closing balance THOR acct ...0961 ({snapshot_label})", bank_thor_close)
    _bs_row(f"Bank closing balance HULK acct ...0881 ({snapshot_label})", bank_hulk_close)
    _bs_row("Total bank balance", _bank_total)
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append([])
    ws.append(["Cash in hand (physical)", ""])
    ws[ws.max_row][0].font = bold
    for name, amt in cash_in_hand.items():
        _bs_row("  " + name, amt)
    _bs_row("Total cash in hand", _cash_total)
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append([])
    _bs_row("Net deposits still owed to active tenants (liability)", _sec_collected)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="9C0006")
    ws.append([])
    _bs_row("True free cash (bank − deposits owed) — excl. cash in hand", _bank_total - _sec_collected)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    _bs_row("True free cash incl. cash in hand", _bank_total + _cash_total - _sec_collected)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    ws.append(["NOTE: negative = deposit money was used to fund early operations (CAPEX+OPEX)"])
    ws.append(["As revenue grows, bank balance will recover and exceed deposit liability"])
    ws.append([])

    # ── 8. FLAGS ───────────────────────────────────────────────────────────────
    ws.append(["⚠ ITEMS NEEDING KIRAN REVIEW", ""])
    ws[ws.max_row][0].font = Font(bold=True, color="FF0000")
    for f in KIRAN_REVIEW_FLAGS:
        ws.append([f, ""])

    # ── Formatting ─────────────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2, max_col=len(header)):
        for i, cell in enumerate(row):
            if i == 0 or cell.value is None or isinstance(cell.value, str):
                continue
            if isinstance(cell.value, (int, float)):
                cell.number_format = INR_NUMBER_FORMAT

    for row in ws.iter_rows(min_row=1, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 6
    from openpyxl.utils import get_column_letter as _gcl
    for _ci in range(3, 3 + len(months) + 1):  # month cols + TOTAL
        ws.column_dimensions[_gcl(_ci)].width = 14


def build_pnl_workbook(dynamic_data: List[dict] | None = None, include_verified: bool = True) -> openpyxl.Workbook:
    """
    Return the P&L workbook.

    With no argument → the canonical verified report (Oct'25 – May'26), byte-identical
    to before. With `dynamic_data` (a list of per-month DB records, oldest→newest) →
    those months are appended as extra columns in the same SOP format, on top of the
    verified history. Tabs 1 & 2 grow; the THOR reconciliation + rules tabs stay verified.

    `include_verified=False` (demo mode only) strips the hardcoded verified history
    entirely — only `dynamic_data` months appear, and the THOR reconciliation + Rules
    tabs (which embed real verified figures) are omitted. Default True reproduces
    existing behaviour byte-for-byte.

    Tabs:
      1. P&L — Full (incl Cash)   — all items including cash income + cash rent
      2. Bank — Digital Only       — same layout; cash income + cash rent excluded
      3. THOR Bank Reconciliation  — omitted when include_verified=False
      4. Rules Applied             — methodology notes; omitted when include_verified=False
    """
    wb = openpyxl.Workbook()

    # Build the (possibly extended) line dicts + snapshot kwargs for tabs 1 & 2.
    if dynamic_data and not include_verified:
        # Demo mode: dynamic DB months only — no verified history embedded.
        inc_pm, opex_pm, excl_pm = [], [], []
        for d in dynamic_data:
            i, o, e = _dynamic_line_values(d)
            inc_pm.append(i); opex_pm.append(o); excl_pm.append(e)

        def _collect(per_month: List[dict]) -> Dict[str, List[int]]:
            keys: List[str] = []
            for pm in per_month:
                for k in pm:
                    if k not in keys:
                        keys.append(k)
            return {k: [pm.get(k, 0) for pm in per_month] for k in keys}

        months   = [d["label"] for d in dynamic_data]
        income   = _collect(inc_pm)
        opex     = _collect(opex_pm)
        excluded = _collect(excl_pm)
        capital: Dict[str, List[int]] = {}
        deposits = {
            _KEY_DEP_SEC:   [d.get("sec_dep", 0) for d in dynamic_data],
            _KEY_DEP_MAINT: [d.get("maint", 0)   for d in dynamic_data],
        }
        last = dynamic_data[-1]
        dyn_kwargs = dict(
            months=months, deposits=deposits, excluded=excluded, capital=capital,
            bank_thor_close=last.get("bank_thor_close", 0),
            bank_hulk_close=last.get("bank_hulk_close", 0),
            cash_in_hand={"Cash in hand (physical count)": last.get("cash_holding", 0)},
            sec_owed=last.get("sec_owed_total", 0),
            snapshot_label=last["label"],
        )
    elif not dynamic_data and not include_verified:
        # Demo mode, no DB months yet — empty shell workbook.
        months, income, opex = [], {}, {}
        dyn_kwargs = {
            "months": [],
            "deposits": {_KEY_DEP_SEC: [], _KEY_DEP_MAINT: []},
            "excluded": {_KEY_EXCL_REFUND: []},
            "capital": {},
        }
    elif dynamic_data:
        # sanity: every translation key must still exist in the verified dicts
        for k in (_KEY_INCOME_THOR, _KEY_INCOME_HULK, _KEY_INCOME_CASH):
            assert k in INCOME, f"income key drift: {k!r}"
        assert _KEY_RENT_CASH in OPEX, "opex key drift (rent cash)"
        for k in (_KEY_EXCL_REFUND, _KEY_EXCL_NONOP):
            assert k in EXCLUDED, f"excluded key drift: {k!r}"
        for k in (_KEY_DEP_SEC, _KEY_DEP_MAINT):
            assert k in DEPOSITS, f"deposits key drift: {k!r}"

        inc_pm, opex_pm, excl_pm = [], [], []
        for d in dynamic_data:
            i, o, e = _dynamic_line_values(d)
            inc_pm.append(i); opex_pm.append(o); excl_pm.append(e)

        months   = MONTHS + [d["label"] for d in dynamic_data]
        income   = _extend_dict(INCOME, inc_pm)
        opex     = _extend_dict(OPEX,   opex_pm)
        excluded = _extend_dict(EXCLUDED, excl_pm)
        capital  = {k: list(v) + [0] * len(dynamic_data) for k, v in CAPITAL_CONTRIBUTIONS.items()}
        deposits = {
            _KEY_DEP_SEC:   list(DEPOSITS[_KEY_DEP_SEC])   + [d.get("sec_dep", 0) for d in dynamic_data],
            _KEY_DEP_MAINT: list(DEPOSITS[_KEY_DEP_MAINT]) + [d.get("maint", 0)   for d in dynamic_data],
        }
        last = dynamic_data[-1]
        dyn_kwargs = dict(
            months=months, deposits=deposits, excluded=excluded, capital=capital,
            bank_thor_close=last.get("bank_thor_close", 0),
            bank_hulk_close=last.get("bank_hulk_close", 0),
            cash_in_hand={"Cash in hand (physical count)": last.get("cash_holding", 0)},
            sec_owed=last.get("sec_owed_total", 0),
            snapshot_label=last["label"],
        )
    else:
        income, opex = INCOME, OPEX
        dyn_kwargs = {}

    # ── Tab 1: Full P&L (canonical — all items) ────────────────────────────────
    ws_full = wb.active
    assert ws_full is not None
    ws_full.title = "P&L — Full (incl Cash)"
    _write_pnl_tab(ws_full, income, opex, **dyn_kwargs)

    # ── Tab 2: Bank / Digital-only P&L ────────────────────────────────────────
    # Removes: cash income from tenants + cash rent paid to owners
    # Useful for: loan applications, bank-verified reporting, digital economy view
    ws_bank = wb.create_sheet("Bank — Digital Only")
    income_bank = {k: v for k, v in income.items() if _CASH_INCOME_KEY not in k}
    opex_bank   = {k: v for k, v in opex.items()
                   if _CASH_RENT_KEY not in k and "paid in cash" not in k}
    _write_pnl_tab(
        ws_bank,
        income_bank,
        opex_bank,
        tab_note="Bank / Digital Only — excl. cash income from tenants + cash rent paid to owners",
        **dyn_kwargs,
    )

    if not include_verified:
        # Demo mode: skip the THOR reconciliation + Rules tabs — both embed real
        # verified figures (bank balances, staff names/phone numbers) that must
        # never appear in a public demo instance.
        return wb

    # ── Tab 3: THOR Bank Reconciliation ───────────────────────────────────────
    # Shows actual bank cash flow — Opening + bank credits - bank debits = Closing
    # Implied outflows = everything that left THOR bank acct (OPEX + refunds + repayments + withdrawals)
    ws_recon = wb.create_sheet("THOR Bank Reconciliation")

    # Capital flows that actually hit the THOR bank account (not cash/personal advances)
    _thor_cap_in  = THOR_CAP_IN
    _thor_cap_out = THOR_CAP_OUT

    _thor_income  = INCOME["THOR — Bank Income (UPI + NEFT)"]
    _thor_open    = [BANK_BALANCE_THOR[m][0] for m in MONTHS]
    _thor_close   = [BANK_BALANCE_THOR[m][1] for m in MONTHS]

    # Balancing figure — all debits from THOR bank (OPEX from bank + deposit refunds + repayments + withdrawals)
    _thor_out = [
        _thor_open[i] + _thor_income[i] + _thor_cap_in[i] - _thor_cap_out[i] - _thor_close[i]
        for i in range(len(MONTHS))
    ]

    R_HDR  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    R_OPEN = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    R_INC  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    R_CAP  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    R_OUT  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    R_CLOSE= PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    R_NOTE = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    WBOLD  = Font(bold=True, color="FFFFFF")

    def _recon_row(ws, label, sign, values, fill, bold=False, total=True):
        row_vals = [label, sign] + values + ([sum(v for v in values if v is not None)] if total else [None])
        ws.append(row_vals)
        for c in ws[ws.max_row]:
            c.fill = fill
            if bold:
                c.font = Font(bold=True)
        for i, c in enumerate(ws[ws.max_row]):
            if i > 1 and isinstance(c.value, (int, float)):
                c.number_format = INR_NUMBER_FORMAT

    # Header
    recon_hdr = ["THOR acct ...0961 — Bank Cash Flow Reconciliation", ""] + MONTHS + ["TOTAL"]
    ws_recon.append(recon_hdr)
    for c in ws_recon[1]:
        c.fill = R_HDR; c.font = WBOLD; c.alignment = Alignment(horizontal="center")

    ws_recon.append(["How the bank balance changed each month — Opening + Credits − Debits = Closing", ""] + [""] * (len(MONTHS) + 1))
    ws_recon[ws_recon.max_row][0].font = Font(italic=True, color="595959")
    ws_recon.append([])

    _recon_row(ws_recon, "Opening Balance",                   "",  _thor_open,   R_OPEN,  bold=True,  total=False)
    _recon_row(ws_recon, "+ Bank Income (UPI + NEFT)",        "+", _thor_income, R_INC,   bold=False)
    _recon_row(ws_recon, "+ Capital injected into bank",      "+", _thor_cap_in, R_CAP,   bold=False)
    _recon_row(ws_recon, "− Capital repaid / transferred out","−", _thor_cap_out,R_CAP,   bold=False)

    # Available subtotal
    _avail = [_thor_open[i] + _thor_income[i] + _thor_cap_in[i] - _thor_cap_out[i] for i in range(len(MONTHS))]
    _recon_row(ws_recon, "= Available before outflows",       "=", _avail,       R_OPEN,  bold=True)

    ws_recon.append([])
    _recon_row(ws_recon, "− Implied Bank Outflows (see note below)", "−", _thor_out, R_OUT, bold=True)
    ws_recon.append([])
    _recon_row(ws_recon, "= Closing Balance (verified from statement)", "=", _thor_close, R_CLOSE, bold=True, total=False)

    ws_recon.append([])
    ws_recon.append([])

    # Explanation note
    notes = RECON_NOTES
    for label, detail in notes:
        ws_recon.append([label, detail])
        row = ws_recon[ws_recon.max_row]
        row[0].fill = R_NOTE
        row[1].fill = R_NOTE
        if label.startswith("Why") or label.startswith("Implied"):
            row[0].font = Font(bold=True)

    ws_recon.column_dimensions["A"].width = 55
    ws_recon.column_dimensions["B"].width = 6
    for i, col in enumerate("CDEFGHIJ"):
        ws_recon.column_dimensions[col].width = 14
    ws_recon.row_dimensions[1].height = 20
    ws_recon.freeze_panes = "C2"

    # ── Tab 4: Rules Applied ───────────────────────────────────────────────────
    ws_rules = wb.create_sheet("Rules Applied")
    rules = RULES_APPLIED
    ws_rules.append(["Rule", "Detail"])
    for c in ws_rules[1]:
        c.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")
    for k, v in rules:
        ws_rules.append([k, v])
    ws_rules.column_dimensions["A"].width = 40
    ws_rules.column_dimensions["B"].width = 120
    for row in ws_rules.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def build_pnl_bytes(dynamic_data: List[dict] | None = None, include_verified: bool = True) -> bytes:
    """Return the P&L workbook as bytes (for streaming from FastAPI).

    Pass `dynamic_data` (per-month DB records) to append new months in SOP format.
    `include_verified=False` (demo mode) strips the hardcoded verified history."""
    buf = io.BytesIO()
    build_pnl_workbook(dynamic_data, include_verified=include_verified).save(buf)
    buf.seek(0)
    return buf.read()
