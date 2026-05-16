"""
src/api/v2/finance.py
Finance endpoints — CSV upload, P&L dashboard, Excel download.
Owner-only (admin role required on every endpoint).
"""
from __future__ import annotations

import hashlib
import io
import logging
import re as _re
import calendar as _calendar
from collections import defaultdict
from datetime import date, datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import extract, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from src.api.v2.auth import AppUser, get_current_user
from src.database.db_manager import get_session
from src.database.models import (
    BankTransaction, BankUpload, CashCount, CashExpense, CheckoutRecord,
    InvestmentExpense, Payment, PaymentFor, PaymentMode, Tenancy, Tenant, UpiCollectionEntry,
)
from src.parsers.yes_bank import read_yes_bank_csv
from src.reports.pnl_builder import build_pnl_bytes
from src.rules.pnl_classify import classify_txn
from src.utils.inr_format import INR_NUMBER_FORMAT

logger = logging.getLogger(__name__)
router = APIRouter()

EXPENSE_CATS = [
    "Property Rent", "Electricity", "Water", "IT & Software", "Internet & WiFi",
    "Food & Groceries", "Fuel & Diesel", "Staff & Labour", "Furniture & Fittings",
    "Maintenance & Repairs", "Cleaning Supplies", "Waste Disposal",
    "Shopping & Supplies", "Operational Expenses", "Marketing",
    "Govt & Regulatory", "Tenant Deposit Refund", "Bank Charges",
    "Capital Investment", "Non-Operating", "Other Expenses",
]

# Only these categories reduce operating profit
_OPEX_CATS_JSON = {
    "Property Rent", "Electricity", "Water", "IT & Software", "Internet & WiFi",
    "Food & Groceries", "Fuel & Diesel", "Staff & Labour",
    "Maintenance & Repairs", "Cleaning Supplies", "Waste Disposal",
    "Shopping & Supplies", "Operational Expenses", "Marketing",
    "Govt & Regulatory", "Bank Charges", "Other Expenses",
}
# Shown separately below operating profit
_CAPEX_CATS_JSON = {"Furniture & Fittings", "Capital Investment"}
# Balance-sheet items — shown for info only, never deducted from any profit line
_EXCL_CATS_JSON  = {"Tenant Deposit Refund", "Non-Operating"}


def _require_admin(user: AppUser):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")


def _make_hash(txn_date: date, amount: float, desc: str) -> str:
    key = f"{txn_date}|{(desc or '').strip().lower()[:80]}|{round(float(amount), 2)}"
    return hashlib.sha256(key.encode()).hexdigest()


async def _auto_reconcile(session):
    """
    For every bank_transaction with category='Tenant Deposit Refund' and no
    reconciled_checkout_id yet, try to match to a checkout_records row by:
      - amount within ±1 rupee
      - deposit_refund_date within ±7 days of txn_date
    """
    import datetime as _dt

    try:
        unmatched = (await session.scalars(
            select(BankTransaction).where(
                BankTransaction.category == "Tenant Deposit Refund",
                BankTransaction.reconciled_checkout_id == None,
            )
        )).all()

        for txn in unmatched:
            window_start = txn.txn_date - _dt.timedelta(days=7)
            window_end   = txn.txn_date + _dt.timedelta(days=7)
            match = await session.scalar(
                select(CheckoutRecord).where(
                    CheckoutRecord.deposit_refunded_amount.between(
                        float(txn.amount) - 1, float(txn.amount) + 1
                    ),
                    CheckoutRecord.deposit_refund_date.between(window_start, window_end),
                )
            )
            if match:
                txn.reconciled_checkout_id = match.id

        await session.flush()
    except Exception:
        logger.exception("_auto_reconcile failed — upload committed without reconciliation")


def _validate_month(m: str, param: str = "month"):
    if not _re.fullmatch(r"\d{4}-\d{2}", m):
        raise HTTPException(status_code=400, detail=f"{param} must be YYYY-MM")


# ── Cash position ─────────────────────────────────────────────────────────────

@router.get("/finance/cash")
async def get_cash_position(
    month: str = Query(..., description="YYYY-MM"),
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    _validate_month(month)
    year, month_num = int(month[:4]), int(month[5:7])

    async with get_session() as session:
        collected = float(await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.payment_mode == PaymentMode.cash,
                Payment.for_type == PaymentFor.rent,
                Payment.is_void == False,
                extract("year", Payment.payment_date) == year,
                extract("month", Payment.payment_date) == month_num,
            )
        ) or 0)

        expense_rows = (await session.scalars(
            select(CashExpense).where(
                extract("year", CashExpense.date) == year,
                extract("month", CashExpense.date) == month_num,
                CashExpense.is_void == False,
            ).order_by(CashExpense.date.desc(), CashExpense.id.desc())
        )).all()

        expenses_total = sum(float(e.amount) for e in expense_rows)
        balance = collected - expenses_total

        last_count_row = await session.scalar(
            select(CashCount).where(
                extract("year", CashCount.date) == year,
                extract("month", CashCount.date) == month_num,
            ).order_by(CashCount.created_at.desc())
        )

        history = []
        for i in range(6):
            hm = month_num - i
            hy = year
            while hm <= 0:
                hm += 12
                hy -= 1
            h_col = float(await session.scalar(
                select(func.coalesce(func.sum(Payment.amount), 0)).where(
                    Payment.payment_mode == PaymentMode.cash,
                    Payment.for_type == PaymentFor.rent,
                    Payment.is_void == False,
                    extract("year", Payment.payment_date) == hy,
                    extract("month", Payment.payment_date) == hm,
                )
            ) or 0)
            h_exp = float(await session.scalar(
                select(func.coalesce(func.sum(CashExpense.amount), 0)).where(
                    extract("year", CashExpense.date) == hy,
                    extract("month", CashExpense.date) == hm,
                    CashExpense.is_void == False,
                )
            ) or 0)
            history.append({
                "month": f"{hy}-{hm:02d}",
                "collected": h_col,
                "expenses": h_exp,
                "balance": h_col - h_exp,
            })

        return {
            "month": month,
            "collected": collected,
            "expenses_total": expenses_total,
            "balance": balance,
            "last_count": {
                "id": last_count_row.id,
                "date": str(last_count_row.date),
                "amount": float(last_count_row.amount),
                "counted_by": last_count_row.counted_by,
                "variance": balance - float(last_count_row.amount),
            } if last_count_row else None,
            "expenses": [
                {
                    "id": e.id,
                    "date": str(e.date),
                    "description": e.description,
                    "amount": float(e.amount),
                    "paid_by": e.paid_by,
                    "is_void": e.is_void,
                }
                for e in expense_rows
            ],
            "history": history,
        }


@router.post("/finance/cash/expenses")
async def add_cash_expense(
    body: dict,
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    missing = {"date", "description", "amount", "paid_by"} - set(body.keys())
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing fields: {missing}")
    try:
        from datetime import date as _date_type
        exp_date = _date_type.fromisoformat(body["date"])
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    try:
        amount = float(body["amount"])
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="amount must be a positive number")
    if body["paid_by"] not in ("Prabhakaran", "Lakshmi", "Other"):
        raise HTTPException(status_code=400, detail="paid_by must be Prabhakaran, Lakshmi, or Other")

    async with get_session() as session:
        expense = CashExpense(
            date=exp_date,
            description=str(body["description"]).strip(),
            amount=amount,
            paid_by=body["paid_by"],
            created_by=user.phone,
        )
        session.add(expense)
        await session.flush()
        result = {
            "id": expense.id,
            "date": str(expense.date),
            "description": expense.description,
            "amount": float(expense.amount),
            "paid_by": expense.paid_by,
            "is_void": expense.is_void,
        }
        await session.commit()
    return result


@router.patch("/finance/cash/expenses/{expense_id}")
async def edit_cash_expense(
    expense_id: int,
    body: dict,
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    async with get_session() as session:
        expense = await session.get(CashExpense, expense_id)
        if not expense:
            raise HTTPException(status_code=404, detail="Expense not found")
        if expense.is_void:
            raise HTTPException(status_code=400, detail="Cannot edit a voided expense")
        if "date" in body:
            try:
                from datetime import date as _date_type
                expense.date = _date_type.fromisoformat(body["date"])
            except ValueError:
                raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        if "description" in body:
            desc = str(body["description"]).strip()
            if not desc:
                raise HTTPException(status_code=400, detail="description cannot be empty")
            expense.description = desc
        if "amount" in body:
            try:
                amount = float(body["amount"])
                if amount <= 0:
                    raise ValueError
                expense.amount = amount
            except (ValueError, TypeError):
                raise HTTPException(status_code=400, detail="amount must be a positive number")
        if "paid_by" in body:
            if body["paid_by"] not in ("Prabhakaran", "Lakshmi", "Other"):
                raise HTTPException(status_code=400, detail="paid_by must be Prabhakaran, Lakshmi, or Other")
            expense.paid_by = body["paid_by"]
        await session.commit()
    return {
        "id": expense.id,
        "date": str(expense.date),
        "description": expense.description,
        "amount": float(expense.amount),
        "paid_by": expense.paid_by,
        "is_void": expense.is_void,
    }


@router.delete("/finance/cash/expenses/{expense_id}")
async def void_cash_expense(
    expense_id: int,
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    from datetime import datetime, timezone
    async with get_session() as session:
        expense = await session.get(CashExpense, expense_id)
        if not expense:
            raise HTTPException(status_code=404, detail="Expense not found")
        if expense.is_void:
            raise HTTPException(status_code=400, detail="Already voided")
        expense.is_void = True
        expense.voided_at = datetime.now(timezone.utc)
        await session.commit()
    return {"ok": True, "id": expense_id}


@router.post("/finance/cash/counts")
async def log_cash_count(
    body: dict,
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    missing = {"date", "amount", "counted_by"} - set(body.keys())
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing fields: {missing}")
    try:
        from datetime import date as _date_type
        count_date = _date_type.fromisoformat(body["date"])
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    try:
        amount = float(body["amount"])
        if amount < 0:
            raise ValueError
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="amount must be a non-negative number")
    if body["counted_by"] not in ("Prabhakaran", "Lakshmi"):
        raise HTTPException(status_code=400, detail="counted_by must be Prabhakaran or Lakshmi")

    async with get_session() as session:
        count = CashCount(
            date=count_date,
            amount=amount,
            counted_by=body["counted_by"],
            notes=body.get("notes"),
        )
        session.add(count)
        await session.flush()
        result = {
            "id": count.id,
            "date": str(count.date),
            "amount": amount,
            "counted_by": count.counted_by,
            "notes": count.notes,
        }
        await session.commit()
    return result


# ── Upload ────────────────────────────────────────────────────────────────────

@router.post("/finance/upload")
async def upload_bank_csv(
    files: list[UploadFile] = File(...),
    account_name: str = Form(...),
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    if account_name not in ("THOR", "HULK"):
        raise HTTPException(status_code=400, detail="account_name must be THOR or HULK")

    total_new = 0
    total_dup = 0
    months_affected: set[str] = set()

    async with get_session() as session:
        for upload_file in files:
            content = await upload_file.read()

            rows = read_yes_bank_csv(content)
            if not rows:
                continue

            from_date = min(r[0] for r in rows)
            to_date   = max(r[0] for r in rows)

            upload = BankUpload(
                phone=user.phone,
                file_path=upload_file.filename or "",
                row_count=len(rows),
                new_count=0,
                from_date=from_date,
                to_date=to_date,
                status="processed",
                account_name=account_name,
            )
            session.add(upload)
            await session.flush()

            new_count = 0
            for txn_date, desc, txn_type, amount in rows:
                # Normalise description FIRST, then hash — so stored value always
                # matches the hash. This is the canonical dedup key.
                norm_desc = (desc or "").strip()
                uhash = _make_hash(txn_date, amount, norm_desc)
                cat, sub = classify_txn(norm_desc, txn_type)

                # DB-atomic insert: ON CONFLICT on both the hash index AND the
                # composite content index (uq_btxn_content) → silently skip dups
                stmt = (
                    pg_insert(BankTransaction)
                    .values(
                        upload_id=upload.id,
                        txn_date=txn_date,
                        description=norm_desc,
                        amount=amount,
                        txn_type=txn_type,
                        category=cat,
                        sub_category=sub,
                        unique_hash=uhash,
                        account_name=account_name,
                    )
                    .on_conflict_do_nothing(index_elements=["unique_hash"])
                    .returning(BankTransaction.id)
                )
                inserted_id = await session.scalar(stmt)
                if inserted_id:
                    months_affected.add(txn_date.strftime("%Y-%m"))
                    new_count += 1
                else:
                    total_dup += 1

            upload.new_count = new_count
            total_new += new_count

        await _auto_reconcile(session)
        await session.commit()

    return {
        "months_affected": sorted(months_affected),
        "new_count": total_new,
        "duplicate_count": total_dup,
    }


# ── P&L ───────────────────────────────────────────────────────────────────────

@router.get("/finance/pnl")
async def get_pnl(
    month: Optional[str] = Query(None, description="YYYY-MM, e.g. 2026-05"),
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    if month:
        _validate_month(month)

    async with get_session() as session:
        month_rows = await session.execute(
            select(func.to_char(BankTransaction.txn_date, "YYYY-MM"))
            .distinct()
            .order_by(func.to_char(BankTransaction.txn_date, "YYYY-MM"))
        )
        available_months = [r[0] for r in month_rows]

        target_months = [month] if month else available_months
        if not target_months:
            return {"months": [], "data": {}}

        result = {}
        for m in target_months:
            y, mo = int(m[:4]), int(m[5:7])
            start = date(y, mo, 1)
            end = date(y, mo, _calendar.monthrange(y, mo)[1])
            period_date = date(y, mo, 1)  # Payment.period_month is stored as 1st of month

            inc_rows = await session.execute(
                select(BankTransaction.category, func.sum(BankTransaction.amount))
                .where(
                    BankTransaction.txn_type == "income",
                    BankTransaction.txn_date.between(start, end),
                    BankTransaction.category.notin_(["Capital Investment", "Advance Deposit"]),
                )
                .group_by(BankTransaction.category)
            )
            bank_income = {row[0]: float(row[1]) for row in inc_rows}
            # "Rent Income" = UPI collection settlements + direct UPI from tenants
            # "Other Income" = NEFT/RTGS inward, cashback — deposits excluded above
            upi_batch   = bank_income.get("Rent Income", 0.0)
            direct_neft = sum(v for k, v in bank_income.items() if k != "Rent Income")

            cash_rows = await session.execute(
                select(func.sum(Payment.amount))
                .where(
                    Payment.payment_mode == PaymentMode.cash,
                    Payment.for_type == PaymentFor.rent,
                    Payment.is_void == False,
                    Payment.period_month == period_date,
                )
            )
            cash_db = float(cash_rows.scalar() or 0)

            cap_rows = await session.execute(
                select(func.sum(BankTransaction.amount))
                .where(
                    BankTransaction.category == "Capital Investment",
                    BankTransaction.txn_date.between(start, end),
                )
            )
            capital = float(cap_rows.scalar() or 0)

            exp_rows = await session.execute(
                select(BankTransaction.category, func.sum(BankTransaction.amount))
                .where(
                    BankTransaction.txn_type == "expense",
                    BankTransaction.txn_date.between(start, end),
                )
                .group_by(BankTransaction.category)
            )
            expense_by_cat = {row[0]: float(row[1]) for row in exp_rows}

            expenses = [
                {"category": cat, "amount": expense_by_cat.get(cat, 0.0)}
                for cat in EXPENSE_CATS
                if expense_by_cat.get(cat, 0.0) > 0 and cat in _OPEX_CATS_JSON
            ]
            capex_items = [
                {"category": cat, "amount": expense_by_cat.get(cat, 0.0)}
                for cat in EXPENSE_CATS
                if expense_by_cat.get(cat, 0.0) > 0 and cat in _CAPEX_CATS_JSON
            ]
            excluded_items = [
                {"category": cat, "amount": expense_by_cat.get(cat, 0.0)}
                for cat in EXPENSE_CATS
                if expense_by_cat.get(cat, 0.0) > 0 and cat in _EXCL_CATS_JSON
            ]

            # Security deposits collected this month (active tenants, by check-in month)
            sec_dep_row = await session.execute(
                select(func.sum(Tenancy.security_deposit))
                .where(
                    func.extract("year",  Tenancy.checkin_date) == y,
                    func.extract("month", Tenancy.checkin_date) == mo,
                    Tenancy.status == "active",
                )
            )
            security_deposits = float(sec_dep_row.scalar() or 0)

            total_expense    = sum(e["amount"] for e in expenses)
            total_capex      = sum(e["amount"] for e in capex_items)
            total_gross      = upi_batch + direct_neft + cash_db
            true_revenue     = total_gross - security_deposits
            operating_profit = true_revenue - total_expense
            net_profit       = operating_profit - total_capex
            margin           = round(operating_profit / true_revenue * 100, 1) if true_revenue else 0.0

            result[m] = {
                "month": m,
                "income": {
                    "upi_batch":         upi_batch,
                    "direct_neft":       direct_neft,
                    "cash_db":           cash_db,
                    "total":             total_gross,
                    "security_deposits": security_deposits,
                    "true_revenue":      true_revenue,
                },
                "capital":          capital,
                "expenses":         expenses,
                "capex_items":      capex_items,
                "excluded_items":   excluded_items,
                "total_expense":    total_expense,
                "total_capex":      total_capex,
                "operating_profit": operating_profit,
                "net_profit":       net_profit,
                "margin_pct":       margin,
            }

    return {"months": target_months, "data": result}


# ── Excel Download — P&L Summary ──────────────────────────────────────────────

# Categories that belong in operating expenses (reduce operating profit)
_OPEX_CATS = [
    "Property Rent", "Electricity", "Water", "IT & Software", "Internet & WiFi",
    "Food & Groceries", "Fuel & Diesel", "Staff & Labour", "Maintenance & Repairs",
    "Cleaning Supplies", "Waste Disposal", "Shopping & Supplies",
    "Operational Expenses", "Marketing", "Govt & Regulatory", "Bank Charges",
    "Other Expenses",
]
# Balance-sheet items — shown for reference but NOT deducted from operating profit
_EXCLUDED_CATS = ["Tenant Deposit Refund", "Non-Operating"]
# Capital expenditure — shown below operating profit
_CAPEX_CATS = ["Furniture & Fittings", "Capital Investment"]


@router.get("/finance/pnl/excel")
async def download_pnl_excel(user: AppUser = Depends(get_current_user)):
    """Verified canonical P&L (Oct'25–Apr'26) — same as local export script."""
    _require_admin(user)
    buf = io.BytesIO(build_pnl_bytes())
    filename = f"PnL_Cozeevo_Verified_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/finance/pnl/live")
async def download_pnl_live(user: AppUser = Depends(get_current_user)):
    """
    Live P&L recomputed from DB — picks up any new CSV uploads (HULK, new months).
    Reclassifies every transaction with the current classifier rules.
    """
    _require_admin(user)

    async with get_session() as session:
        txn_rows = (await session.scalars(
            select(BankTransaction).order_by(BankTransaction.txn_date)
        )).all()

        all_months = sorted({r.txn_date.strftime("%Y-%m") for r in txn_rows})
        if not all_months:
            raise HTTPException(status_code=404, detail="No bank transactions in DB")

        by_cat_month: dict = defaultdict(lambda: defaultdict(float))
        by_sub_month: dict = defaultdict(float)
        bank_upi_batch: dict = defaultdict(float)
        bank_direct:    dict = defaultdict(float)

        for r in txn_rows:
            m = r.txn_date.strftime("%Y-%m")
            cat, sub = classify_txn(r.description or "", r.txn_type or "")
            if r.txn_type == "income":
                if cat == "Rent Income":
                    bank_upi_batch[m] += float(r.amount)
                elif cat != "Advance Deposit":
                    bank_direct[m] += float(r.amount)
            else:
                by_cat_month[cat][m] += float(r.amount)
                by_sub_month[(cat, sub, m)] += float(r.amount)

        cash_rows = await session.execute(
            select(
                func.to_char(Payment.period_month, "YYYY-MM"),
                func.sum(Payment.amount),
            )
            .where(
                Payment.payment_mode == PaymentMode.cash,
                Payment.for_type == PaymentFor.rent,
                Payment.is_void == False,
            )
            .group_by(func.to_char(Payment.period_month, "YYYY-MM"))
        )
        cash_by_month: dict = {row[0]: float(row[1]) for row in cash_rows}

        # Deposits HELD — active tenants only, split by check-in month
        held_sec_rows = await session.execute(
            select(
                func.to_char(func.date_trunc("month", Tenancy.checkin_date), "YYYY-MM"),
                func.sum(Tenancy.security_deposit),
                func.sum(Tenancy.maintenance_fee),
            )
            .where(Tenancy.status == "active")
            .group_by(func.date_trunc("month", Tenancy.checkin_date))
        )
        dep_held_sec:   dict = {}
        dep_held_maint: dict = {}
        for row in held_sec_rows:
            if row[0]:
                dep_held_sec[row[0]]   = float(row[1] or 0)
                dep_held_maint[row[0]] = float(row[2] or 0)

    wb = _build_pnl_excel(all_months, bank_upi_batch, bank_direct, cash_by_month,
                          by_cat_month, by_sub_month,
                          dep_held_sec, dep_held_maint)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"PnL_Live_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _month_label(m: str) -> str:
    """'2025-10' → \"Oct'25\""""
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    y, mo = int(m[:4]), int(m[5:7])
    return f"{months[mo-1]}'{str(y)[2:]}"


def _build_pnl_excel(
    all_months: list,
    bank_upi: dict, bank_direct: dict, cash_db: dict,
    by_cat: dict, by_sub: dict,
    dep_held_sec: dict | None = None,
    dep_held_maint: dict | None = None,
) -> openpyxl.Workbook:
    HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    SEC_FILL  = PatternFill("solid", fgColor="2E3F5C")
    SEC_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EEF4FF")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
    GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")
    RED_FILL  = PatternFill("solid", fgColor="FCE4D6")
    EXC_FILL  = PatternFill("solid", fgColor="FFF2CC")
    CAP_FILL  = PatternFill("solid", fgColor="EDEDED")
    BOLD      = Font(bold=True)
    BOLD_GRN  = Font(bold=True, color="375623")
    BOLD_RED  = Font(bold=True, color="9C0006")
    CTR       = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    NUM_FMT   = INR_NUMBER_FORMAT

    nc = len(all_months) + 2  # label col + month cols + total col

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Summary"

    def _hdr(col, val, width=14):
        c = ws.cell(1, col, val)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
        ws.column_dimensions[get_column_letter(col)].width = width

    # Header row
    _hdr(1, "Category", width=36)
    for ci, m in enumerate(all_months, 2):
        _hdr(ci, _month_label(m))
    _hdr(nc, "TOTAL")
    ws.row_dimensions[1].height = 22

    ri = [2]  # mutable row counter

    def _section(label):
        for col in range(1, nc + 1):
            c = ws.cell(ri[0], col)
            c.fill = SEC_FILL
        ws.cell(ri[0], 1, label).font = SEC_FONT
        ws.cell(ri[0], 1).alignment = LEFT
        ws.row_dimensions[ri[0]].height = 18
        ri[0] += 1

    def _data_row(label, vals_by_month, fill, font=None, indent=False, num_fmt=NUM_FMT):
        lbl = f"  {label}" if indent else label
        c = ws.cell(ri[0], 1, lbl)
        c.fill = fill
        if font: c.font = font
        else: c.font = Font(size=9)
        c.alignment = LEFT
        total = 0.0
        for ci, m in enumerate(all_months, 2):
            v = vals_by_month.get(m, 0.0)
            total += v
            cell = ws.cell(ri[0], ci, round(v) if v else None)
            cell.fill = fill
            if v: cell.number_format = num_fmt
            if font: cell.font = font
            else: cell.font = Font(size=9)
            cell.alignment = RIGHT
        tc = ws.cell(ri[0], nc, round(total) if total else None)
        tc.fill = fill
        if total: tc.number_format = num_fmt
        if font: tc.font = font
        else: tc.font = Font(size=9, bold=True)
        tc.alignment = RIGHT
        ri[0] += 1
        return total

    def _total_row(label, totals_by_month, fill, font):
        c = ws.cell(ri[0], 1, label)
        c.fill = fill; c.font = font; c.alignment = LEFT
        grand = 0.0
        for ci, m in enumerate(all_months, 2):
            v = totals_by_month.get(m, 0.0)
            grand += v
            cell = ws.cell(ri[0], ci, round(v) if v else None)
            cell.fill = fill; cell.font = font; cell.alignment = RIGHT
            if v: cell.number_format = NUM_FMT
        tc = ws.cell(ri[0], nc, round(grand) if grand else None)
        tc.fill = fill; tc.font = font; tc.alignment = RIGHT
        if grand: tc.number_format = NUM_FMT
        ws.row_dimensions[ri[0]].height = 16
        ri[0] += 1
        return grand

    def _blank():
        ri[0] += 1

    # ── GROSS INFLOWS ─────────────────────────────────────────────────────────
    _section("GROSS INFLOWS")
    gross_month: dict = defaultdict(float)
    for m in all_months:
        gross_month[m] = bank_upi.get(m, 0) + bank_direct.get(m, 0) + cash_db.get(m, 0)

    fill = ALT_FILL
    for label, src in [
        ("Bank — UPI batch settlements (merchant QR)", bank_upi),
        ("Bank — direct UPI / NEFT payments",          bank_direct),
        ("Cash collected (not deposited to bank)",     cash_db),
    ]:
        if sum(src.values()):
            _data_row(label, src, fill, indent=True)
            fill = WHT_FILL if fill == ALT_FILL else ALT_FILL

    _total_row("Total Gross Inflows", gross_month, GRN_FILL, BOLD_GRN)

    # ── Security deposit deduction (refundable — must return at exit) ─────────
    dep_held_sec   = dep_held_sec   or {}
    dep_held_maint = dep_held_maint or {}
    LIA_FILL = PatternFill("solid", fgColor="FCE4D6")
    ITLC     = Font(size=9, italic=True)

    sec_dep_neg = {m: -dep_held_sec.get(m, 0) for m in all_months}
    if any(dep_held_sec.values()):
        _data_row("  Less: Security Deposits Received (refundable — must return at exit)",
                  sec_dep_neg, LIA_FILL, font=ITLC)
        true_rev_month: dict = {m: gross_month[m] + sec_dep_neg.get(m, 0) for m in all_months}
        _total_row("True Rent Revenue (excl. refundable deposits)", true_rev_month, GRN_FILL, BOLD_GRN)
        if any(dep_held_maint.values()):
            _data_row("  Note: Maintenance Fee retained (non-refundable, included above)",
                      dep_held_maint, LIA_FILL, font=ITLC)
    else:
        true_rev_month = gross_month
    _blank()

    # ── OPERATING EXPENSES ────────────────────────────────────────────────────
    _section("OPERATING EXPENSES")
    opex_month: dict = defaultdict(float)
    for cat in _OPEX_CATS:
        cat_data = by_cat.get(cat, {})
        if not any(cat_data.values()):
            continue
        fill = ALT_FILL if ri[0] % 2 == 0 else WHT_FILL
        _data_row(cat, cat_data, fill, indent=True)
        for m, v in cat_data.items():
            opex_month[m] += v

    _total_row("Total Operating Expenses", opex_month, RED_FILL, BOLD_RED)
    _blank()

    # ── EBITDA / OPERATING PROFIT (true rent revenue − opex) ─────────────────
    op_profit_month: dict = {m: true_rev_month.get(m, 0) - opex_month.get(m, 0) for m in all_months}
    _total_row("EBITDA / Operating Profit", op_profit_month, GRN_FILL, BOLD_GRN)
    _blank()

    # ── EXCLUDED (balance sheet items — for reference only) ───────────────────
    has_excluded = any(by_cat.get(c, {}) for c in _EXCLUDED_CATS)
    if has_excluded:
        _section("BALANCE SHEET ITEMS (not deducted above)")
        for cat in _EXCLUDED_CATS:
            cat_data = by_cat.get(cat, {})
            if not any(cat_data.values()):
                continue
            fill = ALT_FILL if ri[0] % 2 == 0 else WHT_FILL
            _data_row(cat, cat_data, EXC_FILL, indent=True)
        _blank()

    # ── CAPEX ─────────────────────────────────────────────────────────────────
    capex_month: dict = defaultdict(float)
    has_capex = any(by_cat.get(c, {}) for c in _CAPEX_CATS)
    if has_capex:
        _section("CAPITAL EXPENDITURE (one-time investments)")
        for cat in _CAPEX_CATS:
            cat_data = by_cat.get(cat, {})
            if not any(cat_data.values()):
                continue
            _data_row(cat, cat_data, CAP_FILL, indent=True)
            for m, v in cat_data.items():
                capex_month[m] += v
        _total_row("Total CAPEX", capex_month, CAP_FILL, BOLD)
        _blank()

    # ── NET PROFIT ────────────────────────────────────────────────────────────
    net_month = {m: op_profit_month.get(m, 0) - capex_month.get(m, 0) for m in all_months}
    _total_row("Net Profit after CAPEX", net_month, GRN_FILL, BOLD_GRN)
    _blank()

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 22

    # ── Sheet 2: Sub-category Breakdown ──────────────────────────────────────
    ws2 = wb.create_sheet("Sub-category Detail")
    sub_hdrs = ["Category", "Sub-category"] + [_month_label(m) for m in all_months] + ["TOTAL"]
    for col, h in enumerate(sub_hdrs, 1):
        c = ws2.cell(1, col, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 38
    for i in range(3, len(all_months) + 4):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    ri2 = 2
    all_cats = _OPEX_CATS + _EXCLUDED_CATS + _CAPEX_CATS
    for cat in all_cats:
        subs = sorted({s for (c2, s, m), v in by_sub.items() if c2 == cat and v > 0})
        if not subs:
            continue
        c = ws2.cell(ri2, 1, cat)
        c.fill = SEC_FILL; c.font = SEC_FONT
        for col in range(2, len(all_months) + 4):
            ws2.cell(ri2, col).fill = SEC_FILL
        ri2 += 1
        for sub in subs:
            fill = ALT_FILL if ri2 % 2 == 0 else WHT_FILL
            ws2.cell(ri2, 2, sub).fill = fill
            sub_total = 0.0
            for ci, m in enumerate(all_months, 3):
                v = by_sub.get((cat, sub, m), 0)
                sub_total += v
                c2 = ws2.cell(ri2, ci, round(v) if v else None)
                c2.fill = fill
                if v: c2.number_format = NUM_FMT
            tc = ws2.cell(ri2, len(all_months) + 3, round(sub_total) if sub_total else None)
            tc.fill = fill
            if sub_total: tc.number_format = NUM_FMT; tc.font = Font(bold=True)
            ri2 += 1
    ws2.freeze_panes = "C2"

    return wb


def _build_excel(classified: list) -> openpyxl.Workbook:
    """Build same 3-sheet Excel as export_classified.py."""
    HDR_FILL = PatternFill("solid", fgColor="1a1a2e")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F8F9FA")
    WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
    CTR      = Alignment(horizontal="center", vertical="center")
    TOT_FONT_G = Font(bold=True, size=11, color="008B00")
    TOT_FONT_R = Font(bold=True, size=11, color="FF0000")
    TOT_FONT   = Font(bold=True, size=11)
    SEC_FILL   = PatternFill("solid", fgColor="2d2d44")
    SEC_FONT   = Font(bold=True, color="FFFFFF", size=11)

    months = sorted(set(dt.strftime("%Y-%m") for dt, *_ in classified))
    monthly_exp = defaultdict(lambda: defaultdict(float))
    monthly_inc = defaultdict(lambda: defaultdict(float))
    monthly_sub: dict = defaultdict(float)

    INCOME_CATS = ["Rent Income", "Advance Deposit", "Other Income"]

    for dt, desc, typ, amt, cat, sub in classified:
        m = dt.strftime("%Y-%m")
        if typ == "expense":
            monthly_exp[m][cat] += amt
            monthly_sub[(m, cat, sub)] += amt
        else:
            monthly_inc[m][cat] += amt

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Monthly P&L ──────────────────────────────────────────────────
    ws = wb.create_sheet("Monthly P&L")
    hdr_row = [""] + months + ["TOTAL"]
    for col, h in enumerate(hdr_row, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    ws.row_dimensions[1].height = 22

    def _write_row(ws, ri, label, vals, font=None, fill=None, indent=False):
        lbl = ("  " + label) if indent else label
        c = ws.cell(row=ri, column=1, value=lbl)
        if font: c.font = font
        if fill: c.fill = fill
        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=ri, column=ci, value=round(v) if v else "")
            c.number_format = INR_NUMBER_FORMAT
            if font: c.font = font
            if fill: c.fill = fill

    ri = 2
    c = ws.cell(row=ri, column=1, value="INCOME")
    c.fill = SEC_FILL; c.font = SEC_FONT
    for col in range(2, len(months) + 3):
        ws.cell(row=ri, column=col).fill = SEC_FILL
    ri += 1

    inc_totals = [0.0] * (len(months) + 1)
    for cat in INCOME_CATS:
        vals = [monthly_inc[m].get(cat, 0) for m in months]
        vals.append(sum(vals))
        if sum(vals) == 0:
            continue
        for i, v in enumerate(vals):
            inc_totals[i] += v
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        _write_row(ws, ri, cat, vals, indent=True, fill=fill)
        ri += 1
    _write_row(ws, ri, "TOTAL INCOME", inc_totals, font=TOT_FONT_G)
    ri += 2

    c = ws.cell(row=ri, column=1, value="OPERATING EXPENSES")
    c.fill = SEC_FILL; c.font = SEC_FONT
    for col in range(2, len(months) + 3):
        ws.cell(row=ri, column=col).fill = SEC_FILL
    ri += 1

    exp_totals = [0.0] * (len(months) + 1)
    op_cats = [c for c in EXPENSE_CATS if c != "Non-Operating"]
    for cat in op_cats:
        vals = [monthly_exp[m].get(cat, 0) for m in months]
        vals.append(sum(vals))
        if sum(vals) == 0:
            continue
        for i, v in enumerate(vals):
            exp_totals[i] += v
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        _write_row(ws, ri, cat, vals, indent=True, fill=fill)
        ri += 1
    _write_row(ws, ri, "TOTAL OPERATING EXPENSES", exp_totals, font=TOT_FONT_R)
    ri += 2

    op_profit = [inc_totals[i] - exp_totals[i] for i in range(len(inc_totals))]
    _write_row(ws, ri, "OPERATING PROFIT (EBITDA)", op_profit, font=TOT_FONT)
    ri += 2

    nonop = [monthly_exp[m].get("Non-Operating", 0) for m in months]
    nonop.append(sum(nonop))
    if sum(nonop) > 0:
        c = ws.cell(row=ri, column=1, value="NON-OPERATING EXPENSES")
        c.fill = SEC_FILL; c.font = SEC_FONT
        for col in range(2, len(months) + 3):
            ws.cell(row=ri, column=col).fill = SEC_FILL
        ri += 1
        _write_row(ws, ri, "Loan Repayment / Transfers", nonop, indent=True, fill=ALT_FILL)
        ri += 2

    net = [op_profit[i] - nonop[i] for i in range(len(op_profit))]
    _write_row(ws, ri, "NET PROFIT / (LOSS)", net, font=Font(bold=True, size=12))
    ri += 2

    c = ws.cell(row=ri, column=1, value="OPERATING MARGIN %")
    c.font = Font(bold=True, italic=True)
    for ci, m in enumerate(months, 2):
        inc = inc_totals[ci - 2]
        margin = (op_profit[ci - 2] / inc * 100) if inc else 0
        ws.cell(row=ri, column=ci, value=f"{margin:.1f}%").font = Font(italic=True)
    ws.column_dimensions["A"].width = 32
    for i in range(2, len(months) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "B2"

    # ── Sheet 2: Sub-category Breakdown ──────────────────────────────────────
    ws2 = wb.create_sheet("Sub-category Breakdown")
    sub_hdrs = ["Category", "Sub-category"] + months + ["TOTAL"]
    for col, h in enumerate(sub_hdrs, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR

    ri = 2
    for cat in EXPENSE_CATS:
        all_subs = sorted({s for (m, c2, s), v in monthly_sub.items() if c2 == cat and v > 0})
        if not all_subs:
            continue
        cat_total = sum(monthly_exp[m].get(cat, 0) for m in months)
        if cat_total == 0:
            continue
        c = ws2.cell(row=ri, column=1, value=cat)
        c.fill = PatternFill("solid", fgColor="2d2d44")
        c.font = Font(bold=True, color="FFFFFF")
        for col in range(2, len(months) + 3):
            ws2.cell(row=ri, column=col).fill = PatternFill("solid", fgColor="2d2d44")
        ri += 1
        for sub in all_subs:
            sub_total = sum(monthly_sub.get((m, cat, sub), 0) for m in months)
            if sub_total == 0:
                continue
            fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
            ws2.cell(row=ri, column=1, value="").fill = fill
            ws2.cell(row=ri, column=2, value=sub).fill = fill
            for ci, m in enumerate(months, 3):
                v = monthly_sub.get((m, cat, sub), 0)
                c = ws2.cell(row=ri, column=ci, value=round(v) if v else 0)
                c.fill = fill; c.number_format = INR_NUMBER_FORMAT
            c = ws2.cell(row=ri, column=len(months) + 3, value=round(sub_total))
            c.fill = fill; c.number_format = INR_NUMBER_FORMAT; c.font = Font(bold=True)
            ri += 1
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 38
    for i in range(3, len(months) + 4):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = "C2"

    # ── Sheet 3: All Transactions ─────────────────────────────────────────────
    ws3 = wb.create_sheet("All Transactions")
    txn_hdrs = ["Date", "Month", "Type", "Category", "Sub-category", "Amount (Rs)", "Description"]
    for col, h in enumerate(txn_hdrs, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR
    for ri, (dt, desc, typ, amt, cat, sub) in enumerate(sorted(classified, key=lambda x: x[0]), 2):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        ws3.cell(row=ri, column=1, value=dt.strftime("%d %b %Y")).fill = fill
        ws3.cell(row=ri, column=2, value=dt.strftime("%Y-%m")).fill = fill
        ws3.cell(row=ri, column=3, value=typ).fill = fill
        ws3.cell(row=ri, column=4, value=cat).fill = fill
        ws3.cell(row=ri, column=5, value=sub).fill = fill
        c = ws3.cell(row=ri, column=6, value=amt)
        c.fill = fill; c.number_format = INR_NUMBER_FORMAT
        ws3.cell(row=ri, column=7, value=desc).fill = fill
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 26
    ws3.column_dimensions["E"].width = 38
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 80
    ws3.auto_filter.ref = "A1:G1"
    ws3.freeze_panes = "A2"

    return wb


# ── Deposit Reconciliation ─────────────────────────────────────────────────────

@router.get("/finance/reconcile")
async def get_deposit_reconciliation(
    month: Optional[str] = Query(None),
    user: AppUser = Depends(get_current_user),
):
    _require_admin(user)
    if month:
        _validate_month(month)

    async with get_session() as session:
        q = select(BankTransaction).where(
            BankTransaction.category == "Tenant Deposit Refund"
        ).order_by(BankTransaction.txn_date.desc())

        if month:
            y, mo = int(month[:4]), int(month[5:7])
            start = date(y, mo, 1)
            end   = date(y, mo, _calendar.monthrange(y, mo)[1])
            q = q.where(BankTransaction.txn_date.between(start, end))

        txns = (await session.scalars(q)).all()

        rows = []
        for t in txns:
            tenant_name = None
            if t.reconciled_checkout_id:
                cr = await session.scalar(
                    select(CheckoutRecord).where(CheckoutRecord.id == t.reconciled_checkout_id)
                )
                if cr:
                    tenancy = await session.scalar(
                        select(Tenancy).where(Tenancy.id == cr.tenancy_id)
                    )
                    if tenancy:
                        tenant = await session.scalar(
                            select(Tenant).where(Tenant.id == tenancy.tenant_id)
                        )
                        if tenant:
                            tenant_name = tenant.name

            rows.append({
                "txn_id":      t.id,
                "txn_date":    t.txn_date.isoformat(),
                "amount":      float(t.amount),
                "status":      "matched" if t.reconciled_checkout_id else "unmatched",
                "tenant":      tenant_name,
                "checkout_id": t.reconciled_checkout_id,
            })

    return {"rows": rows}


@router.get("/finance/unit-economics")
async def get_unit_economics_api(
    month: Optional[str] = Query(None, description="YYYY-MM format"),
    user: AppUser = Depends(get_current_user),
):
    """Unit economics KPIs — revenue per bed, cost per bed, avg rent, collection rate."""
    _require_admin(user)
    today = date.today()
    if month:
        _validate_month(month)
        m = date(int(month[:4]), int(month[5:7]), 1)
    else:
        m = date(today.year, today.month, 1)

    from src.services.unit_economics import get_unit_economics
    async with get_session() as session:
        kpis = await get_unit_economics(m, session)
    return {"month": m.strftime("%Y-%m"), **kpis}


# ── UPI Auto-Reconcile ────────────────────────────────────────────────────────

@router.post("/finance/upi-reconcile")
async def upi_reconcile(
    files: list[UploadFile] = File(...),
    account_name: str = Form(...),
    period_month: Optional[str] = Form(None, description="YYYY-MM — defaults to each txn's own month"),
    user: AppUser = Depends(get_current_user),
):
    """
    Upload HULK or THOR Lakshmi UPI export (XLSX or CSV).
    Auto-matches transactions to tenants, creates Payment records.
    Re-uploading the same file is safe — RRN dedup skips existing rows.
    """
    _require_admin(user)
    if account_name not in ("THOR", "HULK"):
        raise HTTPException(status_code=400, detail="account_name must be THOR or HULK")

    mon: Optional[date] = None
    if period_month:
        _validate_month(period_month)
        mon = date(int(period_month[:4]), int(period_month[5:7]), 1)

    from src.services.upi_reconciliation import reconcile_upi_file

    all_matched, all_unmatched, total_skipped = [], [], 0

    async with get_session() as session:
        for upload_file in files:
            content  = await upload_file.read()
            filename = upload_file.filename or "upload.xlsx"
            result   = await reconcile_upi_file(session, content, filename, account_name, mon)
            all_matched.extend([
                {"rrn": e.rrn, "amount": e.amount, "payer": e.payer_name,
                 "tenant": e.tenant_name, "room": e.room, "matched_by": e.matched_by}
                for e in result.matched
            ])
            all_unmatched.extend([
                {"rrn": e.rrn, "amount": e.amount, "payer": e.payer_name, "vpa": e.payer_vpa}
                for e in result.unmatched
            ])
            total_skipped += result.skipped_dup

    return {
        "account_name":     account_name,
        "matched_count":    len(all_matched),
        "matched_amount":   sum(e["amount"] for e in all_matched),
        "unmatched_count":  len(all_unmatched),
        "unmatched_amount": sum(e["amount"] for e in all_unmatched),
        "skipped_duplicate":total_skipped,
        "matched":          all_matched,
        "unmatched":        all_unmatched,
    }


@router.post("/finance/upi-reconcile/assign")
async def assign_upi_entry(
    rrn:          str  = Form(...),
    tenancy_id:   int  = Form(...),
    period_month: str  = Form(..., description="YYYY-MM"),
    user: AppUser = Depends(get_current_user),
):
    """Manually link an unmatched UPI bank entry to a tenant."""
    _require_admin(user)
    _validate_month(period_month)
    mon = date(int(period_month[:4]), int(period_month[5:7]), 1)

    from src.services.upi_reconciliation import assign_upi_entry as _assign
    async with get_session() as session:
        payment_id = await _assign(session, rrn, tenancy_id, mon)
    return {"payment_id": payment_id, "rrn": rrn, "tenancy_id": tenancy_id}


@router.get("/finance/investments")
async def get_investments(user: AppUser = Depends(get_current_user)):
    """Return Whitefield investment tracker — grouped by investor, ordered by date."""
    _require_admin(user)
    from sqlalchemy import select as sa_select

    async with get_session() as session:
        rows = await session.execute(
            sa_select(InvestmentExpense)
            .where(InvestmentExpense.is_void == False)
            .order_by(InvestmentExpense.paid_by, InvestmentExpense.transaction_date)
        )
        txns = rows.scalars().all()

    # Group by investor
    groups: dict[str, dict] = {}
    for t in txns:
        investor = str(t.paid_by) if t.paid_by else "Unknown"
        if investor not in groups:
            groups[investor] = {"investor": investor, "total": 0.0, "rows": []}
        amt = float(str(t.amount))
        groups[investor]["total"] += amt
        groups[investor]["rows"].append({
            "id": t.id,
            "date": str(t.transaction_date) if t.transaction_date else "",
            "purpose": t.purpose,
            "vendor": t.paid_to or "",
            "amount": amt,
            "utr": t.transaction_id or "",
            "property": t.property or "Whitefield",
            "notes": t.notes or "",
        })

    grand_total = sum(g["total"] for g in groups.values())
    return {
        "groups": list(groups.values()),
        "grand_total": round(grand_total, 2),
        "count": len(txns),
    }


@router.get("/finance/upi-reconcile/unmatched")
async def get_unmatched_upi(
    month: Optional[str] = Query(None, description="YYYY-MM"),
    user: AppUser = Depends(get_current_user),
):
    """Return unmatched UPI entries — the daily review queue."""
    _require_admin(user)
    from sqlalchemy import select as sa_select

    async with get_session() as session:
        q = sa_select(UpiCollectionEntry).where(UpiCollectionEntry.tenancy_id.is_(None))
        if month:
            _validate_month(month)
            mon = date(int(month[:4]), int(month[5:7]), 1)
            next_mon = date(mon.year + (mon.month // 12), (mon.month % 12) + 1, 1)
            q = q.where(UpiCollectionEntry.txn_date >= mon)
            q = q.where(UpiCollectionEntry.txn_date < next_mon)
        q = q.order_by(UpiCollectionEntry.txn_date.desc())
        rows = await session.execute(q)
        entries = rows.scalars().all()

    return {"unmatched": [
        {"rrn": e.rrn, "account": e.account_name, "date": str(e.txn_date),
         "amount": round(float(str(e.amount)), 2), "payer": e.payer_name, "vpa": e.payer_vpa}
        for e in entries
    ]}
