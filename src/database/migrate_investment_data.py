"""
Migration + Import: investment_expenses and pg_contacts tables.
================================================================
SAFE TO RE-RUN: every operation is idempotent via unique_hash + ON CONFLICT DO NOTHING.

Source files (must be in project root):
  - Whitefield PG Expense Tracker 14TH NOV-25 (1).xlsx
  - Contacts.xlsx

IMPORTANT - which sheets are imported:
  investment_expenses  <- 'White Field PG Expenses' ONLY (consolidated sheet)
                          ASHOKAN / JITENDRA / NARENDRA / OUR SIDE are subsets of this
                          sheet -- skipping them prevents double-counting.
  pg_contacts          <- 'Sheet1' in Contacts.xlsx (all 62 rows)

Usage:
    python -m src.database.migrate_investment_data          # create tables + import
    python -m src.database.migrate_investment_data --status # show counts only
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from sqlalchemy.ext.asyncio import create_async_engine
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncConnection

load_dotenv()

DB_URL = os.getenv("DATABASE_URL", "").replace("postgresql://", "postgresql+asyncpg://")
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

EXPENSE_FILE = PROJECT_ROOT / "Whitefield PG Expense Tracker 14TH NOV-25 (1).xlsx"
CONTACTS_FILE = PROJECT_ROOT / "Contacts.xlsx"

# ── Table definitions ──────────────────────────────────────────────────────────

CREATE_INVESTMENT_EXPENSES = """
CREATE TABLE IF NOT EXISTS investment_expenses (
    id               SERIAL PRIMARY KEY,
    sno              INTEGER,
    purpose          VARCHAR(500) NOT NULL,
    amount           NUMERIC(15,2) NOT NULL,
    paid_by          VARCHAR(120),
    transaction_date DATE,
    transaction_id   VARCHAR(300),
    paid_to          VARCHAR(300),
    property         VARCHAR(60)  DEFAULT 'Whitefield',
    unique_hash      VARCHAR(64)  UNIQUE NOT NULL,
    is_void          BOOLEAN      DEFAULT FALSE,
    notes            TEXT,
    created_at       TIMESTAMP    DEFAULT NOW()
);
"""

CREATE_PG_CONTACTS = """
CREATE TABLE IF NOT EXISTS pg_contacts (
    id          SERIAL PRIMARY KEY,
    name        VARCHAR(200),
    contact_for TEXT,
    referred_by VARCHAR(120),
    phone       VARCHAR(30),
    comments    TEXT,
    amount_paid NUMERIC(12,2),
    remaining   TEXT,
    category    VARCHAR(80),
    visible_to  VARCHAR(50)  DEFAULT 'owner,staff',
    property    VARCHAR(60)  DEFAULT 'Whitefield',
    unique_hash VARCHAR(64)  UNIQUE NOT NULL,
    created_at  TIMESTAMP    DEFAULT NOW()
);
"""

CREATE_INDEXES = [
    "CREATE INDEX IF NOT EXISTS ix_investment_expenses_paid_by   ON investment_expenses(paid_by)",
    "CREATE INDEX IF NOT EXISTS ix_investment_expenses_date       ON investment_expenses(transaction_date)",
    "CREATE INDEX IF NOT EXISTS ix_pg_contacts_phone             ON pg_contacts(phone)",
    "CREATE INDEX IF NOT EXISTS ix_pg_contacts_category          ON pg_contacts(category)",
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _sha256(*parts) -> str:
    combined = "|".join(str(p or "").strip() for p in parts)
    return hashlib.sha256(combined.encode()).hexdigest()


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _clean_phone(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # Remove unicode direction marks, non-breaking spaces
    s = re.sub(r"[\u202a\u202c\xa0\u200e\u200f]", "", s)
    s = s.replace(" ", "").strip()
    if not s or s.lower() in ("none", "nan", ""):
        return None
    return s


def _derive_category(contact_for: str | None) -> str | None:
    if not contact_for:
        return None
    cf = contact_for.lower()
    if any(w in cf for w in ("plumber", "plumbing", "water tank", "hwt", "hws")):
        return "plumber"
    if any(w in cf for w in ("electrician", "electrical", "lighting", "lights")):
        return "electrician"
    if any(w in cf for w in ("carpenter", "headboard", "cupboard", "ply", "plywood", "laminat")):
        return "carpenter"
    if any(w in cf for w in ("wifi", "internet", "broadband", "connection")):
        return "internet"
    if any(w in cf for w in ("camera", "cc camera", "cctv", "unisol")):
        return "security"
    if any(w in cf for w in ("architect", "design", "logo", "banner", "signage")):
        return "design"
    if any(w in cf for w in ("sofa", "curtain", "cushion", "furniture", "mattress", "bed", "chair", "table", "cot", "stool")):
        return "furniture"
    if any(w in cf for w in ("paint", "painter", "wallpaper", "fall ceil")):
        return "painting"
    if any(w in cf for w in ("vegetable", "chicken", "milk", "curd", "egg", "food", "gas", "cylinder", "grocery")):
        return "food_supply"
    if any(w in cf for w in ("gym", "pool", "tt ", "carrom", "foosball")):
        return "gym_sports"
    if any(w in cf for w in ("police", "bbmp", "emergency")):
        return "government"
    if any(w in cf for w in ("plant", "garden", "flooring", "vinyl", "rubber mat")):
        return "decor"
    if any(w in cf for w in ("garbage", "lift", "water", "diesel", "generator")):
        return "facility"
    if any(w in cf for w in ("marketing", "digital", "t-shirt", "gifting")):
        return "marketing"
    if any(w in cf for w in ("scaffolding", "labour", "unload")):
        return "construction"
    return "vendor"


# ── Load Excel data ────────────────────────────────────────────────────────────

def _load_investment_expenses() -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXPENSE_FILE), read_only=True, data_only=True)
    ws = wb["White Field PG Expenses"]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        sno, purpose, amount, paid_by, txn_date, txn_id, paid_to = (row[j] for j in range(7))

        # Skip blank/formula/total rows
        if not purpose and not amount:
            continue
        if not isinstance(amount, (int, float)):
            continue
        if not purpose:
            continue

        purpose_s   = str(purpose).strip()
        paid_by_s   = str(paid_by).strip() if paid_by else ""
        txn_id_s    = str(txn_id).strip()[:300] if txn_id else None
        paid_to_s   = str(paid_to).strip()[:300] if paid_to else None
        sno_val     = int(sno) if sno and isinstance(sno, (int, float)) else None

        unique_hash = _sha256(str(sno_val), purpose_s, str(amount), paid_by_s)

        rows.append({
            "sno":              sno_val,
            "purpose":          purpose_s[:500],
            "amount":           float(amount),
            "paid_by":          paid_by_s[:120] if paid_by_s else None,
            "transaction_date": _parse_date(txn_date),
            "transaction_id":   txn_id_s,
            "paid_to":          paid_to_s,
            "unique_hash":      unique_hash,
        })

    wb.close()
    return rows


def _load_contacts() -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed.")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(CONTACTS_FILE), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if not any(v is not None for v in row):
            continue

        def _cell(idx):
            return row[idx] if len(row) > idx else None

        name        = str(_cell(0)).strip() if _cell(0) else None
        contact_for = str(_cell(1)).strip() if _cell(1) else None
        referred_by = str(_cell(2)).strip() if _cell(2) else None
        phone_raw   = _cell(3)
        comments    = str(_cell(4)).strip() if _cell(4) else None
        amount_paid_raw = _cell(5)
        remaining   = str(_cell(6)).strip() if _cell(6) else None

        # Skip formula-only rows (amount_paid was a formula string, skip if no usable data)
        if not name and not phone_raw and not contact_for:
            continue

        phone = _clean_phone(phone_raw)
        category = _derive_category(contact_for)

        # amount_paid: may be numeric or a string like "Fully Settled"
        amount_paid = None
        if amount_paid_raw is not None and isinstance(amount_paid_raw, (int, float)):
            amount_paid = float(amount_paid_raw)

        # Unique hash: name + phone + contact_for (handles same person, different services)
        unique_hash = _sha256(
            name or "Unknown",
            phone or "",
            (contact_for or "")[:100],
        )

        rows.append({
            "name":        (name or "")[:200] if name else None,
            "contact_for": contact_for,
            "referred_by": referred_by[:120] if referred_by else None,
            "phone":       phone[:30] if phone else None,
            "comments":    comments,
            "amount_paid": amount_paid,
            "remaining":   remaining[:500] if remaining else None,
            "category":    category,
            "unique_hash": unique_hash,
        })

    wb.close()
    return rows


# ── DB operations ──────────────────────────────────────────────────────────────

async def create_tables(conn: AsyncConnection) -> None:
    print("\n== Creating tables ==")
    await conn.execute(text(CREATE_INVESTMENT_EXPENSES))
    print("  [ok] investment_expenses")
    await conn.execute(text(CREATE_PG_CONTACTS))
    print("  [ok] pg_contacts")
    for idx_sql in CREATE_INDEXES:
        await conn.execute(text(idx_sql))
    print("  [ok] indexes")


async def import_expenses(conn: AsyncConnection) -> None:
    print("\n== Importing investment_expenses ==")
    print(f"  Source: White Field PG Expenses sheet only (consolidated)")

    rows = _load_investment_expenses()
    print(f"  Rows to import: {len(rows)}")

    inserted = 0
    skipped  = 0
    for r in rows:
        result = await conn.execute(text("""
            INSERT INTO investment_expenses
                (sno, purpose, amount, paid_by, transaction_date, transaction_id, paid_to, property, unique_hash)
            VALUES
                (:sno, :purpose, :amount, :paid_by, :txn_date, :txn_id, :paid_to, 'Whitefield', :unique_hash)
            ON CONFLICT (unique_hash) DO NOTHING
        """), {
            "sno":         r["sno"],
            "purpose":     r["purpose"],
            "amount":      r["amount"],
            "paid_by":     r["paid_by"],
            "txn_date":    r["transaction_date"],
            "txn_id":      r["transaction_id"],
            "paid_to":     r["paid_to"],
            "unique_hash": r["unique_hash"],
        })
        if result.rowcount == 1:
            inserted += 1
        else:
            skipped += 1

    print(f"  Inserted: {inserted} | Skipped (already exist): {skipped}")

    # Quick total check
    total_result = await conn.execute(text(
        "SELECT SUM(amount) FROM investment_expenses WHERE property='Whitefield'"
    ))
    total = total_result.scalar() or 0
    print(f"  Total investment amount in DB: Rs {total:,.0f}")


async def import_contacts(conn: AsyncConnection) -> None:
    print("\n== Importing pg_contacts ==")

    rows = _load_contacts()
    print(f"  Rows to import: {len(rows)}")

    inserted = 0
    skipped  = 0
    for r in rows:
        result = await conn.execute(text("""
            INSERT INTO pg_contacts
                (name, contact_for, referred_by, phone, comments, amount_paid, remaining,
                 category, visible_to, property, unique_hash)
            VALUES
                (:name, :contact_for, :referred_by, :phone, :comments, :amount_paid, :remaining,
                 :category, 'owner,staff', 'Whitefield', :unique_hash)
            ON CONFLICT (unique_hash) DO NOTHING
        """), {
            "name":        r["name"],
            "contact_for": r["contact_for"],
            "referred_by": r["referred_by"],
            "phone":       r["phone"],
            "comments":    r["comments"],
            "amount_paid": r["amount_paid"],
            "remaining":   r["remaining"],
            "category":    r["category"],
            "unique_hash": r["unique_hash"],
        })
        if result.rowcount == 1:
            inserted += 1
        else:
            skipped += 1

    print(f"  Inserted: {inserted} | Skipped (already exist): {skipped}")


async def show_status(conn: AsyncConnection) -> None:
    print("\n== Status ==")
    for table in ("investment_expenses", "pg_contacts"):
        r = await conn.execute(text(
            f"SELECT COUNT(*) FROM {table}"
        ))
        count = r.scalar()
        print(f"  {table}: {count} rows")

    # Per-investor totals
    r2 = await conn.execute(text("""
        SELECT paid_by, COUNT(*) as txns, SUM(amount) as total
        FROM investment_expenses
        WHERE is_void = FALSE
        GROUP BY paid_by
        ORDER BY total DESC NULLS LAST
    """))
    rows = r2.fetchall()
    if rows:
        print("\n  Per-investor breakdown:")
        grand = 0
        for row in rows:
            paid_by, txns, total = row
            total = total or 0
            grand += total
            print(f"    {paid_by or 'Unknown':20s}  {txns:3d} txns  Rs {total:>12,.0f}")
        print(f"    {'TOTAL':20s}       Rs {grand:>12,.0f}")

    # Contact categories
    r3 = await conn.execute(text("""
        SELECT category, COUNT(*) as cnt
        FROM pg_contacts
        GROUP BY category
        ORDER BY cnt DESC
    """))
    cat_rows = r3.fetchall()
    if cat_rows:
        print("\n  Contact categories:")
        for cat, cnt in cat_rows:
            print(f"    {cat or 'uncategorized':20s} {cnt:3d}")


async def main(args: argparse.Namespace) -> None:
    if not DB_URL or DB_URL == "+asyncpg://":
        print("ERROR: DATABASE_URL not set in .env")
        sys.exit(1)

    if not EXPENSE_FILE.exists():
        print(f"ERROR: Expense file not found: {EXPENSE_FILE}")
        sys.exit(1)
    if not CONTACTS_FILE.exists():
        print(f"ERROR: Contacts file not found: {CONTACTS_FILE}")
        sys.exit(1)

    engine = create_async_engine(DB_URL, echo=False)
    async with engine.begin() as conn:
        await show_status(conn) if args.status else None
        if not args.status:
            await create_tables(conn)
            await import_expenses(conn)
            await import_contacts(conn)
            await show_status(conn)
    await engine.dispose()
    print("\nDone.\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Create + import investment_expenses and pg_contacts")
    parser.add_argument("--status", action="store_true", help="Show counts only, no changes")
    asyncio.run(main(parser.parse_args()))
