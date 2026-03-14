"""
Excel → Supabase import script.
Reads 'Cozeevo Monthly stay (2).xlsx' and populates all tables.

Run: python -m src.database.excel_import

What it imports:
  - Sheet 'History'    → monthly stay tenants, tenancies, rent_schedule, payments
  - Sheet 'Daily Basis'→ short-stay tenancies

Order:
  1. staff           (from Assigned Staff column)
  2. rooms           (from Room No + BLOCK columns)
  3. tenants         (deduplicated by phone)
  4. tenancies       (one per tenant-room-checkin row)
  5. rent_schedule   (one per tenancy per month tracked)
  6. payments        (from PAID month columns with Cash/UPI amounts)
"""
import asyncio
import os
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

load_dotenv()

from src.database.models import (
    AuthorizedUser, Expense, ExpenseCategory, FoodPlan, Payment,
    PaymentFor, PaymentMode, Property, RateCard, RefundStatus,
    RentSchedule, RentStatus, Room, RoomType, Staff, StayType,
    Tenancy, TenancyStatus, Tenant, UserRole,
)

DATABASE_URL = os.environ["DATABASE_URL"]
EXCEL_PATH   = "Cozeevo Monthly stay (3).xlsx"

# ── Month columns in the History sheet ──────────────────────────────────────
# (col_index, period_date, cash_col_index, upi_col_index)
# Dec has no separate cash/UPI columns — the "until jan Cash/UPI" cols (23,24)
# are cumulative for Dec+Jan combined, so we attribute them to January only.
MONTH_COLS = [
    (20, date(2025, 12, 1), None, None),  # DEC RENT — status only, no payment amounts
    (21, date(2026,  1, 1), 23,   24),    # JAN RENT — "until jan Cash/UPI" (cols 23,24)
    (25, date(2026,  2, 1), 28,   29),    # FEB RENT, FEB Cash, FEB UPI
    (26, date(2026,  3, 1), 31,   32),    # MARCH RENT, March Cash, March UPI
]

# ── Normalization helpers ────────────────────────────────────────────────────

def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def norm_name(v) -> str:
    return clean_str(v).title()

def clean_phone(v) -> str:
    """Normalize phone to +91XXXXXXXXXX format."""
    if not v:
        return ""
    s = re.sub(r"[^0-9]", "", str(v))
    if len(s) == 10:
        return f"+91{s}"
    if len(s) == 12 and s.startswith("91"):
        return f"+{s}"
    if len(s) > 10:
        return f"+{s}"
    return s

def to_decimal(v, default=Decimal("0")) -> Decimal:
    """Extract first number from messy cell values like '3102/2400' or '903*10=...'"""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    s = str(v).strip()
    if not s or s in ("-", "–", "NA", "na"):
        return default
    # Extract first integer/decimal sequence
    m = re.search(r"[\d]+(?:\.\d+)?", s)
    if m:
        try:
            return Decimal(m.group()).quantize(Decimal("0.01"))
        except InvalidOperation:
            pass
    return default

def norm_room_type(v) -> RoomType:
    s = clean_str(v).lower()
    if "premium" in s:
        return RoomType.premium
    if "triple" in s:
        return RoomType.triple
    if "double" in s:
        return RoomType.double
    return RoomType.single

def norm_status(v) -> TenancyStatus:
    s = clean_str(v).upper()
    if s == "CHECKIN":
        return TenancyStatus.active
    if s == "EXIT":
        return TenancyStatus.exited
    if s == "CANCELLED":
        return TenancyStatus.cancelled
    return TenancyStatus.no_show

def norm_rent_status(v) -> RentStatus:
    s = clean_str(v).upper()
    if s == "PAID":
        return RentStatus.paid
    if "PARTIAL" in s:
        return RentStatus.partial
    if s in ("EXIT", "VACATE"):
        return RentStatus.exit
    if s in ("NO SHOW", "NOSHOW"):
        return RentStatus.na
    if s in ("CANCELLED", "CANCEL"):
        return RentStatus.na
    return RentStatus.na

def norm_staff_name(v) -> str:
    """Deduplicate staff names: lokesh lk → Lokesh, kiran → Kiran."""
    if not v:
        return ""
    s = clean_str(v).strip().title()
    # Merge known variants
    mappings = {
        "Lokesh Lk": "Lokesh",
        "Imrana": "Imrana",
    }
    for k, val in mappings.items():
        if s.lower() == k.lower():
            return val
    return s.split()[0]  # take first word only for short names

def placeholder_phone(room_no, name) -> str:
    safe = re.sub(r"[^a-zA-Z0-9]", "", str(name))[:12]
    return f"NOPHONE_{room_no}_{safe}"


# ── Main import ──────────────────────────────────────────────────────────────

async def run_import():
    engine = create_async_engine(DATABASE_URL, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    async with Session() as session:
        # ── Load existing properties ─────────────────────────────────────
        result = await session.execute(select(Property))
        props = {p.name: p for p in result.scalars().all()}
        thor = props.get("Cozeevo THOR")
        hulk = props.get("Cozeevo HULK")
        if not thor or not hulk:
            print("ERROR: Run seed.py first — properties not found.")
            return

        # ── Load existing food plans ─────────────────────────────────────
        result = await session.execute(select(FoodPlan))
        food_map = {fp.name: fp for fp in result.scalars().all()}

        print("\n=== IMPORTING HISTORY SHEET ===")
        ws = wb["History"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r[1]]  # skip empty rows

        # ── PASS 1: Collect unique staff names → create Staff records ────
        print("\n[1/6] Importing staff...")
        staff_names_seen = set()
        for row in rows:
            sn = norm_staff_name(row[15])
            if sn:
                staff_names_seen.add(sn)

        staff_map = {}  # name → Staff object
        for sname in sorted(staff_names_seen):
            result = await session.execute(select(Staff).where(Staff.name == sname))
            existing = result.scalars().first()
            if existing:
                staff_map[sname] = existing
            else:
                s = Staff(name=sname, role="Staff", active=True)
                session.add(s)
                await session.flush()
                staff_map[sname] = s
                print(f"  + Staff: {sname}")

        # ── PASS 2: Collect unique rooms → create Room records ────────────
        print("\n[2/6] Importing rooms...")
        room_map = {}  # (property_id, room_number) → Room

        for row in rows:
            room_raw = row[0]
            block    = clean_str(row[17]).upper()
            floor    = row[18]
            rtype    = norm_room_type(row[12])

            if not room_raw:
                continue

            room_num = str(room_raw).rstrip(".0") if isinstance(room_raw, float) else str(room_raw)
            prop = hulk if block == "HULK" else thor
            key  = (prop.id, room_num)

            if key not in room_map:
                result = await session.execute(
                    select(Room).where(Room.property_id == prop.id, Room.room_number == room_num)
                )
                existing = result.scalars().first()
                if existing:
                    room_map[key] = existing
                else:
                    try:
                        floor_int = int(float(floor)) if floor else 1
                    except (ValueError, TypeError):
                        floor_int = 1

                    r = Room(
                        property_id=prop.id,
                        room_number=room_num,
                        floor=floor_int,
                        room_type=rtype,
                        max_occupancy={"single": 1, "double": 2, "triple": 3, "premium": 1}[rtype.value],
                        active=True,
                    )
                    session.add(r)
                    await session.flush()
                    room_map[key] = r
                    print(f"  + Room {room_num} ({block}, {rtype.value})")

        await session.commit()

        # ── PASS 3: Tenants ───────────────────────────────────────────────
        print("\n[3/6] Importing tenants...")
        tenant_map = {}  # phone → Tenant

        for row in rows:
            name  = norm_name(row[1])
            phone = clean_phone(row[3])
            room_raw = row[0]

            if not name:
                continue

            if not phone:
                room_num = str(room_raw).rstrip(".0") if isinstance(room_raw, float) else str(room_raw or "X")
                phone = placeholder_phone(room_num, name)

            if phone in tenant_map:
                continue

            result = await session.execute(select(Tenant).where(Tenant.phone == phone))
            existing = result.scalars().first()
            if existing:
                tenant_map[phone] = existing
                continue

            gender_raw = clean_str(row[2]).lower()
            gender = "female" if gender_raw == "female" else "male"

            t = Tenant(name=name, phone=phone, gender=gender)
            session.add(t)
            await session.flush()
            tenant_map[phone] = t

        await session.commit()
        print(f"  Total tenants: {len(tenant_map)}")

        # ── PASS 4: Tenancies ─────────────────────────────────────────────
        print("\n[4/6] Importing tenancies...")
        tenancy_list = []

        for row in rows:
            name     = norm_name(row[1])
            phone    = clean_phone(row[3])
            room_raw = row[0]
            block    = clean_str(row[17]).upper()
            checkin  = row[4]
            booking  = to_decimal(row[5])
            deposit  = to_decimal(row[6])
            maint    = to_decimal(row[7])
            monthly  = to_decimal(row[9])
            status   = norm_status(row[16])
            staff_nm = norm_staff_name(row[15])
            food_raw = clean_str(row[34]).lower() if len(row) > 34 else ""

            if not name or not room_raw or not checkin:
                continue

            room_num = str(room_raw).rstrip(".0") if isinstance(room_raw, float) else str(room_raw)
            if not phone:
                phone = placeholder_phone(room_num, name)

            prop = hulk if block == "HULK" else thor
            room = room_map.get((prop.id, room_num))
            tenant = tenant_map.get(phone)

            if not room or not tenant:
                print(f"  SKIP {name}: room or tenant not found")
                continue

            if isinstance(checkin, datetime):
                checkin_date = checkin.date()
            elif isinstance(checkin, date):
                checkin_date = checkin
            else:
                continue

            # Map food plan
            fp = None
            for fp_key in food_map:
                if any(w in food_raw for w in fp_key.split()):
                    fp = food_map[fp_key]
                    break

            staff_obj = staff_map.get(staff_nm)

            t = Tenancy(
                tenant_id=tenant.id,
                room_id=room.id,
                stay_type=StayType.monthly,
                status=status,
                checkin_date=checkin_date,
                booking_amount=booking,
                security_deposit=deposit,
                maintenance_fee=maint,
                agreed_rent=monthly,
                food_plan_id=fp.id if fp else None,
                assigned_staff_id=staff_obj.id if staff_obj else None,
                notes=clean_str(row[14]) if len(row) > 14 else None,
            )
            session.add(t)
            await session.flush()
            tenancy_list.append((t, row))

        await session.commit()
        print(f"  Total tenancies: {len(tenancy_list)}")

        # ── PASS 5: Rent Schedule ─────────────────────────────────────────
        print("\n[5/6] Importing rent schedule...")
        rs_count = 0

        for tenancy, row in tenancy_list:
            for col_idx, period, cash_col, upi_col in MONTH_COLS:
                status_val = row[col_idx] if len(row) > col_idx else None
                if status_val is None:
                    continue

                rs = norm_rent_status(status_val)
                maint = tenancy.maintenance_fee or Decimal("0")

                result = await session.execute(
                    select(RentSchedule).where(
                        RentSchedule.tenancy_id == tenancy.id,
                        RentSchedule.period_month == period,
                    )
                )
                if result.scalars().first():
                    continue

                rs_row = RentSchedule(
                    tenancy_id=tenancy.id,
                    period_month=period,
                    rent_due=tenancy.agreed_rent,
                    maintenance_due=maint,
                    status=rs,
                    due_date=period,
                    notes=clean_str(status_val) if rs != RentStatus.paid else None,
                )
                session.add(rs_row)
                rs_count += 1

        await session.commit()
        print(f"  Total rent_schedule rows: {rs_count}")

        # ── PASS 6: Payments (Cash + UPI from month columns) ──────────────
        print("\n[6/6] Importing payments...")
        pay_count = 0

        for tenancy, row in tenancy_list:
            for col_idx, period, cash_col, upi_col in MONTH_COLS:
                status_val = row[col_idx] if len(row) > col_idx else None
                if status_val is None:
                    continue

                cash_val = to_decimal(row[cash_col]) if (cash_col is not None and len(row) > cash_col) else Decimal("0")
                upi_val  = to_decimal(row[upi_col])  if (upi_col  is not None and len(row) > upi_col)  else Decimal("0")

                if cash_val > 0:
                    session.add(Payment(
                        tenancy_id=tenancy.id,
                        amount=cash_val,
                        payment_date=period,
                        payment_mode=PaymentMode.cash,
                        for_type=PaymentFor.rent,
                        period_month=period,
                        notes="Imported from Excel",
                    ))
                    pay_count += 1

                if upi_val > 0:
                    session.add(Payment(
                        tenancy_id=tenancy.id,
                        amount=upi_val,
                        payment_date=period,
                        payment_mode=PaymentMode.upi,
                        for_type=PaymentFor.rent,
                        period_month=period,
                        notes="Imported from Excel",
                    ))
                    pay_count += 1

                # Deposit as payment (for active tenants)
                if col_idx == 20 and tenancy.security_deposit > 0:
                    session.add(Payment(
                        tenancy_id=tenancy.id,
                        amount=tenancy.security_deposit,
                        payment_date=tenancy.checkin_date,
                        payment_mode=PaymentMode.cash,
                        for_type=PaymentFor.deposit,
                        period_month=None,
                        notes="Security deposit — imported from Excel",
                    ))
                    pay_count += 1

        await session.commit()
        print(f"  Total payment rows: {pay_count}")

        # ── Daily Basis sheet ─────────────────────────────────────────────
        print("\n=== IMPORTING DAILY BASIS SHEET ===")
        ws2 = wb["Daily Basis"]
        daily_count = 0

        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue

            name     = norm_name(row[1])
            phone    = clean_phone(row[2]) if len(row) > 2 else ""
            checkin  = row[3]
            booking  = to_decimal(row[4])
            maint    = to_decimal(row[6])
            day_rate = to_decimal(row[7])
            n_days   = to_decimal(row[8], default=Decimal("1"))
            room_raw = row[0]

            if not name or not checkin:
                continue

            room_num = str(room_raw).rstrip(".0") if isinstance(room_raw, float) else str(room_raw or "DAILY")

            # Try to find room in thor first, then hulk
            room = room_map.get((thor.id, room_num)) or room_map.get((hulk.id, room_num))

            if not room:
                # Create a placeholder room for unknown short-stay rooms
                r = Room(
                    property_id=thor.id,
                    room_number=room_num,
                    floor=0,
                    room_type=RoomType.double,
                    max_occupancy=2,
                    active=True,
                    notes="Daily stay room — imported",
                )
                session.add(r)
                await session.flush()
                room = r
                room_map[(thor.id, room_num)] = room

            # Tenant
            if not phone:
                phone = placeholder_phone(room_num, name)

            tenant = tenant_map.get(phone)
            if not tenant:
                result = await session.execute(select(Tenant).where(Tenant.phone == phone))
                tenant = result.scalars().first()
                if not tenant:
                    tenant = Tenant(name=name, phone=phone, gender="male")
                    session.add(tenant)
                    await session.flush()
                    tenant_map[phone] = tenant

            if isinstance(checkin, datetime):
                checkin_date = checkin.date()
            elif isinstance(checkin, date):
                checkin_date = checkin
            else:
                continue

            total_days = int(n_days) if n_days > 0 else 1
            checkout_date = date(checkin_date.year, checkin_date.month,
                                 min(checkin_date.day + total_days, 28))

            ten = Tenancy(
                tenant_id=tenant.id,
                room_id=room.id,
                stay_type=StayType.daily,
                status=TenancyStatus.exited,
                checkin_date=checkin_date,
                checkout_date=checkout_date,
                booking_amount=booking,
                maintenance_fee=maint,
                agreed_rent=day_rate * Decimal(str(total_days)),
                notes=clean_str(row[12]) if len(row) > 12 else None,
            )
            session.add(ten)
            await session.flush()
            daily_count += 1

            # Payment for daily stay
            if booking > 0:
                session.add(Payment(
                    tenancy_id=ten.id,
                    amount=booking,
                    payment_date=checkin_date,
                    payment_mode=PaymentMode.cash,
                    for_type=PaymentFor.booking,
                    notes="Daily stay — imported from Excel",
                ))

        await session.commit()
        print(f"  Daily stay tenancies: {daily_count}")

    await engine.dispose()

    print("\n" + "="*50)
    print("IMPORT COMPLETE")
    print("="*50)


if __name__ == "__main__":
    asyncio.run(run_import())
